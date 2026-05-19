"""
가나초콜릿하우스 부산 시즌2 — CatchTable 매장 후기 ingest + 브랜드/서비스 필터
입력: Dropbox의 '캐치테이블 리뷰_ 가나초콜릿하우스.xlsx'
출력: 80-r-tech/85-analysis-results/gana-chocolate-house/gana-chocolate-house-catchtable-ingested.json
"""
import sys, io, os, json, re, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
SRC_XLSX = r'C:\Users\insig\Dropbox\02-프로젝트렌트RENT\03_R-lab\02.RXR-store data\test data- 가나2\캐치테이블 리뷰_ 가나초콜릿하우스.xlsx'
OUT_DIR = os.path.join(ROOT, '80-r-tech', '85-analysis-results', 'gana-chocolate-house')
OUT_JSON = os.path.join(OUT_DIR, 'gana-chocolate-house-catchtable-ingested.json')
os.makedirs(OUT_DIR, exist_ok=True)

EVENT_START = '20230212'
EVENT_END = '20230314'

# ============================================================
# 브랜드 / 서비스 키워드 사전
# ============================================================
BRAND_KEYWORDS = {
    # 초콜릿/맛
    '초콜릿','초콜렛','카카오','가나','롯데','달콤','달달','녹아','녹았','사르르','진한','부드러','풍미','향이','향기','맛있','단맛','쌉싸','고소',
    # 위스키/페어링
    '위스키','페어링','바','칵테일','애프터눈','애프터눈티','티세트','스콘','디저트','케이크','에클레어','마카롱',
    # 공간/분위기
    '공간','분위기','인테리어','디자인','포토존','감성','무드','조명','룸','인더스트리','비주얼','꾸며','꾸밈','플레이팅',
    # 고급/브랜드
    '고급','럭셔리','프리미엄','감동','특별','인생','브랜드','경험','기념','추억','퀄리티','완성도','가치','정성',
    # 굿즈/패키지
    '굿즈','선물','패키지','박스','에코백','머그',
    # 클래스/프로그램
    '클래스','원데이','만들기','체험','시음','테이스팅','코스',
}

SERVICE_KEYWORDS = {
    # 직원/응대
    '직원','응대','친절','불친절','서비스','매너','설명해','설명을','안내','안내해','무뚝뚝','퉁명',
    # 대기/예약
    '웨이팅','대기','예약','캐치테이블','만석','줄','기다렸','기다리','기다림','오래 걸','오래걸',
    # 운영
    '시간 제한','시간제한','입장','차단','규칙','제한',
}


def parse_star(star_str):
    """'3.0/4.0/2.0' → (3.0, 4.0, 2.0)"""
    try:
        parts = str(star_str).split('/')
        return float(parts[0]), float(parts[1]), float(parts[2])
    except Exception:
        return 0.0, 0.0, 0.0


def parse_date(date_str):
    """'2023.02.12' → '20230212' (yyyymmdd)"""
    s = str(date_str).strip()
    s = s.replace('.', '').replace('-', '').replace('/', '')
    return s[:8] if len(s) >= 8 else ''


def classify_period(yyyymmdd):
    if not yyyymmdd:
        return '기타'
    if yyyymmdd < EVENT_START:
        return '사전'
    if yyyymmdd <= EVENT_END:
        return '팝업기간'
    return '팝업후'


def count_keywords(text, keywords):
    if not text:
        return 0
    return sum(1 for kw in keywords if kw in text)


def brand_service_filter(full_content, star_taste, star_vibe, star_service):
    brand_hits = count_keywords(full_content, BRAND_KEYWORDS)
    service_hits = count_keywords(full_content, SERVICE_KEYWORDS)
    taste_vibe_avg = (star_taste + star_vibe) / 2.0
    star_gap = round(taste_vibe_avg - star_service, 2)

    if brand_hits >= service_hits + 1 or star_gap >= 1.0:
        return True, 'brand_signal', brand_hits, service_hits, star_gap
    if service_hits >= 2 and brand_hits == 0:
        return False, 'service_only', brand_hits, service_hits, star_gap
    return True, 'ambiguous_kept', brand_hits, service_hits, star_gap


# ============================================================
# Load xlsx
# ============================================================
print("=" * 60)
print("STEP 1: CatchTable xlsx 로드")
print("=" * 60)
print(f"  소스: {SRC_XLSX}")

wb = load_workbook(SRC_XLSX, read_only=True, data_only=True)
ws = wb['Result 1']
print(f"  시트: Result 1 / {ws.max_row}행 x {ws.max_column}열")

rows = list(ws.iter_rows(min_row=2, values_only=True))  # R1은 헤더
raw_records = []
for idx, row in enumerate(rows):
    if not row or not row[0]:
        continue
    date_str, nickname, star_str, star_avg, review, photo_url = (row + (None,)*6)[:6]
    if not review:
        continue
    raw_records.append({
        'row_idx': idx + 2,
        'date': parse_date(date_str),
        'nickname': str(nickname) if nickname else '',
        'star_str': str(star_str) if star_str else '',
        'star_avg': float(star_avg) if star_avg else 0.0,
        'review': str(review).strip(),
        'photo_url': str(photo_url) if photo_url else '',
    })

print(f"  로드: {len(raw_records)}건")


# ============================================================
# Flat record 변환 + 브랜드/서비스 필터
# ============================================================
print()
print("=" * 60)
print("STEP 2: flat record 변환 + 브랜드/서비스 필터")
print("=" * 60)

flat_records = []
for r in raw_records:
    st, sv, ss = parse_star(r['star_str'])
    star_avg = r['star_avg'] or round((st + sv + ss) / 3.0, 1)

    # full_content 구성
    # - 별점 prefix로 detail 토큰 보강
    # - 사진 있으면 토큰 추가
    prefix = f'[맛 {st} 분위기 {sv} 서비스 {ss}] '
    content_body = r['review']
    if r['photo_url']:
        content_body += ' 사진 첨부'
    full_content = prefix + content_body

    # 브랜드/서비스 필터 (원본 리뷰 기준 — prefix는 제외해 키워드 오검출 방지)
    brand_focus, filter_reason, brand_hits, service_hits, star_gap = brand_service_filter(
        r['review'], st, sv, ss
    )

    rec = {
        # --- 파이프라인 표준 필드 ---
        'postdate': r['date'],                       # two_layer가 쓰는 필드
        'date': r['date'],                           # 편의용
        'period': classify_period(r['date']),
        'title': '',
        'description': '',                           # 빈값 — full_content가 충분히 길어지도록
        'full_content': full_content,
        'bloggername': r['nickname'],
        'blogger': r['nickname'],
        'link': f'catchtable://row{r["row_idx"]}',
        # --- CatchTable 특수 필드 ---
        'source': 'catchtable',
        'platform': 'catchtable',
        'star_taste': st,
        'star_vibe': sv,
        'star_service': ss,
        'star_avg': star_avg,
        'image_urls': [r['photo_url']] if r['photo_url'] else [],
        # --- 필터 결과 ---
        'brand_focus': brand_focus,
        'filter_reason': filter_reason,
        'brand_hits': brand_hits,
        'service_hits': service_hits,
        'star_gap': star_gap,
        # --- 원본 리뷰 (리포트용) ---
        'raw_review': r['review'],
    }
    flat_records.append(rec)

# 필터 통계
from collections import Counter
fr = Counter(r['filter_reason'] for r in flat_records)
bf = Counter(r['brand_focus'] for r in flat_records)
print(f"  필터 분포:")
for reason in ['brand_signal', 'ambiguous_kept', 'service_only']:
    print(f"    {reason}: {fr.get(reason, 0)}건")
print(f"  brand_focus True: {bf.get(True, 0)}건 / False: {bf.get(False, 0)}건")

# 제외된 service_only 샘플 3건
excluded = [r for r in flat_records if r['filter_reason'] == 'service_only']
if excluded:
    print(f"\n  [제외 샘플] service_only 리뷰 예시 (최대 3건):")
    for r in excluded[:3]:
        print(f"    - {r['date']} · {r['blogger']} · 별점({r['star_taste']}/{r['star_vibe']}/{r['star_service']})")
        print(f"      B{r['brand_hits']} S{r['service_hits']} gap{r['star_gap']}")
        print(f"      '{r['raw_review'][:80]}'")

# 기간별
pb = Counter(r['period'] for r in flat_records)
print(f"\n  기간별: 사전 {pb.get('사전',0)} / 팝업기간 {pb.get('팝업기간',0)} / 팝업후 {pb.get('팝업후',0)} / 기타 {pb.get('기타',0)}")

# 저장
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(flat_records, f, ensure_ascii=False, indent=2)
print(f"\n저장: {OUT_JSON}")
print(f"✓ ingest 완료: 총 {len(flat_records)}건, brand_focus {bf.get(True,0)}건")
