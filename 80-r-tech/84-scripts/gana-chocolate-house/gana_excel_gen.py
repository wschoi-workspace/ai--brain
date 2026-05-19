"""가나초콜릿하우스 부산 시즌2 — Excel 데이터 파일 생성"""
import json, os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

BASE = os.path.dirname(__file__)
with open(os.path.join(BASE, "../../85-analysis-results/gana-chocolate-house/gana-chocolate-house-2layer-results.json"), "r", encoding="utf-8") as f:
    data = json.load(f)

OUT = os.path.join(BASE, "../../82-case-reports/gana-chocolate-house/gana-chocolate-house-rxr-sns-data.xlsx")

# Colors
PURPLE = "6666FF"
GREEN = "10B981"
RED = "FF5050"
ORANGE = "F59E0B"
BLUE = "4D93F7"
GRAY = "94A3B8"
LIGHT = "F6F6F6"

CLS_COLORS = {"A": "DCFCE7", "B": "F6F6F6", "C": "FEF9C3", "D": "E2E8F0", "E": "FEE2E2", "F": "FEE2E2", "G": "FEE2E2"}
SENT_COLORS = {"긍정": "DCFCE7", "부정": "FEE2E2", "혼합": "DBEAFE", "중립": "F1F5F9"}

HEADERS = ["#", "날짜", "기간", "감성", "Auth", "Clout", "Freshness", "RQL", "토픽", "Content Class", "Trust Score", "협찬", "제목", "블로거", "URL"]

def style_header(ws, row=1):
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def write_data_sheet(ws, items, name=""):
    ws.title = name
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws)

    for idx, r in enumerate(items, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=f"{r['date'][:4]}-{r['date'][4:6]}-{r['date'][6:8]}" if len(r.get('date',''))>=8 else r.get('date',''))
        ws.cell(row=row, column=3, value=r.get("period", ""))
        ws.cell(row=row, column=4, value=r.get("sentiment", ""))
        ws.cell(row=row, column=5, value=r.get("authenticity", 0))
        ws.cell(row=row, column=6, value=r.get("clout", 0))
        ws.cell(row=row, column=7, value=r.get("freshness_index", 0))
        ws.cell(row=row, column=8, value=r.get("rql", ""))
        ws.cell(row=row, column=9, value=r.get("primary_topic", ""))
        ws.cell(row=row, column=10, value=r.get("sincerity_class", "B"))
        ws.cell(row=row, column=11, value=r.get("trust_score", 0))
        ws.cell(row=row, column=12, value="O" if r.get("is_sponsored") else "")
        ws.cell(row=row, column=13, value=r.get("title", ""))
        ws.cell(row=row, column=14, value=r.get("blogger", ""))
        ws.cell(row=row, column=15, value=r.get("link", ""))

        # Color coding
        cls = r.get("sincerity_class", "B")
        sent = r.get("sentiment", "중립")
        ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=CLS_COLORS.get(cls, LIGHT))
        ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=SENT_COLORS.get(sent, "F1F5F9"))
        if r.get("is_sponsored"):
            ws.cell(row=row, column=12).fill = PatternFill("solid", fgColor="FEF9C3")

        # Auth color
        auth = r.get("authenticity", 0)
        if auth >= 75:
            ws.cell(row=row, column=5).font = Font(bold=True, color=GREEN)
        elif auth <= 30:
            ws.cell(row=row, column=5).font = Font(bold=True, color=RED)

    # Column widths
    widths = [5, 12, 10, 8, 8, 8, 10, 12, 14, 14, 12, 8, 50, 15, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(items)+1}"

def write_summary(ws):
    ws.title = "요약 대시보드"
    N = len(data)

    # Title
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="가나초콜릿하우스 부산 시즌2 — RXR SNS 분석 요약")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=PURPLE)

    # Key metrics
    labels = ["총 수집", "Gate 3 유효", "긍정률(전체)", "긍정률(Gate3)", "Avg Auth", "Avg Clout"]
    gate3 = [r for r in data if r.get("sincerity_class") not in ("C","E","F","G") and r.get("authenticity",0)>=60]
    g3_pos = round(sum(1 for r in gate3 if r["sentiment"]=="긍정")/max(len(gate3),1)*100,1)
    values = [N, len(gate3), f"{round(sum(1 for r in data if r['sentiment']=='긍정')/N*100,1)}%", f"{g3_pos}%",
              round(sum(r["authenticity"] for r in data)/N,1), round(sum(r["clout"] for r in data)/N,1)]
    for i, (l, v) in enumerate(zip(labels, values)):
        ws.cell(row=3, column=i+1, value=l)
        ws.cell(row=3, column=i+1).font = Font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=3, column=i+1).fill = PatternFill("solid", fgColor=PURPLE)
        ws.cell(row=4, column=i+1, value=v)
        ws.cell(row=4, column=i+1).font = Font(bold=True, size=12)
        ws.cell(row=4, column=i+1).alignment = Alignment(horizontal="center")

    # 기간별
    ws.cell(row=6, column=1, value="기간별 분석").font = Font(bold=True, size=11, color=PURPLE)
    period_headers = ["기간", "건수", "긍정률", "Auth", "Clout", "Freshness"]
    for i, h in enumerate(period_headers):
        ws.cell(row=7, column=i+1, value=h)
        ws.cell(row=7, column=i+1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=7, column=i+1).fill = PatternFill("solid", fgColor=PURPLE)
    row = 8
    for p in ["사전", "팝업기간", "팝업후"]:
        sub = [r for r in data if r["period"] == p]
        if sub:
            ws.cell(row=row, column=1, value=p)
            ws.cell(row=row, column=2, value=len(sub))
            ws.cell(row=row, column=3, value=f"{round(sum(1 for r in sub if r['sentiment']=='긍정')/len(sub)*100,1)}%")
            ws.cell(row=row, column=4, value=round(sum(r["authenticity"] for r in sub)/len(sub),1))
            ws.cell(row=row, column=5, value=round(sum(r["clout"] for r in sub)/len(sub),1))
            ws.cell(row=row, column=6, value=round(sum(r["freshness_index"] for r in sub)/len(sub),1))
            row += 1

    # 토픽 분포
    ws.cell(row=12, column=1, value="토픽 분포").font = Font(bold=True, size=11, color=PURPLE)
    topics = Counter(r["primary_topic"] for r in data).most_common()
    ws.cell(row=13, column=1, value="토픽").font = Font(bold=True)
    ws.cell(row=13, column=2, value="건수").font = Font(bold=True)
    ws.cell(row=13, column=3, value="비율").font = Font(bold=True)
    for i, (t, c) in enumerate(topics):
        ws.cell(row=14+i, column=1, value=t)
        ws.cell(row=14+i, column=2, value=c)
        ws.cell(row=14+i, column=3, value=f"{round(c/N*100,1)}%")

    # Sincerity Gate
    ws.cell(row=12, column=5, value="Sincerity Gate").font = Font(bold=True, size=11, color=PURPLE)
    for i, (g, items) in enumerate([("Gate 1", [r for r in data if r.get("sincerity_class")!="G"]),
                                     ("Gate 2", [r for r in data if r.get("sincerity_class") not in ("C","E","F","G")]),
                                     ("Gate 3", gate3)]):
        ws.cell(row=13+i, column=5, value=g)
        ws.cell(row=13+i, column=6, value=len(items))
        pos = round(sum(1 for r in items if r["sentiment"]=="긍정")/max(len(items),1)*100,1)
        ws.cell(row=13+i, column=7, value=f"{pos}%")

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14


# Build workbook
wb = Workbook()

# Sheet 1: 요약
write_summary(wb.active)

# Sheet 2: 전체 데이터
ws2 = wb.create_sheet()
write_data_sheet(ws2, sorted(data, key=lambda x: x.get("date","")), "전체 데이터")

# Sheet 3: 팝업기간만
ws3 = wb.create_sheet()
write_data_sheet(ws3, [r for r in data if r["period"]=="팝업기간"], "팝업기간")

# Sheet 4: 사전
ws4 = wb.create_sheet()
write_data_sheet(ws4, [r for r in data if r["period"]=="사전"], "사전")

# Sheet 5: 팝업후
ws5 = wb.create_sheet()
write_data_sheet(ws5, [r for r in data if r["period"]=="팝업후"], "팝업후")

# Sheet 6: 협찬만
ws6 = wb.create_sheet()
write_data_sheet(ws6, [r for r in data if r.get("is_sponsored")], "협찬")

wb.save(OUT)
print(f"Excel 생성 완료: {OUT}")
print(f"파일 크기: {os.path.getsize(OUT):,} bytes")
print(f"시트: 요약 / 전체 {len(data)}건 / 팝업기간 {sum(1 for r in data if r['period']=='팝업기간')}건 / 사전 {sum(1 for r in data if r['period']=='사전')}건 / 팝업후 {sum(1 for r in data if r['period']=='팝업후')}건 / 협찬 {sum(1 for r in data if r.get('is_sponsored'))}건")
