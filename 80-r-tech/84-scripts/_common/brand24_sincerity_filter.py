"""Brand24 데이터에 Sincerity Filter 적용"""
import sys, json, re
from collections import Counter
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\insig\do-better-workspace\30-knowledge\eclipse-brand24-raw.xlsx')
ws = wb['Mentions']

posts = []
for r in range(3, ws.max_row + 1):
    row_id = ws.cell(row=r, column=2).value  # B열부터 시작
    if not row_id:
        continue
    date = str(ws.cell(row=r, column=3).value or '')
    title = str(ws.cell(row=r, column=5).value or '')
    content = str(ws.cell(row=r, column=6).value or '')
    source = str(ws.cell(row=r, column=7).value or '')
    domain = str(ws.cell(row=r, column=8).value or '')
    category = str(ws.cell(row=r, column=9).value or '')
    sentiment = ws.cell(row=r, column=10).value

    channel = 'other'
    dl = domain.lower()
    if 'instagram' in dl: channel = 'instagram'
    elif 'tiktok' in dl: channel = 'tiktok'
    elif 'x.com' in dl or 'twitter' in dl: channel = 'twitter'
    elif 'blog' in dl or 'tistory' in dl: channel = 'blog'
    elif any(n in dl for n in ['news', 'nate', 'topstar', 'itbiz']): channel = 'news'
    elif 'youtube' in dl: channel = 'youtube'
    elif 'bsky' in dl: channel = 'bluesky'

    posts.append({
        'id': str(row_id), 'date': date, 'title': title, 'content': content,
        'source': source, 'domain': domain, 'category': category,
        'sentiment_raw': sentiment, 'channel': channel,
    })

print(f"Brand24 데이터: {len(posts)}건")
print(f"채널별: {dict(Counter(p['channel'] for p in posts))}")

# Sincerity Filter
sponsored_markers = ['제공받', '협찬', '원고료', '체험단', '#광고', '#ad', '서포터즈', '광고 성수']
press_phrases = ['틴케이스 내부를 하나의 거대한', 'MZ세대의 취향을 저격',
                 '포도 에너지를 테마로 한', '소장 가치가 높은 한정판', '레트로 오락실 컨셉']
biz_signals = ['공식', '판매처', '구매링크', '바로가기', '할인코드', '쿠폰', '입점']
eclipse_popup_kw = ['팝업', '성수', '틴케이스', '포도향', '이클립스월드', 'tin case', '이클립스 월드',
                    '이클립스 팝', 'eclipse popup', 'eclipsemints']

for p in posts:
    full = (p['title'] + ' ' + p['content']).lower()
    flags = []
    trust = 100

    # 해시태그만
    hashtags = re.findall(r'#\S+', full)
    non_hash = re.sub(r'#\S+', '', full).strip()
    if len(non_hash) < 30 and len(hashtags) > 3:
        flags.append('HASHTAG_ONLY'); trust -= 25

    # 보도자료
    press_match = sum(1 for ph in press_phrases if ph in full)
    if press_match >= 2:
        flags.append('PRESS_COPY'); trust -= 35

    # 비즈니스
    if sum(1 for s in biz_signals if s in full) >= 2:
        flags.append('BIZ_SUSPECT'); trust -= 25

    # 협찬
    if any(m in full for m in sponsored_markers):
        flags.append('SPONSORED'); trust -= 15

    # 나열형
    if any(s in full for s in ['총정리', '리스트', 'BEST', '모음']):
        flags.append('LISTING'); trust -= 20

    # 팝업 무관 노이즈
    if not any(kw in full for kw in eclipse_popup_kw):
        flags.append('OFF_TOPIC'); trust -= 30

    # 극히 짧은 텍스트
    if len(full) < 20:
        flags.append('TOO_SHORT'); trust -= 20

    # 개인경험 (가점)
    personal = ['다녀왔', '갔다왔', '방문했', '먹어봤', '줄 서', '웨이팅', '친구랑', '갔는데', '왔어', '해봤']
    if sum(1 for s in personal if s in full) >= 1:
        flags.append('PERSONAL_EXP'); trust += 15

    trust = max(0, min(100, trust))

    if 'PRESS_COPY' in flags or 'BIZ_SUSPECT' in flags:
        cls = 'F_비즈니스/보도자료'
    elif 'HASHTAG_ONLY' in flags or 'TOO_SHORT' in flags:
        cls = 'E_리그램/단순공유'
    elif 'OFF_TOPIC' in flags:
        cls = 'G_노이즈(팝업무관)'
    elif 'LISTING' in flags:
        cls = 'D_나열형정보'
    elif 'SPONSORED' in flags:
        cls = 'C_협찬'
    elif 'PERSONAL_EXP' in flags:
        cls = 'A_진성후기'
    else:
        cls = 'B_일반포스트'

    p['content_class'] = cls
    p['trust_score'] = trust
    p['flags'] = flags

# === 전체 통계 ===
print("\n" + "="*60)
print("=== Content Class 분류 결과 ===")
print("="*60)
class_dist = Counter(p['content_class'] for p in posts)
for cls in sorted(class_dist.keys()):
    cnt = class_dist[cls]
    bar = chr(9608) * cnt
    print(f"  {cls}: {bar} {cnt}건 ({cnt/len(posts)*100:.0f}%)")

# === 채널별 분류 ===
print("\n=== 채널별 분류 ===")
for ch in sorted(set(p['channel'] for p in posts)):
    subset = [p for p in posts if p['channel'] == ch]
    ch_cls = Counter(p['content_class'] for p in subset)
    top = ch_cls.most_common(3)
    top_str = ", ".join(f"{c}:{n}" for c, n in top)
    print(f"  [{ch}] {len(subset)}건 | {top_str}")

# === 인스타만 ===
print("\n" + "="*60)
print("=== 인스타그램 상세 ===")
print("="*60)
insta = [p for p in posts if p['channel'] == 'instagram']
print(f"인스타 샘플: {len(insta)}건")
insta_cls = Counter(p['content_class'] for p in insta)
for cls in sorted(insta_cls.keys()):
    cnt = insta_cls[cls]
    print(f"  {cls}: {cnt}건")

# === Trust Score ===
print("\n=== Trust Score 분포 ===")
for label, lo, hi in [('높음(70+)', 70, 101), ('중간(40~69)', 40, 70), ('낮음(~39)', 0, 40)]:
    cnt = sum(1 for p in posts if lo <= p['trust_score'] < hi)
    print(f"  {label}: {cnt}건 ({cnt/len(posts)*100:.0f}%)")

# === Flag 빈도 ===
print("\n=== Flag 빈도 ===")
all_flags = []
for p in posts:
    all_flags.extend(p['flags'])
for f, c in Counter(all_flags).most_common():
    print(f"  {f}: {c}건")

# === Sincerity Gate ===
print("\n" + "="*60)
print("=== Sincerity Gate 적용 ===")
print("="*60)

def gate_stats(subset, label):
    n = len(subset)
    if n == 0:
        print(f"  [{label}] N=0")
        return
    sent = Counter(p.get('sentiment_raw') for p in subset)
    pos = sent.get(1, 0)
    neg = sent.get(-1, 0)
    neu = n - pos - neg
    print(f"  [{label}] N={n} | 긍정={pos}({pos/n*100:.0f}%) 부정={neg}({neg/n*100:.0f}%) 중립={neu}({neu/n*100:.0f}%)")

gate1 = posts
gate2 = [p for p in posts if p['content_class'] not in ('C_협찬', 'E_리그램/단순공유', 'F_비즈니스/보도자료', 'G_노이즈(팝업무관)')]

gate_stats(gate1, "Gate 1 전체")
gate_stats(gate2, "Gate 2 정제 (협찬+리그램+노이즈 제거)")

# 인스타만
print("\n=== 인스타만 Gate ===")
ig1 = [p for p in posts if p['channel'] == 'instagram']
ig2 = [p for p in ig1 if p['content_class'] not in ('C_협찬', 'E_리그램/단순공유', 'F_비즈니스/보도자료', 'G_노이즈(팝업무관)')]
gate_stats(ig1, "인스타 Gate 1 전체")
gate_stats(ig2, "인스타 Gate 2 정제")

# === 필터된 샘플 ===
print("\n=== 필터된 포스트 샘플 ===")
for cls in ['C_협찬', 'E_리그램/단순공유', 'G_노이즈(팝업무관)']:
    filtered = [p for p in posts if p['content_class'] == cls]
    if filtered:
        print(f"\n--- {cls} ({len(filtered)}건) ---")
        for p in filtered[:3]:
            t = p['title'][:70] if p['title'] else p['content'][:70]
            print(f"  [{p['channel']}] {t} | trust:{p['trust_score']} | flags:{','.join(p['flags'])}")

# 저장
out = r'C:\Users\insig\do-better-workspace\30-knowledge\eclipse-brand24-sincerity-filtered.json'
with open(out, 'w', encoding='utf-8') as f:
    json.dump(posts, f, ensure_ascii=False, indent=2)
print(f"\nSaved: {out}")
