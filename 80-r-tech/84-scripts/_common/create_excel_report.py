"""Eclipse Popup Buzz Analysis → Excel Report"""
import json, sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open(r'C:\Users\insig\do-better-workspace\30-knowledge\eclipse-popup-analyzed.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()

# Styles
hdr_font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill(start_color='6666FF', end_color='6666FF', fill_type='solid')
hdr_green = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
hdr_orange = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
sub_font = Font(name='맑은 고딕', size=9, color='666666')
bold_font = Font(name='맑은 고딕', bold=True, size=10)
title_font = Font(name='맑은 고딕', bold=True, size=14, color='6666FF')

sent_fills = {
    '긍정': PatternFill(start_color='D1FAE5', fill_type='solid'),
    '부정': PatternFill(start_color='FEE2E2', fill_type='solid'),
    '중립': PatternFill(start_color='F1F5F9', fill_type='solid'),
    '정보성': PatternFill(start_color='E1E1FF', fill_type='solid'),
}

sorted_data = sorted(data, key=lambda x: x['date'])
total = len(sorted_data)
periods = Counter(d['period'] for d in data)
sentiments = Counter(d['sentiment'] for d in data)
types = Counter(d['post_type'] for d in data)
date_dist = Counter(d['date'] for d in data)

# ===== Sheet 1: 요약 대시보드 =====
ws1 = wb.active
ws1.title = '요약 대시보드'

ws1.merge_cells('A1:H1')
ws1['A1'] = '이클립스 월드 — Into the Tin Case | 팝업스토어 버즈 분석'
ws1['A1'].font = title_font

ws1.merge_cells('A2:H2')
ws1['A2'] = '분석 기간: 2026.02.26 ~ 2026.03.31 | 소스: 네이버 블로그 | Project RENT R-lab'
ws1['A2'].font = sub_font

# 기간별
ws1['A4'] = '■ 기간별 버즈량'
ws1['A4'].font = bold_font
for i, h in enumerate(['기간', '건수', '일평균', '비율']):
    c = ws1.cell(row=5, column=i+1, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')

period_rows = [
    ('팝업전(~3/4)', periods.get('팝업전(~3/4)', 0), 7),
    ('사전홍보(3/5~11)', periods.get('사전홍보(3/5~11)', 0), 7),
    ('팝업진행(3/12~15)', periods.get('팝업진행(3/12~15)', 0), 4),
    ('팝업후(3/16~)', periods.get('팝업후(3/16~)', 0), 16),
]
for r, (name, cnt, days) in enumerate(period_rows):
    ws1.cell(row=6+r, column=1, value=name)
    ws1.cell(row=6+r, column=2, value=cnt).font = bold_font
    ws1.cell(row=6+r, column=3, value=round(cnt/max(days,1), 1))
    ws1.cell(row=6+r, column=4, value=f'{cnt/max(total,1)*100:.0f}%')

ws1.cell(row=10, column=1, value='합계').font = bold_font
ws1.cell(row=10, column=2, value=total).font = Font(name='맑은 고딕', bold=True, size=12, color='6666FF')

# 감성 분포
ws1['A12'] = '■ 감성 분포'
ws1['A12'].font = bold_font
for i, h in enumerate(['감성', '건수', '비율']):
    c = ws1.cell(row=13, column=i+1, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')

for r, (s, cnt) in enumerate(sorted(sentiments.items(), key=lambda x: -x[1])):
    ws1.cell(row=14+r, column=1, value=s)
    ws1.cell(row=14+r, column=2, value=cnt).font = bold_font
    ws1.cell(row=14+r, column=3, value=f'{cnt/max(total,1)*100:.0f}%')

# 포스트 유형
ws1['A19'] = '■ 포스트 유형'
ws1['A19'].font = bold_font
for i, h in enumerate(['유형', '건수', '비율']):
    c = ws1.cell(row=20, column=i+1, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')

for r, (t, cnt) in enumerate(sorted(types.items(), key=lambda x: -x[1])):
    ws1.cell(row=21+r, column=1, value=t)
    ws1.cell(row=21+r, column=2, value=cnt).font = bold_font
    ws1.cell(row=21+r, column=3, value=f'{cnt/max(total,1)*100:.0f}%')

# 날짜별 차트
ws1['E4'] = '■ 날짜별 버즈량'
ws1['E4'].font = bold_font
for i, h in enumerate(['날짜', '건수', '차트', '비고']):
    c = ws1.cell(row=5, column=5+i, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')

row = 6
for date in sorted(date_dist.keys()):
    cnt = date_dist[date]
    label = ''
    if date in ('20260312', '20260313', '20260314', '20260315'):
        label = '★ 팝업 진행일'
    elif date <= '20260304':
        label = '팝업 전'
    elif date <= '20260311':
        label = '사전홍보'
    else:
        label = '팝업 후'

    ws1.cell(row=row, column=5, value=f'{date[4:6]}/{date[6:]}')
    ws1.cell(row=row, column=6, value=cnt).font = bold_font
    bar_cell = ws1.cell(row=row, column=7, value='█' * cnt)
    bar_cell.font = Font(name='맑은 고딕', size=8, color='6666FF')
    note_cell = ws1.cell(row=row, column=8, value=label)
    if '팝업 진행' in label:
        note_cell.font = Font(name='맑은 고딕', size=9, color='FF0000', bold=True)
    else:
        note_cell.font = Font(name='맑은 고딕', size=9, color='999999')
    row += 1

for col, w in [(1, 20), (2, 8), (3, 10), (4, 8), (5, 10), (6, 6), (7, 50), (8, 14)]:
    ws1.column_dimensions[get_column_letter(col)].width = w


def write_data_sheet(ws, items, fill):
    headers = ['날짜', '기간구분', '포스트유형', '감성', '긍정점수', '부정점수',
               '제목', '블로거', '내용 써머리', '세부 내용', '검색키워드', '채널', 'URL']
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i+1, value=h)
        c.font = hdr_font
        c.fill = fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for r, d in enumerate(items):
        row = r + 2
        dt = d['date']
        ws.cell(row=row, column=1, value=f'{dt[:4]}-{dt[4:6]}-{dt[6:]}')
        ws.cell(row=row, column=2, value=d['period'])
        ws.cell(row=row, column=3, value=d['post_type'])
        sc = ws.cell(row=row, column=4, value=d['sentiment'])
        sc.fill = sent_fills.get(d['sentiment'], PatternFill())
        ws.cell(row=row, column=5, value=d['pos_score'])
        ws.cell(row=row, column=6, value=d['neg_score'])
        ws.cell(row=row, column=7, value=d['title'])
        ws.cell(row=row, column=8, value=d['blogger'])
        ws.cell(row=row, column=9, value=d['summary'])
        ws.cell(row=row, column=10, value=d['content_preview'])
        ws.cell(row=row, column=11, value=d['keyword'])
        ws.cell(row=row, column=12, value=d['channel'])
        ws.cell(row=row, column=13, value=d['link'])

    widths = [12, 16, 10, 8, 8, 8, 50, 15, 50, 60, 15, 12, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.auto_filter.ref = f'A1:M{len(items)+1}'


# Sheet 2: 전체 데이터
ws2 = wb.create_sheet('전체 데이터')
write_data_sheet(ws2, sorted_data, hdr_fill)

# Sheet 3: 팝업기간
ws3 = wb.create_sheet('팝업기간(3.12-15)')
popup_only = [d for d in sorted_data if d['period'] == '팝업진행(3/12~15)']
write_data_sheet(ws3, popup_only, hdr_green)

# Sheet 4: 팝업전
ws4 = wb.create_sheet('팝업전(~3.4)')
pre_only = [d for d in sorted_data if d['period'] == '팝업전(~3/4)']
write_data_sheet(ws4, pre_only, hdr_orange)

# Sheet 5: 사전홍보
ws5 = wb.create_sheet('사전홍보(3.5-11)')
pre_promo = [d for d in sorted_data if d['period'] == '사전홍보(3/5~11)']
write_data_sheet(ws5, pre_promo, PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid'))

# Save
out_path = r'C:\Users\insig\do-better-workspace\30-knowledge\eclipse-popup-buzz-analysis.xlsx'
wb.save(out_path)
print(f'Excel saved: {out_path}')
print(f'  Sheet 1: 요약 대시보드')
print(f'  Sheet 2: 전체 데이터 ({len(sorted_data)}건)')
print(f'  Sheet 3: 팝업기간 ({len(popup_only)}건)')
print(f'  Sheet 4: 팝업전 ({len(pre_only)}건)')
print(f'  Sheet 5: 사전홍보 ({len(pre_promo)}건)')
