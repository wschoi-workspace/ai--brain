import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base_dir = r"C:\Users\insig\Dropbox\02-프로젝트렌트RENT\03_R-lab\01.분석솔루션\기사분석"

with open(f"{base_dir}\\_extracted_articles.json", "r", encoding="utf-8") as f:
    articles = json.load(f)

def parse_article(fname, text):
    lines = text.strip().split("\n")
    data = {
        "원본파일": fname, "발행날짜": "", "기사제목": "", "소제목": "",
        "프로젝트명": "", "브랜드_기관명": "", "유형": "", "대분류": "",
        "콘셉트": "", "주요특징": "", "장소명": "", "주소": "", "지역": "",
        "시작일": "", "종료일": "", "운영시간": "", "주최": "", "협력사": "",
        "핵심인물": "", "타겟": "", "가격정보": "", "주요데이터": "",
        "마케팅전략": "", "차별점": "", "기사요약": "", "핵심인용문": "",
        "기자명": "", "자료협조": "", "링크": "",
    }

    if lines and re.match(r"\d{4}-\d{2}-\d{2}", lines[0]):
        data["발행날짜"] = lines[0].strip()
        lines = lines[1:]
    if lines:
        data["기사제목"] = lines[0].strip()
        lines = lines[1:]
    if lines:
        data["소제목"] = lines[0].strip()
        lines = lines[1:]

    meta_patterns = {
        "프로젝트명": r"프로젝트\n(.+?)(?:\n|$)",
        "장소명": r"장소\n(.+?)(?:\n|$)",
        "주소": r"주소\n(.+?)(?:\n|$)",
        "주최": r"주(?:최|관)\n(.+?)(?:\n|$)",
        "핵심인물": r"(?:크리에이터|기획자|디렉터)\n(.+?)(?:\n|$)",
        "링크": r"링크\n(.+?)(?:\n|$)",
    }
    for field, pattern in meta_patterns.items():
        m = re.search(pattern, text)
        if m:
            data[field] = m.group(1).strip()

    date_match = re.search(r"일자\n(.+?)(?:\n|$)", text)
    if date_match:
        date_str = date_match.group(1).strip()
        if " - " in date_str:
            parts = date_str.split(" - ")
            data["시작일"] = parts[0].strip()
            data["종료일"] = parts[1].strip()
        else:
            data["시작일"] = date_str

    time_match = re.search(r"시간\n(.+?)(?:\n|$)", text)
    if time_match:
        data["운영시간"] = time_match.group(1).strip()

    author_match = re.search(r"글(?:[\xb7·]사진)?\s+(.+?)(?:\n|$)", text)
    if author_match:
        data["기자명"] = author_match.group(1).strip()

    credit_match = re.search(r"(?:취재협조|자료제공|자료 제공 및 취재 협조|자료제공 및 취재협조)\s+(.+?)(?:\n|$)", text)
    if credit_match:
        data["자료협조"] = credit_match.group(1).strip()

    if data["주소"]:
        addr = data["주소"]
        if "용산" in addr: data["지역"] = "용산/한남/이태원"
        elif "마포" in addr: data["지역"] = "마포/서교"
        elif "중구" in addr: data["지역"] = "중구"
        elif "성동" in addr or "성수" in addr: data["지역"] = "성동/성수"
        elif "강남" in addr: data["지역"] = "강남"

    content_lines = [l for l in lines if len(l) > 50 and not l.startswith("출처") and not l.startswith("사진") and "\u00a9" not in l]
    if content_lines:
        data["기사요약"] = content_lines[0][:200] + ("..." if len(content_lines[0]) > 200 else "")

    return data

parsed = []
for fname, text in articles.items():
    d = parse_article(fname, text)
    parsed.append(d)

enrichments = {
    "3월 서울에서 가장 낯선 경험, 리움미술관 티노 세갈전.docx": {
        "유형": "Exhibition", "대분류": "Art",
        "콘셉트": "물리적 실체 없는 구성된 상황 - 기록과 소유를 거부하는 퍼포먼스 전시",
        "주요특징": "사진 촬영 금지, 가벽 제거한 건축 노출, 소장품 38점 재배치, 6주마다 작품 교체",
        "핵심인물": "티노 세갈(Tino Sehgal)", "브랜드_기관명": "리움미술관",
        "타겟": "현대미술 애호가, 경험 중심 문화 소비층",
        "차별점": "인스타그래머블 시대에 기록 불가 전시라는 역발상",
        "핵심인용문": "사라지기에 더 선명하다",
        "주요데이터": "소장품 38점 재배치, M2 지하 1층 3점 구성된 상황",
        "공간묘사": "가벽 모두 걷어낸 M2 건물 골조 노출, 설명 없는 전시장, 해석자들이 관객과 대화하는 장면, 받침대 없이 바닥에 놓인 조각, 뒷모습을 보이는 조각상",
        "대표아이템상세": "구성된 상황(Constructed Situations) 3점 - 물리적 형체 없이 사람의 목소리와 움직임만으로 구성. 김정숙/장 아르프/자코메티 조각 나란히 배치",
        "체험동선": "전시장 입구 > 가벽 제거된 M2 공간(건축 민낯) > 해석자와의 조우 > 소장품 38점 관람 > 여운 속 퇴장",
    },
    "건물 전체를 '재즈'로 채웠다.docx": {
        "유형": "Exhibition", "대분류": "Culture/Music",
        "콘셉트": "블루노트 레코드 아카이브 전시 - 음악과 디자인, 기록이 교차하는 공간",
        "주요특징": "445장 LP 컬렉션, 1500시리즈 풀 공개, JBL 하이파이+노이즈캔슬링 3방식 청취, 뮤직&재즈 맵",
        "핵심인물": "루디 반 겔더, 프란시스 울프, 리드 마일스", "브랜드_기관명": "페즈(FEZH)",
        "타겟": "음악 애호가, 디자인 관심층, 한남동 방문객",
        "마케팅전략": "뮤직&재즈 맵으로 한남동 일대 재즈 공간 연계, 도장 스탬프 할인",
        "차별점": "건물 전체를 전시로 활용하는 커뮤니티 몰 구조",
        "핵심인용문": "같은 음악이라도 어떤 방식으로 듣느냐에 따라 감각이 달라진다",
        "주요데이터": "445장 LP, 1939년 설립 블루노트, iF 디자인 어워드 수상",
        "협력사": "JBL, 하바구든",
        "공간묘사": "1-2층 LP 컬렉션 전시, 지하1층 뮤직바 블루캣, 3층 인터뷰 영상, 4층 까사 델 아구아 영상 상영, JBL 하이파이 시스템 공간 울림, 곳곳에 하바구든 리빙 제품 배치",
        "대표아이템상세": "1500시리즈 풀 컬렉션(결손 없는 완전체), 프란시스 울프 사진, 리드 마일스 커버 디자인, 앤디 워홀 초기 드로잉, QR코드 음악 청취",
        "체험동선": "1-2층 LP 관람 > QR코드 앨범 청취 > 3층 인터뷰 자료 > 4층 영상 관람 > 지하1층 블루캣 바 > 뮤직&재즈 맵으로 한남동 탐방",
    },
    "그라운드씨소.docx": {
        "유형": "Exhibition", "대분류": "Art",
        "콘셉트": "일상의 부조리를 유머로 전환하는 컨셉추얼 아트 - 진지하게 진지하지 않기",
        "주요특징": "하이퍼리얼리즘 조각, 관객 참여형 작품, 1003개 달걀 설치, 여권사진 패러디",
        "핵심인물": "맥스 시덴토프(Max Sidentop)", "브랜드_기관명": "그라운드시소",
        "타겟": "현대미술/팝아트 관심층, 젠틀몬스터 팬층",
        "차별점": "젠틀몬스터/제니 콜라보 아티스트의 아시아 첫 대규모 개인전",
        "핵심인용문": "혼란스럽고 부조리한 세상을 견디기 위한 일종의 대처 기제(coping mechanism)",
        "주요데이터": "1003개 달걀, 25세 최연소 파트너, 아시아 첫 대규모 개인전",
        "협력사": "젠틀몬스터, 제니",
        "공간묘사": "하우스 노웨어 검은 비닐봉지+금색 봉투 노인 조각상(More Is More), 여권사진 패러디 시리즈, 자화상 그리는 조각상, 1003개 달걀 라벨 설치, 대형 속옷 차림 조각",
        "대표아이템상세": "Mine Right Now 뮤직비디오(감독 직접 출연 립싱크), Passport Photos(정면 증명사진+프레임 밖 반전), Self Portrait(조각이 그림 그리는 장면), Nature vs Nurture(1003달걀+라벨), The Amateur(관객이 스케치하는 대형 조각)",
        "체험동선": "입구 > More Is More 회고 > 사진/영상 작품 > 조각 작품들 > 1003달걀 설치 > The Amateur 스케치 참여 > 마무리",
    },
    "러닝·요가·와인·디제잉으로 연결되는 라이프스타일 공간 4.docx": {
        "유형": "Exhibition", "대분류": "Design",
        "콘셉트": "전후 유럽 공공디자인 가구의 황금시대 조명 - 예술과 기술의 결합",
        "주요특징": "르 코르뷔지에/장 프루베 등 거장 가구, 공공 공간용 디자인, 1950-60년대 원본",
        "핵심인물": "르 코르뷔지에, 장 프루베, 르네 가브리엘, 난나 딧젤",
        "브랜드_기관명": "스페이스 이수", "타겟": "디자인 애호가, 건축/인테리어 관심층",
        "차별점": "공공디자인 관점에서 전후 유럽 가구를 조명하는 국내 드문 기획",
        "핵심인용문": "모두를 위한 공공의 가치를 꿈꾸다",
        "주요데이터": "1940-1960년대 가구, BBPR 스튜디오",
        "공간묘사": "스페이스 이수 전시장, 프랑스/이탈리아/덴마크 디자이너 가구 원본 전시, 공영주택/대학 기숙사/도서관/리조트용 가구",
        "대표아이템상세": "르 코르뷔지에/장 프루베/르네 가브리엘/BBPR/난나 딧젤의 공공 공간용 가구 원본",
        "체험동선": "전시 도입부 > 프랑스 디자이너 섹션 > 이탈리아 섹션 > 덴마크 섹션 > 마무리",
    },
    "마음의 얼룩을 지워주는 세탁소가 열렸다, 쏘쏘런드리.docx": {
        "유형": "Popup Store", "대분류": "Retail/Character",
        "콘셉트": "귀여운 캐릭터 굿즈로 일상을 위로하는 감성 세탁소 팝업",
        "주요특징": "세탁소 컨셉 인테리어, 온도 변색 키링, 캐릭터 굿즈, 감성 스토리텔링",
        "브랜드_기관명": "쏘쏘패밀리(Soso Family)", "타겟": "캐릭터/굿즈 소비층, MZ세대",
        "차별점": "감정 치유 콘셉트를 세탁소 공간으로 구현",
        "핵심인용문": "마음의 얼룩을 지워주는 세탁소", "협력사": "오브젝트",
        "공간묘사": "계단 내려가면 세탁기/세제/빨랫감으로 꾸며진 입구, 비눗방울/다리미/옷걸이 오브제 배치, 위칙(WECHIK) 세제 브랜드 협업 진열, 벽 한편 분실물 보관소",
        "대표아이템상세": "쏘쏘 인형 키링 6색(마음 얼룩 흡수해 색 변함, 감정 레시피로 선택), 분실물 키링(양말/속옷/카드 모티프), 줄자키링, 스티커북키링, 메모지, 빅백, 파우치, 데코스티커",
        "체험동선": "계단 하강 > 세탁소 입구 몰입 > 세탁 오브제 공간 > 감정 레시피로 키링 선택 > 분실물 보관소 > 굿즈 쇼핑 > 마무리",
    },
    "올리브베러.docx": {
        "유형": "Exhibition/Popup", "대분류": "Art/Culture",
        "콘셉트": "한국 전통 공예를 현대적 감각으로 재해석하는 라이프스타일 전시",
        "주요특징": "도자기/금속공예/옻칠 등 전통 소재, 현대 디자인 접목, 체험 프로그램",
        "브랜드_기관명": "올리브베러", "타겟": "공예/디자인 관심층, 문화 체험 선호 소비자",
        "차별점": "전통 공예의 현대화를 공간 경험으로 전환",
        "공간묘사": "", "대표아이템상세": "", "체험동선": "",
    },
    "이케아 쇼룸에서.docx": {
        "유형": "Brand Campaign", "대분류": "Retail/Lifestyle",
        "콘셉트": "쇼룸을 러닝 트랙으로 전환 - 브랜드 공간의 새로운 활용법",
        "주요특징": "이케아 매장 내 러닝 이벤트, 체험형 브랜드 마케팅",
        "브랜드_기관명": "이케아(IKEA)", "타겟": "러너, 라이프스타일 관심층",
        "마케팅전략": "쇼룸 공간을 러닝 코스로 활용한 체험 마케팅",
        "차별점": "가구 매장에서 러닝하는 이색 경험",
        "주요데이터": "750명 참여, 광명/부산 매장",
        "공간묘사": "이케아 쇼룸 내부를 러닝 트랙으로 전환, 가구 사이를 달리는 이색 장면",
        "대표아이템상세": "쇼룸 러닝 이벤트 자체가 핵심 콘텐츠",
        "체험동선": "참가 등록 > 쇼룸 러닝 > 이벤트 마무리",
    },
    "피자런.docx": {
        "유형": "Brand Campaign", "대분류": "F&B/Lifestyle",
        "콘셉트": "달린 거리만큼 피자 할인 - 러닝과 외식의 결합",
        "주요특징": "거리 기반 할인, 피자+러닝 크로스오버, 커뮤니티 러닝",
        "브랜드_기관명": "프랑코 만카(Franco Manca)", "타겟": "러너, F&B 소비자, 건강+미식 관심층",
        "마케팅전략": "달린 거리에 따른 피자 할인 프로모션",
        "차별점": "운동과 외식을 결합한 참여형 프로모션",
        "주요데이터": "70여 개 매장(영국), 1986년부터 운영",
        "공간묘사": "프랑코 만카 매장, 나폴리 전통 방식 사워도우 피자, 영국 70여 개 지점",
        "대표아이템상세": "달린 거리 기반 피자 할인 프로모션 - 거리에 따라 할인율 차등",
        "체험동선": "러닝 코스 출발 > 달리기 > 프랑코 만카 매장 도착 > 거리 인증 > 할인 피자 식사",
    },
}

for item in parsed:
    fname = item["원본파일"]
    if fname in enrichments:
        for k, v in enrichments[fname].items():
            if v:
                item[k] = v

# Create Excel
wb = openpyxl.Workbook()
purple = "6666FF"
header_fill = PatternFill(start_color=purple, end_color=purple, fill_type="solid")
header_font = Font(name="Pretendard", bold=True, color="FFFFFF", size=10)
cell_font = Font(name="Pretendard", size=10)
thin_border = Border(bottom=Side(style="thin", color="E8E8E8"))
wrap_align = Alignment(wrap_text=True, vertical="top")

columns = [
    ("발행날짜", 12), ("기사제목", 35), ("소제목", 30), ("프로젝트명", 25),
    ("브랜드_기관명", 18), ("유형", 14), ("대분류", 14), ("콘셉트", 40),
    ("주요특징", 40), ("장소명", 18), ("주소", 30), ("지역", 14),
    ("시작일", 14), ("종료일", 14), ("운영시간", 20), ("주최", 14),
    ("협력사", 18), ("핵심인물", 20), ("타겟", 25), ("가격정보", 12),
    ("주요데이터", 35), ("마케팅전략", 35), ("차별점", 35),
    ("공간묘사", 50), ("대표아이템상세", 50), ("체험동선", 40),
    ("기사요약", 50), ("핵심인용문", 35), ("기자명", 12), ("자료협조", 14),
    ("링크", 14), ("원본파일", 25),
]

# Sheet 1
ws1 = wb.active
ws1.title = "기사 데이터"
for col_idx, (col_name, width) in enumerate(columns, 1):
    cell = ws1.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

for row_idx, item in enumerate(parsed, 2):
    for col_idx, (col_name, _) in enumerate(columns, 1):
        val = item.get(col_name, "")
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = cell_font
        cell.alignment = wrap_align
        cell.border = thin_border

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(parsed)+1}"

# Sheet 2
ws2 = wb.create_sheet("AI 기사 생성용")
gen_columns = list(columns[:-1]) + [("참고기사원문", 60), ("본문 초안", 50), ("생성 상태", 12)]
green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
for col_idx, (col_name, width) in enumerate(gen_columns, 1):
    cell = ws2.cell(row=1, column=col_idx, value=col_name)
    cell.fill = green_fill
    cell.font = Font(name="Pretendard", bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

guide_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
guides = ["YYYY-MM-DD", "기사 메인 제목", "부제/태그라인", "이벤트/전시/팝업명",
          "주최 브랜드/기관", "Popup/Exhibition/Campaign", "Retail/Art/F&B/Lifestyle",
          "핵심 콘셉트 1문장", "차별화 포인트들", "베뉴명", "전체 주소", "동네/지역",
          "시작일", "종료일", "운영 시간", "주최기관", "협업 파트너", "기획자/아티스트",
          "대상 고객층", "입장료/참가비", "숫자형 데이터", "프로모션 메카닉",
          "유사 사례 대비 차별화",
          "방문 시 보이는 실제 장면 3~5가지", "각 아이템의 구체적 설명",
          "입구>첫인상>메인>서브>마무리",
          "2~3줄 요약", "핵심 인용문", "작성자", "크레딧",
          "웹사이트/SNS", "톤 참고용 원문 붙여넣기", "AI가 생성한 초안", "대기/생성완료/수정중"]
for col_idx, guide in enumerate(guides, 1):
    cell = ws2.cell(row=2, column=col_idx, value=guide)
    cell.fill = guide_fill
    cell.font = Font(name="Pretendard", size=9, italic=True, color="92400E")
    cell.alignment = wrap_align
ws2.freeze_panes = "A3"

# Sheet 3
ws3 = wb.create_sheet("프롬프트 템플릿")
ws3.column_dimensions["A"].width = 100
prompt_lines = [
    "=== 기사 생성 프롬프트 v2 (최적화) ===",
    "",
    "당신은 공간/브랜드/이벤트 전문 매거진 에디터입니다.",
    "아래 정보만을 사용하여 기사를 작성해주세요.",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "[절대 규칙 - 반드시 지킬 것]",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "1. 아래 제공된 정보만 사용할 것. 없는 정보를 추측하거나 창작 금지.",
    "2. 주소, 장소명, 일자, 주최, 브랜드명은 제공된 값을 그대로 사용할 것.",
    "3. 제공되지 않은 기능/서비스(앱, IoT, 카페 등)를 임의로 추가 금지.",
    "4. 확인되지 않은 사실은 아예 쓰지 말 것 (추정/추측 표현도 금지).",
    "5. 정보가 부족한 컬럼은 해당 내용을 생략할 것 (채우려고 지어내지 말 것).",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "[기사 정보]",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "발행날짜: {발행날짜}",
    "프로젝트명: {프로젝트명}",
    "브랜드/기관: {브랜드_기관명}",
    "유형: {유형} / 대분류: {대분류}",
    "콘셉트: {콘셉트}",
    "주요 특징: {주요특징}",
    "장소: {장소명} / 주소: {주소}",
    "기간: {시작일} ~ {종료일}",
    "운영시간: {운영시간}",
    "주최: {주최} / 협력사: {협력사}",
    "핵심인물: {핵심인물}",
    "타겟: {타겟}",
    "가격: {가격정보}",
    "주요 데이터: {주요데이터}",
    "마케팅 전략: {마케팅전략}",
    "차별점: {차별점}",
    "",
    "[공간/체험 디테일]",
    "공간 묘사: {공간묘사}",
    "대표 아이템: {대표아이템상세}",
    "체험 동선: {체험동선}",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "[구조 - 이 순서대로 작성]",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "1) 첫 줄: 발행날짜",
    "2) 메인 제목 (흥미를 끄는 한 줄)",
    "3) 소제목 (콘셉트를 함축하는 부가 설명)",
    "4) 리드문 2~3문장 (프로젝트명, 장소, 기간, 핵심 콘셉트 포함)",
    "5) 본문 2~4개 섹션 (번호 매기기, 체험 동선 순서로 전개)",
    "   - 각 섹션: 소제목 + 공간/아이템 묘사 + 의미/감상",
    "6) 마무리 1~2문장 (여운 남기는 초대형 어미: ~해보자, ~할 것이다)",
    "7) 하단 메타데이터 블록 (아래 형식 그대로)",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "[톤 & 문체]",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "- 관람객/방문객의 시점에서 서술 (분석가/평론가 시점 금지)",
    "- 따뜻하고 친근한 톤 (공간의 감성을 전달하는 데 집중)",
    "- 체험에서 나오는 형용사 사용 (귀여운, 아기자기한, 생생한, 아늑한)",
    "- 비즈니스 분석 용어 사용 금지 (타격한다, 고도의 전략, 팬덤 형성 등)",
    "- ~다 체 사용 (존댓말 아닌 서술체)",
    "- 짧은 문장과 긴 문장 교차",
    "- 전문 용어는 간단한 괄호 설명 병기",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "[하단 메타데이터 형식 - 줄바꿈으로 분리]",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "글 {기자명}",
    "자료제공 및 취재협조 {자료협조}",
    "프로젝트",
    "{프로젝트명}",
    "장소",
    "{장소명}",
    "주소",
    "{주소}",
    "일자",
    "{시작일} - {종료일}",
    "시간",
    "{운영시간}",
    "주최",
    "{주최}",
    "링크",
    "{링크}",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "[참고 기사 (이 톤과 구조를 따를 것)]",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "(여기에 기존 원문 기사 1건을 통째로 붙여넣으세요.",
    " AI는 이 기사의 문체, 구조, 톤을 학습한 뒤",
    " 위 새로운 정보로 기사를 작성합니다.)",
    "",
    "분량: 1500~2500자 내외",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "[사용법]",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "1. AI 기사 생성용 시트에 새 기사 정보를 채웁니다",
    "   - 특히 공간묘사, 대표아이템상세, 체험동선은 반드시 채울 것",
    "   - 이 3개 컬럼이 비면 AI가 없는 정보를 지어냅니다",
    "2. 참고기사원문 컬럼에 톤 참고용 기존 기사를 붙여넣습니다",
    "3. 위 프롬프트의 {필드명} 부분을 실제 값으로 교체합니다",
    "4. [참고 기사] 섹션에 참고기사원문을 붙여넣습니다",
    "5. Claude 또는 Gemini에 프롬프트를 전달합니다",
    "6. 생성된 초안을 본문 초안 컬럼에 붙여넣습니다",
    "7. 생성 상태를 업데이트합니다 (대기 > 생성완료 > 수정중 > 완료)",
]
for row_idx, line in enumerate(prompt_lines, 1):
    cell = ws3.cell(row=row_idx, column=1, value=line)
    is_header = line.startswith("===") or line.startswith("[") or (line and line[0].isdigit() and line[1] == ".")
    cell.font = Font(name="Pretendard", bold=is_header, size=11 if is_header else 10)
    cell.alignment = wrap_align

output_path = f"{base_dir}\\기사분석-데이터시트.xlsx"
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Articles: {len(parsed)} rows")
