"""헬로키티×지수 팝업스토어 RXR SNS 내부 리포트 HTML + Excel 생성"""
import json
import sys
from collections import Counter
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

BASE = Path(r'C:\Users\insig\do-better-workspace\30-knowledge')

# ========== 데이터 로드 ==========
with open(BASE / 'hellokitty-jisoo-popup-2layer-results.json', 'r', encoding='utf-8') as f:
    posts = json.load(f)
with open(BASE / 'hellokitty-jisoo-popup-final-stats.json', 'r', encoding='utf-8') as f:
    stats_all = json.load(f)
with open(BASE / 'hellokitty-jisoo-popup-naver-raw.json', 'r', encoding='utf-8') as f:
    raw = json.load(f)

stats = stats_all['stats']
gates = stats_all['gates']
class_dist = stats_all['class_dist']
gate3_posts = stats_all['gate3_posts']

TOTAL = len(posts)
sponsored = [p for p in posts if p['is_sponsored']]
organic = [p for p in posts if not p['is_sponsored']]

# 기간 분류 (2026-01-14 ~ 01-22 팝업기간)
def period_of(date):
    if date <= '20260113':
        return '이벤트전'
    elif date <= '20260122':
        return '팝업기간'
    else:
        return '팝업후'

for p in posts:
    p['period_en'] = period_of(p['date'])

pre = [p for p in posts if p['period_en'] == '이벤트전']
during = [p for p in posts if p['period_en'] == '팝업기간']
post_ev = [p for p in posts if p['period_en'] == '팝업후']

# 감성
sent_pct = stats['sentiment_pct']
SENT_POS = sent_pct.get('긍정', 40.0)
SENT_MIX = sent_pct.get('혼합', 12.2)
SENT_NEU = sent_pct.get('중립', 44.1)
SENT_NEG = sent_pct.get('부정', 3.7)

TOPIC = stats['topic_dist']
RQL = stats['rql_dist']
AVG_AUTH = stats['avg_auth']
AVG_CLOUT = stats['avg_clout']
AVG_FRESH = stats['avg_freshness']
SPON_N = stats['sponsored_count']
ORG_N = stats['organic_count']
SPON_AUTH = stats['sponsored_avg_auth']
ORG_AUTH = stats['organic_avg_auth']

PERIOD_STATS = stats['period_stats']
PRE_STATS = PERIOD_STATS.get('이벤트전')
DURING_STATS = PERIOD_STATS.get('팝업기간')
POST_STATS = PERIOD_STATS.get('팝업후')

G1 = gates['gate1']
G2 = gates['gate2']
G3 = gates['gate3']

# ========== HTML 생성 ==========

CSS = """@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/variable/pretendardvariable.css');
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&family=DIN+Alternate&display=swap');
:root{--primary:#6666FF;--grad:linear-gradient(135deg,#5353FF,#8A8AFF);--black:#000;--dark:#262626;--body:#424D61;--light:#F6F6F6;--white:#FFF;--red:#C00000;--coral:#FF5050;--blue:#4D93F7;--green:#10b981;--orange:#f59e0b;--pink:#d946ef;--primary-pale:#E1E1FF;}
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Pretendard Variable',system-ui,sans-serif;background:var(--white);color:var(--dark);font-size:14px;line-height:1.6;max-width:1200px;margin:0 auto;padding:40px 32px;}
h1{font-size:28px;font-weight:800;margin-bottom:4px;} h2{font-size:20px;font-weight:800;margin:32px 0 12px;padding-bottom:8px;border-bottom:2px solid var(--primary-pale);} h3{font-size:15px;font-weight:700;margin:14px 0 8px;}
.sub{font-size:13px;color:var(--body);margin-bottom:20px;}
.card{background:var(--light);border:1px solid #e8e8e8;border-radius:12px;padding:16px 18px;}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px;} .grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;} .grid4{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;}
table{width:100%;border-collapse:collapse;font-size:12px;margin:10px 0;}
th{background:var(--primary);color:#fff;padding:6px 10px;text-align:left;font-size:11px;} td{padding:5px 10px;border-bottom:1px solid #eee;}
tr:nth-child(even){background:#fafbff;}
.insight{padding:14px 18px;background:var(--primary-pale);border-radius:10px;border-left:4px solid var(--primary);margin:10px 0;font-size:13px;line-height:1.7;} .insight strong{color:var(--primary);}
.insight-warn{background:#fef2f2;border-left-color:var(--coral);} .insight-warn strong{color:var(--coral);}
.ba-container{display:grid;grid-template-columns:1fr 50px 1fr;gap:0;align-items:stretch;margin:16px 0;}
.ba-box{padding:20px;border-radius:14px;} .ba-before{background:#f1f5f9;border:2px solid #cbd5e1;} .ba-after{background:#eef2ff;border:2px solid var(--primary);}
.ba-vs{display:flex;align-items:center;justify-content:center;font-size:24px;font-weight:900;color:#94a3b8;}
.ba-label{font-size:12px;font-weight:700;letter-spacing:1px;margin-bottom:10px;} .ba-title{font-size:16px;font-weight:800;margin-bottom:12px;}
.gauge-row{display:flex;align-items:center;gap:8px;margin-bottom:8px;}
.gauge-label{width:120px;font-size:12px;font-weight:600;text-align:right;flex-shrink:0;}
.gauge-track{flex:1;height:24px;background:#f0f0f0;border-radius:5px;overflow:hidden;}
.gauge-fill{height:100%;border-radius:5px;display:flex;align-items:center;padding-left:8px;font-size:10px;font-weight:700;color:#fff;}
.gauge-val{width:45px;font-family:'DIN Alternate','Poppins',sans-serif;font-size:15px;font-weight:700;text-align:center;flex-shrink:0;}
.rawbox{padding:12px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;margin-top:10px;}
.concl{padding:12px 16px;background:#eef2ff;border-left:4px solid var(--primary);border-radius:8px;margin-top:10px;font-size:13px;line-height:1.7;color:var(--dark);}
.concl strong{color:var(--primary);}
.footer{margin-top:40px;padding-top:16px;border-top:2px solid var(--primary);font-size:11px;color:#94a3b8;text-align:center;}
@media print{
  @page{size:A4 portrait;margin:14mm 12mm;}
  body{max-width:100%!important;padding:0!important;margin:0!important;}
  .pdf-panel{display:none!important;}
  .card,.ba-container,.grid2,.grid3,.grid4,.insight,.gauge-row,table{break-inside:avoid!important;page-break-inside:avoid!important;}
  h2,h3{break-after:avoid;page-break-after:avoid;}
  h1{font-size:22px!important;} h2{font-size:16px!important;margin-top:20px!important;}
  .card{padding:10px 12px!important;font-size:10px!important;} table{font-size:10px!important;} th,td{padding:4px 6px!important;}
  .ba-container{gap:0!important;} .ba-box{padding:14px!important;font-size:11px!important;}
}
.pdf-panel{position:fixed;bottom:24px;right:24px;z-index:1000;}
.pdf-panel-toggle{display:flex;align-items:center;gap:8px;padding:12px 22px;background:var(--grad);color:#fff;border-radius:50px;font-size:14px;font-weight:600;cursor:pointer;box-shadow:0 4px 20px rgba(102,102,255,0.4);border:none;font-family:'Pretendard Variable',sans-serif;}
.pdf-panel-body{display:none;position:absolute;bottom:56px;right:0;width:220px;background:#fff;border-radius:16px;box-shadow:0 8px 32px rgba(0,0,0,0.18);padding:16px;}
.pdf-panel-body.open{display:block;}
.pdf-orient-group{display:flex;gap:8px;margin-bottom:14px;}
.pdf-orient-btn{flex:1;padding:12px 8px;background:#f9fafb;border:2px solid #e2e8f0;border-radius:12px;cursor:pointer;font-family:'Pretendard Variable',sans-serif;color:#475569;font-size:11px;font-weight:600;text-align:center;}
.pdf-orient-btn.active{border-color:var(--primary);background:#eef2ff;color:var(--primary);}
.pdf-print-btn{width:100%;padding:10px;background:var(--grad);color:#fff;border:none;border-radius:10px;font-size:13px;font-weight:600;cursor:pointer;font-family:'Pretendard Variable',sans-serif;}"""


def make_date_bars():
    dates = Counter(p['date'] for p in posts)
    ordered = sorted(dates.items())
    mx = max(dates.values())
    out = ""
    for raw_d, cnt in ordered:
        y, m, d = raw_d[:4], raw_d[4:6], raw_d[6:]
        label = f"{m}/{d}"
        pct = cnt / mx * 100
        # 이벤트 전: ~20260113, 기간: 0114-0122, 이후: 0123~
        if raw_d <= '20260113':
            color = 'var(--orange)'
            note = '이벤트 전' if raw_d == '20251231' else ''
        elif raw_d <= '20260122':
            color = 'var(--primary)'
            if raw_d == '20260114':
                note = '★ 오픈'
            elif raw_d == '20260122':
                note = '★ 종료'
            elif cnt == mx:
                note = '★ 피크'
            else:
                note = ''
        else:
            color = '#94a3b8'
            note = '팝업 후' if raw_d == '20260123' else ''
        ns = f'color:{color};font-weight:700;' if note else ''
        bt = str(cnt) if cnt >= 10 else ''
        out += (f'<div style="display:flex;align-items:center;gap:5px;margin-bottom:2px;">'
                f'<div style="width:40px;font-size:9px;color:var(--body);text-align:right;">{label}</div>'
                f'<div style="flex:1;height:15px;background:#f0f0f0;border-radius:3px;overflow:hidden;">'
                f'<div style="width:{pct:.0f}%;height:100%;background:{color};border-radius:3px;display:flex;align-items:center;padding-left:4px;font-size:8px;font-weight:600;color:#fff;">{bt}</div></div>'
                f'<div style="width:20px;font-size:9px;font-weight:600;">{cnt}</div>'
                f'<div style="width:70px;font-size:8px;{ns}">{note}</div></div>\n')
        if raw_d == '20260113':
            out += '<div style="height:3px;border-bottom:2px solid var(--primary);margin:2px 0;"></div>\n'
        elif raw_d == '20260122':
            out += '<div style="height:3px;border-bottom:2px dashed #ddd;margin:2px 0;"></div>\n'
    return out


def make_spon_details():
    top_spon = sorted(sponsored, key=lambda x: -x['authenticity'])[:12]
    out = ""
    for d in top_spon:
        a = d['authenticity']
        ac = '#059669' if a >= 55 else '#f59e0b' if a >= 30 else '#dc2626'
        bg = '#ecfdf5' if a >= 55 else '#fef3c7' if a >= 30 else '#fee2e2'
        title = d['title'][:42].replace('&', '&amp;').replace('<', '&lt;')
        dt = f"{d['date'][4:6]}/{d['date'][6:]}"
        out += (f'<div style="padding:5px 8px;background:{bg};border-radius:6px;margin-bottom:3px;font-size:11px;">'
                f'<strong style="color:{ac};">Auth {a}</strong> — "{title}" ({dt}) · 느낌표 {d.get("exclamation",0)}개 · Clout {d["clout"]}</div>\n')
    return out


def make_appendix():
    # Sort gate3_posts by auth desc
    g3 = sorted(gate3_posts, key=lambda x: -x['authenticity'])
    out = ""
    for i, d in enumerate(g3):
        title = d['title'][:55].replace('&', '&amp;').replace('<', '&lt;')
        df = f"{d['date'][:4]}-{d['date'][4:6]}-{d['date'][6:]}"
        sent = d.get('sentiment', '')
        sc = {'긍정': '#059669', '부정': '#dc2626', '혼합': '#6666FF', '중립': '#94a3b8'}.get(sent, '#94a3b8')
        ac = '#059669' if d['authenticity'] >= 75 else '#6666FF'
        bg = '#f0fdf4' if i % 2 == 0 else '#fff'
        blogger = d.get('blogger', '')[:14].replace('&', '&amp;')
        period = period_of(d['date'])
        rql = d.get('rql', '').replace('_', ' ')
        out += (f'<tr style="background:{bg};"><td style="padding:3px 4px;text-align:center;">{i+1}</td>'
                f'<td style="padding:3px 4px;">{df}</td>'
                f'<td style="padding:3px 4px;color:{sc};font-weight:600;">{sent}</td>'
                f'<td style="padding:3px 4px;text-align:center;"><strong style="color:{ac};">{d["authenticity"]}</strong></td>'
                f'<td style="padding:3px 4px;text-align:center;">{d["clout"]}</td>'
                f'<td style="padding:3px 4px;text-align:center;">{d.get("freshness_index",0)}</td>'
                f'<td style="padding:3px 4px;">{rql}</td>'
                f'<td style="padding:3px 4px;">{d.get("primary_topic","")}</td>'
                f'<td style="padding:3px 4px;">{period}</td>'
                f'<td style="padding:3px 4px;"><a href="{d.get("link","")}" target="_blank" style="color:#6666FF;text-decoration:none;">{title}</a></td>'
                f'<td style="padding:3px 4px;color:#94a3b8;">{blogger}</td></tr>\n')
    return out


# 토픽 비중
topic_total = sum(TOPIC.values())
def topic_row(name, count, color):
    pct = count / topic_total * 100
    return (f'<div style="display:flex;align-items:center;gap:5px;margin-bottom:3px;">'
            f'<div style="width:110px;font-size:10px;text-align:right;color:{color};font-weight:700;">{name} {pct:.0f}%</div>'
            f'<div style="flex:1;height:16px;background:#f0f0f0;border-radius:3px;overflow:hidden;">'
            f'<div style="width:{pct:.1f}%;height:100%;background:{color};border-radius:3px;display:flex;align-items:center;padding-left:4px;font-size:8px;color:#fff;font-weight:600;">{count}건</div></div></div>')

topic_bars = ""
topic_colors = ['var(--primary)', 'var(--green)', 'var(--pink)', 'var(--blue)', 'var(--orange)', '#8b5cf6', '#94a3b8', '#ec4899']
topic_order = ['헬로키티IP', '굿즈/제품', '지수/팬덤', '공간/분위기', '웨이팅/예약', '컨셉/스토리', '크림/CJ', '가격/혜택']
for i, tname in enumerate(topic_order):
    if tname in TOPIC:
        topic_bars += topic_row(tname, TOPIC[tname], topic_colors[i % len(topic_colors)])

# RQL 바
rql_total = sum(RQL.values())
def rql_row(label, key, color):
    cnt = RQL.get(key, 0)
    pct = cnt / rql_total * 100
    show = f'{cnt}건' if cnt >= 30 else ''
    return (f'<div style="display:flex;align-items:center;gap:5px;margin-bottom:3px;">'
            f'<div style="width:100px;font-size:10px;text-align:right;color:{color};font-weight:700;">{label}</div>'
            f'<div style="flex:1;height:16px;background:#f0f0f0;border-radius:3px;overflow:hidden;">'
            f'<div style="width:{pct:.1f}%;height:100%;background:{color};border-radius:3px;display:flex;align-items:center;padding-left:4px;font-size:8px;color:#fff;font-weight:600;">{show}</div></div>'
            f'<div style="width:30px;font-size:9px;">{cnt}</div></div>')

rql_bars = (rql_row('Q5 서사형 ×2.0', 'Q5_서사형', 'var(--primary)')
            + rql_row('Q4 분석형 ×1.5', 'Q4_분석형', 'var(--blue)')
            + rql_row('Q3 경험형 ×1.0', 'Q3_경험형', 'var(--green)')
            + rql_row('Q2 감상형 ×0.5', 'Q2_감상형', 'var(--orange)')
            + rql_row('Q1 단문형 ×0.2', 'Q1_단문형', '#94a3b8'))

# Class 분포
CLASS_LABELS = {
    'A': ('자발적 상세', 'var(--green)'),
    'B': ('자발적 일반', 'var(--primary)'),
    'C': ('협찬 표기', 'var(--orange)'),
    'D': ('의심 자발', '#94a3b8'),
    'E': ('리그램', 'var(--coral)'),
    'F': ('단문 저품질', 'var(--coral)'),
    'G': ('광고성', 'var(--red)'),
}
class_html = ""
for k in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    cnt = class_dist.get(k, 0)
    label, color = CLASS_LABELS[k]
    pct = cnt / TOTAL * 100
    class_html += (f'<div style="display:flex;align-items:center;gap:5px;margin-bottom:3px;">'
                   f'<div style="width:30px;font-size:12px;font-weight:800;color:{color};text-align:center;">{k}</div>'
                   f'<div style="width:90px;font-size:10px;">{label}</div>'
                   f'<div style="flex:1;height:16px;background:#f0f0f0;border-radius:3px;overflow:hidden;">'
                   f'<div style="width:{pct:.1f}%;height:100%;background:{color};border-radius:3px;display:flex;align-items:center;padding-left:4px;font-size:8px;color:#fff;font-weight:600;">{cnt}건</div></div>'
                   f'<div style="width:40px;font-size:10px;text-align:right;font-weight:700;">{pct:.1f}%</div></div>')

# Gate1->3 긍정률 하락
GATE13_DELTA = G3['positive_pct'] - G1['positive_pct']

html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>헬로키티×지수 팝업스토어 — RXR SNS 분석 (내부용)</title>
<style>{CSS}</style><style id="pageOrientStyle"></style></head><body>

<div style="display:flex;align-items:center;gap:12px;margin-bottom:4px;">
  <span style="background:var(--grad);color:#fff;padding:4px 14px;border-radius:6px;font-family:'Poppins';font-size:11px;font-weight:600;letter-spacing:1px;">RXR SNS ANALYSIS · INTERNAL</span>
</div>
<h1>헬로키티×지수 팝업스토어 — 교환일기</h1>
<div class="sub">KREAM 도산 플래그십 · 2026.01.14~01.22 · CJ온스타일×KREAM | 네이버 블로그 {TOTAL}건 | Project RENT · R-lab</div>

<!-- 리서치 개요 -->
<div style="padding:20px;background:var(--light);border-radius:14px;margin-bottom:24px;">
  <div class="grid2">
    <div>
      <div style="font-size:11px;font-family:'Poppins';letter-spacing:2px;color:var(--body);margin-bottom:6px;">ABOUT THIS RESEARCH</div>
      <h3 style="margin-top:0;">분석 대상</h3>
      <table style="font-size:12px;">
        <tr><td style="color:var(--body);width:70px;padding:4px 0;">이벤트</td><td style="padding:4px 0;font-weight:700;">헬로키티×지수 팝업스토어 (교환일기)</td></tr>
        <tr><td style="color:var(--body);padding:4px 0;border-top:1px solid #e8e8e8;">기간</td><td style="padding:4px 0;border-top:1px solid #e8e8e8;">2026.01.14(수) ~ 01.22(목), <strong>9일간</strong></td></tr>
        <tr><td style="color:var(--body);padding:4px 0;border-top:1px solid #e8e8e8;">장소</td><td style="padding:4px 0;border-top:1px solid #e8e8e8;">KREAM 도산 플래그십 스토어 (강남구 신사동)</td></tr>
        <tr><td style="color:var(--body);padding:4px 0;border-top:1px solid #e8e8e8;">주최</td><td style="padding:4px 0;border-top:1px solid #e8e8e8;">CJ온스타일 × KREAM</td></tr>
        <tr><td style="color:var(--body);padding:4px 0;border-top:1px solid #e8e8e8;">컨셉</td><td style="padding:4px 0;border-top:1px solid #e8e8e8;">"헬로키티와 지수의 교환일기" — 지수 친필 캐릭터 '슈몬' + 오더시트 텍스트힙</td></tr>
        <tr><td style="color:var(--body);padding:4px 0;border-top:1px solid #e8e8e8;">굿즈</td><td style="padding:4px 0;border-top:1px solid #e8e8e8;">머그컵·텀블러·가방·인형·키링 (오픈 행사 지수 참석, 최대 5시간 대기)</td></tr>
        <tr><td style="color:var(--body);padding:4px 0;border-top:1px solid #e8e8e8;">분석 기간</td><td style="padding:4px 0;border-top:1px solid #e8e8e8;">2025.12.31 ~ 2026.02.19 (전 2주 + 중 9일 + 후 4주)</td></tr>
      </table>
    </div>
    <div>
      <div style="font-size:11px;font-family:'Poppins';letter-spacing:2px;color:var(--body);margin-bottom:6px;">RESEARCH PERSPECTIVE</div>
      <h3 style="margin-top:0;">리서치 관점</h3>
      <div style="font-size:12px;color:var(--body);line-height:1.7;">
        <div style="padding:4px 0;"><strong style="color:var(--primary);">핵심 질문:</strong> 헬로키티 IP의 화력이 '지수×교환일기' 고유 컨셉의 침투를 압도했는가?</div>
        <div style="padding:4px 0;border-top:1px solid #e8e8e8;"><strong style="color:var(--primary);">데이터:</strong> 네이버 블로그 {TOTAL}건 (Sincerity Filter 통과)</div>
        <div style="padding:4px 0;border-top:1px solid #e8e8e8;"><strong style="color:var(--primary);">분석:</strong> Content Class A~G + Psyche(Auth/Clout/Fresh) + Sincerity Gate 3단계</div>
        <div style="padding:4px 0;border-top:1px solid #e8e8e8;"><strong style="color:var(--primary);">비교:</strong> 새로중앙박물관(비IP형) 대비 IP 파워 차이 관찰</div>
      </div>
    </div>
  </div>
  <div style="margin-top:12px;padding:10px 14px;background:var(--primary-pale);border-radius:8px;font-size:12px;color:var(--body);">
    📖 <strong>읽는 법:</strong> 각 섹션은 <strong>BEFORE(기존 도구)</strong> vs <strong>AFTER(RXR)</strong> 비교입니다. AFTER에는 📊 Raw Data 근거 박스 + 결론 해석 박스가 포함됩니다.
  </div>
</div>

<!-- 0. 데이터 수집 개요 -->
<h2>0. 데이터 수집 개요</h2>
<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin:12px 0;">
  <div class="card" style="text-align:center;border-top:3px solid var(--primary);padding:12px;"><div style="font-size:24px;font-weight:800;color:var(--primary);">{TOTAL}</div><div style="font-size:10px;color:var(--body);">네이버 블로그</div></div>
  <div class="card" style="text-align:center;border-top:3px solid var(--green);padding:12px;"><div style="font-size:24px;font-weight:800;color:var(--green);">{len(during)}</div><div style="font-size:10px;color:var(--body);">팝업 기간</div></div>
  <div class="card" style="text-align:center;border-top:3px solid var(--orange);padding:12px;"><div style="font-size:24px;font-weight:800;color:var(--orange);">{len(pre)}</div><div style="font-size:10px;color:var(--body);">이벤트 전</div></div>
  <div class="card" style="text-align:center;border-top:3px solid var(--pink);padding:12px;"><div style="font-size:24px;font-weight:800;color:var(--pink);">{len(post_ev)}</div><div style="font-size:10px;color:var(--body);">팝업 후</div></div>
  <div class="card" style="text-align:center;border-top:3px solid var(--blue);padding:12px;"><div style="font-size:24px;font-weight:800;color:var(--blue);">9일</div><div style="font-size:10px;color:var(--body);">운영 기간</div></div>
</div>
<h3>네이버 블로그 분해</h3>
<div class="grid4" style="margin:10px 0;">
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:10px;color:var(--body);">Class A 자발적 상세</div><div style="font-size:20px;font-weight:800;color:var(--green);">{class_dist.get('A',0)}</div></div>
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:10px;color:var(--body);">Class B 자발적 일반</div><div style="font-size:20px;font-weight:800;color:var(--primary);">{class_dist.get('B',0)}</div></div>
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:10px;color:var(--body);">Class C 협찬</div><div style="font-size:20px;font-weight:800;color:var(--orange);">{class_dist.get('C',0)}</div></div>
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:10px;color:var(--body);">Class D~G 저신뢰</div><div style="font-size:20px;font-weight:800;color:var(--coral);">{class_dist.get('D',0)+class_dist.get('E',0)+class_dist.get('F',0)+class_dist.get('G',0)}</div></div>
</div>

<h3>날짜별 버즈 분포</h3>
{make_date_bars()}

<div class="grid2" style="margin-top:12px;">
  <div><h3>기간별 비교</h3><table><thead><tr><th>기간</th><th>건수</th><th>일평균</th><th>비율</th></tr></thead><tbody>
    <tr><td style="color:var(--orange);font-weight:700;">이벤트 전 (2주)</td><td><strong>{len(pre)}</strong></td><td>{len(pre)/14:.1f}</td><td>{len(pre)/TOTAL*100:.0f}%</td></tr>
    <tr style="background:#eef2ff;"><td style="color:var(--primary);font-weight:700;">팝업 기간 (9일)</td><td><strong>{len(during)}</strong></td><td>{len(during)/9:.1f}</td><td>{len(during)/TOTAL*100:.0f}%</td></tr>
    <tr><td style="color:#94a3b8;">팝업 후 (4주)</td><td><strong>{len(post_ev)}</strong></td><td>{len(post_ev)/28:.1f}</td><td>{len(post_ev)/TOTAL*100:.0f}%</td></tr>
    <tr style="background:var(--primary-pale);"><td style="font-weight:700;">합계</td><td style="font-weight:700;color:var(--primary);">{TOTAL}</td><td></td><td>100%</td></tr>
  </tbody></table></div>
  <div><h3>핵심 패턴</h3><div style="font-size:12px;line-height:1.7;">
    <div style="padding:5px 8px;background:var(--light);border-radius:6px;margin-bottom:4px;border-left:3px solid var(--primary);"><strong style="color:var(--primary);">팝업 후 {len(post_ev)/TOTAL*100:.0f}% 최다</strong> — 종료 후 리셀·후기 유입이 본격화 (전형적 IP 팝업 패턴)</div>
    <div style="padding:5px 8px;background:var(--light);border-radius:6px;margin-bottom:4px;border-left:3px solid var(--orange);"><strong style="color:var(--orange);">사전 버즈 {len(pre)/TOTAL*100:.0f}%</strong> — 지수 티징 + 오픈 임박 이슈</div>
    <div style="padding:5px 8px;background:var(--light);border-radius:6px;border-left:3px solid var(--green);"><strong style="color:var(--green);">01/15 피크</strong> — 오픈 직후 첫 주말 방문·후기 폭발</div>
  </div></div>
</div>

<!-- 1. 버즈 규모 -->
<h2>1. 버즈 규모 — "총량"에서 "진심 유효량"으로</h2>
<div class="ba-container">
  <div class="ba-box ba-before">
    <div class="ba-label" style="color:#94a3b8;">BEFORE — 기존 도구</div>
    <div class="ba-title" style="color:#475569;">단순 카운팅</div>
    <div style="text-align:center;margin:16px 0;"><div style="font-size:48px;font-weight:800;color:#475569;">{TOTAL}건</div><div style="font-size:13px;color:#94a3b8;">Total Blog Posts</div></div>
    <div style="padding:8px;background:#e2e8f0;border-radius:6px;font-size:12px;color:#475569;text-align:center;">"{TOTAL}건 후기입니다. 대박이네요."<br><strong>그 중 진짜는 몇 건?</strong></div>
  </div>
  <div class="ba-vs">→</div>
  <div class="ba-box ba-after">
    <div class="ba-label" style="color:var(--primary);">AFTER — RXR Sincerity Gate</div>
    <div class="ba-title" style="color:var(--primary);">진심 유효량 {G3['n']}건</div>
    <div style="display:flex;gap:6px;margin:10px 0;">
      <div style="flex:1;text-align:center;padding:8px;background:#fff;border-radius:8px;"><div style="font-size:18px;font-weight:800;">{G1['n']}</div><div style="font-size:9px;">G1 전체</div></div>
      <div style="flex:1;text-align:center;padding:8px;background:#fff;border-radius:8px;"><div style="font-size:18px;font-weight:800;color:var(--blue);">{G2['n']}</div><div style="font-size:9px;">G2 자발적</div></div>
      <div style="flex:1;text-align:center;padding:8px;background:#eef2ff;border-radius:8px;border:2px solid var(--primary);"><div style="font-size:18px;font-weight:800;color:var(--primary);">{G3['n']}</div><div style="font-size:9px;">G3 진심</div></div>
    </div>
    <div style="padding:8px;background:#eef2ff;border-radius:6px;font-size:12px;color:var(--primary);">→ "{TOTAL}건 중 <strong>Gate 3 통과 {G3['n']}건 ({G3['n']/TOTAL*100:.0f}%)</strong>"<br>→ "<strong>과장·저신뢰 {TOTAL-G3['n']}건 ({(1-G3['n']/TOTAL)*100:.0f}%)</strong>이 거품"</div>
  </div>
</div>
<div class="rawbox">
  <div style="font-size:12px;font-weight:700;color:var(--primary);margin-bottom:8px;">📊 왜 이런 수치? — Raw Data 근거</div>
  <div class="grid2">
    <div style="padding:8px;background:#ecfdf5;border-radius:8px;border-left:3px solid var(--green);"><div style="font-size:11px;font-weight:700;color:var(--green);">✅ Gate3 유효 {G3['n']}건</div><div style="font-size:10px;color:var(--body);line-height:1.5;">자발적 + Auth 60+ 필터. 구체적 체험·사진·개인 감상을 담은 "진짜 목소리"</div></div>
    <div style="padding:8px;background:#fef2f2;border-radius:8px;border-left:3px solid var(--coral);"><div style="font-size:11px;font-weight:700;color:var(--coral);">⚠️ 탈락 {TOTAL-G3['n']}건</div><div style="font-size:10px;color:var(--body);line-height:1.5;">협찬 {SPON_N}건 + Auth 60미만 {G2['n']-G3['n']}건 + 저신뢰 클래스 {TOTAL-G1['n']}건</div></div>
  </div>
</div>
<div class="concl"><strong>왜 → 뭘 의미 → 차별화:</strong> IP 팝업 특성상 "감성 사진 한장+이모지" 유형이 많아 Gate2→3에서 {G2['n']-G3['n']}건이 탈락. 의미: 공개 버즈의 <strong>1/3은 화제성에 기여하지만 구매 의사결정 근거로는 약함</strong>. 차별화: 기존 도구는 {TOTAL}건으로만 자랑 가능, RXR은 <strong>"진심 {G3['n']}건"</strong>으로 퀄리티 KPI를 제시한다.</div>

<!-- 2. 감성 4단계 -->
<h2>2. 감성 분석 — "좋았나요?"에서 "진심이었나요?"로</h2>
<div class="ba-container">
  <div class="ba-box ba-before">
    <div class="ba-label" style="color:#94a3b8;">BEFORE</div>
    <div class="ba-title" style="color:#475569;">감성 3단계</div>
    <div style="display:flex;gap:8px;margin:12px 0;">
      <div style="flex:1;text-align:center;padding:10px;background:#fff;border-radius:8px;"><div style="font-size:22px;font-weight:800;color:var(--green);">{SENT_POS+SENT_MIX:.0f}%</div><div style="font-size:10px;">긍정(혼합 포함)</div></div>
      <div style="flex:1;text-align:center;padding:10px;background:#fff;border-radius:8px;"><div style="font-size:22px;font-weight:800;color:#94a3b8;">{SENT_NEU:.0f}%</div><div style="font-size:10px;">중립</div></div>
      <div style="flex:1;text-align:center;padding:10px;background:#fff;border-radius:8px;"><div style="font-size:22px;font-weight:800;color:var(--red);">{SENT_NEG:.0f}%</div><div style="font-size:10px;">부정</div></div>
    </div>
    <div style="padding:8px;background:#e2e8f0;border-radius:6px;font-size:12px;color:#475569;text-align:center;">"긍정 절반 이상 → 성공!"<br><strong>그런데 Auth는?</strong></div>
  </div>
  <div class="ba-vs">→</div>
  <div class="ba-box ba-after">
    <div class="ba-label" style="color:var(--primary);">AFTER</div>
    <div class="ba-title" style="color:var(--primary);">감성 4단계 + Auth {AVG_AUTH}</div>
    <div style="display:flex;gap:6px;margin:10px 0;">
      <div style="flex:1;text-align:center;padding:8px;background:#fff;border-radius:8px;"><div style="font-size:18px;font-weight:800;color:var(--green);">{SENT_POS:.1f}%</div><div style="font-size:9px;">긍정</div></div>
      <div style="flex:1;text-align:center;padding:8px;background:#fff;border-radius:8px;"><div style="font-size:18px;font-weight:800;color:var(--primary);">{SENT_MIX:.1f}%</div><div style="font-size:9px;">혼합</div></div>
      <div style="flex:1;text-align:center;padding:8px;background:#fff;border-radius:8px;"><div style="font-size:18px;font-weight:800;color:#94a3b8;">{SENT_NEU:.1f}%</div><div style="font-size:9px;">중립</div></div>
      <div style="flex:1;text-align:center;padding:8px;background:#fff;border-radius:8px;"><div style="font-size:18px;font-weight:800;color:var(--red);">{SENT_NEG:.1f}%</div><div style="font-size:9px;">부정</div></div>
    </div>
    <div style="padding:8px;background:#eef2ff;border-radius:6px;font-size:12px;color:var(--primary);">→ "순수 긍정 <strong>{SENT_POS:.0f}%</strong>, Auth <strong>{AVG_AUTH}</strong> — 절반은 중립·혼합"<br>→ "<strong>혼합 {SENT_MIX:.0f}%</strong>: 좋았지만 5시간 웨이팅 언급"</div>
  </div>
</div>
<div class="rawbox">
  <div style="font-size:12px;font-weight:700;color:var(--primary);margin-bottom:8px;">📊 Raw Data 근거</div>
  <div class="grid3">
    <div style="padding:8px;background:#ecfdf5;border-radius:8px;border-left:3px solid var(--green);"><div style="font-size:11px;font-weight:700;color:var(--green);">긍정 {stats['sentiment_dist'].get('긍정',0)}건</div><div style="font-size:10px;color:var(--body);">"귀엽고 힐링됐다" + 사진 + 지수 애정</div></div>
    <div style="padding:8px;background:#eef2ff;border-radius:8px;border-left:3px solid var(--primary);"><div style="font-size:11px;font-weight:700;color:var(--primary);">혼합 {stats['sentiment_dist'].get('혼합',0)}건</div><div style="font-size:10px;color:var(--body);">"굿즈는 예쁜데 대기가 미쳤다" — 웨이팅·품절</div></div>
    <div style="padding:8px;background:#f1f5f9;border-radius:8px;border-left:3px solid #94a3b8;"><div style="font-size:11px;font-weight:700;color:#94a3b8;">중립 {stats['sentiment_dist'].get('중립',0)}건</div><div style="font-size:10px;color:var(--body);">정보 전달형 (장소·운영시간·예약법)</div></div>
  </div>
</div>
<div class="concl"><strong>왜 → 뭘 의미 → 차별화:</strong> 중립 {SENT_NEU:.0f}%가 긍정 {SENT_POS:.0f}%를 넘어선다 — 이는 <strong>"정보 공유형 포스트"가 방문기 절반</strong>임을 뜻한다. 의미: 실제 감정 회자는 40%뿐이며, 혼합 12%에서 웨이팅 피로가 노출된다. 차별화: 3단계 분류면 "긍정 52%!"로 포장 가능하지만, 4단계는 <strong>혼합을 분리해 불만 신호를 포착</strong>한다.</div>

<!-- 3. Content Class A~G -->
<h2>3. 데이터 분류 — "총량"에서 "A~G 7단계 신뢰 등급"으로</h2>
<div class="ba-container">
  <div class="ba-box ba-before">
    <div class="ba-label" style="color:#94a3b8;">BEFORE</div>
    <div class="ba-title" style="color:#475569;">단일 카운트</div>
    <div style="text-align:center;padding:18px;"><div style="font-size:44px;font-weight:800;color:#475569;">{TOTAL}건</div><div style="font-size:12px;color:#94a3b8;">"협찬? 그건 따로 빼세요"</div></div>
  </div>
  <div class="ba-vs">→</div>
  <div class="ba-box ba-after">
    <div class="ba-label" style="color:var(--primary);">AFTER</div>
    <div class="ba-title" style="color:var(--primary);">Class A~G 스펙트럼</div>
    {class_html}
    <div style="padding:8px;background:#eef2ff;border-radius:6px;font-size:12px;color:var(--primary);margin-top:8px;"><strong>"A+B 자발적 {class_dist.get('A',0)+class_dist.get('B',0)}건 ({(class_dist.get('A',0)+class_dist.get('B',0))/TOTAL*100:.0f}%) vs 협찬 {class_dist.get('C',0)}건 ({class_dist.get('C',0)/TOTAL*100:.0f}%)"</strong></div>
  </div>
</div>
<div class="concl"><strong>왜 → 뭘 의미 → 차별화:</strong> Class B가 {class_dist.get('B',0)}건({class_dist.get('B',0)/TOTAL*100:.0f}%)로 압도적 — IP 파워 덕에 <strong>자발적 "가벼운 후기"가 대량 발생</strong>했다는 뜻. 협찬 {class_dist.get('C',0)}건은 전체의 {class_dist.get('C',0)/TOTAL*100:.1f}%에 불과. 차별화: Class A(상세 서사)가 {class_dist.get('A',0)}건 있다는 사실은 "IP가 깊이 있는 후기를 유도할 수 있다"는 역학을 증명한다.</div>

<!-- 4. 협찬 Auth 스펙트럼 -->
<h2>4. 협찬 분석 — "있다/없다"에서 "진심의 농도"로</h2>
<div class="ba-container">
  <div class="ba-box ba-before">
    <div class="ba-label" style="color:#94a3b8;">BEFORE</div>
    <div class="ba-title" style="color:#475569;">협찬 표기 유무</div>
    <div style="text-align:center;padding:16px;"><div style="display:flex;justify-content:center;gap:16px;margin:10px 0;"><div style="padding:10px 20px;background:#fff;border-radius:8px;">✅ 협찬 {SPON_N}</div><div style="padding:10px 20px;background:#fff;border-radius:8px;">❌ 자발 {ORG_N}</div></div></div>
    <div style="padding:8px;background:#e2e8f0;border-radius:6px;font-size:12px;color:#475569;text-align:center;"><strong>협찬이면 제외? → 데이터 손실</strong></div>
  </div>
  <div class="ba-vs">→</div>
  <div class="ba-box ba-after">
    <div class="ba-label" style="color:var(--primary);">AFTER</div>
    <div class="ba-title" style="color:var(--primary);">진심 스펙트럼 (Auth 격차 {ORG_AUTH-SPON_AUTH:+.1f})</div>
    <div class="gauge-row" style="margin-bottom:6px;"><div style="width:90px;font-size:11px;">자발 ({ORG_N}건)</div><div class="gauge-track" style="height:20px;"><div class="gauge-fill" style="width:{ORG_AUTH:.1f}%;background:var(--green);height:100%;">Auth {ORG_AUTH}</div></div><div style="width:40px;font-size:13px;font-weight:700;color:var(--green);text-align:center;">{ORG_AUTH}</div></div>
    <div class="gauge-row" style="margin-bottom:6px;"><div style="width:90px;font-size:11px;">협찬 ({SPON_N}건)</div><div class="gauge-track" style="height:20px;"><div class="gauge-fill" style="width:{SPON_AUTH:.1f}%;background:var(--orange);height:100%;">Auth {SPON_AUTH}</div></div><div style="width:40px;font-size:13px;font-weight:700;color:var(--orange);text-align:center;">{SPON_AUTH}</div></div>
    <div style="padding:8px;background:#eef2ff;border-radius:6px;font-size:12px;color:var(--primary);margin-top:6px;"><strong>Auth 차이 {ORG_AUTH-SPON_AUTH:+.1f}점</strong> — 같은 "긍정"도 진심 농도가 다름</div>
  </div>
</div>
<div class="rawbox">
  <div style="font-size:12px;font-weight:700;color:var(--primary);margin-bottom:6px;">📊 협찬 {SPON_N}건 상위 Auth Raw Data</div>
  {make_spon_details()}
</div>
<div class="concl"><strong>왜 → 뭘 의미 → 차별화:</strong> 협찬 Auth {SPON_AUTH} vs 자발 Auth {ORG_AUTH} ({ORG_AUTH-SPON_AUTH:+.1f}점). 협찬 글은 느낌표·과잉 형용사가 많고 개인 서사가 약하다. 의미: <strong>"브랜드가 돈 주고 산 긍정"과 "소비자의 진심 긍정"은 수치로 분리 가능</strong>. 차별화: 기존 도구는 "협찬 n건"만 표시, RXR은 협찬 내부에서도 Auth 스펙트럼을 그린다.</div>

<!-- 5. 토픽 -->
<h2>5. 토픽 분석 — "워드클라우드"에서 "컨셉 침투율"로</h2>
<div class="ba-container">
  <div class="ba-box ba-before">
    <div class="ba-label" style="color:#94a3b8;">BEFORE</div>
    <div class="ba-title" style="color:#475569;">워드클라우드</div>
    <div style="text-align:center;padding:14px;background:#fff;border-radius:8px;margin:10px 0;"><span style="font-size:26px;font-weight:800;">헬로키티</span> <span style="font-size:20px;">지수</span> <span style="font-size:16px;color:#94a3b8;">팝업</span> <span style="font-size:18px;">도산</span> <span style="font-size:14px;color:#94a3b8;">KREAM</span> <span style="font-size:13px;color:#94a3b8;">굿즈</span></div>
    <div style="font-size:12px;color:#94a3b8;text-align:center;"><strong style="color:#475569;">교환일기·슈몬은 어디?</strong></div>
  </div>
  <div class="ba-vs">→</div>
  <div class="ba-box ba-after">
    <div class="ba-label" style="color:var(--primary);">AFTER</div>
    <div class="ba-title" style="color:var(--primary);">구조화된 토픽 비중</div>
    {topic_bars}
    <div style="padding:8px;background:#eef2ff;border-radius:6px;font-size:12px;color:var(--primary);margin-top:6px;"><strong>"헬로키티 IP {TOPIC.get('헬로키티IP',0)/topic_total*100:.0f}%가 절반 — 컨셉(교환일기+슈몬) {TOPIC.get('컨셉/스토리',0)/topic_total*100:.0f}%만 회자"</strong></div>
  </div>
</div>
<div class="concl insight-warn" style="background:#fef2f2;border-left-color:var(--coral);"><strong style="color:var(--coral);">왜 → 뭘 의미 → 차별화:</strong> <strong>컨셉(교환일기·슈몬)이 전체의 {TOPIC.get('컨셉/스토리',0)/topic_total*100:.0f}%</strong>밖에 안 된다 — 브랜드 IP(헬로키티)에 완전히 묻혔다. <strong>5시간 웨이팅</strong>이 언론에 회자됐지만 토픽 비중은 {TOPIC.get('웨이팅/예약',0)/topic_total*100:.0f}%뿐. 의미: 소비자는 "헬로키티 팝업"으로 기억할 뿐 "지수의 교환일기"로 각인되지 않았다. 차별화: 워드클라우드로는 "헬로키티·지수"만 크게 보이지만 RXR은 <strong>컨셉 침투율이 실패했음을 수치로 증명</strong>한다.</div>

<!-- 6. 심리 기간별 -->
<h2>6. 심리 분석 — 기존에 불가능했던 영역</h2>
<div style="text-align:center;margin:10px 0;padding:8px;background:#fef2f2;border-radius:8px;font-size:13px;font-weight:700;color:var(--coral);">BEFORE: 이 분석 자체가 불가능 — 어떤 기존 도구도 제공하지 않음</div>

<div class="gauge-row"><div class="gauge-label" style="color:var(--primary);">Authenticity<br><span style="font-size:10px;color:var(--body);">진정성</span></div><div class="gauge-track"><div class="gauge-fill" style="width:{AVG_AUTH}%;background:var(--primary);">{AVG_AUTH}</div></div><div class="gauge-val" style="color:var(--primary);">{AVG_AUTH}</div></div>
<div class="gauge-row"><div class="gauge-label" style="color:var(--green);">Clout<br><span style="font-size:10px;color:var(--body);">추천 확신도</span></div><div class="gauge-track"><div class="gauge-fill" style="width:{AVG_CLOUT}%;background:var(--green);">{AVG_CLOUT}</div></div><div class="gauge-val" style="color:var(--green);">{AVG_CLOUT}</div></div>
<div class="gauge-row"><div class="gauge-label" style="color:var(--orange);">Freshness<br><span style="font-size:10px;color:var(--body);">신선함</span></div><div class="gauge-track"><div class="gauge-fill" style="width:{AVG_FRESH}%;background:var(--orange);">{AVG_FRESH}</div></div><div class="gauge-val" style="color:var(--orange);">{AVG_FRESH}</div></div>

<h3>기간별 심리 변화 (Auth/Clout/Fresh)</h3>
<table><thead><tr><th>기간</th><th>N</th><th>Auth</th><th>Clout</th><th>Fresh</th><th>긍정률</th><th>해석</th></tr></thead><tbody>
  <tr><td style="color:var(--orange);font-weight:700;">이벤트 전</td><td>{PRE_STATS['n']}</td><td>{PRE_STATS['auth']}</td><td>{PRE_STATS['clout']}</td><td>{PRE_STATS['freshness']}</td><td>{PRE_STATS['positive_pct']:.1f}%</td><td style="font-size:11px;">지수 티징·기대감</td></tr>
  <tr style="background:#eef2ff;"><td style="color:var(--primary);font-weight:700;">팝업 기간</td><td>{DURING_STATS['n']}</td><td><strong style="color:var(--green);">{DURING_STATS['auth']}</strong></td><td>{DURING_STATS['clout']}</td><td>{DURING_STATS['freshness']}</td><td>{DURING_STATS['positive_pct']:.1f}%</td><td style="font-size:11px;">실체험 Auth 최고</td></tr>
  <tr><td style="color:#94a3b8;font-weight:700;">팝업 후</td><td>{POST_STATS['n']}</td><td><strong style="color:var(--coral);">{POST_STATS['auth']}</strong></td><td>{POST_STATS['clout']}</td><td>{POST_STATS['freshness']}</td><td>{POST_STATS['positive_pct']:.1f}%</td><td style="font-size:11px;color:var(--coral);font-weight:700;">Auth {POST_STATS['auth']-DURING_STATS['auth']:+.1f}점 하락</td></tr>
</tbody></table>
<div class="insight-warn insight"><strong>Auth 팝업중 {DURING_STATS['auth']} → 팝업후 {POST_STATS['auth']} ({POST_STATS['auth']-DURING_STATS['auth']:+.1f}점).</strong> 새로중앙박물관과 유사 패턴 — <strong>종료 후 리셀·중고거래·"나 다녀왔다" 회상글이 유입되며 진정성이 희석</strong>된다. 긍정률이 오히려 올라간({POST_STATS['positive_pct']:.0f}%) 것은 "후회 없는 추억"으로 미화된 탓.</div>
<div class="rawbox">
  <div style="font-size:12px;font-weight:700;color:var(--primary);margin-bottom:8px;">📊 기간별 차이 근거</div>
  <div class="grid3">
    <div style="padding:8px;background:#fff7ed;border-radius:8px;border-top:3px solid var(--orange);"><div style="font-size:11px;font-weight:700;color:var(--orange);">이벤트 전 (Auth {PRE_STATS['auth']})</div><div style="font-size:10px;color:var(--body);line-height:1.5;">"지수 팝업 가야겠다" 기대형·정보 전달형</div></div>
    <div style="padding:8px;background:#eef2ff;border-radius:8px;border-top:3px solid var(--primary);"><div style="font-size:11px;font-weight:700;color:var(--primary);">팝업 중 (Auth {DURING_STATS['auth']})</div><div style="font-size:10px;color:var(--body);line-height:1.5;">"5시간 대기했지만 귀여웠다" — 체감+균형</div></div>
    <div style="padding:8px;background:#f1f5f9;border-radius:8px;border-top:3px solid #94a3b8;"><div style="font-size:11px;font-weight:700;color:#94a3b8;">팝업 후 (Auth {POST_STATS['auth']})</div><div style="font-size:10px;color:var(--body);line-height:1.5;">리셀 구매 인증·회상 미화글 증가</div></div>
  </div>
</div>
<div class="concl"><strong>왜 → 뭘 의미 → 차별화:</strong> Auth는 팝업 중 최고점 → 종료 후 하락하는 "U자 역전" 구조. 의미: <strong>이벤트의 진정성 피크는 현장 체험 기간</strong>이며, 후속 버즈 195건 중 상당수는 "리셀 인증/회상 미화"다. 차별화: 건수만 보면 팝업 후가 최다라 "롱테일 성공"으로 오독 가능, RXR은 <strong>후속 버즈 품질 하락을 드러낸다</strong>.</div>

<!-- 7. RQL -->
<h2>7. 후기 품질 — "1건=1건"에서 "RQL 5단계"로</h2>
<div class="ba-container">
  <div class="ba-box ba-before">
    <div class="ba-label" style="color:#94a3b8;">BEFORE</div>
    <div class="ba-title" style="color:#475569;">단순 건수</div>
    <div style="text-align:center;padding:18px;"><div style="font-size:44px;font-weight:800;color:#475569;">{TOTAL}건</div><div style="font-size:12px;color:#94a3b8;">"귀여워요 ❤️" 1건 = "3000자 후기" 1건<br><strong style="color:#475569;">같은 1건</strong></div></div>
  </div>
  <div class="ba-vs">→</div>
  <div class="ba-box ba-after">
    <div class="ba-label" style="color:var(--primary);">AFTER</div>
    <div class="ba-title" style="color:var(--primary);">RQL 5단계 가중</div>
    {rql_bars}
    <div style="padding:8px;background:#eef2ff;border-radius:6px;font-size:12px;color:var(--primary);margin-top:6px;"><strong>"Q4+Q5 깊은 후기 {RQL.get('Q4_분석형',0)+RQL.get('Q5_서사형',0)}건 ({(RQL.get('Q4_분석형',0)+RQL.get('Q5_서사형',0))/rql_total*100:.0f}%) — IP 팝업답게 상세 서사 비율 높음"</strong></div>
  </div>
</div>
<div class="concl"><strong>왜 → 뭘 의미 → 차별화:</strong> Q5 서사형 {RQL.get('Q5_서사형',0)}건 + Q4 분석형 {RQL.get('Q4_분석형',0)}건. 이는 <strong>헬로키티 IP + 지수 조합이 블로거의 "긴 글 동기"를 자극</strong>했음을 의미. 의미: 단순 감탄형(Q1~Q2)이 {RQL.get('Q1_단문형',0)+RQL.get('Q2_감상형',0)}건으로 28%에 달해 거품도 공존. 차별화: 기존 도구는 "{TOTAL}건" 단일 숫자, RXR은 <strong>Q5=2.0 가중 적용시 유효 건수를 {sum((1 if k=='Q3_경험형' else 2.0 if k=='Q5_서사형' else 1.5 if k=='Q4_분석형' else 0.5 if k=='Q2_감상형' else 0.2)*v for k,v in RQL.items()):.0f}"건"으로 재계산</strong> 가능하다.</div>

<!-- 8. Sincerity Gate -->
<h2>8. Sincerity Gate — "{TOTAL}건"의 진짜 숫자</h2>
<div style="margin:14px 0;">
  <div style="padding:14px 18px;background:var(--light);border-radius:10px;margin-bottom:4px;display:flex;justify-content:space-between;align-items:center;border:2px solid #e8e8e8;width:100%;">
    <div><strong>Gate 1</strong> <span style="font-size:12px;">전체 수집</span></div>
    <div style="font-size:18px;font-weight:800;">{G1['n']}건 <span style="font-size:12px;color:var(--body);">긍정 {G1['positive_pct']:.1f}% · Auth {G1['auth']}</span></div>
  </div>
  <div style="text-align:center;font-size:10px;color:var(--coral);padding:2px;">▼ 협찬·저신뢰 제거 ({G1['n']-G2['n']}건)</div>
  <div style="padding:14px 18px;background:#eef2ff;border-radius:10px;margin-bottom:4px;display:flex;justify-content:space-between;align-items:center;border:2px solid var(--primary);width:92%;margin-left:auto;margin-right:auto;">
    <div><strong style="color:var(--primary);">Gate 2</strong> <span style="font-size:12px;">자발적</span></div>
    <div style="font-size:18px;font-weight:800;color:var(--primary);">{G2['n']}건 <span style="font-size:12px;">긍정 {G2['positive_pct']:.1f}% ({G2['positive_pct']-G1['positive_pct']:+.1f}%p) · Auth {G2['auth']}</span></div>
  </div>
  <div style="text-align:center;font-size:10px;color:var(--coral);padding:2px;">▼ Auth 60 미만 {G2['n']-G3['n']}건 탈락</div>
  <div style="padding:14px 18px;background:var(--primary);border-radius:10px;display:flex;justify-content:space-between;align-items:center;color:#fff;width:80%;margin-left:auto;margin-right:auto;">
    <div><strong>Gate 3</strong> <span style="font-size:12px;color:rgba(255,255,255,0.7);">진심</span></div>
    <div style="font-size:18px;font-weight:800;">{G3['n']}건 <span style="font-size:12px;color:rgba(255,255,255,0.7);">긍정 {G3['positive_pct']:.1f}% ({GATE13_DELTA:+.1f}%p) · Auth {G3['auth']}</span></div>
  </div>
</div>
<div class="insight"><strong>Gate1→3 긍정률 {G1['positive_pct']:.1f}% → {G3['positive_pct']:.1f}% ({GATE13_DELTA:+.1f}%p).</strong> 즉 "진심으로 좋았다"고 말하는 소비자는 전체의 약 1/3. <strong>{G3['positive_pct']:.0f}%가 실제 만족도의 현실적 수치</strong>이며, 기존 도구의 긍정률 {SENT_POS+SENT_MIX:.0f}%는 과대평가다.</div>

<!-- 9. 협찬 vs 자발 VS -->
<h2>9. 협찬 vs 자발 — 마지막 VS 비교</h2>
<div class="grid2" style="margin:14px 0;">
  <div class="card" style="border-top:4px solid var(--green);padding:18px;">
    <div style="font-size:12px;font-weight:700;color:var(--green);letter-spacing:1px;">ORGANIC</div>
    <div style="font-size:28px;font-weight:800;color:var(--green);margin:6px 0;">{ORG_N}건</div>
    <div class="gauge-row"><div style="width:50px;font-size:10px;">Auth</div><div class="gauge-track" style="height:18px;"><div class="gauge-fill" style="width:{ORG_AUTH}%;background:var(--green);">{ORG_AUTH}</div></div></div>
    <div style="font-size:12px;color:var(--body);margin-top:8px;">개인 블로거가 자기 돈 내고 방문하여 쓴 후기. 구체적 사진·개인 감상·균형 평가 비율↑</div>
  </div>
  <div class="card" style="border-top:4px solid var(--orange);padding:18px;">
    <div style="font-size:12px;font-weight:700;color:var(--orange);letter-spacing:1px;">SPONSORED</div>
    <div style="font-size:28px;font-weight:800;color:var(--orange);margin:6px 0;">{SPON_N}건</div>
    <div class="gauge-row"><div style="width:50px;font-size:10px;">Auth</div><div class="gauge-track" style="height:18px;"><div class="gauge-fill" style="width:{SPON_AUTH}%;background:var(--orange);">{SPON_AUTH}</div></div></div>
    <div style="font-size:12px;color:var(--body);margin-top:8px;">#협찬·#광고 표기. 느낌표·과잉 형용사 다수, 균형 평가 부족. Auth 격차 <strong>{ORG_AUTH-SPON_AUTH:+.1f}점</strong></div>
  </div>
</div>

<!-- 10. 종합 비교표 -->
<h2>10. 종합: Before → After 비교표</h2>
<table style="font-size:13px;">
  <thead><tr><th style="width:120px;">분석 영역</th><th style="background:#64748b;">BEFORE (기존)</th><th>AFTER (RXR)</th><th style="width:80px;">차별화</th></tr></thead>
  <tbody>
    <tr><td style="font-weight:700;">버즈 규모</td><td>총 {TOTAL}건</td><td><strong style="color:var(--primary);">Gate 3 진심 {G3['n']}건 ({TOTAL-G3['n']}건 거품)</strong></td><td style="color:var(--coral);font-weight:700;">진실</td></tr>
    <tr><td style="font-weight:700;">감성</td><td>긍/중/부 3단계</td><td><strong style="color:var(--primary);">4단계 + Auth {AVG_AUTH} + Clout {AVG_CLOUT}</strong></td><td style="color:var(--primary);font-weight:700;">깊이 ×3</td></tr>
    <tr><td style="font-weight:700;">데이터 분류</td><td>건수만</td><td><strong style="color:var(--primary);">A~G 7단계 신뢰등급</strong></td><td style="color:var(--primary);font-weight:700;">정확도</td></tr>
    <tr><td style="font-weight:700;">협찬</td><td>있다/없다</td><td><strong style="color:var(--primary);">진심 스펙트럼 (자발 {ORG_AUTH} vs 협찬 {SPON_AUTH})</strong></td><td style="color:var(--primary);font-weight:700;">유일</td></tr>
    <tr><td style="font-weight:700;">토픽</td><td>워드클라우드</td><td><strong style="color:var(--primary);">컨셉 침투율 — 교환일기 {TOPIC.get('컨셉/스토리',0)/topic_total*100:.0f}%</strong></td><td style="color:var(--primary);font-weight:700;">액셔너블</td></tr>
    <tr><td style="font-weight:700;">심리</td><td style="color:var(--coral);">불가능</td><td><strong style="color:var(--primary);">Auth/Clout/Fresh 기간별 시계열</strong></td><td style="color:var(--coral);font-weight:700;">유일</td></tr>
    <tr><td style="font-weight:700;">후기 품질</td><td>1건=1건</td><td><strong style="color:var(--primary);">RQL Q5 {RQL.get('Q5_서사형',0)}건 ×2.0 가중</strong></td><td style="color:var(--primary);font-weight:700;">유일</td></tr>
    <tr><td style="font-weight:700;">시계열</td><td>버즈량만</td><td><strong style="color:var(--primary);">Auth 팝업후 {POST_STATS['auth']-DURING_STATS['auth']:+.1f} 감지</strong></td><td style="color:var(--primary);font-weight:700;">유일</td></tr>
    <tr><td style="font-weight:700;">긍정률</td><td>{SENT_POS+SENT_MIX:.0f}% (포장 가능)</td><td><strong style="color:var(--primary);">Gate3 {G3['positive_pct']:.1f}% (현실 수치)</strong></td><td style="color:var(--coral);font-weight:700;">진실</td></tr>
  </tbody>
</table>

<!-- 11. BOTTOM LINE 확장형 -->
<h2>11. BOTTOM LINE</h2>
<div style="margin:18px 0;padding:22px;background:var(--primary-pale);border-radius:14px;">
  <div style="font-size:12px;color:var(--body);margin-bottom:10px;font-family:'Poppins';letter-spacing:2px;text-align:center;">BOTTOM LINE · 핵심 수치 대비</div>
  <div class="grid2" style="gap:14px;">
    <div style="padding:14px;background:#fff;border-radius:10px;border-left:4px solid #94a3b8;">
      <div style="font-size:11px;color:#94a3b8;font-weight:700;">기존 도구가 보는 그림</div>
      <div style="font-size:15px;font-weight:700;margin-top:6px;line-height:1.6;">"{TOTAL}건 · 긍정 {SENT_POS+SENT_MIX:.0f}%<br>헬로키티·지수 대박!"</div>
    </div>
    <div style="padding:14px;background:#fff;border-radius:10px;border-left:4px solid var(--primary);">
      <div style="font-size:11px;color:var(--primary);font-weight:700;">RXR이 보는 그림</div>
      <div style="font-size:15px;font-weight:700;margin-top:6px;line-height:1.6;color:var(--primary);">"진심 {G3['n']}건 · 진심 긍정 {G3['positive_pct']:.0f}%<br>컨셉 침투 {TOPIC.get('컨셉/스토리',0)/topic_total*100:.0f}% — IP에 묻힘"</div>
    </div>
  </div>
</div>

<h3>왜 이런 결과? — 4가지 이유</h3>
<div class="grid4" style="margin:10px 0;">
  <div class="card" style="border-top:3px solid var(--primary);padding:14px;"><div style="font-size:11px;color:var(--primary);font-weight:700;">① IP 파워 압도</div><div style="font-size:12px;margin-top:6px;line-height:1.5;">헬로키티 IP가 토픽 {TOPIC.get('헬로키티IP',0)/topic_total*100:.0f}%를 점유 — 지수·교환일기 컨셉은 자체 화력을 만들지 못하고 IP 하위 속성으로 소비됐다.</div></div>
  <div class="card" style="border-top:3px solid var(--orange);padding:14px;"><div style="font-size:11px;color:var(--orange);font-weight:700;">② 웨이팅 = 회자 ≠ 토픽</div><div style="font-size:12px;margin-top:6px;line-height:1.5;">5시간 대기가 언론에선 상징이 됐지만 블로그 토픽 비중은 {TOPIC.get('웨이팅/예약',0)/topic_total*100:.0f}%뿐. 대기가 포기·선망의 감정을 낳을 뿐 후기 소재가 되진 않는다.</div></div>
  <div class="card" style="border-top:3px solid var(--coral);padding:14px;"><div style="font-size:11px;color:var(--coral);font-weight:700;">③ 팝업 후 Auth 하락</div><div style="font-size:12px;margin-top:6px;line-height:1.5;">팝업 후 Auth {POST_STATS['auth']} ({POST_STATS['auth']-DURING_STATS['auth']:+.1f}). 리셀·중고·"나 다녀왔다" 회상 유입 추정 — 새로 팝업과 동일 패턴.</div></div>
  <div class="card" style="border-top:3px solid var(--green);padding:14px;"><div style="font-size:11px;color:var(--green);font-weight:700;">④ 협찬 {ORG_AUTH-SPON_AUTH:+.1f}점 격차</div><div style="font-size:12px;margin-top:6px;line-height:1.5;">협찬 Auth {SPON_AUTH} vs 자발 {ORG_AUTH}. 같은 "긍정"도 진심 농도가 다르다 — 기존 도구는 이를 숨긴다.</div></div>
</div>

<h3>제언 — 3가지 액션</h3>
<div class="grid3" style="margin:10px 0;">
  <div class="card" style="border-left:4px solid var(--primary);padding:14px;"><div style="font-size:11px;color:var(--primary);font-weight:700;">① 컨셉 화제화 리마케팅</div><div style="font-size:12px;margin-top:6px;line-height:1.5;">'교환일기·슈몬' 컨셉 자체를 별도 콘텐츠(지수 친필 인터뷰, 슈몬 제작기)로 분리 리마케팅하여 IP 하위 속성이 아닌 독립 화제로 끌어올린다.</div></div>
  <div class="card" style="border-left:4px solid var(--orange);padding:14px;"><div style="font-size:11px;color:var(--orange);font-weight:700;">② 웨이팅 경험화</div><div style="font-size:12px;margin-top:6px;line-height:1.5;">대기 자체를 후기 소재로 만들기. 대기줄 전용 인터랙션·인증 컨텐츠를 배치해 "5시간 대기했다"가 Q4~Q5 서사형 후기로 전환되도록.</div></div>
  <div class="card" style="border-left:4px solid var(--green);padding:14px;"><div style="font-size:11px;color:var(--green);font-weight:700;">③ 팝업 후 4주 관리</div><div style="font-size:12px;margin-top:6px;line-height:1.5;">종료 직후 Auth 하락 감지. 공식 후기 앵커 콘텐츠·"굿즈 사용 인증" 유도로 리셀·회상 미화 이탈을 방지하고 Auth를 재방어한다.</div></div>
</div>

<!-- 부록 -->
<h2>부록: Gate 3 통과 Raw Data ({G3['n']}건 중 상위 {len(gate3_posts)}건)</h2>
<p style="font-size:11px;color:var(--body);margin-bottom:6px;">Auth 60+, 자발적 포스트만. Authenticity 내림차순.</p>
<div style="margin-bottom:10px;display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:8px;">
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:9px;color:var(--body);">총 건수</div><div style="font-size:18px;font-weight:800;color:var(--primary);">{G3['n']}</div></div>
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:9px;color:var(--body);">평균 Auth</div><div style="font-size:18px;font-weight:800;color:var(--green);">{G3['auth']}</div></div>
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:9px;color:var(--body);">평균 Clout</div><div style="font-size:18px;font-weight:800;color:var(--blue);">{G3['clout']}</div></div>
  <div class="card" style="padding:10px;text-align:center;"><div style="font-size:9px;color:var(--body);">긍정률</div><div style="font-size:18px;font-weight:800;color:var(--orange);">{G3['positive_pct']:.1f}%</div></div>
</div>
<table style="font-size:10px;">
<thead><tr style="background:var(--primary);">
  <th style="padding:3px;color:#fff;text-align:center;width:20px;">#</th>
  <th style="padding:3px;color:#fff;width:60px;">날짜</th>
  <th style="padding:3px;color:#fff;width:30px;">감성</th>
  <th style="padding:3px;color:#fff;width:28px;">Auth</th>
  <th style="padding:3px;color:#fff;width:28px;">Clout</th>
  <th style="padding:3px;color:#fff;width:32px;">Fresh</th>
  <th style="padding:3px;color:#fff;width:56px;">RQL</th>
  <th style="padding:3px;color:#fff;width:62px;">토픽</th>
  <th style="padding:3px;color:#fff;width:48px;">기간</th>
  <th style="padding:3px;color:#fff;">제목</th>
  <th style="padding:3px;color:#fff;width:68px;">블로거</th>
</tr></thead>
<tbody>
{make_appendix()}
</tbody></table>

<div class="footer">Project RENT · R-lab · 2026 | RXR 2-Layer + Sincerity Gate | 헬로키티×지수 팝업스토어 · INTERNAL</div>

<!-- PDF Panel -->
<div class="pdf-panel" id="pdfPanel"><button class="pdf-panel-toggle" onclick="document.getElementById('pdfBody').classList.toggle('open')"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>PDF</button><div class="pdf-panel-body" id="pdfBody"><div style="font-size:12px;font-weight:700;color:#1e293b;margin-bottom:12px;text-align:center;">PDF 출력</div><div class="pdf-orient-group"><button class="pdf-orient-btn" id="btnL" onclick="setOrient('landscape')">가로</button><button class="pdf-orient-btn active" id="btnP" onclick="setOrient('portrait')">세로</button></div><button class="pdf-print-btn" onclick="window.print()">출력하기</button></div></div>
<script>document.addEventListener('click',function(e){{var p=document.getElementById('pdfPanel');if(p&&!p.contains(e.target))document.getElementById('pdfBody').classList.remove('open');}});function setOrient(o){{var s=document.getElementById('pageOrientStyle'),bL=document.getElementById('btnL'),bP=document.getElementById('btnP');if(o==='landscape'){{s.textContent='@media print{{@page{{size:A4 landscape;margin:8mm;}}}}';bL.classList.add('active');bP.classList.remove('active');}}else{{s.textContent='';bP.classList.add('active');bL.classList.remove('active');}}}}</script>
</body></html>"""

OUT_HTML = BASE / 'hellokitty-jisoo-popup-rxr-sns-report.html'
with open(OUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'HTML 생성 완료: {OUT_HTML}')
print(f'  크기: {OUT_HTML.stat().st_size:,} bytes')


# ========== Excel 생성 ==========
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

CLASS_FILL = {
    'A': PatternFill('solid', fgColor='C6EFCE'),
    'B': PatternFill('solid', fgColor='FFFFFF'),
    'C': PatternFill('solid', fgColor='FFF2CC'),
    'D': PatternFill('solid', fgColor='E7E6E6'),
    'E': PatternFill('solid', fgColor='FFC7CE'),
    'F': PatternFill('solid', fgColor='FFC7CE'),
    'G': PatternFill('solid', fgColor='FFC7CE'),
}
SENT_FILL = {
    '긍정': PatternFill('solid', fgColor='C6EFCE'),
    '부정': PatternFill('solid', fgColor='FFC7CE'),
    '혼합': PatternFill('solid', fgColor='DDEBF7'),
    '중립': PatternFill('solid', fgColor='E7E6E6'),
}
HDR_FILL = PatternFill('solid', fgColor='6666FF')
HDR_FONT = Font(color='FFFFFF', bold=True, name='맑은 고딕', size=10)
BODY_FONT = Font(name='맑은 고딕', size=10)

# Build raw dict by link (to get 써머리/본문)
raw_by_link = {r.get('link', ''): r for r in raw}


def write_summary_sheet(ws):
    ws.title = '요약 대시보드'
    ws.append(['헬로키티×지수 팝업스토어 — RXR SNS 분석 요약'])
    ws['A1'].font = Font(bold=True, size=16, color='6666FF', name='맑은 고딕')
    ws.append(['기간: 2026-01-14 ~ 01-22 (9일) · 장소: KREAM 도산 · 주최: CJ온스타일×KREAM'])
    ws.append([])

    ws.append(['■ 기간별 통계'])
    ws['A4'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['기간', '건수', '일평균', 'Auth', 'Clout', 'Freshness', '긍정률(%)'])
    for c in ws[5]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    ws.append(['이벤트 전 (12/31~01/13)', len(pre), round(len(pre)/14, 1), PRE_STATS['auth'], PRE_STATS['clout'], PRE_STATS['freshness'], PRE_STATS['positive_pct']])
    ws.append(['팝업 기간 (01/14~01/22)', len(during), round(len(during)/9, 1), DURING_STATS['auth'], DURING_STATS['clout'], DURING_STATS['freshness'], DURING_STATS['positive_pct']])
    ws.append(['팝업 후 (01/23~02/19)', len(post_ev), round(len(post_ev)/28, 1), POST_STATS['auth'], POST_STATS['clout'], POST_STATS['freshness'], POST_STATS['positive_pct']])
    ws.append(['합계', TOTAL, '', AVG_AUTH, AVG_CLOUT, AVG_FRESH, SENT_POS])
    ws[9][0].font = Font(bold=True)

    ws.append([])
    ws.append(['■ 감성 분포'])
    ws['A11'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['감성', '건수', '비율(%)'])
    for c in ws[12]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    for sent in ['긍정', '혼합', '중립', '부정']:
        ws.append([sent, stats['sentiment_dist'].get(sent, 0), sent_pct.get(sent, 0)])
        ws[ws.max_row][0].fill = SENT_FILL.get(sent, PatternFill())

    ws.append([])
    ws.append(['■ Content Class 분포'])
    r = ws.max_row
    ws[f'A{r}'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['클래스', '라벨', '건수', '비율(%)'])
    for c in ws[ws.max_row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    for k in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cnt = class_dist.get(k, 0)
        label = CLASS_LABELS[k][0]
        ws.append([k, label, cnt, round(cnt/TOTAL*100, 1)])
        ws[ws.max_row][0].fill = CLASS_FILL.get(k, PatternFill())

    ws.append([])
    ws.append(['■ 토픽 분포'])
    r = ws.max_row
    ws[f'A{r}'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['토픽', '건수', '비율(%)'])
    for c in ws[ws.max_row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    for tname, cnt in sorted(TOPIC.items(), key=lambda x: -x[1]):
        ws.append([tname, cnt, round(cnt/topic_total*100, 1)])

    ws.append([])
    ws.append(['■ RQL 분포'])
    r = ws.max_row
    ws[f'A{r}'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['RQL', '건수', '비율(%)', '가중치'])
    for c in ws[ws.max_row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    weights = {'Q5_서사형': 2.0, 'Q4_분석형': 1.5, 'Q3_경험형': 1.0, 'Q2_감상형': 0.5, 'Q1_단문형': 0.2}
    for k in ['Q5_서사형', 'Q4_분석형', 'Q3_경험형', 'Q2_감상형', 'Q1_단문형']:
        cnt = RQL.get(k, 0)
        ws.append([k.replace('_', ' '), cnt, round(cnt/rql_total*100, 1), weights[k]])

    ws.append([])
    ws.append(['■ 협찬 vs 자발 Auth 비교'])
    r = ws.max_row
    ws[f'A{r}'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['유형', '건수', 'Auth 평균'])
    for c in ws[ws.max_row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    ws.append(['자발 (Organic)', ORG_N, ORG_AUTH])
    ws.append(['협찬 (Sponsored)', SPON_N, SPON_AUTH])
    ws.append(['차이', '', round(ORG_AUTH - SPON_AUTH, 1)])

    ws.append([])
    ws.append(['■ Sincerity Gate'])
    r = ws.max_row
    ws[f'A{r}'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['Gate', '건수', '긍정률(%)', 'Auth', 'Clout', 'Freshness'])
    for c in ws[ws.max_row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    for name, g in [('Gate1 전체', G1), ('Gate2 자발', G2), ('Gate3 진심', G3)]:
        ws.append([name, g['n'], g['positive_pct'], g['auth'], g['clout'], g['freshness']])

    ws.append([])
    ws.append(['■ 날짜별 버즈'])
    r = ws.max_row
    ws[f'A{r}'].font = Font(bold=True, size=12, color='6666FF')
    ws.append(['날짜', '건수', '기간', '바'])
    for c in ws[ws.max_row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    dates_c = Counter(p['date'] for p in posts)
    mx = max(dates_c.values())
    for d in sorted(dates_c):
        cnt = dates_c[d]
        per = period_of(d)
        bar = '■' * int(cnt / mx * 30)
        ws.append([f'{d[:4]}-{d[4:6]}-{d[6:]}', cnt, per, bar])

    # Column widths
    for col, w in zip('ABCDEFG', [28, 10, 12, 10, 10, 12, 12]):
        ws.column_dimensions[col].width = w


def write_data_sheet(ws, title, data_list):
    ws.title = title
    headers = ['날짜', '기간', '유형', '감성', '긍정수', '부정수', '제목', '블로거', '써머리',
               '본문_앞400자', 'ContentClass', 'TrustScore', 'Auth', 'Clout', 'Freshness',
               'RQL', '주요토픽', 'URL']
    ws.append(headers)
    for c in ws[1]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    for p in data_list:
        link = p.get('link', '')
        raw_item = raw_by_link.get(link, {})
        summary = raw_item.get('description', '') or raw_item.get('summary', '')
        body = raw_item.get('content', '') or raw_item.get('body', '') or summary
        utype = '협찬' if p['is_sponsored'] else '자발'
        row = [
            f"{p['date'][:4]}-{p['date'][4:6]}-{p['date'][6:]}",
            period_of(p['date']),
            utype,
            p.get('sentiment', ''),
            p.get('pos_count', 0),
            p.get('neg_count', 0),
            p.get('title', ''),
            p.get('blogger', ''),
            (summary or '')[:200],
            (body or '')[:400],
            p.get('sincerity_class', ''),
            p.get('trust_score', 0),
            p.get('authenticity', 0),
            p.get('clout', 0),
            p.get('freshness_index', 0),
            p.get('rql', '').replace('_', ' '),
            p.get('primary_topic', ''),
            link,
        ]
        ws.append(row)
        r = ws.max_row
        # color class col
        cls = p.get('sincerity_class', 'B')
        ws.cell(r, 11).fill = CLASS_FILL.get(cls, PatternFill())
        # color sentiment
        ws.cell(r, 4).fill = SENT_FILL.get(p.get('sentiment', ''), PatternFill())
        for c in ws[r]:
            c.font = BODY_FONT
            c.alignment = Alignment(vertical='center', wrap_text=False)

    widths = [12, 10, 8, 8, 8, 8, 45, 14, 40, 50, 10, 10, 8, 8, 10, 12, 14, 35]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w


# Sheet 1: 요약
ws0 = wb.active
write_summary_sheet(ws0)

# Sheet 2: 전체
ws1 = wb.create_sheet()
write_data_sheet(ws1, '전체 데이터', posts)

# Sheet 3: 이벤트 기간
ws2 = wb.create_sheet()
write_data_sheet(ws2, '팝업기간(01-14~22)', during)

# Sheet 4: 이벤트 전
ws3 = wb.create_sheet()
write_data_sheet(ws3, '이벤트전(~01-13)', pre)

# Sheet 5: 이벤트 후
ws4 = wb.create_sheet()
write_data_sheet(ws4, '팝업후(01-23~)', post_ev)

# Sheet 6: 협찬만
ws5 = wb.create_sheet()
write_data_sheet(ws5, '협찬만(ClassC)', sponsored)

OUT_XLSX = BASE / 'hellokitty-jisoo-popup-rxr-sns-data.xlsx'
wb.save(OUT_XLSX)
print(f'Excel 생성 완료: {OUT_XLSX}')
print(f'  크기: {OUT_XLSX.stat().st_size:,} bytes')
print(f'  시트: {wb.sheetnames}')
