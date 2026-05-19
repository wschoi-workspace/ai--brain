#!/usr/bin/env python3
"""엑셀 → JSON 변환 (CEO Dashboard용)

Usage:
    python3 parse-excel.py <file.xlsx> --type regular --output data/
    python3 parse-excel.py <file.xlsx> --type pos --output data/
    python3 parse-excel.py <file.xlsx> --type project --output data/
    python3 parse-excel.py <file.xlsx> --type attendance --output data/
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl 필요. pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)

# ===== 컬럼명 매핑 (한글 유사어 포함) =====
COLUMN_SYNONYMS = {
    # POS
    "날짜": "date", "일자": "date",
    "매장명": "store_name", "매장": "store_name", "점포명": "store_name", "점포": "store_name",
    "매출액": "sales_amount", "매출": "sales_amount", "일매출": "sales_amount",
    "객수": "customers", "방문객수": "customers", "고객수": "customers",
    "객단가": "avg_per_customer", "평균객단가": "avg_per_customer",
    "voc건수": "voc_count", "voc": "voc_count", "VOC건수": "voc_count", "VOC": "voc_count",
    "재고부족품목": "inventory_shortage", "재고부족": "inventory_shortage", "재고경고": "inventory_shortage",
    "이슈사항": "issues", "이슈": "issues", "특이사항": "issues",
    "업무요청내용": "work_request_content", "업무요청": "work_request_content", "요청사항": "work_request_content",
    "요청긴급도": "work_request_urgency", "긴급도": "work_request_urgency",
    "처리상태": "work_request_status", "상태": "work_request_status",
    # Regular tasks
    "업무명": "name", "업무": "name", "정기업무명": "name",
    "주기": "cycle", "반복주기": "cycle",
    "날짜(일)": "day", "날짜": "day", "일": "day",
    "금액": "amount", "금액(원)": "amount",
    "비고": "note", "메모": "note",
    # Project
    "프로젝트명": "project_name", "프로젝트": "project_name",
    "진행률": "progress", "진행률(%)": "progress",
    "총예산": "total_budget", "예산": "total_budget",
    "집행액": "spent", "지출액": "spent",
    "다음마일스톤": "next_milestone", "마일스톤": "next_milestone",
    "마일스톤일자": "milestone_date", "마일스톤날짜": "milestone_date",
    "오늘논의사항": "today_topics", "논의사항": "today_topics",
    # Attendance
    "월": "month",
    "직원명": "employee_name", "직원": "employee_name", "이름": "employee_name",
    "소속매장": "store", "소속": "store",
    "초과근무(h)": "overtime_hours", "초과근무": "overtime_hours",
    "야근(h)": "night_hours", "야근": "night_hours",
    "특근(h)": "holiday_hours", "특근": "holiday_hours",
    "예상수당(원)": "estimated_pay", "예상수당": "estimated_pay",
    "총연차(일)": "total_leave", "총연차": "total_leave",
    "사용연차(일)": "used_leave", "사용연차": "used_leave",
    "잔여연차(일)": "remaining_leave", "잔여연차": "remaining_leave",
}


def find_header_row(ws):
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(cell is not None and str(cell).strip() for cell in row):
            return idx
    return 0


def handle_merged_cells(ws):
    for merge_range in list(ws.merged_cells.ranges):
        min_row = merge_range.min_row
        min_col = merge_range.min_col
        value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=row, column=col).value = value


def map_headers(ws, header_row_idx):
    """헤더 행의 컬럼명을 내부 키로 매핑"""
    row = list(ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1, values_only=True))[0]
    mapping = {}
    for col_idx, cell in enumerate(row):
        if cell is None:
            continue
        name = str(cell).strip().replace(" ", "")
        key = COLUMN_SYNONYMS.get(name)
        if not key:
            # 괄호 제거 후 재시도: "요청긴급도(상/중/하)" → "요청긴급도"
            name_clean = re.sub(r'\(.*?\)', '', name).strip()
            key = COLUMN_SYNONYMS.get(name_clean)
        if not key:
            # 대소문자 무시 재시도
            for synonym, k in COLUMN_SYNONYMS.items():
                if synonym.lower() == name.lower() or synonym.lower() == name_clean.lower():
                    key = k
                    break
        if key:
            mapping[col_idx] = key
    return mapping


def read_rows(ws, header_row_idx, col_mapping):
    """데이터 행을 딕셔너리 리스트로 변환"""
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx <= header_row_idx:
            continue
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record = {}
        for col_idx, key in col_mapping.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None and isinstance(val, str):
                    val = val.strip()
                record[key] = val
        rows.append(record)
    return rows


def safe_num(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    n = safe_num(val)
    return int(n) if n is not None else None


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


# ===== Type-specific transforms =====

def transform_regular(rows):
    """정기업무 엑셀 → regular-tasks.json"""
    today = datetime.now()
    tasks = []
    for r in rows:
        day = safe_int(r.get("day")) or safe_int(r.get("date"))
        if day is None:
            # "날짜" 컬럼이 없으면 name에서 추출 불가 → 스킵
            continue
        cycle = safe_str(r.get("cycle")) or "매월"

        # 다음 실행일 계산
        if "매월" in cycle:
            next_date = today.replace(day=min(day, 28))
            if next_date < today:
                month = today.month + 1
                year = today.year + (1 if month > 12 else 0)
                month = month if month <= 12 else month - 12
                next_date = next_date.replace(year=year, month=month)
            next_date_str = next_date.strftime("%Y-%m-%d")
        elif "매주" in cycle:
            next_date = today + timedelta(days=(day - today.weekday()) % 7)
            next_date_str = next_date.strftime("%Y-%m-%d")
        else:
            next_date_str = f"{today.strftime('%Y')}-{today.strftime('%m')}-{str(day).zfill(2)}"

        tasks.append({
            "name": safe_str(r.get("name")),
            "cycle": cycle,
            "day": day,
            "amount": safe_int(r.get("amount")),
            "next_date": next_date_str,
            "note": safe_str(r.get("note")),
        })

    return {
        "generated_at": datetime.now().isoformat(),
        "tasks": tasks,
    }


def transform_pos(rows):
    """POS 매출 엑셀 → pos-sales.json"""
    stores = {}
    for r in rows:
        name = safe_str(r.get("store_name"))
        if not name:
            continue
        if name not in stores:
            stores[name] = {
                "name": name,
                "yesterday": 0,
                "same_day_last_week": None,
                "same_day_last_month": None,
                "change_vs_last_week": None,
                "change_vs_last_month": None,
                "customers": 0,
                "avg_per_customer": 0,
                "monthly_target": None,
                "monthly_actual": 0,
                "daily_trend": [],
                "voc_count": 0,
                "voc_unresolved": 0,
                "inventory_alerts": [],
                "issues": "",
                "work_requests": [],
            }

        s = stores[name]
        sales = safe_int(r.get("sales_amount"))
        if sales:
            s["yesterday"] = sales
            s["monthly_actual"] += sales
        customers = safe_int(r.get("customers"))
        if customers:
            s["customers"] = customers
        avg = safe_int(r.get("avg_per_customer"))
        if avg:
            s["avg_per_customer"] = avg
        voc = safe_int(r.get("voc_count"))
        if voc:
            s["voc_count"] += voc
            s["voc_unresolved"] += voc

        inv = safe_str(r.get("inventory_shortage"))
        if inv:
            s["inventory_alerts"].append(inv)

        issue = safe_str(r.get("issues"))
        if issue:
            s["issues"] = (s["issues"] + "; " + issue).strip("; ")

        req_content = safe_str(r.get("work_request_content"))
        if req_content:
            s["work_requests"].append({
                "content": req_content,
                "urgency": safe_str(r.get("work_request_urgency")) or "중",
                "status": safe_str(r.get("work_request_status")) or "대기",
                "date": str(r.get("date", ""))[:10],
            })

    store_list = list(stores.values())
    total_yesterday = sum(s["yesterday"] for s in store_list)
    total_monthly = sum(s["monthly_actual"] for s in store_list)

    return {
        "generated_at": datetime.now().isoformat(),
        "source_file": "",
        "stores": store_list,
        "total": {
            "yesterday": total_yesterday,
            "monthly_target": None,
            "monthly_actual": total_monthly,
            "achievement_rate": None,
        },
    }


def transform_project(rows):
    """프로젝트 엑셀 → projects.json"""
    projects = []
    for r in rows:
        name = safe_str(r.get("project_name"))
        if not name:
            continue
        progress = safe_num(r.get("progress"))
        budget = safe_num(r.get("total_budget"))
        spent = safe_num(r.get("spent"))
        spend_rate = round(spent / budget * 100, 1) if budget and spent else None

        projects.append({
            "name": name,
            "progress": progress,
            "total_budget": safe_int(r.get("total_budget")),
            "spent": safe_int(r.get("spent")),
            "spend_rate": spend_rate,
            "next_milestone": safe_str(r.get("next_milestone")),
            "milestone_date": safe_str(r.get("milestone_date")),
            "today_topics": safe_str(r.get("today_topics")),
        })

    return {"generated_at": datetime.now().isoformat(), "projects": projects}


def transform_attendance(rows):
    """근태 엑셀 → attendance.json"""
    employees = []
    for r in rows:
        name = safe_str(r.get("employee_name"))
        if not name:
            continue
        overtime = safe_num(r.get("overtime_hours")) or 0
        night = safe_num(r.get("night_hours")) or 0
        holiday = safe_num(r.get("holiday_hours")) or 0

        employees.append({
            "name": name,
            "store": safe_str(r.get("store")),
            "month": safe_str(r.get("month")),
            "overtime_hours": overtime,
            "night_hours": night,
            "holiday_hours": holiday,
            "total_extra_hours": round(overtime + night + holiday, 1),
            "estimated_pay": safe_int(r.get("estimated_pay")),
            "total_leave": safe_num(r.get("total_leave")),
            "used_leave": safe_num(r.get("used_leave")),
            "remaining_leave": safe_num(r.get("remaining_leave")),
        })

    return {"generated_at": datetime.now().isoformat(), "employees": employees}


TRANSFORMS = {
    "regular": ("regular-tasks.json", transform_regular),
    "pos": ("pos-sales.json", transform_pos),
    "project": ("projects.json", transform_project),
    "attendance": ("attendance.json", transform_attendance),
}


def main():
    parser = argparse.ArgumentParser(description="엑셀 → JSON (CEO Dashboard)")
    parser.add_argument("file", help="엑셀 파일 경로")
    parser.add_argument("--type", required=True, choices=TRANSFORMS.keys(), help="데이터 유형")
    parser.add_argument("--output", default="data/", help="출력 디렉토리 (기본: data/)")
    parser.add_argument("--sheet", help="특정 시트명 (기본: 첫 번째 시트)")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"ERROR: 파일 없음 — {args.file}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, read_only=False, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    handle_merged_cells(ws)
    header_row_idx = find_header_row(ws)
    col_mapping = map_headers(ws, header_row_idx)

    if not col_mapping:
        print(f"ERROR: 인식 가능한 컬럼명이 없습니다. 헤더 행을 확인해주세요.", file=sys.stderr)
        sys.exit(1)

    print(f"  매핑된 컬럼: {', '.join(f'{v}' for v in col_mapping.values())}")
    rows = read_rows(ws, header_row_idx, col_mapping)
    print(f"  데이터 행: {len(rows)}건")

    filename, transform_fn = TRANSFORMS[args.type]
    result = transform_fn(rows)
    result["source_file"] = os.path.basename(args.file)

    os.makedirs(args.output, exist_ok=True)
    out_path = os.path.join(args.output, filename)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"  ✓ {out_path} 생성 완료")
    wb.close()


if __name__ == "__main__":
    main()
