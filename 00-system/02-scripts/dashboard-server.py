#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""프로젝트 포트폴리오 대시보드 — 공유 서버 (표준 라이브러리만).
화면(포트폴리오_대시보드.html)과 JSON API를 한 서버에서 서빙한다.
PM이 수정하면 서버에 저장되어 대표·팀원 모두 같은 데이터를 본다.

데이터: 00-system/01-templates/_data/  (users.json + projects/{id}.json)
권한:   대표=전체 열람·수정·생성·삭제 / PM=본인 프로젝트 수정 / 직원=배정 프로젝트 열람
실행:   python3 dashboard-server.py [port]   (기본 8770, 127.0.0.1 — Tailscale serve로 노출)
"""
import json, os, re, sys, threading, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, quote
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from shared import report_score as _report_score  # 채점 SSOT (2026-07-20 갭 해소)

# .env 로드(WEEKLY_KEY 등 — 환경 미설정 시 do-better-workspace/.env에서 보충)
_WS = Path(__file__).resolve().parent.parent.parent
for _envp in (_WS / ".env",):
    if _envp.exists():
        for _l in _envp.read_text(encoding="utf-8").splitlines():
            _l = _l.strip()
            if _l and not _l.startswith("#") and "=" in _l:
                _k, _, _v = _l.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip())

# 이관 가능: 경로는 스크립트 기준(상대) + 환경변수로 덮어쓰기 가능
BASE = Path(os.environ.get("DASHBOARD_BASE") or (Path(__file__).resolve().parent.parent / "01-templates"))
HTML = BASE / "포트폴리오_대시보드.html"
DATA = Path(os.environ.get("DASHBOARD_DATA") or (BASE / "_data"))
PROJ_DIR = DATA / "projects"
DOC_DIR = DATA / "project-docs"   # 프로젝트 자료(회의록) 원문 — JSON에는 메타만
HOST = os.environ.get("DASHBOARD_HOST", "127.0.0.1")
PORT = int(os.environ.get("DASHBOARD_PORT", "8770"))
# ARISA 주간 대시보드 서빙 — /weekly. 성장지표는 대표 토큰(?key=WEEKLY_KEY)일 때만 노출.
WEEKLY_DIR = Path(os.environ.get("WEEKLY_DIR") or (_WS / "20-operations" / "23-arisa" / "weekly"))
WEEKLY_KEY = os.environ.get("WEEKLY_KEY", "")
# ARISA 대표 Daily Brief — /brief. HTML 내장 로그인이 대표(admin)만 입장시킴.
BRIEF_DIR = Path(os.environ.get("BRIEF_DIR") or (_WS / "20-operations" / "23-arisa" / "brief"))
# 팀 리더 판정 출처 — arisa-employees.json의 team_leads(팀→리더이름) 역매핑.
EMP_PATH = Path(__file__).resolve().parent / "arisa-employees.json"
# ARISA 프로젝트 메모리(회의록·결정·진행 로그) 링크백 — filament 프로젝트 포커스 반영 (B1, 2026-07-20)
MEMORY_DIR = Path(os.environ.get("ARISA_MEMORY_DIR") or (Path.home() / "arisa-project-memory" / "projects"))
# ARISA 2.0 리버스 프록시 — /arisa2/* → localhost:ARISA2_PORT
ARISA2_PORT = int(os.environ.get("ARISA2_PORT", "8787"))
ARISA2_UPSTREAM = f"http://127.0.0.1:{ARISA2_PORT}"
_lock = threading.Lock()

# ── 주간분장 시트 연동 (개인 대시보드 — 업무분장 입력/조회) ──
import sys as _sys
_sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from shared import gws as _asgws
except Exception:
    _asgws = None
try:
    from shared.decision import load_open_decisions as _load_open_decisions
except Exception:
    _load_open_decisions = None
from shared import status as _ST  # 상태·우선순위 단일출처 (G2) — 배포 시 shared/status.py 동반 필수
from shared import naming as _NM  # 프로젝트 네이밍 규칙 (P2) — 배포 시 shared/naming.py 동반 필수
try:
    from shared.status_log import log_status_change as _log_st  # 상태 이력 (G5) — 실패 무해
    from shared.status_log import load_history as _load_st_history  # 이력 조회 (G7)
except Exception:
    def _log_st(*a, **k):
        return False

    def _load_st_history(*a, **k):
        return []
DAILY_SHEET = os.environ.get("DAILY_REPORT_SHEET_ID", "")
ASSIGN_TAB = "주간분장"


def _week_label(d=None):
    d = d or datetime.date.today()
    return f"W{d.isocalendar().week:02d}"


def _assign_read():
    """주간분장 탭 → dict 리스트. 탭 없거나 gws 실패 시 [] (안전 — 개인탭이 죽지 않게)."""
    if not (_asgws and DAILY_SHEET):
        return []
    try:
        rows = _asgws.values_get(DAILY_SHEET, f"{ASSIGN_TAB}!A2:L5000", retries=2, timeout=20)
    except Exception:
        return []
    out = []
    for i, r in enumerate(rows):
        r = list(r) + [""] * (12 - len(r))
        # 시트 헤더: 날짜·프로젝트명·팀구분·담당자·업무내용·일정(완료예상)·결과물·상태·이해관계자·우선순위·프로젝트ID(K, G1)·등록자(L, 2026-07-20)
        a = {"row": i + 2,  # 시트 실제 행 번호 (A2부터) — 상태 업데이트용
             "date": (r[0] or "").strip(), "project": (r[1] or "").strip(),
             "team": (r[2] or "").strip(), "assignee": (r[3] or "").strip(),
             "task": (r[4] or "").strip(), "deadline": (r[5] or "").strip(),
             "result": (r[6] or "").strip(), "status": _ST.norm_assign_status(r[7]),
             "stakeholder": (r[8] or "").strip(), "priority": _ST.norm_priority(r[9]),
             "pid": (r[10] or "").strip(), "by": (r[11] or "").strip()}
        # filament 반영 — 지연 N일(열린 분장만)을 모든 소비자(내 업무·리더 홈·대표창)에 공급
        a["days_overdue"] = _ST.overdue_days(a["deadline"]) if _ST.is_overdue(a["deadline"], a["status"]) else 0
        out.append(a)
    return out


def _assign_append(assignee, task, deadline, priority, by, project="", result="", stakeholder=""):
    """주간분장 append. 사용자 헤더 순서: 날짜·프로젝트명·팀·담당자·업무내용·일정·결과물·상태·이해관계자·우선순위."""
    if not (_asgws and DAILY_SHEET):
        return False, "시트 미설정"
    team = emp_team(assignee) or ""
    # L열 등록자(by) — 받은 업무 섹션의 출처 표시(대표 지시/리더 이관/본인 등록)용 (2026-07-20)
    row = [datetime.date.today().isoformat(), project, team, assignee, task, deadline,
           result, "미착수", stakeholder, priority, _resolve_pid(project), (by or "").strip()]
    try:
        ok = _asgws.append_to_sheet(DAILY_SHEET, f"{ASSIGN_TAB}!A1", row, timeout=20)
        return bool(ok), "" if ok else "주간분장 탭 없음/append 실패"
    except Exception as e:
        return False, str(e)[:80]


OPENAI_KEY = os.environ.get("OPENAI_API_KEY", "")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")


def _call_llm_json(system_prompt, user_msg, max_tokens=2000,
                   openai_model="gpt-4o", openai_only=False):
    """LLM API 호출 → JSON 파싱. Anthropic 우선, OpenAI fallback. 실패 시 None.

    openai_only=True면 Anthropic을 건너뛴다 — 채점은 봇과 동일 모델(gpt-4o-mini,
    temp 0.3)로 고정해 채널 간 점수 편차를 없앤다 (2026-07-20)."""
    def _parse_json(text):
        m = re.search(r"```json\s*([\s\S]*?)```", text)
        return json.loads(m.group(1) if m else text)
    # 1) Anthropic
    if ANTHROPIC_KEY and not openai_only:
        try:
            payload = json.dumps({
                "model": "claude-sonnet-4-20250514",
                "max_tokens": max_tokens, "temperature": 0.3,
                "system": system_prompt,
                "messages": [{"role": "user", "content": user_msg}]
            }).encode("utf-8")
            req = Request("https://api.anthropic.com/v1/messages", data=payload, headers={
                "x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01",
                "Content-Type": "application/json"
            }, method="POST")
            with urlopen(req, timeout=60) as r:
                d = json.loads(r.read())
            return _parse_json(d["content"][0]["text"])
        except Exception:
            pass  # fallback to OpenAI
    # 2) OpenAI fallback
    if OPENAI_KEY:
        try:
            payload = json.dumps({
                "model": openai_model, "temperature": 0.3,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_msg}
                ]
            }).encode("utf-8")
            req = Request("https://api.openai.com/v1/chat/completions", data=payload,
                          headers={"Authorization": "Bearer " + OPENAI_KEY,
                                   "Content-Type": "application/json"}, method="POST")
            with urlopen(req, timeout=60) as r:
                d = json.loads(r.read())
            return _parse_json(d["choices"][0]["message"]["content"])
        except Exception:
            pass
    return None


# 하위 호환: 기존 _call_claude 호출부 유지
_call_claude = _call_llm_json


# 주간업무계획.xlsx 파싱 폴백용 파이썬 (시스템 3.9에 openpyxl 없을 때 subprocess)
_XLSX_PY_CANDIDATES = [
    str(_WS / "20-operations" / "24-second-brain" / ".venv311" / "bin" / "python"),
    "/usr/bin/python3",
]

_PLAN_PARSE_SNIPPET = r'''
import json, sys
import openpyxl
rows = []
wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
for ws in wb.worksheets:
    hdr = None
    last_proj = ""
    for r in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in r]
        if hdr is None:
            if any("프로젝트명" in c for c in cells):
                hdr = {}
                for i, c in enumerate(cells):
                    if "프로젝트명" in c: hdr["project"] = i
                    elif "금주" in c: hdr["this"] = i
                    elif "차주" in c: hdr["next"] = i
                    elif "계약상황" in c: hdr["status"] = i
                    elif "due" in c.lower() or "런칭" in c: hdr["due"] = i
            continue
        def g(k):
            i = hdr.get(k)
            return cells[i] if i is not None and i < len(cells) else ""
        proj = g("project") or last_proj
        if g("project"):
            last_proj = g("project")
        tw, nw = g("this"), g("next")
        if not (tw or nw):
            continue
        rows.append({"project": proj, "this": tw, "next": nw, "status": g("status"), "due": g("due")})
print(json.dumps(rows, ensure_ascii=False))
'''


def _parse_weekly_plan(path):
    """주간업무계획.xlsx → [{project, this, next, status, due}] — 헤더('프로젝트명') 탐지,
    프로젝트명 빈 행은 직전 상속, 금주·차주 모두 빈 행 skip.
    인프로세스 openpyxl 우선, 없으면 venv 파이썬 subprocess 폴백."""
    try:
        import openpyxl  # noqa: F401
        rows = []
        wb = openpyxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            hdr = None
            last_proj = ""
            for r in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in r]
                if hdr is None:
                    if any("프로젝트명" in c for c in cells):
                        hdr = {}
                        for i, c in enumerate(cells):
                            if "프로젝트명" in c: hdr["project"] = i
                            elif "금주" in c: hdr["this"] = i
                            elif "차주" in c: hdr["next"] = i
                            elif "계약상황" in c: hdr["status"] = i
                            elif "due" in c.lower() or "런칭" in c: hdr["due"] = i
                    continue
                def g(k):
                    i = hdr.get(k)
                    return cells[i] if i is not None and i < len(cells) else ""
                proj = g("project") or last_proj
                if g("project"):
                    last_proj = g("project")
                tw, nw = g("this"), g("next")
                if not (tw or nw):
                    continue
                rows.append({"project": proj, "this": tw, "next": nw,
                             "status": g("status"), "due": g("due")})
        return rows
    except ImportError:
        pass
    import subprocess, tempfile
    for py in _XLSX_PY_CANDIDATES:
        if not Path(py).exists():
            continue
        try:
            r = subprocess.run([py, "-c", _PLAN_PARSE_SNIPPET, str(path)],
                               capture_output=True, text=True, timeout=60)
            if r.returncode == 0 and r.stdout.strip():
                return json.loads(r.stdout)
        except Exception:
            continue
    return []


def _llm_todo(text, max_items=12):
    """자유 텍스트 업무지시 → 실행가능 to-do 항목 리스트. OpenAI(urllib, stdlib). 실패 시 []."""
    text = (text or "").strip()
    if not (OPENAI_KEY and text):
        return []
    _today = datetime.date.today().isoformat()
    sys_p = (f"오늘은 {_today}(YYYY-MM-DD)이다. 다음 업무 지시 텍스트를 담당자가 바로 실행할 수 있는 "
             f"개별 to-do 항목으로 분해한다. 각 항목은 구체적 한 문장. 지시에 없는 내용을 지어내지 말 것. 항목 최대 {max_items}개. "
             "**프로젝트 귀속 필수**: 항목이 특정 프로젝트에 속하면 project에 상위 프로젝트명을 넣는다. "
             "예) '봉은사 프로젝트 … (상세항목 _ BM 세부 리서치 및 아이데이션)' → 봉은사의 하위 상세항목이므로 "
             "그 항목의 project는 '봉은사'. 상위 프로젝트가 명시된 문장의 하위 항목들은 모두 같은 project를 상속한다. 불명확하면 빈칸. "
             "마감(deadline)은 텍스트에 명시적 날짜·요일이 있을 때만 오늘 기준 YYYY-MM-DD로 계산하고, 없거나 불명확하면 반드시 빈칸. "
             "임의 날짜 생성 절대 금지. priority는 '긴급' 근거 없으면 '일반'. "
             '반드시 JSON만 출력: {"items":[{"task":"...","project":"","deadline":"","priority":"일반"}]}')
    payload = json.dumps({"model": "gpt-4o-mini", "temperature": 0.2,
                          "response_format": {"type": "json_object"},
                          "messages": [{"role": "system", "content": sys_p},
                                       {"role": "user", "content": text[:4000]}]}).encode("utf-8")
    req = Request("https://api.openai.com/v1/chat/completions", data=payload,
                  headers={"Authorization": "Bearer " + OPENAI_KEY, "Content-Type": "application/json"}, method="POST")
    try:
        with urlopen(req, timeout=30) as r:
            d = json.loads(r.read())
        items = json.loads(d["choices"][0]["message"]["content"]).get("items", [])
    except Exception:
        return []
    out = []
    for it in items:
        t = (it.get("task") or "").strip()
        if t:
            out.append({"task": t, "project": (it.get("project") or "").strip(),
                        "deadline": (it.get("deadline") or "").strip(),
                        "priority": (it.get("priority") or "일반").strip()})
    return out

# ── 보고 시뮬레이터 AI 프롬프트 ──

# 일일보고 채점 — 루브릭·감점규칙은 shared/report_score.py(SSOT), 여기는 교육용 출력만.
# 2026-07-20: 유형별(A/B/C) 이원 루브릭 + grace/strict 2모드로 봇과 동일 기준.
_SIM_DAILY_OUTPUT = """

## 6-Layer 프레임워크 (피드백 서술 시 참조)
| Layer | 핵심 질문 |
|-------|----------|
| Context | 이 프로젝트가 무엇인가? 왜 하는가? |
| Progress | 현재 어디까지 왔는가? Output/Outcome |
| Thinking | 왜 그렇게 판단했는가? 사실 vs 의견 구분 |
| Priority | 무엇이 가장 중요한가? 오늘/내일 최중요 |
| Risk | 문제·막힌 것·예상 위험 |
| Decision | 대표가 결정해야 할 것, 승인·지원 요청 |

## 안티패턴 감지 (7종) — 감지된 패턴은 반드시 해당 항목 점수에 감점으로 반영하라
1. output_copy: Output 텍스트가 Outcome에 그대로 복사
2. vague_decision: "컨펌 필요"만 — 옵션/기한 없음
3. empty_risk: Risk 빈칸 ("없음" 명시가 아닌 진짜 빈칸)
4. no_why: 중요 업무에 "왜" 설명 없음
5. abstract_tomorrow: "계속 진행", "추가 확인" 등 구체성 없는 내일 계획
6. fact_opinion_mix: 사실/의견 미분리 ("클라이언트가 원하는 것 같다" 등)
7. no_numbers: Output/Outcome에 수치(건수, %, 금액 등) 전무

## ThinkingCost
대표가 이 보고를 읽고 추가로 물어야 할 질문 수를 예측하라.
- 0개 = ★★★★★ (5), 1개 = ★★★★☆ (4), 2개 = ★★★☆☆ (3), 3개 = ★★☆☆☆ (2), 4개 = ★☆☆☆☆ (1), 5개+ = ☆☆☆☆☆ (0)

## Before/After 참고
Bad: "도면 작업 중" → Good: "[세스크멘슬] 주방 도면 v3 — 설비팀 피드백 반영 중. 환기닥트 위치 변경으로 레이아웃 수정. 내일 오전 최종본 공유 예정"
Bad: "컨펌 필요" → Good: "결정 요청: 자재 A안(2,000만원·내구성↑) vs B안(1,500만원·납기↑). 추천: A안 (유지보수비 절감). 기한: 7/15까지"
Bad: "리서치 진행함" → Good: "경쟁사 3곳 벤치마킹 완료 — 가격대/메뉴구성/타깃 비교표 작성(12건). 공유 드라이브에 업로드"

## 출력 형식 (반드시 이 JSON만 출력)
scores에는 **분류된 유형의 항목만** 넣고, 각 항목의 max는 그 유형의 배점을 그대로 써라.
```json
{
  "report_type": "A",
  "scores": {
    "context": {"score": 0, "max": 25, "pass": false, "feedback": "..."},
    "objective": {"score": 0, "max": 15, "pass": false, "feedback": "..."},
    "evidence": {"score": 0, "max": 20, "pass": false, "feedback": "..."},
    "priority": {"score": 0, "max": 25, "pass": false, "feedback": "..."},
    "risk": {"score": 0, "max": 15, "pass": false, "feedback": "..."}
  },
  "total": 0,
  "grade": "D",
  "patterns_detected": [
    {"type": "패턴명", "field": "해당필드", "detail": "구체 설명"}
  ],
  "thinking_cost": {"predicted_questions": 0, "stars": 5, "detail": "예측 근거"},
  "summary": "전체 코멘트 (2~3문장) — 보고 유형과 그 유형에서 무엇이 중요한지 언급",
  "improvement_tips": ["개선 팁1", "개선 팁2", "개선 팁3"]
}
```
등급: S(90~100), A(75~89), B(60~74), C(40~59), D(0~39)"""


def _sim_daily_prompt():
    """공유 코어(grace/strict 자동) + 시뮬레이터 교육용 출력 블록 — 요청 시점마다 조립."""
    return _report_score.build_prompt() + _SIM_DAILY_OUTPUT


SIMULATOR_BRIEF_PROMPT = """너는 ARISA 기획안(1P-Brief) 품질 시뮬레이터다. 직원의 기획안보고를 리뷰하고 100점을 채점한다.

## 1P-Brief 7항목 루브릭 (100점)
| 항목 | 배점 | 만점 기준 |
|------|------|----------|
| title | 15 | "[대상]을 위한 [내용]" 형식. 한 문장으로 무엇인지 알 수 있음 |
| summary | 20 | 대상/문제/방식/결과 4문장. 대표가 이것만 읽고 판단 가능 |
| stakeholder | 10 | P1/P2/P3 이해관계자 + 각각 원하는 것 명시 |
| client_value | 10 | 클라이언트가 얻는 구체적 이익(수치/상태) |
| core_idea | 20 | 형식이 아닌 해결방식(How). "팝업을 한다"는 형식, "유휴공간을 활용한 체험형 쇼룸"이 아이디어 |
| success_criteria | 15 | 측정 가능한 상태/숫자 (KPI). "성공적으로 마친다"는 0점 |
| support_needed | 10 | 예산/인력/파트너/의사결정. "없음" 명시 = 만점, 빈칸 = 0 |

## 기획안 안티패턴
1. abstract_title: "[OO] 프로젝트" 같은 추상적 제목 — 대상·내용 불명
2. no_summary: Summary 없음 또는 1문장 미만
3. format_as_idea: "팝업을 한다", "전시를 한다"처럼 형식만 적고 해결방식(How) 없음
4. unmeasurable_success: "성공적으로", "효과적으로" 등 측정 불가 기준
5. missing_stakeholder: 이해관계자 미식별 또는 "우리 회사"만
6. vague_budget: "추후 협의" 등 구체성 없는 예산/지원
7. no_problem: 왜 이 기획이 필요한지(문제 정의) 빠짐

## ThinkingCost
대표가 "이거 진행해도 되나?"를 판단하기 위해 추가로 물어야 할 질문 수 예측.
별점: 0개=5★, 1개=4★, 2개=3★, 3개=2★, 4개=1★, 5+개=0★

## 출력 형식 (반드시 이 JSON만 출력)
```json
{
  "scores": {
    "title": {"score": 0, "max": 15, "pass": false, "feedback": "..."},
    "summary": {"score": 0, "max": 20, "pass": false, "feedback": "..."},
    "stakeholder": {"score": 0, "max": 10, "pass": false, "feedback": "..."},
    "client_value": {"score": 0, "max": 10, "pass": false, "feedback": "..."},
    "core_idea": {"score": 0, "max": 20, "pass": false, "feedback": "..."},
    "success_criteria": {"score": 0, "max": 15, "pass": false, "feedback": "..."},
    "support_needed": {"score": 0, "max": 10, "pass": false, "feedback": "..."}
  },
  "total": 0,
  "grade": "D",
  "patterns_detected": [
    {"type": "패턴명", "field": "해당필드", "detail": "구체 설명"}
  ],
  "thinking_cost": {"predicted_questions": 0, "stars": 5, "detail": "예측 근거"},
  "summary": "전체 코멘트 (2~3문장)",
  "improvement_tips": ["개선 팁1", "개선 팁2", "개선 팁3"]
}
```
등급: S(90~100), A(75~89), B(60~74), C(40~59), D(0~39)"""

SIMULATOR_DRAFT_PROMPT = """너는 업무 보고 구조화 도우미다. 사용자가 자유롭게 서술한 업무 텍스트를 받아, 지정된 필드 구조에 맞게 정보를 추출·배분한다.

## 규칙
1. 텍스트에 실제로 언급된 정보만 추출한다. 없는 정보를 지어내지 않는다.
2. 빈 필드는 빈 문자열 ""로 반환한다.
3. 구체적 수치·이름·날짜가 있으면 반드시 포함한다.
4. 리뷰/채점/평가를 하지 않는다. 정보 추출·구조화만 수행한다.
5. 반드시 JSON만 출력한다.

## daily 모드 (일일보고 8필드)
```json
{"context":"","output":"","outcome":"","risk":"","tomorrow":"","decision":"","support":"","evidence":""}
```
- context: 프로젝트명 + 왜 중요한지
- output: 오늘 한 일의 구체적 산출물
- outcome: 목표 대비 진행 상황
- risk: 문제·리스크 (없으면 "없음")
- tomorrow: 내일 할 일
- decision: 대표에게 필요한 결정 (없으면 "없음")
- support: 필요한 지원 (없으면 "없음")
- evidence: 사실 vs 의견 구분

## brief 모드 (기획안 7필드)
```json
{"title":"","summary":"","stakeholder":"","client_value":"","core_idea":"","success_criteria":"","support_needed":""}
```
- title: [대상]을 위한 [내용] 형식 제목
- summary: 대상/문제/방식/결과 4문장
- stakeholder: P1/P2/P3 이해관계자
- client_value: 클라이언트가 얻는 이익
- core_idea: 해결방식(How)
- success_criteria: 측정 가능한 KPI
- support_needed: 예산/인력/파트너 (없으면 "없음")

사용자 메시지 첫 줄에 [MODE:daily] 또는 [MODE:brief]가 명시된다."""

_DAILY_FIELDS = [
    ("context", "오늘 가장 중요한 업무 (프로젝트명 + 왜)"),
    ("output", "오늘 얻은 결과 (Output)"),
    ("outcome", "목표 대비 현재 위치 (Outcome)"),
    ("risk", "발견한 문제 / 리스크"),
    ("tomorrow", "내일 가장 중요한 일"),
    ("decision", "대표에게 필요한 결정"),
    ("support", "필요한 지원"),
    ("evidence", "사실·의견 구분 메모"),
]
_BRIEF_FIELDS = [
    ("title", "프로젝트 제목"),
    ("summary", "Executive Summary"),
    ("stakeholder", "Primary Stakeholder (P1/P2/P3)"),
    ("client_value", "Client Value"),
    ("core_idea", "핵심 아이디어 (How)"),
    ("success_criteria", "성공 기준 (KPI)"),
    ("support_needed", "필요한 지원"),
]


def _build_simulator_input(mode, fields):
    """사용자 입력 필드를 LLM 유저 메시지로 조합."""
    items = _BRIEF_FIELDS if mode == "brief" else _DAILY_FIELDS
    lines = []
    for key, label in items:
        val = (fields.get(key) or "").strip()
        lines.append(f"[{label}]\n{val if val else '(빈칸)'}")
    return "\n\n".join(lines)


SIMULATOR_HTML = """<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>보고 시뮬레이터 — ARISA</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable.min.css">
<style>
:root{--bg:#1A1A1A;--bg2:#202020;--bg3:#262626;--fg:#F5F0EB;--fg2:#C4BEB7;--muted:#8A857E;--line:#333;--accent:#6C5CE7;--red:#E17055;--green:#00b894;--amber:#FDCB6E;--coral:#E17055}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--fg);font-family:'Pretendard Variable',sans-serif;font-weight:300;min-height:100vh}
.wrap{max-width:1400px;margin:0 auto;padding:24px 20px}
h1{font-size:22px;font-weight:700;margin-bottom:4px}
.sub{color:var(--muted);font-size:13px;margin-bottom:20px}
.modes{display:flex;gap:8px;margin-bottom:20px}
.mode-btn{background:var(--bg3);border:1px solid var(--line);color:var(--muted);border-radius:8px;padding:10px 24px;font-size:14px;font-weight:600;cursor:pointer;font-family:inherit}
.mode-btn.on{background:var(--accent);color:#fff;border-color:var(--accent)}
.main{display:grid;grid-template-columns:1fr 1fr;gap:20px}
@media(max-width:1200px){.main{grid-template-columns:1fr}}
.panel{background:var(--bg2);border:1px solid var(--line);border-radius:14px;padding:20px}
.panel h2{font-size:16px;font-weight:600;margin-bottom:14px}
.field{margin-bottom:14px}
.field label{display:flex;align-items:center;gap:8px;font-size:13px;font-weight:500;margin-bottom:6px}
.field .tag{font-size:10px;padding:2px 7px;border-radius:4px;font-weight:600}
.tag-ctx{background:rgba(108,92,231,.15);color:var(--accent)}
.tag-prg{background:rgba(0,184,148,.15);color:var(--green)}
.tag-thk{background:rgba(253,203,110,.15);color:var(--amber)}
.tag-pri{background:rgba(253,203,110,.15);color:var(--amber)}
.tag-rsk{background:rgba(225,112,85,.15);color:var(--coral)}
.tag-dec{background:rgba(225,112,85,.15);color:var(--coral)}
.field textarea{width:100%;background:var(--bg3);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:10px 12px;font-size:13px;font-family:inherit;resize:vertical;line-height:1.6;min-height:64px}
.field textarea:focus{border-color:var(--accent);outline:none}
.guide-toggle{font-size:11px;color:var(--accent);cursor:pointer;margin-left:auto;user-select:none}
.guide-box{display:none;margin:6px 0 0;background:var(--bg);border-radius:8px;padding:10px 12px;font-size:12px;line-height:1.6}
.guide-box.open{display:block}
.g-bad{color:var(--coral)}
.g-good{color:var(--green)}
.submit-row{margin-top:16px;display:flex;gap:10px;align-items:center}
.submit-btn{background:var(--accent);color:#fff;border:0;border-radius:10px;padding:14px 32px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit}
.submit-btn:disabled{opacity:.5;cursor:default}
.submit-msg{font-size:12px;color:var(--muted)}
.draft-section{background:var(--bg3);border:1px solid var(--line);border-radius:10px;margin-bottom:16px;overflow:hidden}
.draft-header{display:flex;align-items:center;justify-content:space-between;padding:12px 14px;cursor:pointer;user-select:none}
.draft-header h3{font-size:13px;font-weight:600;color:var(--accent)}
.draft-header .arrow{font-size:12px;color:var(--muted);transition:transform .2s}
.draft-header .arrow.open{transform:rotate(180deg)}
.draft-body{display:none;padding:0 14px 14px}
.draft-body.open{display:block}
.draft-body textarea{width:100%;min-height:80px;background:var(--bg);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:10px 12px;font-size:13px;font-family:inherit;resize:vertical;line-height:1.6}
.draft-body textarea:focus{border-color:var(--accent);outline:none}
.draft-actions{display:flex;gap:8px;margin-top:10px;align-items:center}
.draft-btn{background:var(--accent);color:#fff;border:0;border-radius:8px;padding:10px 18px;font-size:13px;font-weight:600;cursor:pointer;font-family:inherit}
.draft-btn:disabled{opacity:.5;cursor:default}
.draft-msg{font-size:12px;color:var(--muted)}
.draft-hint{font-size:11px;color:var(--muted);margin-bottom:8px;line-height:1.5}
.ai-compare{margin-top:6px;background:rgba(108,92,231,.06);border:1px solid rgba(108,92,231,.2);border-radius:8px;padding:8px 10px;display:none}
.ai-compare.show{display:block}
.ai-compare-label{font-size:10px;color:var(--accent);font-weight:600;margin-bottom:4px}
.ai-compare-text{font-size:12px;color:var(--fg2);line-height:1.5}
.ai-replace-btn{background:transparent;border:1px solid var(--accent);color:var(--accent);border-radius:6px;padding:3px 10px;font-size:11px;cursor:pointer;margin-top:6px;font-family:inherit}
.ai-replace-btn:hover{background:var(--accent);color:#fff}
.result-empty{color:var(--muted);font-size:14px;text-align:center;padding:60px 0}
.score-circle{width:140px;height:140px;margin:0 auto 12px;position:relative}
.score-circle svg{transform:rotate(-90deg)}
.score-circle .val{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center}
.score-circle .num{font-size:36px;font-weight:800}
.score-circle .grade{font-size:18px;font-weight:700;margin-top:2px}
.tc-row{text-align:center;margin-bottom:16px}
.tc-stars{font-size:22px;letter-spacing:2px}
.tc-detail{font-size:12px;color:var(--muted);margin-top:4px}
.item-card{background:var(--bg);border:1px solid var(--line);border-radius:10px;padding:12px 14px;margin-bottom:8px}
.item-head{display:flex;justify-content:space-between;align-items:center}
.item-name{font-size:13px;font-weight:600}
.item-score{font-size:13px;font-weight:700}
.item-bar{height:4px;background:var(--bg3);border-radius:2px;margin:8px 0 6px;overflow:hidden}
.item-bar-fill{height:100%;border-radius:2px;transition:width .4s}
.item-fb{font-size:12px;color:var(--fg2);line-height:1.5}
.pat-card{background:rgba(225,112,85,.08);border:1px solid rgba(225,112,85,.25);border-radius:10px;padding:10px 14px;margin-bottom:8px}
.pat-type{font-size:12px;font-weight:700;color:var(--coral)}
.pat-detail{font-size:12px;color:var(--fg2);margin-top:4px}
.tips{margin-top:14px}
.tips h3{font-size:13px;font-weight:600;margin-bottom:8px}
.tip-item{font-size:12px;color:var(--fg2);padding:4px 0 4px 14px;border-left:2px solid var(--accent);margin-bottom:6px;line-height:1.5}
.summary-box{background:var(--bg);border-radius:10px;padding:14px;margin-bottom:14px;font-size:13px;line-height:1.7}
</style></head><body>
<div class="wrap">
<h1>보고 시뮬레이터</h1>
<p class="sub">Reporting OS v2 — 보고를 쓰고 AI 리뷰를 받아보세요. ReportScore 100점 채점 + ThinkingCost 예측</p>
<div class="modes">
  <button class="mode-btn on" data-m="daily">일일보고</button>
  <button class="mode-btn" data-m="brief">기획안 (1P-Brief)</button>
</div>
<div class="main">
  <div class="panel" id="input-panel">
    <h2 id="input-title">일일보고 입력</h2>
    <div class="draft-section" id="draft-section">
      <div class="draft-header" id="draft-toggle">
        <h3>AI 비교 보기</h3>
        <span class="arrow" id="draft-arrow">▼</span>
      </div>
      <div class="draft-body" id="draft-body">
        <div class="draft-hint">먼저 아래 폼을 직접 작성한 뒤, 같은 내용을 자유 텍스트로 입력하면 AI 드래프트와 비교할 수 있습니다.<br>내 보고와 AI 드래프트의 차이를 보며 빠진 항목을 발견하세요.</div>
        <textarea id="draft-text" placeholder="업무를 자유롭게 설명하세요...&#10;예) 오늘 세스크멘슬 주방 도면 v3 작업했고, 설비팀 피드백 반영 중. 환기닥트 위치가 바뀌어서 레이아웃 수정해야 함. 내일 오전까지 최종본 보내야 하는데, 자재 A안 B안 중에 대표님 결정이 필요함."></textarea>
        <div class="draft-actions">
          <button class="draft-btn" id="draft-btn" disabled>AI 드래프트와 비교</button>
          <span class="draft-msg" id="draft-msg"></span>
        </div>
      </div>
    </div>
    <div id="fields"></div>
    <div class="submit-row">
      <button class="submit-btn" id="submit-btn">AI 리뷰 받기</button>
      <span class="submit-msg" id="submit-msg"></span>
    </div>
  </div>
  <div class="panel" id="result-panel">
    <h2>AI 리뷰 결과</h2>
    <div id="result"><div class="result-empty">왼쪽 폼을 작성하고 "AI 리뷰 받기"를 눌러보세요.</div></div>
  </div>
</div>
</div>
<script>
(function(){
var mode='daily';
var DAILY=[
  {key:'context',label:'오늘 가장 중요한 업무',tag:'Context',tagC:'ctx',ph:'프로젝트명 + 왜 중요한지\\n예) [세스크멘슬] 주방 도면 v3 최종화 — 설비팀 피드백 반영 기한이 오늘',
   bad:'도면 작업 중',good:'[세스크멘슬] 주방 도면 v3 — 설비팀 피드백(환기닥트 위치 변경) 반영. 내일 오전 최종본 공유 예정'},
  {key:'output',label:'오늘 얻은 결과 (Output)',tag:'Progress',tagC:'prg',ph:'구체적 산출물 + 수량\\n예) 경쟁사 3곳 벤치마킹 비교표 작성(12건), 드라이브 업로드 완료',
   bad:'리서치 진행함',good:'경쟁사 3곳 벤치마킹 완료 — 가격대/메뉴구성/타깃 비교표(12건). 드라이브에 업로드'},
  {key:'outcome',label:'목표 대비 현재 위치 (Outcome)',tag:'Thinking',tagC:'thk',ph:'Output이 목표에 어떤 의미인지\\n예) 전체 진행률 70% — 남은 건 자재 발주(A/B안 확정 대기)',
   bad:'(Output을 그대로 복사)',good:'도면 검토 3/5건 완료(60%). 남은 2건은 구조 변경 영향 확인 후 내일 반영 예정'},
  {key:'risk',label:'발견한 문제 / 리스크',tag:'Risk',tagC:'rsk',ph:'문제 + 영향 + 대응\\n없으면 "없음"이라고 명시 (빈칸은 0점)',
   bad:'(빈칸)',good:'자재 납기 1주 지연 통보 — 오픈일 영향 가능. 대안업체 2곳 견적 요청 완료, 내일 비교'},
  {key:'tomorrow',label:'내일 가장 중요한 일',tag:'Priority',tagC:'pri',ph:'★ 최중요 1개 + 기타\\n예) ★ 자재 A/B안 비교표 대표 보고 → 나머지: 도면 잔여 2건 마무리',
   bad:'계속 진행',good:'★ 자재 A/B안 비교표 작성 → 오전 중 대표 보고. 도면 잔여 2건 오후 마무리'},
  {key:'decision',label:'대표에게 필요한 결정',tag:'Decision',tagC:'dec',ph:'옵션 + 추천안 + 기한\\n없으면 "없음"이라고 명시',
   bad:'컨펌 필요',good:'결정 요청: 자재 A안(2,000만·내구성↑) vs B안(1,500만·납기↑). 추천: A안. 기한: 7/15'},
  {key:'support',label:'필요한 지원',tag:'Decision',tagC:'dec',ph:'무엇이·언제까지·왜 필요한지\\n없으면 "없음"이라고 명시',
   bad:'(빈칸)',good:'설비업체 미팅 동석 요청 — 7/16(수) 오후. 환기 설계 변경 최종 확인 필요'},
  {key:'evidence',label:'사실·의견 구분 메모',tag:'Thinking',tagC:'thk',ph:'판단 근거가 직접 확인인지, 추정인지\\n예) 사실: 업체가 메일로 납기지연 통보 / 의견: 대안업체가 더 빠를 것으로 예상',
   bad:'클라이언트가 원하는 것 같다',good:'사실: 7/10 미팅에서 면적 확대 요청 직접 발언. 의견: 예산 내 가능할 것으로 판단(견적 대기 중)'}
];
var BRIEF=[
  {key:'title',label:'프로젝트 제목',tag:'Context',tagC:'ctx',ph:'"[대상]을 위한 [내용]" 형식\\n예) 봉은사 방문객을 위한 도심 명상 체험 프로그램',
   bad:'봉은사 프로젝트',good:'봉은사 방문객을 위한 도심 명상 체험 프로그램'},
  {key:'summary',label:'Executive Summary',tag:'Context',tagC:'ctx',ph:'4문장: 대상 / 문제 / 방식 / 결과',
   bad:'봉은사에서 프로그램을 진행합니다.',good:'대상: 25~45세 도시 직장인. 문제: 봉은사 방문자 90%가 산책만 하고 떠남(체류 20분). 방식: 점심시간 런치명상(30분) 정기 프로그램. 결과: 재방문율 40%↑, 체류시간 2배'},
  {key:'stakeholder',label:'Primary Stakeholder',tag:'Context',tagC:'ctx',ph:'P1/P2/P3 + 각각 원하는 것',
   bad:'우리 회사',good:'P1: 봉은사 주지스님 — 현대적 포교 확대. P2: 직장인 참가자 — 접근성 높은 명상. P3: 기업 HR — 직원 복지 콘텐츠'},
  {key:'client_value',label:'Client Value',tag:'Context',tagC:'ctx',ph:'클라이언트가 얻는 구체적 이익',
   bad:'좋은 프로그램을 제공합니다',good:'봉은사: 25~45세 신규 방문자 월 500명↑, 정기후원 전환 5%. 기업 HR: 직원 스트레스 지수 15% 감소(3개월 측정)'},
  {key:'core_idea',label:'핵심 아이디어 (How)',tag:'Thinking',tagC:'thk',ph:'형식(팝업/전시)이 아닌 해결방식',
   bad:'명상 프로그램을 합니다',good:'출퇴근 동선에 30분 명상을 끼워넣는 "런치명상" — 앱 예약+현장 안내+사후 가이드 3단계 여정 설계'},
  {key:'success_criteria',label:'성공 기준 (KPI)',tag:'Priority',tagC:'pri',ph:'측정 가능한 숫자/상태',
   bad:'성공적으로 마친다',good:'3개월 내 정기참가자 200명, 만족도 4.5/5.0, 재참가율 60%, 후원전환 5%'},
  {key:'support_needed',label:'필요한 지원',tag:'Decision',tagC:'dec',ph:'예산/인력/파트너/의사결정\\n없으면 "없음"',
   bad:'추후 협의',good:'초기 예산 1,500만원 승인(공간조성 800+운영 700). 봉은사 담당 스님 소개 연결. 기한: 8월 초'}
];
var esc=function(s){return String(s==null?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');};
function renderFields(){
  var items=mode==='brief'?BRIEF:DAILY, h='';
  document.getElementById('input-title').textContent=mode==='brief'?'기획안 (1P-Brief) 입력':'일일보고 입력';
  items.forEach(function(f){
    h+='<div class="field"><label>'+esc(f.label)+' <span class="tag tag-'+f.tagC+'">'+esc(f.tag)+'</span>'
      +'<span class="guide-toggle" data-k="'+f.key+'">가이드 ▼</span></label>'
      +'<textarea id="f-'+f.key+'" placeholder="'+esc(f.ph)+'"></textarea>'
      +'<div class="guide-box" id="g-'+f.key+'">'
      +'<div class="g-bad">Bad: '+esc(f.bad)+'</div>'
      +'<div class="g-good">Good: '+esc(f.good)+'</div></div></div>';
  });
  document.getElementById('fields').innerHTML=h;
  document.querySelectorAll('.guide-toggle').forEach(function(el){
    el.onclick=function(){
      var g=document.getElementById('g-'+el.dataset.k);
      g.classList.toggle('open');
      el.textContent=g.classList.contains('open')?'가이드 ▲':'가이드 ▼';
    };
  });
}
function gradeColor(g){
  if(g==='S') return 'var(--green)';
  if(g==='A') return 'var(--accent)';
  if(g==='B') return 'var(--amber)';
  return 'var(--coral)';
}
var ITEM_LABELS={context:'Context (맥락)',objective:'Objective (목표)',evidence:'Evidence (근거)',
  priority:'Priority (우선순위)',risk:'Risk (리스크)',decision:'Decision (결정)',support:'Support (지원)',
  title:'제목',summary:'Executive Summary',stakeholder:'Stakeholder',client_value:'Client Value',
  core_idea:'핵심 아이디어',success_criteria:'성공 기준',support_needed:'필요한 지원'};
function renderResult(d){
  var box=document.getElementById('result');
  if(!d||!d.scores){box.innerHTML='<div class="result-empty">리뷰 결과를 가져오지 못했습니다. 다시 시도해주세요.</div>';return;}
  var pct=d.total, gc=gradeColor(d.grade);
  var r=Math.round(pct/100*251.2); // circumference=2*PI*40
  var h='<div class="score-circle"><svg width="140" height="140"><circle cx="70" cy="70" r="40" fill="none" stroke="var(--bg3)" stroke-width="8"/>'
    +'<circle cx="70" cy="70" r="40" fill="none" stroke="'+gc+'" stroke-width="8" stroke-dasharray="'+r+' 251.2" stroke-linecap="round"/></svg>'
    +'<div class="val"><div class="num" style="color:'+gc+'">'+d.total+'</div><div class="grade" style="color:'+gc+'">'+esc(d.grade)+'</div></div></div>';
  // 보고 유형·채점 모드 배지 (2026-07-20 유형별 루브릭)
  var TL={A:'진행 공유',B:'이슈·리스크',C:'의사결정'};
  if(d.report_type){
    h+='<div style="text-align:center;margin:2px 0 10px;font-size:12px">'
      +'<span style="background:rgba(108,92,231,.15);color:var(--accent);border-radius:6px;padding:2px 10px">'+(TL[d.report_type]||d.report_type)+' 루브릭</span>'
      +(d.mode==='grace'?' <span style="background:rgba(244,196,48,.14);color:#f4c430;border-radius:6px;padding:2px 10px">순화 기준(~10/20)</span>':'')
      +'</div>';
  }
  // ThinkingCost
  var tc=d.thinking_cost||{};
  var stars='';for(var i=0;i<5;i++) stars+=(i<(tc.stars||0))?'★':'☆';
  h+='<div class="tc-row"><div style="font-size:12px;color:var(--muted);margin-bottom:4px">ThinkingCost</div>'
    +'<div class="tc-stars" style="color:'+gc+'">'+stars+'</div>'
    +'<div class="tc-detail">예상 추가질문 '+(tc.predicted_questions||0)+'개'+(tc.detail?(' — '+esc(tc.detail)):'')+'</div></div>';
  // Summary
  if(d.summary) h+='<div class="summary-box">'+esc(d.summary)+'</div>';
  // Item cards
  var scores=d.scores;
  for(var k in scores){
    var s=scores[k], barPct=Math.round(s.score/s.max*100);
    var barC=s.pass?'var(--green)':'var(--coral)';
    h+='<div class="item-card"><div class="item-head"><span class="item-name">'+(s.pass?'✓':'✗')+' '+(ITEM_LABELS[k]||k)+'</span>'
      +'<span class="item-score" style="color:'+barC+'">'+s.score+'/'+s.max+'</span></div>'
      +'<div class="item-bar"><div class="item-bar-fill" style="width:'+barPct+'%;background:'+barC+'"></div></div>'
      +'<div class="item-fb">'+esc(s.feedback||'')+'</div></div>';
  }
  // Patterns
  var pats=d.patterns_detected||[];
  if(pats.length){
    h+='<h3 style="font-size:13px;font-weight:600;margin:14px 0 8px">감지된 안티패턴</h3>';
    pats.forEach(function(p){
      h+='<div class="pat-card"><div class="pat-type">⚠ '+esc(p.type)+(p.field?(' · '+esc(p.field)):'')+'</div>'
        +'<div class="pat-detail">'+esc(p.detail||'')+'</div></div>';
    });
  }
  // Tips
  var tips=d.improvement_tips||[];
  if(tips.length){
    h+='<div class="tips"><h3>개선 팁</h3>';
    tips.forEach(function(t){ h+='<div class="tip-item">'+esc(t)+'</div>'; });
    h+='</div>';
  }
  box.innerHTML=h;
}
document.querySelectorAll('.mode-btn').forEach(function(b){
  b.onclick=function(){
    mode=b.dataset.m;
    document.querySelectorAll('.mode-btn').forEach(function(x){x.classList.toggle('on',x.dataset.m===mode);});
    renderFields();
    document.getElementById('result').innerHTML='<div class="result-empty">왼쪽 폼을 작성하고 "AI 리뷰 받기"를 눌러보세요.</div>';
  };
});
document.getElementById('submit-btn').onclick=function(){
  var btn=this, msg=document.getElementById('submit-msg');
  var items=mode==='brief'?BRIEF:DAILY, fields={}, empty=0;
  items.forEach(function(f){
    var v=(document.getElementById('f-'+f.key)||{}).value||'';
    fields[f.key]=v;
    if(!v.trim()) empty++;
  });
  if(empty===items.length){msg.textContent='최소 1개 항목을 입력해주세요.';return;}
  btn.disabled=true; msg.textContent='AI가 리뷰 중… (10~20초 소요)';
  document.getElementById('result').innerHTML='<div class="result-empty">분석 중…</div>';
  fetch('/api/simulator/review',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({mode:mode,fields:fields})})
    .then(function(r){return r.json();})
    .then(function(d){
      btn.disabled=false; msg.textContent='';
      if(d.ok) renderResult(d.result);
      else{document.getElementById('result').innerHTML='<div class="result-empty">'+(d.error||'리뷰 실패')+'</div>';}
    })
    .catch(function(){btn.disabled=false;msg.textContent='서버 연결 실패';});
};
renderFields();
// --- AI 비교 보기 ---
var draftOpen=false, aiDraft=null;
document.getElementById('draft-toggle').onclick=function(){
  draftOpen=!draftOpen;
  document.getElementById('draft-body').classList.toggle('open',draftOpen);
  document.getElementById('draft-arrow').classList.toggle('open',draftOpen);
};
// draft-btn 활성화: 폼에 1개 이상 입력 + 자유텍스트 입력 시
function checkDraftReady(){
  var items=mode==='brief'?BRIEF:DAILY, hasField=false;
  items.forEach(function(f){if((document.getElementById('f-'+f.key)||{}).value&&(document.getElementById('f-'+f.key)||{}).value.trim()) hasField=true;});
  var hasText=((document.getElementById('draft-text')||{}).value||'').trim().length>0;
  document.getElementById('draft-btn').disabled=!(hasField&&hasText);
}
document.getElementById('draft-text').addEventListener('input',checkDraftReady);
// 폼 필드에도 리스너 (renderFields 후 재부착)
var origRender=renderFields;
renderFields=function(){
  origRender();
  aiDraft=null; clearCompare();
  var items=mode==='brief'?BRIEF:DAILY;
  items.forEach(function(f){
    var el=document.getElementById('f-'+f.key);
    if(el) el.addEventListener('input',checkDraftReady);
  });
  checkDraftReady();
};
renderFields();
function clearCompare(){
  document.querySelectorAll('.ai-compare').forEach(function(el){el.classList.remove('show');el.innerHTML='';});
}
function showCompare(draft){
  var items=mode==='brief'?BRIEF:DAILY;
  items.forEach(function(f){
    var cmp=document.getElementById('cmp-'+f.key);
    if(!cmp) return;
    var val=(draft[f.key]||'').trim();
    if(!val){cmp.classList.remove('show');cmp.innerHTML='';return;}
    cmp.classList.add('show');
    cmp.innerHTML='<div class="ai-compare-label">AI 드래프트</div>'
      +'<div class="ai-compare-text">'+esc(val)+'</div>'
      +'<button class="ai-replace-btn" data-k="'+f.key+'">이 값으로 교체</button>';
    cmp.querySelector('.ai-replace-btn').onclick=function(){
      var ta=document.getElementById('f-'+this.dataset.k);
      if(ta){ta.value=draft[this.dataset.k]||'';ta.style.borderColor='var(--accent)';setTimeout(function(){ta.style.borderColor='';},1500);}
    };
  });
}
// renderFields에서 compare div 삽입
var origRender2=renderFields;
renderFields=function(){
  origRender2();
  // 각 필드 아래에 비교 영역 추가
  var items=mode==='brief'?BRIEF:DAILY;
  items.forEach(function(f){
    var fieldDiv=document.getElementById('f-'+f.key);
    if(!fieldDiv) return;
    var cmpDiv=document.createElement('div');
    cmpDiv.className='ai-compare';cmpDiv.id='cmp-'+f.key;
    fieldDiv.parentNode.appendChild(cmpDiv);
  });
};
renderFields();
document.getElementById('draft-btn').onclick=function(){
  var btn=this, msg=document.getElementById('draft-msg');
  var text=(document.getElementById('draft-text').value||'').trim();
  if(!text){msg.textContent='텍스트를 입력해주세요.';return;}
  btn.disabled=true; msg.textContent='AI가 구조화 중... (5~10초)';
  clearCompare();
  fetch('/api/simulator/draft',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({mode:mode,text:text})})
    .then(function(r){return r.json();})
    .then(function(d){
      btn.disabled=false; msg.textContent='';
      if(d.ok&&d.fields){aiDraft=d.fields;showCompare(d.fields);msg.textContent='비교 결과가 각 필드 아래에 표시됩니다.';}
      else{msg.textContent=d.error||'드래프트 생성 실패';}
      checkDraftReady();
    })
    .catch(function(){btn.disabled=false;msg.textContent='서버 연결 실패';checkDraftReady();});
};
})();
</script></body></html>"""

# / — 역할 인식 통합 셸: 로그인 1회 후 역할별 탭을 iframe으로 전환.
#   대표   = [프로젝트 | 오늘 Brief | 이번 주 | Decision Window] + 팀 스코프 드롭다운
#   리더   = [프로젝트 | 오늘 Brief(팀) | 이번 주(팀)] (2팀 리더는 드롭다운)
#   직원   = [프로젝트]
# iframe 격리라 CSS 충돌 없음. 통합 로그인이 pm_sess(포트폴리오)+brief_sess 등 세션키를 미리
# 세팅해 각 iframe(/projects·/brief·/weekly·/team-*·/arisa2/)이 게이트를 자동 통과.
UNIFIED_SHELL = """<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Project Rent 대시보드</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable.min.css">
<style>
:root{--bg:#1A1A1A;--bg-2:#202020;--bg-3:#262626;--fg:#F5F0EB;--muted:#8A857E;--line:#333;--accent:#6C5CE7;--red:#E17055}
*{box-sizing:border-box;margin:0;padding:0}
html,body{height:100%}
body{background:var(--bg);color:var(--fg);font-family:'Pretendard Variable',sans-serif;font-weight:300}
#login-gate{position:fixed;inset:0;background:var(--bg);display:flex;align-items:center;justify-content:center;z-index:100}
#login-box{background:var(--bg-2);border:1px solid var(--line);border-radius:14px;padding:38px 40px;width:330px;text-align:center}
#login-box h2{font-size:19px;font-weight:600;margin-bottom:6px}
#login-box .lg-sub{color:var(--muted);font-size:13px;margin-bottom:22px}
#login-box input{width:100%;background:var(--bg-3);border:1px solid var(--line);border-radius:8px;padding:12px 14px;color:var(--fg);font-size:14px;margin-bottom:10px;font-family:inherit}
#login-box button{width:100%;background:var(--accent);color:#fff;border:0;border-radius:8px;padding:12px;font-size:14px;font-weight:600;cursor:pointer;margin-top:6px}
#login-err{color:var(--red);font-size:12px;margin-top:12px;min-height:16px}
#shell{display:none;flex-direction:column;height:100vh}
#tabs{display:flex;gap:6px;padding:10px 16px;border-bottom:1px solid var(--line);background:var(--bg-2);align-items:center;flex-shrink:0}
.tab,.tab-ext{background:transparent;border:1px solid var(--line);color:var(--muted);border-radius:8px;padding:8px 18px;font-size:14px;cursor:pointer;font-family:inherit;font-weight:500}
.tab.on{background:var(--accent);color:#fff;border-color:var(--accent)}
#scope-sel{background:var(--bg-3);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:8px 12px;font-size:13px;font-family:inherit}
#tabs .who{margin-left:auto;font-size:12px;color:var(--muted)}
#tabs .who a{color:var(--accent);cursor:pointer;margin-left:8px}
.frame{flex:1;border:0;width:100%;display:none;background:var(--bg)}
.frame.on{display:block}
#f-mywork{overflow:auto;padding:24px}
.mw-wrap{max-width:1000px;margin:0 auto}
.mw-h{font-size:15px;font-weight:600;margin:22px 0 12px}
.mw-form{background:var(--bg-2);border:1px solid var(--line);border-radius:12px;padding:14px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:6px}
.mw-form select,.mw-form input{background:var(--bg-3);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:9px 12px;font-size:13px;font-family:inherit}
.mw-form button{background:var(--accent);color:#fff;border:0;border-radius:8px;padding:9px 18px;font-size:13px;font-weight:600;cursor:pointer}
.mw-assign{background:var(--bg-2);border:1px solid var(--line);border-radius:12px;padding:14px;margin-bottom:8px}
.mw-assign textarea{width:100%;min-height:76px;background:var(--bg-3);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:11px 13px;font-size:13px;font-family:inherit;resize:vertical;line-height:1.5}
.td-actions{display:flex;gap:8px;margin-top:10px;align-items:center}
.td-actions button,.mw-assign button{background:var(--accent);color:#fff;border:0;border-radius:8px;padding:9px 18px;font-size:13px;font-weight:600;cursor:pointer;font-family:inherit}
.td-actions button:disabled{opacity:.5}
button.btn-sec{background:var(--bg-3);color:var(--fg);border:1px solid var(--line)}
.td-row{display:flex;gap:6px;align-items:center;margin-top:7px}
.tg-box{background:var(--bg-2);border:1px solid var(--line);border-radius:12px;padding:12px 14px;margin-top:10px}
.tg-head{display:flex;gap:8px;align-items:center;flex-wrap:wrap;border-bottom:1px solid var(--line);padding-bottom:10px}
.tg-label{font-size:11px;color:var(--muted);flex-shrink:0}
.tg-proj{flex:1;min-width:150px;background:var(--bg-3);border:1px solid var(--accent);color:var(--fg);border-radius:8px;padding:9px 12px;font-size:13px;font-weight:500;font-family:inherit}
.tg-as{background:var(--bg-3);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:9px 10px;font-size:12px;font-family:inherit}
.tg-add{background:transparent;border:1px dashed var(--line);color:var(--muted);border-radius:8px;padding:7px 12px;font-size:12px;cursor:pointer;margin-top:8px;font-family:inherit}
.mw-plan{display:flex;gap:10px;align-items:center;margin-bottom:10px;flex-wrap:wrap}
.mw-file{background:rgba(108,92,231,.12);border:1px solid rgba(108,92,231,.4);color:var(--accent);border-radius:8px;padding:8px 14px;font-size:12.5px;font-weight:600;cursor:pointer}
.mw-file:hover{background:var(--accent);color:#fff}
.pg-head{font-size:13px;font-weight:600;color:var(--accent);margin:16px 0 6px;display:flex;align-items:center;gap:8px}
.pg-cnt{font-size:11px;color:var(--muted);font-weight:400}
.pg-item{margin-left:10px;border-left:2px solid var(--line)}
.td-row .td-proj{width:120px;flex-shrink:0;background:var(--bg-3);border:1px solid var(--accent);color:var(--fg);border-radius:8px;padding:9px 10px;font-size:12px;font-family:inherit}
.td-row .td-task{flex:1;min-width:140px;background:var(--bg-3);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:9px 12px;font-size:13px;font-family:inherit}
.td-row select,.td-row input[type=date]{background:var(--bg-3);border:1px solid var(--line);color:var(--fg);border-radius:8px;padding:9px 10px;font-size:12px;font-family:inherit}
.td-row .td-del{background:transparent;border:1px solid var(--line);color:var(--muted);border-radius:8px;width:32px;height:34px;cursor:pointer;font-size:16px;flex-shrink:0}
.mw-card{background:var(--bg-2);border:1px solid var(--line);border-radius:10px;padding:12px 15px;margin-bottom:7px}
.mw-card .t{font-size:14px;font-weight:500}
.mw-card .m{font-size:12px;color:var(--muted);margin-top:4px}
.mw-badge{font-size:11px;border-radius:5px;padding:1px 7px;margin-left:8px}
.st-act{cursor:pointer;font-size:11px;border:1px solid var(--line);border-radius:6px;padding:2px 9px;margin-left:6px;color:var(--muted);white-space:nowrap}
.st-act:hover{border-color:var(--accent);color:var(--accent)}
.st-wait{font-size:11px;border-radius:5px;padding:1px 7px;margin-left:8px;background:rgba(244,196,48,.14);color:#f4c430}
.st-chk{accent-color:var(--red);width:14px;height:14px;cursor:pointer;margin-left:8px;vertical-align:middle}
#mw-bulkbar{display:none;position:fixed;bottom:26px;left:50%;transform:translateX(-50%);background:var(--bg-2);
  border:1px solid var(--red);border-radius:12px;padding:10px 16px;gap:12px;align-items:center;z-index:60;
  box-shadow:0 6px 24px rgba(0,0,0,.5);font-size:13px}
#mw-bulkbar button{border:0;border-radius:8px;padding:7px 14px;cursor:pointer;font-family:inherit;font-size:13px}
#mw-bulkbar .bb-del{background:var(--red);color:#fff;font-weight:600}
#mw-bulkbar .bb-clear{background:transparent;border:1px solid var(--line);color:var(--muted)}
.mw-urgent{background:rgba(225,112,85,.16);color:var(--red)}
.mw-proj{border-left:2px solid var(--accent);padding-left:14px;margin-bottom:16px}
.mw-proj-name{font-size:14px;font-weight:600;margin-bottom:8px}
.mw-empty{color:var(--muted);font-size:13px;padding:8px 0}
.mw-quick{margin:2px 0 6px;display:flex;gap:8px;flex-wrap:wrap}
.mw-link{display:inline-flex;align-items:center;gap:6px;background:rgba(108,92,231,.12);border:1px solid rgba(108,92,231,.4);color:var(--accent);border-radius:8px;padding:8px 14px;font-size:12.5px;font-weight:600;text-decoration:none}
.mw-link:hover{background:var(--accent);color:#fff}
.mw-done-sec{margin:4px 0 16px;border:1px solid var(--line);border-radius:10px;background:var(--bg-2)}
.mw-done-sec summary{cursor:pointer;color:var(--muted);font-size:13px;padding:10px 14px}
.mw-done-sec[open] summary{border-bottom:1px solid var(--line)}
.mw-done-sec .mw-card{border:0;border-top:1px solid var(--line);border-radius:0;margin:0}
.mw-done-sec .mw-card:first-of-type{border-top:0}
.pf-bar{display:flex;gap:6px;flex-wrap:wrap;margin:0 0 10px}
.pf-chip{cursor:pointer;font-size:12px;border:1px solid var(--line);border-radius:16px;padding:5px 12px;color:var(--muted);background:var(--bg-2)}
.pf-chip.on{border-color:var(--accent);color:var(--accent);background:rgba(108,92,231,.12)}
.il-sel{background:var(--bg-3);border:1px solid var(--line);color:var(--muted);border-radius:6px;padding:2px 6px;font-size:11px;font-family:inherit;margin-left:6px;cursor:pointer;max-width:110px}
.il-sel:hover{border-color:var(--accent);color:var(--accent)}
#arisa2-status{font-size:11px;color:var(--muted);margin-left:6px}
#arisa2-status.ok{color:#00b894}
#arisa2-status.err{color:var(--red)}
</style></head>
<body>
<div id="login-gate"><div id="login-box">
  <h2>Project Rent 대시보드</h2>
  <div class="lg-sub">이름 + PIN (첫 로그인 시 PIN 설정)</div>
  <input id="lg-id" placeholder="이름" autocomplete="username">
  <input id="lg-pin" type="password" placeholder="PIN" autocomplete="current-password">
  <button id="lg-btn">로그인</button>
  <div id="login-err"></div>
</div></div>
<div id="shell">
  <div id="tabs">
    <button class="tab on" data-t="mywork">내 업무</button>
    <button class="tab" data-t="projects">프로젝트</button>
    <button class="tab" data-t="simulator">보고 시뮬레이터</button>
    <button class="tab" data-t="brief" style="display:none">오늘 Brief</button>
    <button class="tab" data-t="weekly" style="display:none">이번 주</button>
    <button class="tab-ext" id="tab-hr" style="display:none" onclick="window.open('https://rent-hr-portal.fly.dev/','_blank')">HR 포털 ↗</button>
    <select id="scope-sel" style="display:none"></select>
    <span class="who" id="who"></span>
  </div>
  <div class="frame on" id="f-mywork"></div>
  <iframe class="frame" id="f-projects"></iframe>
  <iframe class="frame" id="f-brief"></iframe>
  <iframe class="frame" id="f-weekly"></iframe>
  <iframe class="frame" id="f-simulator"></iframe>
</div>
<script>
(function(){
  var gate=document.getElementById('login-gate'), shell=document.getElementById('shell'), who=document.getElementById('who');
  var sel=document.getElementById('scope-sel');
  var frames={mywork:document.getElementById('f-mywork'),projects:document.getElementById('f-projects'),
              brief:document.getElementById('f-brief'),weekly:document.getElementById('f-weekly'),
              simulator:document.getElementById('f-simulator')};
  var SESS=null, curTab='mywork', curScope='', loaded={projects:false,brief:false,weekly:false,simulator:false};
  var MW_ASSIGNEES=[];
  var SESS_KEYS=['pm_sess','brief_sess','weekly_sess','team_brief_sess','team_weekly_sess'];
  var CLEAR_KEYS=SESS_KEYS.concat(['arisa_sess','arisa_token']);
  function tabBtn(t){ return document.querySelector('.tab[data-t="'+t+'"]'); }
  function srcFor(t){
    if(t==='projects') return '/projects';
    if(t==='simulator') return '/simulator';
    if(t==='decision') return '/arisa2/';
    var lt=SESS.lead_teams||[];
    if(t==='brief'){
      if(SESS.admin) return curScope==='' ? '/brief' : '/team-brief?team='+encodeURIComponent(curScope);
      if(lt.length) return curScope==='' ? '/lead-brief?teams='+encodeURIComponent(lt.join(',')) : '/team-brief?team='+encodeURIComponent(curScope);
      return '/my-brief?name='+encodeURIComponent(SESS.name);  // 직원 — 내 카드 + 팀 헤드라인
    }
    if(SESS.admin && curScope==='') return '/weekly';
    return '/team-weekly?team='+encodeURIComponent(curScope||lt[0]||'');
  }
  function showTab(t){
    curTab=t;
    document.querySelectorAll('.tab').forEach(function(b){ b.classList.toggle('on', b.dataset.t===t); });
    Object.keys(frames).forEach(function(k){ frames[k].classList.toggle('on', k===t); });
    if(t==='mywork'){ renderMyWork(); return; }  // div 직접 렌더(iframe 아님) — 매번 최신
    if(!loaded[t]){ frames[t].src=srcFor(t); loaded[t]=true; }
  }
  function loadScope(scope){
    curScope=scope; loaded.brief=false; loaded.weekly=false;
    frames.brief.src='about:blank'; frames.weekly.src='about:blank';
    if(curTab==='brief'||curTab==='weekly'){ frames[curTab].src=srcFor(curTab); loaded[curTab]=true; }
  }
  function arisa2Login(s){
    // PIN 단일화(users.json symlink)로 같은 자격 재사용 → ARISA 2.0 토큰 발급. Promise<bool> 반환(showTab이 await).
    return fetch('/arisa2/api/login?name='+encodeURIComponent(s.name)+'&pin='+encodeURIComponent(s.pin||''),{method:'POST'})
      .then(function(r){ if(!r.ok) throw new Error('login '+r.status); return r.json(); })
      .then(function(d){
        if(!(d && d.ok)) throw new Error('bad response');
        localStorage.setItem('arisa_sess',JSON.stringify(d.user));
        localStorage.setItem('arisa_token',d.token);
        if(loaded.decision){ frames.decision.src=srcFor('decision'); } // 이미 열려 있었으면 재로드
        return true;
      }).catch(function(e){ console.warn('[arisa2Login]', e && e.message); return false; });
  }
  function checkArisa2(){
    fetch('/arisa2/api/health',{method:'GET'}).then(function(r){
      var ok=r.status<500; // 401 등 4xx = 서버 살아있음(인증 게이트)
      a2s.textContent=ok?'ARISA 2.0 ON':'ARISA 2.0 OFF';
      a2s.className=ok?'ok':'err';
    }).catch(function(){ a2s.textContent='ARISA 2.0 OFF'; a2s.className='err'; });
  }
  function esc(s){ return String(s==null?'':s).replace(/[&<>"]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];}); }
  function mwAssignHtml(lv){
    return '<div class="mw-h">업무 분장 <span class="sub2">— 자유롭게 적으면 AI가 항목화 → 편집 후 '+lv+'에게 배분 → 승인'+(lv==='팀원'?' (대표가 배분한 내 업무를 참고해도 됩니다)':'')+'</span></div>'
      +'<div class="mw-assign">'
      +'<div class="mw-plan"><label class="mw-file" for="mw-xlsx">📄 주간업무계획 업로드 (.xlsx)</label>'
      +'<input type="file" id="mw-xlsx" accept=".xlsx" style="display:none">'
      +'<a href="/guide/template.xlsx" style="color:var(--accent);font-size:12px;text-decoration:none">📥 템플릿 받기</a>'
      +'<span class="sub2">— 금주·차주 업무를 프로젝트별 검토 초안으로 자동 변환</span></div>'
      +'<textarea id="mw-text" placeholder="이번주 팀 업무를 자유롭게 적으세요.&#10;예) 봉은사 마스터플랜 검토·정리본 공유, 세스크멘슬 주방도면 정리, KBO 굿즈 발주 최종확인"></textarea>'
      +'<div class="td-actions"><button id="mw-parse">AI로 항목 정리</button></div>'
      +'<div id="mw-todos"></div><div id="mw-msg" class="sub2"></div></div>';
  }
  function mwBindParse(){
    var fx=document.getElementById('mw-xlsx');
    if(fx){ fx.onchange=function(){
      var f=fx.files&&fx.files[0], msg=document.getElementById('mw-msg');
      if(!f) return;
      msg.textContent='📄 '+f.name+' — 주간업무계획을 검토 초안으로 변환 중…';
      var rd=new FileReader();
      rd.onload=function(){
        fetch('/api/assign-from-plan',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:SESS.name,pin:SESS.pin,b64:rd.result})})
          .then(function(r){return r.json();}).then(function(d){
            fx.value='';
            if(d.ok && d.items && d.items.length){
              msg.textContent='검토 초안 '+d.items.length+'건 생성 ('+ (d.rows||0) +'개 행) — 담당자 지정 후 등록하세요.';
              mwRenderTodos(d.items);
            } else { msg.textContent=d.error||'항목을 뽑지 못했습니다.'; }
          }).catch(function(){ fx.value=''; msg.textContent='서버 오류'; });
      };
      rd.readAsDataURL(f);
    };}
    var pb=document.getElementById('mw-parse');
    if(!pb) return;
    pb.onclick=function(){
      var txt=document.getElementById('mw-text').value.trim(), msg=document.getElementById('mw-msg');
      if(!txt){ msg.textContent='업무 내용을 입력하세요'; return; }
      msg.textContent='AI가 항목으로 정리 중…'; pb.disabled=true;
      fetch('/api/assign-parse',{method:'POST',headers:{'Content-Type':'application/json'},
        body:JSON.stringify({user:SESS.name,pin:SESS.pin,text:txt})})
        .then(function(r){return r.json();}).then(function(d){
          pb.disabled=false;
          if(d.ok && d.items && d.items.length){ msg.textContent='검토 후 각 항목에 담당자를 지정하고 승인하세요.'; mwRenderTodos(d.items); }
          else { msg.textContent='항목을 뽑지 못했습니다. 더 구체적으로 적어보세요.'; }
        }).catch(function(){ pb.disabled=false; msg.textContent='서버 오류'; });
    };
  }
  function renderLeadHome(){
    // 리더 홈 — 팀 Todo(대표·리더 분장) + 분장 생성 + 진행중 프로젝트 + 팀원 오늘 보고
    var box=frames.mywork;
    box.innerHTML='<div class="mw-wrap"><div class="mw-empty">불러오는 중…</div></div>';
    var u=encodeURIComponent(SESS.name);
    Promise.all([
      fetch('/api/lead-home?user='+u).then(function(r){return r.json();}),
      fetch('/api/assignees?user='+u).then(function(r){return r.json();})
    ]).then(function(res){
      var lh=res[0]||{}, ac=res[1]||{}, h='<div class="mw-wrap">';
      MW_ASSIGNEES = ac.assignees || [];
      var A=lh.assignments||[];
      // 위계 분리: 받은 업무(대표→리더 본인) vs 팀 Todo(리더→팀원 배분분)
      var _rcSeen={};
      var received=A.filter(function(a){
        if(a.assignee!==SESS.name) return false;
        var k=[a.project,a.task,a.deadline,a.status].join('|');
        if(_rcSeen[k]) return false; _rcSeen[k]=1; return true;
      });
      var teamTodo=A.filter(function(a){ return a.assignee!==SESS.name; });
      h+=mwDailyFocusHtml(lh.daily);
      var lhNew=mwNewTodosHtml(lh.daily);
      if(lhNew){ h+='<div class="mw-h">🆕 새로 해야 할 일 <span class="sub2">어제 보고에서 도출 — [분장 등록]으로 확정</span></div>'+lhNew; }
      h+=mwProjUpdatesHtml(lh.daily);
      h+=mwApprovalsHtml(lh.approvals||[]);
      if(received.length){
        h+='<div class="mw-h">📥 받은 업무 <span class="sub2">— 대표 지시·리더 이관 · [상세 분장]으로 팀원에게 쪼개서 배분하세요</span></div>';
        received.forEach(function(a,i){
          var st=a.status||'미착수';
          var badge='<span class="lh-st '+(st==='완료'?'lh-done':(st==='진행중'?'lh-doing':'lh-todo'))+'">'+esc(st)+'</span>';
          var urg=(a.priority==='긴급')?'<span class="mw-badge mw-urgent">긴급</span>':'';
          var pj=a.project?(esc(a.project)+' · '):'';
          var dl=a.deadline?(' · 마감 '+esc(a.deadline)):'';
          var act='';
          if(a.row){
            if(st==='완료'){ act='<span class="st-wait">⏳ 승인 대기</span>'+mwStBtn(a,'진행중','↩ 되돌리기'); }
            else{ act=mwStBtn(a,'완료','✓ 완료')+mwStBtn(a,'완료','📣 완료·보고',true); }
          }
          h+='<div class="mw-card rc-card"><div class="t">'+badge+' '+esc(a.task)+urg+act
            +'<button class="rc-btn" data-i="'+i+'">→ 팀원에게 상세 분장</button></div>'
            +'<div class="m">'+pj+esc(a.src||'대표 지시')+dl+'</div></div>';
        });
      }
      h+=mwDueHtml(A);                              // A1 — 지난·오늘·내일 마감 모아보기
      h+=mwUnassignedHtml(lh.unassigned||[]);       // A2 — 담당 미지정 배정 큐
      h+='<div class="mw-h">팀 Todo · 이번주 분장 <span class="sub2">'+esc((lh.teams||[]).join(' · '))+'</span></div>';
      if(teamTodo.length){
        // B3 — 담당자별 보기 (filament 팀 보드 사이드바 패턴: 이름 + 열린 건수)
        var pf={}, pfo=[];
        teamTodo.forEach(function(a){ var n=a.assignee||'미지정';
          if(!(n in pf)){ pf[n]=0; pfo.push(n); }
          if(a.status==='미착수'||a.status==='진행중') pf[n]++;
        });
        h+='<div class="pf-bar"><span class="pf-chip on" data-name="">전체 '+teamTodo.length+'</span>';
        pfo.forEach(function(n){ h+='<span class="pf-chip" data-name="'+esc(n)+'">'+esc(n)+' '+pf[n]+'</span>'; });
        h+='</div><div id="lh-todo-list">'+mwAssignListHtml(teamTodo, true)+'</div>';
      }
      else { h+='<div class="mw-empty">팀원에게 배분된 분장이 아직 없습니다. 아래에서 바로 만들 수 있습니다.</div>'; }
      if(ac.canAssign){ h+=mwAssignHtml(ac.level||'팀원'); }
      h+='<div class="mw-h">진행중인 프로젝트 <span class="sub2">'+(lh.projects||[]).length+'건 — 클릭하면 프로젝트 탭</span></div>';
      var P=lh.projects||[];
      if(P.length){ P.forEach(function(p){
        var pr=(p.task_total?(' · 업무 '+p.task_done+'/'+p.task_total+(p.percent!=null?(' · '+p.percent+'%'):'')):'');
        h+='<div class="lh-proj" data-open="projects"><div class="t">'+esc(p.name)+'</div><div class="m">PM '+esc(p.pm||'-')+(p.dday?(' · D-day '+esc(p.dday)):'')+pr+'</div></div>';
      }); } else { h+='<div class="mw-empty">진행중인 팀 프로젝트가 없습니다.</div>'; }
      h+='<div class="mw-h">팀원 오늘 보고 <span class="sub2">'+esc(lh.brief_date||'')+'</span></div>';
      h+=(lh.brief_html||'<div class="mw-empty">보고가 아직 없습니다.</div>');
      h+='</div>'; box.innerHTML=h;
      mwBindParse();
      mwBindStatus(box);
      mwBindBulk(box);
      mwBindSelfAssign(box);
      mwBindEdit(box);
      mwBindInline(box);
      box.querySelectorAll('.pf-chip').forEach(function(ch){
        ch.onclick=function(){
          box.querySelectorAll('.pf-chip').forEach(function(c){ c.classList.toggle('on', c===ch); });
          var n=ch.getAttribute('data-name');
          var list=document.getElementById('lh-todo-list'); if(!list) return;
          list.innerHTML=mwAssignListHtml(n?teamTodo.filter(function(a){return (a.assignee||'미지정')===n;}):teamTodo, true);
          mwBindStatus(list); mwBindBulk(box); mwBindEdit(list); mwBindInline(list);
        };
      });
      box.querySelectorAll('.lh-proj').forEach(function(el){ el.onclick=function(){ showTab('projects'); }; });
      box.querySelectorAll('.rc-btn').forEach(function(b){
        b.onclick=function(){
          var a=received[parseInt(b.getAttribute('data-i'),10)]; if(!a) return;
          var ta=document.getElementById('mw-text'); if(!ta) return;
          ta.value=(a.project?('['+a.project+'] '):'')+a.task+' — 이를 실행하기 위한 상세 업무:\\n- ';
          ta.focus(); ta.scrollIntoView({behavior:'smooth',block:'center'});
        };
      });
    }).catch(function(){ box.innerHTML='<div class="mw-wrap"><div class="mw-empty">불러오기 실패</div></div>'; });
  }
  function renderMyWork(){
    var lt=SESS.lead_teams||[];
    if(!SESS.admin && lt.length){ renderLeadHome(); return; }  // 리더 → 팀 홈
    var box=frames.mywork;
    box.innerHTML='<div class="mw-wrap"><div class="mw-empty">불러오는 중…</div></div>';
    var u=encodeURIComponent(SESS.name);
    Promise.all([
      fetch('/api/my-work?user='+u).then(function(r){return r.json();}),
      fetch('/api/assignees?user='+u).then(function(r){return r.json();}),
      SESS.admin ? fetch('/api/exec-attn?user='+u).then(function(r){return r.json();}) : Promise.resolve(null)
    ]).then(function(res){
      var mw=res[0]||{}, ac=res[1]||{}, ex=res[2], h='<div class="mw-wrap">';
      MW_ASSIGNEES = ac.assignees || [];
      if(SESS.admin){ h+='<div class="mw-quick"><a class="mw-link" href="https://rent-hr-portal.fly.dev/" target="_blank" rel="noopener">🏢 HR 포털 바로가기 ↗</a><a class="mw-link" href="https://rent-hr-portal.fly.dev/onboard/admin/v2" target="_blank" rel="noopener">🧾 입퇴사 온보딩 ↗</a></div>'; }  // 대표 전용 — 독립 서비스 새 창(탭바 tab-hr와 동일 링크), 온보딩=입퇴사 자동화 대시보드 v2
      if(ex && ex.ok){
        var D=ex.decisions||[], O=ex.overdue||[];
        if(D.length){
          h+='<div class="mw-h">🧾 결재·확인 필요 <span class="sub2">'+D.length+'건 — 미해결 결정·승인 요청</span></div>';
          D.forEach(function(d){
            var age=d.age>0?(' · '+d.age+'일 경과'):'';
            var pj=d.project?(' · '+esc(d.project)):'';
            h+='<div class="mw-card ex-red"><div class="t">'+esc(d.title)+'</div><div class="m">'+esc(d.who||'')+pj+age+'</div></div>';
          });
        }
        if(O.length){
          h+='<div class="mw-h">⏰ 지연 업무 <span class="sub2">'+O.length+'건 — 마감 경과·미완료 분장</span></div>';
          O.forEach(function(a){
            var pj=a.project?(esc(a.project)+' · '):'';
            h+='<div class="mw-card ex-amber"><div class="t">'+esc(a.task)+' <span class="mw-badge mw-urgent">D+'+a.days_overdue+'</span>'
              +(a.row?(mwStBtn(a,'삭제','🗑')+mwChk(a)):'')+'</div>'
              +'<div class="m">'+pj+esc(a.assignee||'미지정')+' · 마감 '+esc(a.deadline||'')+' · '+esc(a.status||'미착수')+'</div></div>';
          });
        }
        h+=mwUnassignedHtml(ex.unassigned||[]);   // A2 — 담당 미지정 배정 큐 (대표: 리더급에게 배정)
        h+=mwApprovalsHtml(ex.approvals||[]);
        if(!D.length && !O.length && !(ex.approvals||[]).length && !(ex.unassigned||[]).length){ h+='<div class="mw-h">✅ 지연·결재·승인 대기 없음</div>'; }
      }
      if(ac.canAssign){ h+=mwAssignHtml(ac.level||'담당자'); }
      h+=mwDailyFocusHtml(mw.daily);
      h+='<div class="mw-h">✅ 오늘 할 일 · 내 분장 <span class="sub2">🆕 제안 = 어제 보고에서 도출 — [분장 등록]으로 확정</span></div>';
      var A=mw.assignments||[];
      if(A.length){ h+=mwAssignListHtml(A, false); }
      h+=mwNewTodosHtml(mw.daily);
      if(!A.length && !((mw.daily||{}).new_todos||[]).length){ h+='<div class="mw-empty">배정된 분장이 없습니다.</div>'; }
      h+=mwProjUpdatesHtml(mw.daily);
      h+='<div class="mw-h">내 프로젝트 일정</div>';
      var P=mw.projects||[];
      if(P.length){ P.forEach(function(p){
        h+='<div class="mw-proj"><div class="mw-proj-name">'+esc(p.name)+(p.dday?(' <span style="color:var(--muted);font-weight:400;font-size:12px">D-day '+esc(p.dday)+'</span>'):'')+'</div>';
        (p.tasks||[]).forEach(function(t){
          var per=(t.start||'')+(t.end?('~'+t.end):'');
          h+='<div class="mw-card"><div class="t">'+esc(t.task||'')+'</div><div class="m">'+esc(t.status||'')+' · '+esc(t.division||'')+(per?(' · '+esc(per)):'')+'</div></div>';
        });
        h+='</div>';
      }); } else { h+='<div class="mw-empty">배정된 프로젝트 업무가 없습니다.</div>'; }
      h+='</div>'; box.innerHTML=h;
      mwBindParse();
      mwBindStatus(box);
      mwBindBulk(box);
      mwBindSelfAssign(box);
      mwBindEdit(box);
      mwBindInline(box);
    }).catch(function(){ box.innerHTML='<div class="mw-wrap"><div class="mw-empty">불러오기 실패</div></div>'; });
  }
  function mwAsOpts(sel){ return MW_ASSIGNEES.map(function(a){var lbl=a.name+' ('+(a.team||'')+(a.leader?' 리더 · 이관':'')+')';return '<option value="'+esc(a.name)+'"'+(a.name===sel?' selected':'')+'>'+esc(lbl)+'</option>';}).join(''); }
  function mwTodoRow(it){
    it=it||{};
    return '<div class="td-row"><input class="td-task" value="'+esc(it.task||'')+'" placeholder="업무 내용">'
      +'<select class="td-as"><option value="">담당자</option>'+mwAsOpts(it.assignee||'')+'</select>'
      +'<input class="td-dl" type="date" value="'+esc(it.deadline||'')+'" title="마감">'
      +'<select class="td-pr"><option'+(it.priority==='긴급'?'':' selected')+'>일반</option><option'+(it.priority==='긴급'?' selected':'')+'>긴급</option></select>'
      +'<button class="td-del" title="삭제">×</button></div>';
  }
  function mwBindRow(row){ row.querySelector('.td-del').onclick=function(){ row.remove(); }; }
  function mwGroupBox(project, items){
    // 프로젝트 그룹 에디터 — 프로젝트명은 그룹당 1개 입력, 담당자는 항목별 개별 지정
    return '<div class="tg-box"><div class="tg-head">'
      +'<span class="tg-label">프로젝트</span><input class="tg-proj" value="'+esc(project)+'" placeholder="(비우면 기타)">'
      +'<span class="tg-label">일괄 담당자</span><select class="tg-as"><option value="">선택 시 빈 담당자 채움</option>'+mwAsOpts('')+'</select>'
      +'</div><div class="tg-rows">'+items.map(mwTodoRow).join('')+'</div>'
      +'<button class="tg-add">+ 이 프로젝트에 항목 추가</button></div>';
  }
  function mwBindGroup(g){
    g.querySelectorAll('.td-row').forEach(mwBindRow);
    g.querySelector('.tg-add').onclick=function(){
      var d=document.createElement('div'); d.innerHTML=mwTodoRow({}); var row=d.firstChild;
      mwBindRow(row); g.querySelector('.tg-rows').appendChild(row);
    };
    g.querySelector('.tg-as').onchange=function(){
      var v=this.value; if(!v) return;
      g.querySelectorAll('.td-as').forEach(function(s){ if(!s.value) s.value=v; });
      this.selectedIndex=0;
    };
  }
  function mwRenderTodos(items){
    var box=document.getElementById('mw-todos');
    var groups={}, order=[];
    (items||[]).forEach(function(it){
      var p=(it.project||'').trim();
      if(!(p in groups)){ groups[p]=[]; order.push(p); }
      groups[p].push(it);
    });
    if(!order.length) order.push('');
    box.innerHTML=order.map(function(p){ return mwGroupBox(p, groups[p]||[{}]); }).join('')
      +'<div class="td-actions"><button id="mw-addgrp" class="btn-sec">+ 새 프로젝트 그룹</button>'
      +'<button id="mw-commit">최종 승인 · 등록</button></div>';
    box.querySelectorAll('.tg-box').forEach(mwBindGroup);
    document.getElementById('mw-addgrp').onclick=function(){
      var d=document.createElement('div'); d.innerHTML=mwGroupBox('', [{}]); var g=d.firstChild;
      mwBindGroup(g); box.insertBefore(g, box.querySelector('.td-actions'));
    };
    document.getElementById('mw-commit').onclick=mwCommit;
  }
  function mwCommit(){
    var items=[], msg=document.getElementById('mw-msg');
    document.querySelectorAll('#mw-todos .tg-box').forEach(function(g){
      if(g.id==='mw-projconf') return;
      var proj=g.querySelector('.tg-proj'); if(!proj) return;
      var pv=proj.value.trim();
      g.querySelectorAll('.td-row').forEach(function(r){
        var task=r.querySelector('.td-task').value.trim();
        if(task) items.push({assignee:r.querySelector('.td-as').value, task:task, project:pv,
          deadline:r.querySelector('.td-dl').value, priority:r.querySelector('.td-pr').value});
      });
    });
    if(!items.length){ msg.textContent='등록할 항목이 없습니다'; return; }
    if(items.some(function(i){return !i.assignee;})){ msg.textContent='모든 항목에 담당자를 지정하세요'; return; }
    var names=[]; items.forEach(function(i){ var p=(i.project||'').trim(); if(p && names.indexOf(p)<0) names.push(p); });
    if(!names.length){ mwSend(items, []); return; }
    msg.textContent='프로젝트 확인 중…';
    fetch('/api/assign-project-check',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({user:SESS.name,pin:SESS.pin,names:names})})
      .then(function(r){return r.json();}).then(function(d){
        if(!d.ok){ msg.textContent=d.error||'프로젝트 확인 실패'; return; }
        var news=(d.results||[]).filter(function(x){return x.isNew;});
        if(!news.length){ mwSend(items, []); return; }
        mwRenderProjConfirm(items, news, d.all||[], d.results||[]);
      }).catch(function(){ msg.textContent='서버 오류'; });
  }
  function pcPmOpts(){
    var names=MW_ASSIGNEES.map(function(a){return a.name;});
    if(names.indexOf(SESS.name)<0) names.unshift(SESS.name);
    return names.map(function(n){return '<option'+(n===SESS.name?' selected':'')+'>'+esc(n)+'</option>';}).join('');
  }
  function mwRenderProjConfirm(items, news, all, results){
    // 신규 프로젝트 확인 패널 — 그룹별 [신규 생성(PM·일정)] 또는 [기존에 합치기] 선택 후 확정
    var msg=document.getElementById('mw-msg'); msg.textContent='';
    var old=document.getElementById('mw-projconf'); if(old) old.remove();
    var today=new Date().toISOString().slice(0,10);
    var box=document.createElement('div'); box.id='mw-projconf'; box.className='tg-box';
    var h='<div class="tg-head"><span class="tg-label">🆕 신규 프로젝트 확인</span>'
      +'<span class="sub2">포트폴리오에 없는 이름입니다 — 신규 생성 또는 기존 프로젝트에 합치세요</span></div>';
    var matchedInfo=(results||[]).filter(function(x){return x.matched && (x.name!==x.matched.name || x.matched.archived);})
      .map(function(x){ return '“'+esc(x.name)+'” → 기존 “'+esc(x.matched.name)+'”에 반영됩니다'
        +(x.matched.archived?' <span style="color:#f0a05a">(📦 아카이브된 프로젝트 — 재개하려면 대표가 복원 필요)</span>':''); });
    if(matchedInfo.length){ h+='<div class="sub2" style="padding:4px 2px">'+matchedInfo.join(' · ')+'</div>'; }
    news.forEach(function(n,i){
      var opts='<option value="">— 기존 프로젝트 선택 —</option>', candIds={};
      (n.candidates||[]).forEach(function(c){ candIds[c.id]=1; opts+='<option value="'+esc(c.id)+'">'+esc(c.name)+' (유사)</option>'; });
      (all||[]).forEach(function(c){ if(!candIds[c.id]) opts+='<option value="'+esc(c.id)+'">'+esc(c.name)+'</option>'; });
      var dls=items.filter(function(it){return (it.project||'').trim()===n.name && it.deadline;})
        .map(function(it){return it.deadline;}).sort();
      var endDef=dls.length?dls[dls.length-1]:'';
      var nameNote='';
      if(n.nameError){ nameNote='<div class="sub2" style="color:#f0a05a;margin-bottom:4px">⚠️ '+esc(n.nameError)+' — 이름을 수정한 뒤 등록하세요</div>'; }
      else if(n.cleaned){ nameNote='<div class="sub2" style="margin-bottom:4px">이름 규칙 자동정리 → “'+esc(n.cleaned)+'”로 등록됩니다 (괄호·특수문자 제거)</div>'; }
      h+='<div class="pc-row" data-name="'+esc(n.name)+'" style="padding:10px 2px;border-top:1px solid rgba(255,255,255,.08)">'
        +'<div style="font-weight:600;margin-bottom:6px">🆕 '+esc(n.name)+'</div>'+nameNote
        +'<label><input type="radio" name="pc-act-'+i+'" value="create" checked> 신규 생성</label>'
        +'<label style="margin-left:12px"><input type="radio" name="pc-act-'+i+'" value="merge"> 기존에 합치기</label>'
        +'<div class="pc-create" style="margin-top:6px">'
          +'<span class="tg-label">PM(리드)</span><select class="pc-pm">'+pcPmOpts()+'</select> '
          +'<span class="tg-label">시작</span><input type="date" class="pc-start" value="'+today+'"> '
          +'<span class="tg-label">종료</span><input type="date" class="pc-end" value="'+esc(endDef)+'">'
        +'</div>'
        +'<div class="pc-merge" style="margin-top:6px;display:none"><select class="pc-target">'+opts+'</select></div>'
        +'</div>';
    });
    h+='<div class="td-actions" style="margin-top:8px"><button id="pc-ok">확정 등록</button>'
      +'<button id="pc-cancel" class="btn-sec">취소</button></div>';
    box.innerHTML=h;
    var todos=document.getElementById('mw-todos'); todos.parentNode.insertBefore(box, todos.nextSibling);
    box.querySelectorAll('.pc-row').forEach(function(row){
      row.querySelectorAll('input[type=radio]').forEach(function(rb){ rb.onchange=function(){
        var m=this.value==='merge';
        row.querySelector('.pc-create').style.display=m?'none':'';
        row.querySelector('.pc-merge').style.display=m?'':'none';
      };});
    });
    document.getElementById('pc-cancel').onclick=function(){ box.remove(); };
    document.getElementById('pc-ok').onclick=function(){
      var acts=[], bad=null;
      box.querySelectorAll('.pc-row').forEach(function(row){
        var name=row.getAttribute('data-name');
        var act=row.querySelector('input[type=radio]:checked').value;
        if(act==='merge'){
          var t=row.querySelector('.pc-target').value;
          if(!t){ bad='“'+name+'”: 합칠 기존 프로젝트를 선택하세요'; return; }
          acts.push({name:name, action:'merge', mergeId:t});
        } else {
          acts.push({name:name, action:'create', pm:row.querySelector('.pc-pm').value||SESS.name,
            start:row.querySelector('.pc-start').value, end:row.querySelector('.pc-end').value});
        }
      });
      if(bad){ msg.textContent=bad; return; }
      box.remove();
      mwSend(items, acts);
    };
  }
  function mwSend(items, projectActions){
    var msg=document.getElementById('mw-msg');
    msg.textContent='등록 중…';
    fetch('/api/assign-commit',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({user:SESS.name,pin:SESS.pin,items:items,projectActions:projectActions||[]})})
      .then(function(r){return r.json();}).then(function(d){
        if(d.added){
          var extra=[];
          if(d.created&&d.created.length) extra.push('신규 프로젝트 생성: '+d.created.join(', '));
          if(d.merged&&d.merged.length) extra.push('기존 프로젝트에 합침: '+d.merged.join(', '));
          if(extra.length) alert('등록 완료 ('+d.added+'건) — '+extra.join(' / '));
          renderMyWork();
        } else { msg.textContent=(d.errors&&d.errors.join(', '))||'등록 실패(주간분장 탭 확인)'; }
      }).catch(function(){ msg.textContent='서버 오류'; });
  }
  function mwDailyFocusHtml(d){
    if(!d||!d.focus) return '';
    var h='<div class="mw-h">⚡ 오늘 포커스 <span class="sub2">'+esc(d.date||'')+' 보고 기반 자동 생성</span></div>'
      +'<div class="mw-card" style="border-left:3px solid var(--accent)"><div class="t">'+esc(d.focus)+'</div></div>';
    if((d.completed||[]).length){
      h+='<div class="mw-h">✅ 보고 기반 자동 완료 <span class="sub2">'+d.completed.length+'건 — 어제 보고에서 완료 확인, 승인 대기로 전환</span></div>';
      d.completed.forEach(function(c){
        h+='<div class="mw-card"><div class="t"><span class="st-wait">⏳ 승인 대기</span> '+esc(c.task)+'</div>'
          +'<div class="m">'+(c.project?esc(c.project)+' · ':'')+esc(c.basis||'')+'</div></div>';
      });
    }
    return h;
  }
  function mwNewTodosHtml(d){
    if(!d||!(d.new_todos||[]).length) return '';
    return d.new_todos.map(function(t){
      var pj=t.project?('['+esc(t.project)+'] '):'';
      var dl=t.deadline?(' · 마감 '+esc(t.deadline)):'';
      return '<div class="mw-card" style="border-style:dashed" title="'+esc(t.basis||'')+'">'
        +'<div class="t"><span class="mw-badge" style="background:rgba(108,92,231,.16);color:var(--accent)">🆕 제안</span> '
        +esc(t.task)
        +'<a class="ns-add st-act" style="border-color:var(--accent);color:var(--accent)" data-task="'+esc(t.task)
        +'" data-project="'+esc(t.project||'')+'" data-deadline="'+esc(t.deadline||'')+'">+ 분장 등록</a></div>'
        +'<div class="m">'+pj+esc(t.basis||'어제 보고에서 도출')+dl+'</div></div>';
    }).join('');
  }
  function mwProjUpdatesHtml(d){
    if(!d||!(d.project_updates||[]).length) return '';
    var h='<div class="mw-h">📁 프로젝트 업데이트 <span class="sub2">어제 보고 기반 프로젝트 단위 요약</span></div>';
    d.project_updates.forEach(function(u){
      h+='<div class="mw-card"><div class="t">'+esc(u.project||'기타')+'</div><div class="m">'+esc(u.update)+'</div></div>';
    });
    return h;
  }
  function mwBindSelfAssign(scope){
    scope.querySelectorAll('.ns-add').forEach(function(el){
      el.onclick=function(){
        el.textContent='등록 중…';
        fetch('/api/assign-self',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:SESS.name,pin:SESS.pin,task:el.getAttribute('data-task'),
            project:el.getAttribute('data-project'),deadline:el.getAttribute('data-deadline')})})
        .then(function(r){return r.json();}).then(function(d){
          if(!d.ok) alert(d.error||'등록 실패');
          renderMyWork();
        }).catch(function(){ alert('서버 오류'); renderMyWork(); });
      };
    });
  }
  function mwOvBadge(a){
    // 지연 N일 배지(A1 — filament) — 서버 _assign_read가 열린 분장에만 days_overdue 부여
    return (a.days_overdue>0)?' <span class="mw-badge mw-urgent">지연 '+a.days_overdue+'일</span>':'';
  }
  function mwLocalDate(off){
    var d=new Date(Date.now()+(off||0)*86400000);
    return d.getFullYear()+'-'+('0'+(d.getMonth()+1)).slice(-2)+'-'+('0'+d.getDate()).slice(-2);
  }
  function mwDueHtml(A){
    // ⏰ 마감 임박(A1 — filament '오늘' 탭: 지난·오늘·내일 모아보기)
    var t0=mwLocalDate(0), t1=mwLocalDate(1);
    var open=(A||[]).filter(function(a){ return a.status==='미착수'||a.status==='진행중'; });
    var late=open.filter(function(a){ return a.days_overdue>0; });
    var soon=open.filter(function(a){ return !a.days_overdue && ((a.deadline||'').slice(0,10)===t0||(a.deadline||'').slice(0,10)===t1); });
    if(!late.length && !soon.length) return '';
    var h='<div class="mw-h">⏰ 마감 임박 <span class="sub2">지난 '+late.length+'건 · 오늘·내일 '+soon.length+'건 — 아래 팀 Todo에서 처리</span></div>';
    late.concat(soon).forEach(function(a){
      var pj=a.project?(esc(a.project)+' · '):'';
      var tag=a.days_overdue>0?mwOvBadge(a):' <span class="mw-badge" style="background:rgba(244,196,48,.14);color:#f4c430">'+((a.deadline||'').slice(0,10)===t0?'오늘':'내일')+' 마감</span>';
      h+='<div class="mw-card ex-amber"><div class="t">'+esc(a.task)+tag+'</div>'
        +'<div class="m">'+pj+esc(a.assignee||'미지정')+' · 마감 '+esc(a.deadline||'')+' · '+esc(a.status||'')+'</div></div>';
    });
    return h;
  }
  function mwUnassignedHtml(U){
    // ◆ 담당 미지정 배정 큐(A2 — filament 팀 보드 패턴)
    if(!U||!U.length) return '';
    var h='<div class="mw-h">◆ 담당 미지정 배정 <span class="sub2">'+U.length+'건 — 담당자 지정 시 시트에 바로 반영</span></div>';
    U.forEach(function(a){
      var pj=a.project?(esc(a.project)+' · '):'';
      var dl=a.deadline?('마감 '+esc(a.deadline)):'마감 미정';
      h+='<div class="mw-card ex-amber"><div class="t">'+esc(a.task)+mwOvBadge(a)
        +' <select class="il-sel ua-as" data-row="'+a.row+'" data-task="'+esc(a.task)+'">'
        +'<option value="">담당자 선택…</option>'+mwAsOpts('')+'</select>'
        +'<a class="st-act ua-go">✓ 배정</a></div>'
        +'<div class="m">'+pj+dl+'</div></div>';
    });
    return h;
  }
  function mwInlineSel(a){
    // 리더·대표 인라인 편집(B2 — filament 상태·담당자 칩) — 변경 즉시 시트 반영
    var st=a.status||'미착수';
    var s='<select class="il-sel il-st" data-row="'+a.row+'" data-task="'+esc(a.task)+'" data-assignee="'+esc(a.assignee||'')+'">';
    ['미착수','진행중','완료'].forEach(function(o){ s+='<option'+(o===st?' selected':'')+'>'+o+'</option>'; });
    s+='</select><select class="il-sel il-as" data-row="'+a.row+'" data-task="'+esc(a.task)+'" data-assignee="'+esc(a.assignee||'')+'">'
      +'<option value="">담당자…</option>'+mwAsOpts(a.assignee||'')+'</select>';
    return s;
  }
  function mwBindInline(scope){
    scope.querySelectorAll('.il-st').forEach(function(el){
      el.onchange=function(){
        fetch('/api/assign-status',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:SESS.name,pin:SESS.pin,row:+el.getAttribute('data-row'),
            task:el.getAttribute('data-task'),assignee:el.getAttribute('data-assignee'),status:el.value})})
        .then(function(r){return r.json();}).then(function(d){
          if(!d.ok) alert(d.error||'상태 변경 실패');
          renderMyWork();
        }).catch(function(){ alert('서버 오류'); renderMyWork(); });
      };
    });
    scope.querySelectorAll('.il-as').forEach(function(el){
      el.onchange=function(){
        var v=el.value; if(!v||v===el.getAttribute('data-assignee')) return;
        el.disabled=true;
        fetch('/api/assign-edit',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:SESS.name,pin:SESS.pin,row:+el.getAttribute('data-row'),
            task:el.getAttribute('data-task'),assignee:el.getAttribute('data-assignee'),
            new_task:el.getAttribute('data-task'),new_assignee:v})})
        .then(function(r){return r.json();}).then(function(d){
          if(!d.ok) alert(d.error||'담당자 변경 실패');
          renderMyWork();
        }).catch(function(){ alert('서버 오류'); renderMyWork(); });
      };
    });
    scope.querySelectorAll('.ua-go').forEach(function(btn){
      btn.onclick=function(){
        var sel=btn.closest('.mw-card').querySelector('.ua-as'); var v=sel.value;
        if(!v){ alert('담당자를 선택하세요'); return; }
        btn.textContent='배정 중…';
        fetch('/api/assign-edit',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:SESS.name,pin:SESS.pin,row:+sel.getAttribute('data-row'),
            task:sel.getAttribute('data-task'),assignee:'',new_task:sel.getAttribute('data-task'),new_assignee:v})})
        .then(function(r){return r.json();}).then(function(d){
          if(!d.ok) alert(d.error||'배정 실패');
          renderMyWork();
        }).catch(function(){ alert('서버 오류'); renderMyWork(); });
      };
    });
  }
  function mwStBtn(a, st, label, notify){
    return '<a class="st-act" data-row="'+a.row+'" data-task="'+esc(a.task)+'" data-assignee="'+esc(a.assignee||'')
      +'" data-st="'+st+'"'+(notify?' data-notify="1"':'')+'>'+label+'</a>';
  }
  function mwChk(a){
    return '<input type="checkbox" class="st-chk" title="일괄 삭제 선택" data-row="'+a.row+'" data-task="'+esc(a.task)+'" data-assignee="'+esc(a.assignee||'')+'">';
  }
  function mwEditBtn(a){
    return '<a class="ed-btn st-act" title="내용 수정" data-row="'+a.row+'" data-task="'+esc(a.task)
      +'" data-assignee="'+esc(a.assignee||'')+'" data-project="'+esc(a.project||'')
      +'" data-deadline="'+esc(a.deadline||'')+'">✏️</a>';
  }
  function mwBindEdit(scope){
    var inS='background:var(--bg-3);border:1px solid var(--line);border-radius:7px;color:var(--fg);padding:7px 9px;font-family:inherit;font-size:12px';
    scope.querySelectorAll('.ed-btn').forEach(function(el){
      el.onclick=function(){
        var card=el.closest('.mw-card');
        var nx=card.nextElementSibling;
        if(nx && nx.classList.contains('ed-form')){ nx.remove(); return; }
        var d=document.createElement('div'); d.className='mw-card ed-form';
        d.innerHTML='<div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center">'
          +'<input class="ef-task" value="'+esc(el.getAttribute('data-task'))+'" placeholder="업무 내용" style="flex:2;min-width:220px;'+inS+'">'
          +'<input class="ef-proj" value="'+esc(el.getAttribute('data-project')||'')+'" placeholder="프로젝트" style="flex:1;min-width:130px;'+inS+'">'
          +'<input class="ef-dl" type="date" value="'+esc(el.getAttribute('data-deadline')||'')+'" style="'+inS+'">'
          +'<button class="ef-save" style="background:var(--accent);border:0;border-radius:7px;color:#fff;padding:7px 14px;font-size:12px;cursor:pointer;font-family:inherit">저장</button>'
          +'<button class="ef-cancel" style="background:transparent;border:1px solid var(--line);border-radius:7px;color:var(--muted);padding:7px 12px;font-size:12px;cursor:pointer;font-family:inherit">취소</button></div>';
        card.parentNode.insertBefore(d, card.nextSibling);
        d.querySelector('.ef-cancel').onclick=function(){ d.remove(); };
        d.querySelector('.ef-save').onclick=function(){
          var nt=d.querySelector('.ef-task').value.trim();
          if(!nt){ alert('업무 내용을 입력하세요'); return; }
          this.disabled=true; this.textContent='저장 중…';
          fetch('/api/assign-edit',{method:'POST',headers:{'Content-Type':'application/json'},
            body:JSON.stringify({user:SESS.name,pin:SESS.pin,row:+el.getAttribute('data-row'),
              task:el.getAttribute('data-task'),assignee:el.getAttribute('data-assignee'),
              new_task:nt,new_project:d.querySelector('.ef-proj').value.trim(),
              new_deadline:d.querySelector('.ef-dl').value})})
          .then(function(r){return r.json();}).then(function(res){
            if(!res.ok) alert(res.error||'수정 실패');
            renderMyWork();
          }).catch(function(){ alert('서버 오류'); renderMyWork(); });
        };
      };
    });
  }
  function mwBindBulk(scope){
    var chks=scope.querySelectorAll('.st-chk');
    var bar=document.getElementById('mw-bulkbar');
    if(!bar){ bar=document.createElement('div'); bar.id='mw-bulkbar'; document.body.appendChild(bar); }
    bar.style.display='none';
    if(!chks.length) return;
    function upd(){
      var sel=scope.querySelectorAll('.st-chk:checked');
      if(!sel.length){ bar.style.display='none'; return; }
      bar.style.display='flex';
      bar.innerHTML='<span>🗑 <b>'+sel.length+'건</b> 선택됨</span>'
        +'<button class="bb-del">선택 삭제</button><button class="bb-clear">선택 해제</button>';
      bar.querySelector('.bb-clear').onclick=function(){
        [].forEach.call(sel,function(c){ c.checked=false; }); upd();
      };
      bar.querySelector('.bb-del').onclick=function(){
        if(!confirm('선택한 '+sel.length+'건 분장을 삭제할까요? 목록·집계·포트폴리오에서 제외됩니다. (시트 행은 삭제 표시로 보존)')) return;
        var items=[].map.call(sel,function(c){ return {row:+c.getAttribute('data-row'),
          task:c.getAttribute('data-task'), assignee:c.getAttribute('data-assignee')}; });
        this.disabled=true; this.textContent='삭제 중…';
        fetch('/api/assign-bulk-delete',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:SESS.name,pin:SESS.pin,items:items})})
        .then(function(r){return r.json();}).then(function(d){
          if(d.errors&&d.errors.length) alert((d.deleted||0)+'건 삭제, 일부 실패: '+d.errors.join(' / '));
          bar.style.display='none'; renderMyWork();
        }).catch(function(){ alert('서버 오류'); renderMyWork(); });
      };
    }
    [].forEach.call(chks,function(c){ c.onchange=upd; });
  }
  function mwBindStatus(scope){
    scope.querySelectorAll('.st-act').forEach(function(el){
      el.onclick=function(){
        var st=el.getAttribute('data-st');
        var notify=el.getAttribute('data-notify')==='1';
        if(st==='승인' && !confirm('승인하면 내 업무에서 정리되고 프로젝트 포트폴리오에 완료로 기록됩니다.')) return;
        if(st==='삭제' && !confirm('이 분장을 삭제할까요? 목록·집계에서 제외되고 포트폴리오 기록도 제거됩니다. (시트 행은 삭제 표시로 보존)')) return;
        if(notify && !confirm('완료 처리하고 리더·대표에게 완료 보고 알림(텔레그램)을 보낼까요?')) return;
        el.textContent='…';
        fetch('/api/assign-status',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:SESS.name,pin:SESS.pin,row:+el.getAttribute('data-row'),
            task:el.getAttribute('data-task'),assignee:el.getAttribute('data-assignee'),status:st,notify:notify})})
        .then(function(r){return r.json();}).then(function(d){
          if(!d.ok) alert(d.error||'상태 변경 실패');
          else if(notify) alert(d.notified?('✅ 완료 처리 + 보고 알림 '+d.notified+'명 발송'):'완료 처리됨 — 알림 발송 실패(승인 대기에는 정상 반영)');
          renderMyWork();
        }).catch(function(){ alert('서버 오류'); renderMyWork(); });
      };
    });
  }
  function mwApprovalsHtml(AP){
    if(!AP||!AP.length) return '';
    var h='<div class="mw-h">✅ 완료 승인 대기 <span class="sub2">'+AP.length+'건 — 승인 시 프로젝트 포트폴리오에 완료 기록</span></div>';
    AP.forEach(function(a){
      var pj=a.project?(esc(a.project)+' · '):'';
      var dl=a.deadline?(' · 마감 '+esc(a.deadline)):'';
      h+='<div class="mw-card"><div class="t">'+esc(a.task)
        +mwStBtn(a,'승인','✓ 승인')+mwStBtn(a,'진행중','↩ 반려')+mwStBtn(a,'삭제','🗑')+mwChk(a)
        +'</div><div class="m">'+pj+esc(a.assignee||'미지정')+dl+'</div></div>';
    });
    return h;
  }
  function mwAssignCard(a, withAssignee){
    var st=a.status||'미착수';
    var badge='<span class="lh-st '+(st==='완료'?'lh-done':(st==='진행중'?'lh-doing':'lh-todo'))+'">'+esc(st)+'</span>';
    var urg=(a.priority==='긴급')?'<span class="mw-badge mw-urgent">긴급</span>':'';
    var dl=a.deadline?(' · 마감 '+esc(a.deadline)):'';
    var who=withAssignee?esc(a.assignee||'미지정'):'';
    var act='';
    var canDel = SESS.admin || (SESS.lead_teams||[]).length;
    if(!withAssignee && a.row){
      if(st==='완료'){ act='<span class="st-wait">⏳ 승인 대기</span>'+mwStBtn(a,'진행중','↩ 되돌리기'); }
      else{
        if(st==='미착수') act+=mwStBtn(a,'진행중','▶ 진행');
        act+=mwStBtn(a,'완료','✓ 완료');
        act+=mwStBtn(a,'완료','📣 완료·보고',true);
      }
      act+=mwEditBtn(a)+mwStBtn(a,'삭제','🗑')+mwChk(a);  // 본인 분장은 수정·삭제 가능 (잘못 등록 정정)
    }
    if(withAssignee && a.row && canDel){ act+=mwInlineSel(a)+mwEditBtn(a)+mwStBtn(a,'삭제','🗑')+mwChk(a); }  // B2 — 인라인 상태·담당자
    return '<div class="mw-card pg-item"><div class="t">'+badge+' '+esc(a.task)+urg+mwOvBadge(a)+act+'</div><div class="m">'+who+dl+'</div></div>';
  }
  function mwAssignListHtml(A, withAssignee){
    // 열린 분장 중심(A3 — filament 완료/보관함 분리) — 완료(승인 대기)는 하단 접힘.
    // 프로젝트 단위 그룹 + 완전 동일 행 표시 중복 제거(시트 무변경)
    var seen={}, groups={}, order=[], done=[];
    (A||[]).forEach(function(a){
      var key=[(a.project||''),(a.task||''),(a.assignee||''),(a.deadline||''),(a.status||'')].join('|');
      if(seen[key]) return; seen[key]=1;
      if((a.status||'')==='완료'){ done.push(a); return; }
      var p=(a.project||'').trim()||'기타';
      if(!(p in groups)){ groups[p]=[]; order.push(p); }
      groups[p].push(a);
    });
    order.sort(function(a,b){ if(a==='기타') return 1; if(b==='기타') return -1; return a.localeCompare(b,'ko'); });
    var h='';
    order.forEach(function(p){
      var items=groups[p];
      h+='<div class="pg-head">📁 '+esc(p)+' <span class="pg-cnt">'+items.length+'건</span></div>';
      items.forEach(function(a){ h+=mwAssignCard(a, withAssignee); });
    });
    if(done.length){
      h+='<details class="mw-done-sec"><summary>▸ 완료·승인 대기 '+done.length+'건</summary>';
      done.forEach(function(a){ h+=mwAssignCard(a, withAssignee); });
      h+='</details>';
    }
    return h;
  }
  function changePin(){
    var cur=prompt('현재 PIN을 입력하세요'); if(!cur) return;
    var nw=prompt('새 PIN (4자 이상)'); if(!nw) return;
    fetch('/api/set-pin',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:SESS.name,pin:cur,new_pin:nw})})
      .then(function(r){return r.json();}).then(function(d){
        if(d.ok){ SESS.pin=nw; var j=JSON.stringify(SESS); SESS_KEYS.forEach(function(k){localStorage.setItem(k,j);}); alert('PIN이 변경되었습니다'); }
        else alert(d.error||'변경 실패');
      });
  }
  function enter(s){
    SESS=s; var j=JSON.stringify(s);
    SESS_KEYS.forEach(function(k){ localStorage.setItem(k,j); });
    gate.style.display='none'; shell.style.display='flex';
    who.innerHTML = s.name+' ('+(s.role||'')+') <a id="lg-guide">가이드</a> <a id="lg-pin-c">PIN변경</a> <a id="lg-out">로그아웃</a>';
    document.getElementById('lg-out').onclick=function(){ CLEAR_KEYS.forEach(function(k){localStorage.removeItem(k);}); location.reload(); };
    document.getElementById('lg-pin-c').onclick=changePin;
    document.getElementById('lg-guide').onclick=function(){ window.open('/guide-os','_blank'); };
    var lt=s.lead_teams||[];
    if(s.admin){
      tabBtn('brief').style.display=''; tabBtn('weekly').style.display='';
      document.getElementById('tab-hr').style.display='';  // HR 포털 새 창 — 대표 전용 (독립 서비스 링크, 프록시 아님)
      // Decision Window(ARISA 2.0)는 데이터 취합 백단 창구 — 대시보드 전면 노출 안 함(탭 숨김)
      sel.style.display='inline-block'; sel.innerHTML='<option value="">전체</option>';
      lt.forEach(function(t){ var o=document.createElement('option'); o.value=t; o.textContent=t; sel.appendChild(o); });
      sel.onchange=function(){ loadScope(sel.value); };
      curScope='';
    } else if(lt.length){
      tabBtn('mywork').textContent='팀 홈';  // 리더 — 팀 Todo·분장·프로젝트·팀원 보고 통합
      tabBtn('brief').style.display=''; tabBtn('weekly').style.display='';
      if(lt.length>1){
        // 겸임 리더 — 기본 '담당팀 전체'(병합 한 페이지), 개별 팀은 집중 보기
        curScope='';
        sel.style.display='inline-block'; sel.innerHTML='<option value="">담당팀 전체</option>';
        lt.forEach(function(t){ var o=document.createElement('option'); o.value=t; o.textContent=t; sel.appendChild(o); });
        sel.onchange=function(){ loadScope(sel.value); };
      } else { curScope=lt[0]; sel.style.display='none'; }
    } else {
      tabBtn('brief').style.display='';  // 직원 — 내 브리프(본인 카드+팀 헤드라인)
      sel.style.display='none';
    }
    showTab('mywork');
  }
  document.querySelectorAll('.tab').forEach(function(b){ b.onclick=function(){ showTab(b.dataset.t); }; });
  var sess=null; try{ sess=JSON.parse(localStorage.getItem('pm_sess')||'null'); }catch(e){}
  if(sess && sess.name) enter(sess);
  function doLogin(){
    var id=document.getElementById('lg-id').value.trim(), pin=document.getElementById('lg-pin').value.trim();
    var err=document.getElementById('login-err'); err.textContent='';
    fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:id,pin:pin})})
      .then(function(r){return r.json();})
      .then(function(d){
        if(d.ok){ var s=Object.assign({},d,{id:id,pin:pin}); enter(s); if(d.pin_set) alert('PIN이 설정되었습니다'); }
        else { err.textContent=d.error||'로그인 실패'; }
      })
      .catch(function(){ err.textContent='서버 연결 실패'; });
  }
  document.getElementById('lg-btn').onclick=doLogin;
  document.getElementById('lg-pin').addEventListener('keydown',function(e){ if(e.key==='Enter') doLogin(); });
})();
</script>
</body></html>"""

def load_users():
    """users.json 로드 — dict 스키마(대시보드)와 list 스키마(ARISA 2.0, symlink 단일화)를 모두 지원."""
    try:
        u = json.loads((DATA/"users.json").read_text(encoding="utf-8")).get("users", {})
    except Exception:
        return {}
    if isinstance(u, list):  # ARISA 2.0 스키마: [{name, role(admin/leader/member), team, pin}]
        return {x["name"]: {"pin": x.get("pin", ""),
                            "role": "대표" if x.get("role") == "admin" else (x.get("role") or "직원")}
                for x in u if x.get("name")}
    return u

def load_projects():
    out = []
    if PROJ_DIR.exists():
        for f in sorted(PROJ_DIR.glob("*.json")):
            try: out.append(json.loads(f.read_text(encoding="utf-8")))
            except Exception: pass
    return out

def get_project(pid):
    f = PROJ_DIR / f"{_safe(pid)}.json"
    if f.exists():
        try: return json.loads(f.read_text(encoding="utf-8"))
        except Exception: return None
    return None

def save_project(p):
    PROJ_DIR.mkdir(parents=True, exist_ok=True)
    pid = _safe(p["id"])
    with _lock:
        (PROJ_DIR / f"{pid}.json").write_text(json.dumps(p, ensure_ascii=False, indent=2), encoding="utf-8")

def _safe(pid):
    return re.sub(r"[^A-Za-z0-9가-힣_\-]", "_", str(pid))[:80] or "proj"

def auth(uid, pin):
    u = load_users().get(uid)
    if u and str(u.get("pin")) == str(pin): return u
    return None

def is_admin(uid):
    u = load_users().get(uid)
    return bool(u) and u.get("role") in ("대표", "admin")

def load_emp():
    try: return json.loads(EMP_PATH.read_text(encoding="utf-8"))
    except Exception: return {}

def lead_teams_of(uid):
    """이 사람이 리더인 팀 목록(team_leads 역매핑). 대표(admin)는 전체 팀."""
    tl = load_emp().get("team_leads", {})
    if is_admin(uid):
        return sorted(set(tl.keys()))
    return sorted([team for team, leader in tl.items() if leader == uid])

def set_pin(uid, new_pin):
    """users.json에 PIN 저장(첫 설정/변경) — dict/list 스키마 모두 지원. 성공 True."""
    p = DATA / "users.json"
    with _lock:
        try:
            doc = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return False
        users = doc.get("users")
        if isinstance(users, list):  # ARISA 2.0 스키마
            for x in users:
                if x.get("name") == uid:
                    x["pin"] = str(new_pin)
                    break
            else:
                return False
        else:
            if uid not in (users or {}):
                return False
            users[uid]["pin"] = str(new_pin)
        p.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    return True

def emp_team(name):
    """직원의 소속 팀(arisa-employees.json by_name[…].team)."""
    return (load_emp().get("by_name", {}).get(name, {}) or {}).get("team")

def is_leader(uid):
    """팀 리더 여부(대표 제외한 순수 리더도 True, 대표도 True)."""
    return bool(lead_teams_of(uid))


_PROJ_STOP = {"프로젝트", "행사", "팝업", "기획", "관련", "운영", "브랜드", "상세"}


def _proj_tokens(s):
    """프로젝트명 → 매칭용 토큰 집합 (영숫자·한글 단어, 일반어 제외)."""
    toks = set()
    for t in re.findall(r"[A-Za-z0-9]+|[가-힣]+", (s or "").lower()):
        if len(t) >= 2 and t not in _PROJ_STOP:
            toks.add(t)
    return toks


def _match_project(ap, pname):
    """분장 프로젝트명(ap) vs 포트폴리오명(pname) — 표기 변형 허용 매칭.
    ① 정규화 상호 포함 ② 유의 토큰 교집합 (영문 3자+/한글 2자+, 일반어 제외)."""
    na = re.sub(r"[^a-z0-9가-힣]", "", (ap or "").lower())
    nb = re.sub(r"[^a-z0-9가-힣]", "", (pname or "").lower())
    if na and nb and (na in nb or nb in na):
        return True
    common = _proj_tokens(ap) & _proj_tokens(pname)
    return any(len(t) >= 3 or re.fullmatch(r"[가-힣]{2,}", t) for t in common)


def _match_project_p(ap, p):
    """분장 프로젝트명(ap) vs 프로젝트 dict — 정식 명칭 + aliases(별칭·구표기)까지 매칭."""
    if _match_project(ap, p.get("name") or ""):
        return True
    return any(_match_project(ap, al) for al in (p.get("aliases") or []))


def _resolve_pid(project):
    """분장 프로젝트명 → 프로젝트 ID (G1 — 등록 시 1회 확정, 이후 소비는 ID 우선)."""
    pn = (project or "").strip()
    if not pn:
        return ""
    try:
        # P1 — 아카이브 프로젝트는 신규 귀속 대상에서 제외 (재개하려면 대표가 복원)
        hit = next((p for p in load_projects()
                    if not p.get("archived") and _match_project_p(pn, p)), None)
        return (hit.get("id") or "") if hit else ""
    except Exception:
        return ""


_MEMORY_FILES = (("00_project_brief.md", "📌 프로젝트 브리프"),
                 ("02_decisions.md", "⚖️ 결정 기록"),
                 ("03_todos.md", "✅ 할 일"),
                 ("05_strategy_report.md", "🧭 전략 리포트"),
                 ("06_progress_log.md", "📈 진행 로그"))


def _memory_dir_for(p):
    """포트폴리오 프로젝트 → arisa-project-memory/projects/<폴더> 매칭 (이름·별칭)."""
    try:
        if not MEMORY_DIR.exists():
            return None
        names = [p.get("name") or ""] + [a for a in (p.get("aliases") or []) if a]
        for d in sorted(MEMORY_DIR.iterdir()):
            if not d.is_dir() or d.name.startswith("_"):
                continue
            if any(_match_project(d.name, n) for n in names if n):
                return d
    except Exception:
        pass
    return None


def _memory_hub(p):
    """프로젝트 상세용 ARISA 메모리 링크백(B1) — 존재 파일 + 최근 회의록 5건."""
    d = _memory_dir_for(p)
    if not d:
        return None
    files = [{"file": f, "label": lab} for f, lab in _MEMORY_FILES if (d / f).exists()]
    ml = d / "01_meeting_logs"
    meetings = sorted((f.name for f in ml.glob("*.md")), reverse=True)[:5] if ml.is_dir() else []
    return {"folder": d.name, "files": files, "meetings": meetings}


def _find_project_for_assign(assign):
    """분장 → 프로젝트 dict. pid(ID Relation) 우선, 없으면 이름 토큰 매칭 폴백 (G1)."""
    pid = (assign.get("pid") or "").strip()
    if pid:
        p = get_project(pid)
        if p:
            return p
    pn = (assign.get("project") or "").strip()
    if not pn:
        return None
    return next((p for p in load_projects() if _match_project_p(pn, p)), None)


def _project_assignments(pname, aliases=None, pid=""):
    """최근 2주 주간분장 중 프로젝트가 매칭되는 항목 — 프로젝트 상세 '분장 업무' 섹션용.
    (엄격한 '이번주' 필터는 주가 바뀌면 미완료 할일이 사라져 2주 윈도우 사용.)
    완전 동일 행은 표시 중복 제거. G1: 행에 pid 있으면 ID 일치 우선, 없으면 이름 토큰 매칭."""
    pname = (pname or "").strip()
    if not pname:
        return []
    names = [pname] + list(aliases or [])
    today = datetime.date.today()
    ws = today - datetime.timedelta(days=today.weekday() + 7)  # 지난주 월요일부터
    we = today
    out, seen = [], set()
    for a in _assign_read():
        ds = (a.get("date") or "").strip()[:10]
        try:
            d = datetime.date.fromisoformat(ds)
        except ValueError:
            continue
        if not (ws <= d <= we):
            continue
        if (a.get("status") or "") == "삭제":
            continue
        ap = (a.get("project") or "").strip()
        apid = (a.get("pid") or "").strip()
        if pid and apid:
            if apid != pid:  # ID Relation — 확정 연결 (G1)
                continue
        elif not ap or not any(_match_project(ap, n) for n in names):
            continue
        key = (ap, a.get("task"), a.get("assignee"), a.get("deadline"), a.get("status"))
        if key in seen:
            continue
        seen.add(key)
        out.append(a)
    out.sort(key=lambda a: (a.get("status") == "완료", a.get("deadline") or "9999"))
    return out


# 자료 기반 AI 갱신 대상 브리프 필드 (화이트리스트 + 표시 라벨)
_BRIEF_AI_FIELDS = {
    "summary": "프로젝트 정의", "status": "컨디션", "start": "시작일", "end": "종료일", "dday": "D-Day",
    "req": "클라이언트 요구사항", "goal": "목표", "kpi": "KPI",
    "critical": "실행 주요포인트", "deliverables": "산출물", "risk": "주의사항",
}

# ── 분장 ↔ 프로젝트 포트폴리오 연동 ──────────────────────────
# 상태·우선순위 정의 단일출처: shared/status.py (G2 — 값 불변, 정의만 이동)
_ASSIGN_ST_MAP = _ST.ASSIGN_TO_TASK
_ASSIGN_PROG_MAP = _ST.ASSIGN_TO_PROGRESS
_ASSIGN_DONE_STATES = _ST.ASSIGN_DONE_STATES   # 분장 완료 판정
_ASSIGN_HIDDEN_STATES = _ST.ASSIGN_HIDDEN_STATES  # 내 업무·팀 목록에서 숨김
_TASK_DONE_STATES = _ST.TASK_DONE_STATES       # 프로젝트 tasks 완료 판정 (영문/국문 혼재)


def _remove_done_task(assign):
    """삭제된 분장을 프로젝트 tasks에서 제거 (akey 매칭). 프로젝트는 pid 우선(G1)."""
    tp = _find_project_for_assign(assign)
    if not tp:
        return False
    k = _akey(assign.get("date"), assign.get("task"), assign.get("assignee"))
    before = len(tp.get("tasks") or [])
    tp["tasks"] = [t for t in (tp.get("tasks") or []) if t.get("akey") != k]
    if len(tp["tasks"]) < before:
        save_project(tp)
        return True
    return False


def _can_approve(uid, assignee):
    """완료 승인 권한 — 대표 전체, 리더는 자기 팀원."""
    if is_admin(uid):
        return True
    return is_leader(uid) and emp_team(assignee) in set(lead_teams_of(uid))


def _record_done_task(assign, approver):
    """승인된 분장을 프로젝트 포트폴리오 tasks에 Done으로 영구 기록. 프로젝트는 pid 우선(G1)."""
    tp = _find_project_for_assign(assign)
    if not tp:
        return False
    k = _akey(assign.get("date"), assign.get("task"), assign.get("assignee"))
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    hit = next((t for t in (tp.get("tasks") or []) if t.get("akey") == k), None)
    if hit:
        hit.update({"status": "Done", "progress": 100, "approved_by": approver, "approved_at": now})
    else:
        tp.setdefault("tasks", []).append({
            "division": emp_team(assign.get("assignee")) or "", "task": assign.get("task") or "",
            "owner": assign.get("assignee") or "", "start": (assign.get("date") or "")[:10],
            "end": (assign.get("deadline") or "").strip(), "status": "Done", "progress": 100,
            "akey": k, "approved_by": approver, "approved_at": now})
    save_project(tp)
    return True


def _akey(date, task, assignee):
    """분장 행 식별키(등록일|업무|담당자) — 프로젝트 tasks 중복 방지·상태 동기화용."""
    return "|".join([(date or "").strip()[:10], (task or "").strip(), (assignee or "").strip()])


def _open_assigns(p):
    """이 프로젝트의 열린 분장(미착수·진행중) — 날짜 무관 전 기간 (P1 성급 아카이브 방지).

    KBO 사례(2026-07-19): 종료일 경과만으로 아카이브했다가 후속 분장 4건 발견 → 복원.
    아카이브 판단은 종료일이 아니라 열린 분장 유무가 먼저다. pid 일치 우선, 없으면 이름 매칭.
    반환: [{"task","assignee","deadline","status"}] (마감 오름차순)
    """
    out = []
    pid = (p.get("id") or "").strip()
    try:
        for a in _assign_read():
            if (a.get("status") or "미착수") not in _ST.ASSIGN_OPEN_STATES:
                continue
            apid = (a.get("pid") or "").strip()
            if apid:
                if apid != pid:
                    continue
            elif not _match_project_p((a.get("project") or "").strip(), p):
                continue
            out.append({"task": a.get("task") or "", "assignee": a.get("assignee") or "",
                        "deadline": a.get("deadline") or "", "status": a.get("status") or "미착수"})
    except Exception:
        return []
    return sorted(out, key=lambda x: x["deadline"] or "9999")


def _archive_log(pid, name, by, action, retro=None):
    """아카이브 전이 기록 (P1) — DATA/archive-log.jsonl append. 90-archive 풀 스크립트의 소스."""
    try:
        entry = {"ts": datetime.datetime.now().isoformat(timespec="seconds"),
                 "pid": pid, "name": name or "", "by": by, "action": action}
        if retro:
            entry["retro"] = retro
        with _lock:
            with open(DATA / "archive-log.jsonl", "a", encoding="utf-8") as f:
                f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass  # 기록 실패는 본흐름 무영향


def _similar_projects(name, projects, limit=3):
    """유사 프로젝트 후보 — 토큰 교집합 + n-gram 유사도(붙임 표기 감지) 결합, 최대 limit개."""
    toks = _proj_tokens(name)
    scored = []
    for p in projects:
        pn = p.get("name") or ""
        common = len(toks & _proj_tokens(pn))
        sim = _NM.name_similarity(name, pn)  # 붙임 표기('중기제품팝업스토')는 토큰 0이어도 잡힘
        if common or sim >= 0.45:
            scored.append((common + sim, p))
    scored.sort(key=lambda x: -x[0])
    return [{"id": p.get("id"), "name": p.get("name") or p.get("id")} for _, p in scored[:limit]]


def _find_dup_project(name, projects):
    """신규 생성 전 중복 의심 프로젝트 — 엄격 매칭 또는 n-gram 임계 초과 (활성만).
    사람 확인(confirm/force) 전제의 경고용 — pid 귀속에는 사용하지 않는다."""
    best, best_sim = None, 0.0
    for p in projects:
        if p.get("archived"):
            continue
        if _match_project_p(name, p):
            return p
        sim = _NM.name_similarity(name, p.get("name") or "")
        if sim >= _NM.SIMILARITY_DUP and sim > best_sim:
            best, best_sim = p, sim
    return best


def _append_assign_tasks(p, items):
    """분장 항목을 프로젝트 tasks에 영구 반영 (akey 동일 항목 skip). 추가 건수 반환."""
    tasks = p.setdefault("tasks", [])
    seen = {t.get("akey") for t in tasks if t.get("akey")}
    today = datetime.date.today().isoformat()
    added = 0
    for it in items:
        k = _akey(today, it.get("task"), it.get("assignee"))
        if k in seen:
            continue
        seen.add(k)
        tasks.append({"division": emp_team(it.get("assignee")) or "", "task": (it.get("task") or "").strip(),
                      "owner": (it.get("assignee") or "").strip(), "start": today,
                      "end": (it.get("deadline") or "").strip(), "status": "Not Started",
                      "progress": 0, "akey": k})
        added += 1
    return added


def _sync_assign_status(p):
    """akey 있는 프로젝트 tasks 상태를 주간분장 시트(SSOT) 최신 상태로 갱신. 변경 여부 반환."""
    keyed = [t for t in (p.get("tasks") or []) if t.get("akey")]
    if not keyed:
        return False
    sheet = {}
    for a in _assign_read():
        sheet[_akey(a.get("date"), a.get("task"), a.get("assignee"))] = a.get("status") or "미착수"
    changed = False
    removed = []
    for t in keyed:
        st = sheet.get(t["akey"])
        if not st:
            continue
        if st == "삭제":
            removed.append(t["akey"])
            continue
        new_st = _ASSIGN_ST_MAP.get(st, "Not Started")
        if t.get("status") != new_st:
            t["status"] = new_st
            t["progress"] = _ASSIGN_PROG_MAP.get(st, 0)
            changed = True
    if removed:
        p["tasks"] = [t for t in (p.get("tasks") or []) if t.get("akey") not in removed]
        changed = True
    return changed


def _notify_approvers(assignee, msg):
    """담당 팀 리더 + 대표에게 텔레그램 멘션. 발송 성공 수 반환."""
    team = emp_team(assignee) or ""
    leader = (load_emp().get("team_leads") or {}).get(team)
    targets = set()
    if leader and leader != assignee:
        targets.add(leader)
    for n, u in (load_users() or {}).items():
        if u.get("role") in ("대표", "admin") and n != assignee:
            targets.add(n)
    sent = 0
    for n in targets:
        cid = _tg_chat_id(n)
        if cid and _tg_send(cid, msg):
            sent += 1
    return sent


def _person_workbrief(uid, mine):
    """개인 아침 브리프(배치 산출 my-brief-{date}-{name}.json) 최신본 로드.
    new_todos 중 이미 분장에 등록된 항목(동일 task)은 제외. 없으면 None."""
    pdir = BRIEF_DIR / "person"
    if not pdir.exists():
        return None
    files = sorted(pdir.glob(f"my-brief-2*-{uid}.json"))
    if not files:
        return None
    try:
        d = json.loads(files[-1].read_text(encoding="utf-8"))
    except Exception:
        return None
    existing = {(a.get("task") or "").strip() for a in (mine or [])}
    d["new_todos"] = [t for t in (d.get("new_todos") or [])
                      if (t.get("task") or "").strip() not in existing]
    return d


# ── 브리프 코멘트 (대표·리더 → 보고자 피드백 + 텔레그램 회신) ──
CMT_DIR = BRIEF_DIR / "comments"


def _load_comments(date_str):
    f = CMT_DIR / f"{date_str}.json"
    try:
        return json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save_comments(date_str, arr):
    CMT_DIR.mkdir(parents=True, exist_ok=True)
    with _lock:
        (CMT_DIR / f"{date_str}.json").write_text(json.dumps(arr, ensure_ascii=False, indent=1), encoding="utf-8")


def _tg_chat_id(name):
    """직원명 → 텔레그램 chat_id (arisa-employees.json by_telegram_id 역매핑)."""
    for cid, nm in (load_emp().get("by_telegram_id", {}) or {}).items():
        if nm == name:
            return cid
    return None


def _brief_item_detail(date_str, item, src):
    """브리프 JSON에서 항목(title)+보고자 매칭 → 보고 원문(detail). 대표 브리프 → 팀 브리프 순."""
    def _cands(doc):
        out = list(doc.get("items") or []) + list(doc.get("decision_summary") or []) + list(doc.get("top5") or [])
        for t in (doc.get("teams") or []):
            out += list(t.get("top") or [])
        return out
    files = [BRIEF_DIR / f"daily-brief-{date_str}.json"] + sorted(BRIEF_DIR.glob(f"daily-brief-{date_str}-*.json"))
    for f in files:
        try:
            doc = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        for c in _cands(doc):
            if not isinstance(c, dict):
                continue
            title = (c.get("title") or "").strip()
            if not title:
                continue
            if title != item and title not in item and item not in title:
                continue
            if src and (c.get("source_employee") or "").strip() not in ("", src):
                continue
            return (c.get("detail") or "").strip()
    return ""


def _report_raw_lookup(date_str, item, src):
    """일일보고 시트(SSOT)에서 담당자 보고 원문을 그대로 조회 — 축약 없음.
    메타 탭(요청한 결정 H·지원 요청 I·블로커 F·질문 L) 필드와 핵심업무 행 중
    항목과 토큰이 가장 겹치는 것을 반환. 겹침 2토큰 미만이면 '' (오매칭 방지)."""
    if not (_asgws and DAILY_SHEET and src):
        return ""
    try:
        from shared.normalize import normalize_date as _nd
    except Exception:
        _nd = lambda s: (s or "").strip()[:10]
    # 브리프는 전날 저녁 제출 보고로 만들어짐 — 브리프 날짜 + 직전 3영업일까지 조회
    dates = {date_str}
    try:
        d = datetime.date.fromisoformat(date_str)
        while len(dates) < 4:
            d -= datetime.timedelta(days=1)
            if d.weekday() < 5:
                dates.add(d.isoformat())
    except ValueError:
        pass
    toks = _proj_tokens(item)
    def _score(t):
        return len(toks & _proj_tokens(t))
    def _mine(r, width):
        r = list(r) + [""] * (width - len(r))
        return r if ((_nd(r[0]) or "") in dates and (r[1] or "").strip() == src) else None
    best_txt, best = "", 0
    try:
        meta = _asgws.values_get(DAILY_SHEET, "메타!A2:O5000", retries=2, timeout=20)
    except Exception:
        meta = []
    for r in meta:
        r = _mine(r, 15)
        if not r:
            continue
        for lab, txt in (("요청한 결정", r[7]), ("지원 요청", r[8]), ("블로커", r[5]), ("질문", r[11])):
            txt = (txt or "").strip()
            if txt and _score(txt) > best:
                best, best_txt = _score(txt), f"{lab}: {txt}"
    try:
        core = _asgws.values_get(DAILY_SHEET, "핵심업무!A2:L5000", retries=2, timeout=20)
    except Exception:
        core = []
    for r in core:
        r = _mine(r, 12)
        if not r:
            continue
        blob = " ".join([(r[5] or ""), (r[9] or ""), (r[10] or ""), (r[11] or "")])
        s = _score(blob)
        if s > best:
            parts = []
            if (r[5] or "").strip():
                parts.append(f"[업무] {r[5].strip()}")
            for lab, v in (("성과", r[11]), ("이슈", r[10]), ("산출물", r[9])):
                if (v or "").strip():
                    parts.append(f"{lab}: {v.strip()}")
            if parts:
                best, best_txt = s, "\n".join(parts)
    # 메타 raw(K열 — 보고 원문 전체)에서 블록 매칭: '['헤더/빈 줄로 블록 분리 후 최다 겹침 블록
    for r in meta:
        r = _mine(r, 15)
        if not r or not (r[10] or "").strip():
            continue
        block, blocks = [], []
        for ln in (r[10] or "").splitlines():
            if (ln.strip().startswith("[") or not ln.strip()) and block:
                blocks.append("\n".join(block)); block = []
            if ln.strip():
                block.append(ln.rstrip())
        if block:
            blocks.append("\n".join(block))
        for blk in blocks:
            s = _score(blk)
            if s > best:
                best, best_txt = s, blk.strip()
    # Basket 매장보고 시트 (송금·승인/구매 등 결재성 요청은 여기 기록됨)
    bsid = os.environ.get("BASKET_REPORT_SHEET_ID", "")
    if bsid:
        try:
            brows = _asgws.values_get(bsid, "일일보고!A2:O5000", retries=2, timeout=20)
        except Exception:
            brows = []
        blabs = {3: "매출", 4: "지출", 5: "송금·승인", 6: "특이사항", 7: "장비",
                 8: "업무보고", 9: "대관", 10: "스태프", 11: "구매", 12: "입점제안", 13: "복기"}
        for r in brows:
            r = list(r) + [""] * (15 - len(r))
            if (_nd(r[1]) or "") not in dates or (r[2] or "").strip() != src:
                continue
            for idx, lab in blabs.items():
                txt = (r[idx] or "").strip()
                if txt and _score(txt) > best:
                    best, best_txt = _score(txt), f"[매장보고 · {lab}] {txt}"
    return best_txt if best >= 2 else ""


def _brief_item_source(date_str, item, src):
    """담당자가 보고한 원문 — ① 일일보고 시트 원문 → ② 브리프 JSON people.core →
    ③ AI 요약 detail 순 폴백. 원문은 축약 없이 그대로."""
    raw = _report_raw_lookup(date_str, item, src)
    if raw:
        return raw
    toks_item = _proj_tokens(item)
    best, best_score, n_cands = None, 0, 0
    files = [BRIEF_DIR / f"daily-brief-{date_str}.json"] + sorted(BRIEF_DIR.glob(f"daily-brief-{date_str}-*.json"))
    for f in files:
        try:
            doc = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        for p in (doc.get("people") or []):
            if src and (p.get("name") or "").strip() != src:
                continue
            for e in (p.get("core") or []):
                if not isinstance(e, dict):
                    continue
                n_cands += 1
                blob = " ".join(str(e.get(k) or "") for k in ("task", "issue", "detail", "outcome", "output"))
                score = len(toks_item & _proj_tokens(blob))
                if score > best_score:
                    best, best_score = e, score
    # 겹침 2토큰↑ 신뢰, 보고 행이 1건뿐이면 1토큰도 허용 (오매칭 방지)
    if best and (best_score >= 2 or (best_score >= 1 and n_cands == 1)):
        parts = []
        if (best.get("task") or "").strip():
            parts.append(f"[업무] {best['task'].strip()}")
        for lab, k in (("성과", "outcome"), ("이슈", "issue"), ("상세", "detail"), ("산출물", "output")):
            if (best.get(k) or "").strip():
                parts.append(f"{lab}: {best[k].strip()}")
        if parts:
            return "\n".join(parts)
    return _brief_item_detail(date_str, item, src)


def _tg_send(chat_id, text):
    """일일보고 봇으로 메시지 발송 — 실패해도 예외 없이 False (코멘트 저장은 유지)."""
    tok = os.environ.get("DAILY_REPORT_BOT_TOKEN", "")
    if not (tok and chat_id):
        return False
    import urllib.request
    try:
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{tok}/sendMessage",
            data=json.dumps({"chat_id": chat_id, "text": text}).encode("utf-8"),
            headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as r:
            return bool(json.loads(r.read().decode()).get("ok"))
    except Exception:
        return False


# 브리프 페이지 주입용 코멘트 위젯 — .vz-item(제목+보고자) 단위로 코멘트 표시·입력.
# __DATE__ 는 서빙 시 해당 브리프 날짜로 치환. 세션은 셸이 심어둔 localStorage 키 재사용.
BRIEF_CMT_JS = """<script>(function(){
var DS='__DATE__', sess=null;
['brief_sess','team_brief_sess','pm_sess'].some(function(k){
  try{var s=JSON.parse(localStorage.getItem(k)||'null'); if(s&&s.name&&s.pin){sess=s;return true;}}catch(e){}
  return false;});
if(!sess) return;
function esc(s){return String(s||'').replace(/[&<>"']/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];});}
function cline(c){return '<div style="color:var(--muted,#8A857E);padding:3px 0">💬 <b style="color:var(--fg,#eee)">'
  +esc(c.author)+'</b> '+esc(c.text)+' <span style="opacity:.6">'+esc(c.ts)+(c.tg_sent?' · TG✓':'')+'</span></div>';}
fetch('/api/brief-comments?user='+encodeURIComponent(sess.name)+'&date='+DS)
.then(function(r){return r.json();}).then(function(d){
  if(!d.ok) return;
  var by={};(d.comments||[]).forEach(function(c){var k=c.item+'|'+(c.src||'');(by[k]=by[k]||[]).push(c);});
  [].slice.call(document.querySelectorAll('.vz-item, .tb-chip')).forEach(function(el){
    var t=el.querySelector('.vz-t'), s=el.querySelector('.vz-src');
    if(!t) return;
    var item=t.textContent.trim(), src=s?s.textContent.trim():'';
    var wrap=document.createElement('div'); wrap.style.cssText='margin:2px 0 8px 22px;font-size:12px';
    var list=document.createElement('div'); wrap.appendChild(list);
    (by[item+'|'+src]||[]).forEach(function(c){ list.innerHTML+=cline(c); });
    var btn=document.createElement('a'); btn.textContent='💬 코멘트';
    btn.style.cssText='cursor:pointer;color:var(--accent,#6C5CE7);font-size:11px';
    btn.onclick=function(){
      if(wrap.querySelector('.bc-in')) return;
      var box=document.createElement('div'); box.className='bc-in'; box.style.cssText='display:flex;gap:6px;margin-top:4px';
      box.innerHTML='<input style="flex:1;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.15);'
        +'border-radius:6px;padding:6px 8px;color:inherit;font:inherit;font-size:12px" placeholder="코멘트 입력 — 등록 시 '
        +esc(src||'보고자')+' 텔레그램으로 회신">'
        +'<button style="background:var(--accent,#6C5CE7);border:0;border-radius:6px;color:#fff;padding:6px 12px;'
        +'font-size:12px;cursor:pointer">등록</button>';
      var inp=box.querySelector('input'), sb=box.querySelector('button');
      function go(){
        var v=inp.value.trim(); if(!v) return; sb.disabled=true;
        fetch('/api/brief-comment',{method:'POST',headers:{'Content-Type':'application/json'},
          body:JSON.stringify({user:sess.name,pin:sess.pin,date:DS,item:item,src:src,text:v})})
        .then(function(r){return r.json();}).then(function(d2){
          if(d2.ok){ box.remove(); list.innerHTML+=cline(d2.comment); }
          else{ sb.disabled=false; alert(d2.error||'등록 실패'); }
        }).catch(function(){ sb.disabled=false; alert('서버 오류'); });
      }
      sb.onclick=go; inp.onkeydown=function(e){ if(e.key==='Enter') go(); };
      wrap.appendChild(box); inp.focus();
    };
    wrap.appendChild(btn);
    el.parentNode.insertBefore(wrap, el.nextSibling);
  });
});
})();</script>"""


# ── 계정별 브리프 뷰 (/my-brief 직원 · /lead-brief 겸임리더) — 브리프 JSON 서버렌더 ──
_DBA = None

def _dba():
    """daily-brief-aggregate.py 모듈(하이픈 파일명) — 개인카드 렌더러 재사용."""
    global _DBA
    if _DBA is None:
        import importlib.util
        p = Path(__file__).resolve().parent / "daily-brief-aggregate.py"
        spec = importlib.util.spec_from_file_location("dba_mod", str(p))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        _DBA = m
    return _DBA


# 브리프 카드 스타일 (daily-brief-aggregate.py 렌더와 동일 팔레트 — pp-*/as-*/tb-* 발췌)
# 셸(리더 홈)에도 주입되므로 페이지 골격(BRIEF_PAGE_CSS)과 분리.
BRIEF_CARD_CSS = """
.muted{color:var(--muted);font-size:12px}
.urg{font-size:11px;border-radius:5px;padding:2px 8px;white-space:nowrap;flex-shrink:0}
.urg-high{background:rgba(225,112,85,.16);color:var(--red);border:1px solid rgba(225,112,85,.4)}
.urg-mid{background:rgba(217,163,75,.14);color:var(--amber);border:1px solid rgba(217,163,75,.35)}
.urg-low{background:var(--bg-3);color:var(--muted);border:1px solid var(--line)}
.card{background:var(--bg-3);border:1px solid var(--line);border-radius:12px;padding:16px}
.pp-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(330px,1fr));gap:12px}
.pp-head{display:flex;align-items:center;gap:8px;margin-bottom:10px}
.pp-name{font-size:15px;font-weight:600}
.pp-team{font-size:11px;color:var(--accent);background:rgba(108,92,231,.12);border:1px solid rgba(108,92,231,.35);border-radius:5px;padding:1px 7px}
.pp-task{font-size:13px;line-height:1.5;margin-bottom:8px}
.pp-task .pt-out{display:block;font-size:12px;color:var(--muted);margin-top:2px;padding-left:18px}
.pp-task .pt-detail{display:block;font-size:12px;color:var(--fg);opacity:.85;line-height:1.6;margin-top:3px;padding-left:18px}
.pp-issue{font-size:12px;color:var(--amber);margin-top:2px;padding-left:18px}
.pp-meta{border-top:1px solid var(--line);margin-top:10px;padding-top:8px;font-size:12px;line-height:1.7}
.pp-meta .pm-red{color:var(--red)}
.pp-meta .pm-amber{color:var(--amber)}
.pp-asg{border-top:1px solid var(--line);margin-top:10px;padding-top:8px}
.as-label{font-size:11px;font-weight:600;color:var(--accent);margin-bottom:6px}
.as-item{display:flex;align-items:baseline;gap:7px;font-size:12px;line-height:1.55;margin-bottom:4px}
.as-st{font-size:10px;border-radius:4px;padding:1px 6px;white-space:nowrap;flex-shrink:0}
.as-todo{background:rgba(225,112,85,.14);color:var(--red);border:1px solid rgba(225,112,85,.35)}
.as-doing{background:rgba(217,163,75,.14);color:var(--amber);border:1px solid rgba(217,163,75,.35)}
.as-done{background:var(--bg-3);color:var(--muted);border:1px solid var(--line)}
.as-done + .as-task{text-decoration:line-through;color:var(--muted)}
.as-task{flex:1;min-width:0}
.as-dl{font-size:11px;color:var(--muted);white-space:nowrap}
.tb-block{background:var(--bg-2);border:1px solid var(--line);border-radius:14px;padding:18px 20px;margin-bottom:16px}
.tb-head{font-size:15px;font-weight:600;display:flex;align-items:center;gap:10px;margin-bottom:8px}
.tb-head .cnt{margin-left:auto;color:var(--muted);font-size:12px;font-weight:400}
.tb-hl{font-size:13px;color:var(--muted);line-height:1.55;margin-bottom:10px}
.tb-chips{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:14px}
.tb-chip{background:var(--bg-3);border:1px solid var(--line);border-radius:8px;padding:6px 10px;font-size:12px;display:inline-flex;align-items:center;gap:7px;line-height:1.4}
.lh-st{font-size:10px;border-radius:4px;padding:1px 6px;white-space:nowrap;margin-right:4px}
.lh-todo{background:rgba(225,112,85,.14);color:var(--red);border:1px solid rgba(225,112,85,.35)}
.lh-doing{background:rgba(217,163,75,.14);color:var(--amber);border:1px solid rgba(217,163,75,.35)}
.lh-done{background:var(--bg-3);color:var(--muted);border:1px solid var(--line)}
.lh-proj{background:var(--bg-2);border:1px solid var(--line);border-radius:12px;padding:13px 15px;margin-bottom:8px;cursor:pointer}
.lh-proj:hover{border-color:var(--accent)}
.lh-proj .t{font-size:14px;font-weight:500}
.lh-proj .m{font-size:12px;color:var(--muted);margin-top:3px}
.rc-card{border-left:3px solid var(--accent)}
.ex-red{border-left:3px solid var(--red, #E17055)}
.ex-amber{border-left:3px solid var(--amber, #D9A34B)}
.rc-btn{float:right;background:rgba(108,92,231,.12);border:1px solid rgba(108,92,231,.4);color:var(--accent);border-radius:7px;padding:4px 11px;font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;margin-left:10px}
.rc-btn:hover{background:var(--accent);color:#fff}
"""

BRIEF_PAGE_CSS = """
:root{--bg:#1A1A1A;--bg-2:#202020;--bg-3:#262626;--fg:#F5F0EB;--muted:#8A857E;--line:#333;
--accent:#6C5CE7;--green:#8FA37E;--amber:#D9A34B;--red:#E17055;--blue:#6F8AA3}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--fg);font-family:'Pretendard Variable',sans-serif;font-weight:300;line-height:1.5;padding:32px 20px 64px}
.wrap{max-width:1100px;margin:0 auto}
header h1{font-weight:600;font-size:24px;letter-spacing:-.02em}
header .sub{color:var(--muted);margin-top:4px;font-size:13px}
#bv-gate{position:fixed;inset:0;background:var(--bg);display:flex;align-items:center;justify-content:center;z-index:100}
#bv-box{background:var(--bg-2);border:1px solid var(--line);border-radius:14px;padding:38px 40px;width:330px;text-align:center}
#bv-box h2{font-size:19px;font-weight:600;margin-bottom:6px}
#bv-box .lg-sub{color:var(--muted);font-size:13px;margin-bottom:22px}
#bv-box input{width:100%;background:var(--bg-3);border:1px solid var(--line);border-radius:8px;padding:12px 14px;color:var(--fg);font-size:14px;margin-bottom:10px;font-family:inherit}
#bv-box button{width:100%;background:var(--accent);color:#fff;border:0;border-radius:8px;padding:12px;font-size:14px;font-weight:600;cursor:pointer;margin-top:6px}
#bv-err{color:var(--red);font-size:12px;margin-top:12px;min-height:16px}
"""


def _brief_json(date_str, team):
    """daily-brief-{date}-{team}.json 로드 (없으면 None)."""
    p = BRIEF_DIR / f"daily-brief-{date_str}-{team}.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def _brief_biz_dates(team):
    """해당 팀 브리프 JSON이 존재하는 영업일 목록(오름차순)."""
    pat = re.compile(r"daily-brief-(\d{4}-\d{2}-\d{2})-" + re.escape(team) + r"$")
    out = []
    for f in sorted(BRIEF_DIR.glob(f"daily-brief-2*-{team}.json")):
        m = pat.fullmatch(f.stem)
        if not m:
            continue
        try:
            if datetime.datetime.strptime(m.group(1), "%Y-%m-%d").weekday() < 5:
                out.append(m.group(1))
        except ValueError:
            pass
    return out


def _brief_nav(base_url, dates, sel):
    """최근 7영업일 날짜 버튼 nav — 기존 /brief 패턴과 동일 스타일."""
    wk = ["월", "화", "수", "목", "금", "토", "일"]
    btns = []
    for ds in dates[-7:]:
        d = datetime.datetime.strptime(ds, "%Y-%m-%d")
        lab = f"{wk[d.weekday()]} {d.strftime('%m/%d')}"
        on = "background:var(--accent);color:#fff;border-color:var(--accent)" if ds == sel else "color:var(--muted)"
        btns.append(f'<a href="{base_url}&date={ds}" style="text-decoration:none;border:1px solid var(--line);'
                    f'border-radius:7px;padding:5px 12px;font-size:12px;{on}">{lab}</a>')
    return ('<div style="display:flex;gap:6px;margin:0 0 18px;align-items:center;flex-wrap:wrap">'
            '<span style="color:var(--muted);font-size:11px;margin-right:4px">최근 영업일</span>'
            + "".join(btns) + '</div>')


def _team_block_html(dba, td, only_person=None, chips_max=3):
    """팀 브리프 JSON → 팀 블록 HTML. only_person 지정 시 그 사람 카드만(직원 뷰)."""
    esc = dba._esc
    hl = f'<div class="tb-hl">{esc(td.get("headline") or "")}</div>' if td.get("headline") else ""
    chips = ""
    tops = (td.get("top5") or [])[:chips_max]
    if tops:
        cs = []
        for it in tops:
            col = dba.CAT_META.get(it.get("category"), ("", "var(--accent)"))[1]
            src = esc(it.get("source_employee") or "")
            cs.append(f'<span class="tb-chip" style="border-left:3px solid {col}">'
                      f'{dba._urg_badge(it.get("urgency", "low"))}<span class="vz-t">{esc(it.get("title", ""))}</span>'
                      + (f' <span class="vz-src" style="color:var(--muted);font-size:11px">{src}</span>' if src else "")
                      + '</span>')
        chips = f'<div class="tb-chips">{"".join(cs)}</div>'
    people = td.get("people") or []
    if only_person is not None:
        people = [p for p in people if p.get("name") == only_person]
    cards = (f'<div class="pp-grid">{"".join(dba._person_card(p) for p in people)}</div>'
             if people else "")
    return hl, chips, cards


def _brief_view_page(title, h1, sub, body, allow_js, deny_msg):
    """계정별 브리프 뷰 공통 페이지 — 셸 프리셋 세션(pm_sess) 게이트 + 자체 로그인 폴백.
    allow_js: 세션 객체 s를 받아 허용 여부를 리턴하는 JS 표현식 (예: "s.name===NAME||s.admin")"""
    return f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable.min.css">
<style>{BRIEF_PAGE_CSS}{BRIEF_CARD_CSS}</style></head>
<body>
<div id="bv-gate"><div id="bv-box">
  <h2>{h1}</h2><div class="lg-sub">{sub}</div>
  <input id="bv-id" placeholder="이름" autocomplete="username">
  <input id="bv-pin" type="password" placeholder="PIN" autocomplete="current-password">
  <button id="bv-btn">로그인</button><div id="bv-err"></div>
</div></div>
<div id="content" style="display:none"><div class="wrap">
<header><h1>{h1}</h1><div class="sub">{sub}</div></header>
{body}
</div></div>
<script>
(function(){{
  var gate=document.getElementById('bv-gate'), content=document.getElementById('content');
  function allowed(s){{ return !!(s && ({allow_js})); }}
  function enter(){{ gate.style.display='none'; content.style.display='block'; }}
  var s=null;
  ['pm_sess','brief_sess','team_brief_sess'].some(function(k){{
    try{{ var v=JSON.parse(localStorage.getItem(k)||'null'); if(v&&v.name){{ s=v; return true; }} }}catch(e){{}} return false;
  }});
  if(allowed(s)){{ enter(); return; }}
  if(s){{ document.getElementById('bv-err').textContent={deny_msg!r}; }}
  function doLogin(){{
    var id=document.getElementById('bv-id').value.trim(), pin=document.getElementById('bv-pin').value.trim();
    var err=document.getElementById('bv-err'); err.textContent='';
    fetch('/api/login',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{id:id,pin:pin}})}})
      .then(function(r){{return r.json();}})
      .then(function(d){{
        if(d.ok && allowed(d)){{ localStorage.setItem('pm_sess',JSON.stringify(Object.assign({{}},d,{{id:id,pin:pin}}))); enter(); }}
        else if(d.ok){{ err.textContent={deny_msg!r}; }}
        else {{ err.textContent=d.error||'로그인 실패'; }}
      }}).catch(function(){{ err.textContent='서버 연결 실패'; }});
  }}
  document.getElementById('bv-btn').onclick=doLogin;
  document.getElementById('bv-pin').addEventListener('keydown',function(e){{ if(e.key==='Enter') doLogin(); }});
}})();
</script>
</body></html>"""

def project_teams(p):
    """프로젝트에 걸린 팀 집합 = PM + 멤버들의 소속 팀."""
    people = list(p.get("members") or [])
    if p.get("pm"): people.append(p["pm"])
    return {t for t in (emp_team(x) for x in people) if t}

def can_manage(uid, p):
    """멤버 배정 권한: 대표 / 담당 PM / (본인 리더팀이 이 프로젝트에 걸린) 팀 리더."""
    if is_admin(uid) or uid == p.get("pm"):
        return True
    return bool(set(lead_teams_of(uid)) & project_teams(p))

def can_view(uid, p):
    # 리더는 '본인 팀이 걸린' 프로젝트를 멤버가 아니어도 열람(멤버 배정 위해)
    return is_admin(uid) or uid == p.get("pm") or uid in (p.get("members") or []) or can_manage(uid, p)

def can_edit(uid, p):
    # 내용(업무·이슈·브리프) 편집은 기존대로 담당 PM·대표만 — 리더는 멤버 배정만
    return is_admin(uid) or uid == p.get("pm")

class H(BaseHTTPRequestHandler):
    server_version = "PRDashboard/1.0"
    def log_message(self, *a): pass

    def _proxy_arisa2(self, method="GET"):
        """리버스 프록시: /arisa2/* → ARISA2_UPSTREAM (stdlib only)."""
        # /arisa2/foo → /foo 로 변환 (/arisa2 자체는 / 로)
        upstream_path = self.path[len("/arisa2"):] or "/"
        url = ARISA2_UPSTREAM + upstream_path
        body = None
        if method in ("POST", "PUT", "PATCH"):
            n = int(self.headers.get("Content-Length", 0) or 0)
            body = self.rfile.read(n) if n else b""
        headers = {}
        for h in ("Content-Type", "Authorization"):  # Bearer 토큰 전달(ARISA 2.0 인증)
            v = self.headers.get(h)
            if v:
                headers[h] = v
        try:
            req = Request(url, data=body, headers=headers, method=method)
            with urlopen(req, timeout=30) as resp:
                resp_body = resp.read()
                ctype = resp.headers.get("Content-Type", "application/octet-stream")
                if "text/html" in ctype:
                    # SPA의 API 베이스를 프록시 프리픽스로 재작성 → /arisa2/api/* 로 호출
                    resp_body = resp_body.replace(b"var API='';", b"var API='/arisa2';")
                self.send_response(resp.status)
                self.send_header("Content-Type", ctype)
                self.send_header("Content-Length", str(len(resp_body)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(resp_body)
        except HTTPError as e:
            # upstream이 4xx/5xx를 반환해도 상태·본문 그대로 전달(401 인증게이트 = 서버 살아있음)
            resp_body = e.read()
            self.send_response(e.code)
            self.send_header("Content-Type", e.headers.get("Content-Type", "application/json; charset=utf-8"))
            self.send_header("Content-Length", str(len(resp_body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(resp_body)
        except URLError:
            err = b'{"ok":false,"error":"ARISA 2.0 unavailable"}'
            self.send_response(502)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(err)))
            self.end_headers()
            self.wfile.write(err)

    def _send(self, code, body, ctype="application/json; charset=utf-8"):
        data = body if isinstance(body, (bytes, bytearray)) else json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def _body(self):
        n = int(self.headers.get("Content-Length", 0) or 0)
        if not n: return {}
        try: return json.loads(self.rfile.read(n).decode("utf-8") or "{}")
        except Exception: return {}

    def do_GET(self):
        u = urlparse(self.path); path = u.path; q = parse_qs(u.query)
        if path.startswith("/arisa2"):
            return self._proxy_arisa2("GET")
        if path in ("/", "/index.html"):
            # 통합 셸 — 로그인 1회, 역할별 탭(프로젝트/Brief/Weekly/Decision)
            # 리더 홈(팀 Todo·팀원 보고 카드)이 브리프 카드 스타일을 쓰므로 CARD_CSS 주입
            html = UNIFIED_SHELL.replace("</style>", BRIEF_CARD_CSS + "</style>", 1)
            return self._send(200, html.encode("utf-8"), "text/html; charset=utf-8")
        if path == "/projects":
            # 포트폴리오 HTML (구 / 핸들러) — 셸 iframe 또는 직접 접속(자체 로그인)
            try:
                html = HTML.read_text(encoding="utf-8")
            except Exception:
                return self._send(500, "<h1>대시보드 파일이 없습니다. generate-portfolio.py 실행 필요</h1>".encode("utf-8"), "text/html; charset=utf-8")
            inject = "<script>window.SERVED=true;window.API_BASE='/api';</script>\n</head>"
            html = html.replace("</head>", inject, 1)
            return self._send(200, html.encode("utf-8"), "text/html; charset=utf-8")
        if path in ("/dashboard", "/team"):
            # 구 대표/리더 셸 → 통합 셸로 (기존 알림 링크 호환)
            self.send_response(301)
            self.send_header("Location", "/")
            self.send_header("Content-Length", "0")
            self.end_headers()
            return
        if path == "/weekly":
            # 로그인은 HTML 내장 JS가 /api/login으로 처리. 성장지표(.growth)는 admin(대표)만
            # CSS로 노출(직원 로그인 시 숨김 — 측정설계 v2). 서버는 최신 HTML만 서빙.
            # 날짜 형식만 매칭(팀 suffix 파일 `…-W26-공간팀.html` 격리)
            files = sorted(f for f in WEEKLY_DIR.glob("weekly-report-2*.html")
                           if re.fullmatch(r"weekly-report-\d{4}-W\d{2}", f.stem))
            if not files:
                return self._send(404, "<h1>주간 대시보드가 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            return self._send(200, files[-1].read_text(encoding="utf-8").encode("utf-8"), "text/html; charset=utf-8")
        if path == "/brief":
            # 대표 Daily Brief — 로그인은 HTML 내장 JS가 /api/login으로 처리(대표 admin만 입장).
            # 날짜 네비: 업무 맥락 위해 오늘 + 직전 3영업일(주말 제외)을 버튼으로 주입. ?date=로 전환.
            # 날짜 형식만 매칭(팀 suffix 파일 `…-06-30-공간팀.html` 격리)
            files = sorted(f for f in BRIEF_DIR.glob("daily-brief-2*.html")
                           if re.fullmatch(r"daily-brief-\d{4}-\d{2}-\d{2}", f.stem))
            if not files:
                return self._send(404, "<h1>오늘 브리프가 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            biz = []  # 영업일 brief 날짜(월~금)
            for f in files:
                ds = f.stem.replace("daily-brief-", "")
                try:
                    if datetime.datetime.strptime(ds, "%Y-%m-%d").weekday() < 5:
                        biz.append(ds)
                except ValueError:
                    pass
            recent = biz[-7:]  # 오늘 + 직전 6영업일 (최근 일주일 이력)
            sel = (q.get("date") or [""])[0]
            target = BRIEF_DIR / f"daily-brief-{sel}.html"
            if not (sel and target.exists()):
                target = files[-1]
                sel = files[-1].stem.replace("daily-brief-", "")
            wk = ["월", "화", "수", "목", "금", "토", "일"]
            btns = []
            for ds in recent:
                d = datetime.datetime.strptime(ds, "%Y-%m-%d")
                lab = f"{wk[d.weekday()]} {d.strftime('%m/%d')}"
                on = "background:var(--accent);color:#fff;border-color:var(--accent)" if ds == sel else "color:var(--muted)"
                btns.append(f'<a href="/brief?date={ds}" style="text-decoration:none;border:1px solid var(--line);'
                            f'border-radius:7px;padding:5px 12px;font-size:12px;{on}">{lab}</a>')
            nav = ('<div style="display:flex;gap:6px;margin:0 0 18px;align-items:center">'
                   '<span style="color:var(--muted);font-size:11px;margin-right:4px">최근 영업일</span>'
                   + "".join(btns) + '</div>')
            html = target.read_text(encoding="utf-8").replace("</header>", "</header>" + nav, 1)
            html = html.replace("</body>", BRIEF_CMT_JS.replace("__DATE__", sel) + "</body>", 1)
            return self._send(200, html.encode("utf-8"), "text/html; charset=utf-8")
        if path == "/team-brief":
            # 팀 Brief — 로그인은 셸이 처리(team_brief_sess). team별 파일 + 날짜네비.
            team = (q.get("team") or [""])[0]
            if not team:
                return self._send(400, "<h1>team 파라미터가 필요합니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            pre = "daily-brief-"
            pat = re.compile(r"daily-brief-\d{4}-\d{2}-\d{2}-" + re.escape(team) + r"$")
            files = sorted(f for f in BRIEF_DIR.glob(f"daily-brief-2*-{team}.html") if pat.fullmatch(f.stem))
            if not files:
                return self._send(404, f"<h1>{team} 팀 브리프가 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            def _ds(f):  # stem에서 날짜부분(YYYY-MM-DD) 추출
                return f.stem[len(pre):-(len(team) + 1)]
            biz = []
            for f in files:
                ds = _ds(f)
                try:
                    if datetime.datetime.strptime(ds, "%Y-%m-%d").weekday() < 5:
                        biz.append(ds)
                except ValueError:
                    pass
            recent = biz[-7:]  # 최근 일주일 이력 (팀)
            sel = (q.get("date") or [""])[0]
            target = BRIEF_DIR / f"daily-brief-{sel}-{team}.html"
            if not (sel and target.exists()):
                target = files[-1]
                sel = _ds(files[-1])
            wk = ["월", "화", "수", "목", "금", "토", "일"]
            btns = []
            for ds in recent:
                d = datetime.datetime.strptime(ds, "%Y-%m-%d")
                lab = f"{wk[d.weekday()]} {d.strftime('%m/%d')}"
                on = "background:var(--accent);color:#fff;border-color:var(--accent)" if ds == sel else "color:var(--muted)"
                btns.append(f'<a href="/team-brief?team={quote(team)}&date={ds}" style="text-decoration:none;border:1px solid var(--line);'
                            f'border-radius:7px;padding:5px 12px;font-size:12px;{on}">{lab}</a>')
            nav = ('<div style="display:flex;gap:6px;margin:0 0 18px;align-items:center">'
                   '<span style="color:var(--muted);font-size:11px;margin-right:4px">최근 영업일</span>'
                   + "".join(btns) + '</div>')
            html_str = target.read_text(encoding="utf-8").replace("</header>", "</header>" + nav, 1)
            html_str = html_str.replace("</body>", BRIEF_CMT_JS.replace("__DATE__", sel) + "</body>", 1)
            return self._send(200, html_str.encode("utf-8"), "text/html; charset=utf-8")
        if path == "/my-brief":
            # 직원용 개인 브리프 — 내 카드(상세+분장) + 팀 헤드라인·핵심칩만 (동료 카드 비공개).
            name = (q.get("name") or [""])[0]
            if not name:
                return self._send(400, "<h1>name 파라미터가 필요합니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            team = emp_team(name) or ""
            try:
                dba = _dba()
            except Exception as e:
                return self._send(500, f"<h1>브리프 렌더러 로드 실패: {e}</h1>".encode("utf-8"), "text/html; charset=utf-8")
            dates = _brief_biz_dates(team) if team else []
            if not dates:
                body = '<div class="muted">아직 생성된 팀 브리프가 없습니다.</div>'
                page = _brief_view_page(f"내 브리프 · {name}", "내 오늘 브리프",
                                        f"{name} · {team or '팀 미배정'}", body,
                                        f"s.name==={json.dumps(name, ensure_ascii=False)}||s.admin",
                                        "본인 전용 화면입니다")
                return self._send(200, page.encode("utf-8"), "text/html; charset=utf-8")
            sel = (q.get("date") or [""])[0]
            if sel not in dates:
                sel = dates[-1]
            td = _brief_json(sel, team) or {}
            hl, chips, cards = _team_block_html(dba, td, only_person=name)
            if not cards:
                cards = ('<div class="card"><div class="muted">이 날짜에 제출한 보고가 없습니다. '
                         '저녁 20:00·22:00 리마인드에 맞춰 보고하면 다음날 아침 브리프에 반영됩니다.</div></div>')
            nav = _brief_nav(f"/my-brief?name={quote(name)}", dates, sel)
            body = (nav + f'<div class="tb-block"><div class="tb-head">{dba._esc(team)}'
                    f'<span class="cnt">{dba._esc(sel)}</span></div>{hl}{chips}{cards}</div>')
            page = _brief_view_page(f"내 브리프 · {name}", "내 오늘 브리프",
                                    f"{name} · {team}", body,
                                    f"s.name==={json.dumps(name, ensure_ascii=False)}||s.admin",
                                    "본인 전용 화면입니다")
            return self._send(200, page.encode("utf-8"), "text/html; charset=utf-8")
        if path == "/lead-brief":
            # 겸임 리더용 — 담당팀들을 한 페이지로 병합 (팀 headline+핵심칩+개인카드 전체).
            teams_p = (q.get("teams") or [""])[0]
            teams = [t.strip() for t in teams_p.split(",") if t.strip()]
            if not teams:
                return self._send(400, "<h1>teams 파라미터가 필요합니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            try:
                dba = _dba()
            except Exception as e:
                return self._send(500, f"<h1>브리프 렌더러 로드 실패: {e}</h1>".encode("utf-8"), "text/html; charset=utf-8")
            all_dates = sorted(set().union(*[set(_brief_biz_dates(t)) for t in teams]))
            if not all_dates:
                return self._send(404, "<h1>팀 브리프가 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            sel = (q.get("date") or [""])[0]
            if sel not in all_dates:
                sel = all_dates[-1]
            blocks = []
            for t in teams:
                td = _brief_json(sel, t)
                if not td:
                    blocks.append(f'<div class="tb-block"><div class="tb-head">{_dba()._esc(t)}</div>'
                                  '<div class="muted">이 날짜 브리프 없음</div></div>')
                    continue
                hl, chips, cards = _team_block_html(dba, td)
                if not cards:
                    cards = '<div class="muted">오늘 보고 없음</div>'
                blocks.append(f'<div class="tb-block"><div class="tb-head">{dba._esc(t)}'
                              f'<span class="cnt">보고 {td.get("active_people", 0)}명</span></div>{hl}{chips}{cards}</div>')
            nav = _brief_nav(f"/lead-brief?teams={quote(teams_p)}", all_dates, sel)
            teams_js = json.dumps(teams, ensure_ascii=False)
            allow = f"(s.admin || {teams_js}.every(function(t){{return (s.lead_teams||[]).indexOf(t)>=0;}}))"
            page = _brief_view_page(f"팀 Brief · {' · '.join(teams)}", f"{' · '.join(teams)} 오늘 브리프",
                                    f"{sel} · 팀 리더", nav + "".join(blocks), allow,
                                    "담당 팀 리더 전용 화면입니다")
            page = page.replace("</body>", BRIEF_CMT_JS.replace("__DATE__", sel) + "</body>", 1)
            return self._send(200, page.encode("utf-8"), "text/html; charset=utf-8")
        if path == "/team-weekly":
            # 팀 주간 — 로그인은 셸이 처리(team_weekly_sess). team별 최신 파일.
            team = (q.get("team") or [""])[0]
            if not team:
                return self._send(400, "<h1>team 파라미터가 필요합니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            pat = re.compile(r"weekly-report-\d{4}-W\d{2}-" + re.escape(team) + r"$")
            files = sorted(f for f in WEEKLY_DIR.glob(f"weekly-report-2*-{team}.html") if pat.fullmatch(f.stem))
            if not files:
                return self._send(404, f"<h1>{team} 팀 주간이 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            return self._send(200, files[-1].read_text(encoding="utf-8").encode("utf-8"), "text/html; charset=utf-8")
        if path == "/simulator":
            return self._send(200, SIMULATOR_HTML.encode("utf-8"), "text/html; charset=utf-8")
        if path == "/guide":
            # 일일업무보고 가이드 — 직원 열람용(무인증, 공개).
            gp = _WS / "20-operations" / "23-arisa" / "guide" / "daily-report-guide.html"
            if not gp.exists():
                return self._send(404, "<h1>가이드가 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            return self._send(200, gp.read_text(encoding="utf-8").encode("utf-8"), "text/html; charset=utf-8")
        if path == "/guide-os":
            # ARISA OS 통합 사용 가이드 (무인증, 공개)
            gp = _WS / "20-operations" / "23-arisa" / "guide" / "arisa-os-guide.html"
            if not gp.exists():
                return self._send(404, "<h1>가이드가 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            return self._send(200, gp.read_text(encoding="utf-8").encode("utf-8"), "text/html; charset=utf-8")
        if path == "/guide-flow":
            # 업무분장 플로우 구조도 (무인증, 공개)
            gp = _WS / "20-operations" / "23-arisa" / "assignment-flow-구조도.html"
            if not gp.exists():
                return self._send(404, "<h1>구조도가 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            return self._send(200, gp.read_text(encoding="utf-8").encode("utf-8"), "text/html; charset=utf-8")
        if path == "/guide/template.xlsx":
            # 주간업무계획 템플릿 다운로드
            gp = _WS / "20-operations" / "23-arisa" / "guide" / "주간업무계획_템플릿.xlsx"
            if not gp.exists():
                return self._send(404, "<h1>템플릿이 아직 없습니다.</h1>".encode("utf-8"), "text/html; charset=utf-8")
            data = gp.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename*=UTF-8''%EC%A3%BC%EA%B0%84%EC%97%85%EB%AC%B4%EA%B3%84%ED%9A%8D_%ED%85%9C%ED%94%8C%EB%A6%BF.xlsx")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if path == "/api/health":
            return self._send(200, {"ok": True})
        if path == "/api/assignees":
            # 계층적 분장 후보 — 대표=리더급, 리더=자기 팀원 + 타팀 리더(이관)
            uid = (q.get("user") or [""])[0]
            emp = load_emp().get("by_name", {})
            tl = load_emp().get("team_leads", {})
            if is_admin(uid):
                lead_team = {}
                for team, ldr in tl.items():
                    lead_team.setdefault(ldr, []).append(team)
                data = [{"name": n, "team": "·".join(lead_team[n])} for n in sorted(lead_team.keys())]
                if uid not in lead_team:  # 대표 본인도 담당자 후보(직접 챙기는 업무)
                    data.insert(0, {"name": uid, "team": emp_team(uid) or "경영"})
                level = "리더"
            else:
                my = set(lead_teams_of(uid))
                data = [{"name": n, "team": (v or {}).get("team", "")} for n, v in sorted(emp.items())
                        if (v or {}).get("team") in my]
                # 타팀 리더 이관 후보 (2026-07-20): 리더→타팀 리더→그 팀원 위계 경로.
                # 이관받은 리더의 팀 스코프로 기록돼 그 팀 홈에 잡힌다 (재분장은 기존 기능).
                seen = {d["name"] for d in data}
                other_leads = {}
                for team, ldr in tl.items():
                    if team not in my:
                        other_leads.setdefault(ldr, []).append(team)
                for n in sorted(other_leads):
                    if n != uid and n not in seen:
                        data.append({"name": n, "team": "·".join(other_leads[n]),
                                     "leader": True})
                level = "팀원"
            return self._send(200, {"ok": True, "canAssign": is_admin(uid) or is_leader(uid),
                                    "level": level, "assignees": data})
        if path == "/api/my-work":
            # 개인 대시보드 — 내 분장(미완) + 내 프로젝트 일정(owner)
            uid = (q.get("user") or [""])[0]
            mine = [a for a in _assign_read() if a["assignee"] == uid and a["status"] not in _ASSIGN_HIDDEN_STATES]
            mine.sort(key=lambda a: (a["status"] == "완료", a["status"] != "진행중", a.get("deadline") or "9999"))
            projs = []
            for p in load_projects():
                ts = [t for t in (p.get("tasks") or []) if t.get("owner") == uid
                      and (t.get("status") or "") not in _TASK_DONE_STATES]
                if ts:
                    projs.append({"id": p.get("id"), "name": p.get("name"), "dday": p.get("dday"),
                                  "tasks": [{"task": t.get("task"), "start": t.get("start"), "end": t.get("end"),
                                             "status": t.get("status"), "priority": t.get("priority"),
                                             "division": t.get("division")} for t in ts]})
            return self._send(200, {"ok": True, "user": uid, "assignments": mine, "projects": projs,
                                    "daily": _person_workbrief(uid, mine)})
        if path == "/api/exec-attn":
            # 대표창 — 지연 업무(마감 경과·미완료 분장) + 결재·확인 필요(미해결 결정)
            uid = (q.get("user") or [""])[0]
            if not is_admin(uid):
                return self._send(403, {"ok": False, "error": "대표 전용"})
            today = datetime.date.today()
            reg_cut = today - datetime.timedelta(days=30)
            overdue, seen = [], set()
            approvals = []
            unassigned, useen = [], set()  # filament 반영 — 담당 미지정 배정 큐
            for a in _assign_read():
                st = (a.get("status") or "미착수")
                if st == "완료":
                    approvals.append(a)   # 승인 대기 큐 (대표 승인용)
                if st in _ASSIGN_DONE_STATES or st == "삭제":
                    continue
                ds = (a.get("date") or "").strip()[:10]
                try:
                    if datetime.date.fromisoformat(ds) < reg_cut:
                        continue
                except ValueError:
                    pass
                if not (a.get("assignee") or "").strip():
                    ukey = (a.get("project"), a.get("task"), a.get("deadline"))
                    if ukey not in useen:
                        useen.add(ukey)
                        unassigned.append(a)
                if not a.get("days_overdue"):
                    continue
                key = (a.get("project"), a.get("task"), a.get("assignee"), (a.get("deadline") or "").strip()[:10])
                if key in seen:
                    continue
                seen.add(key)
                overdue.append(a)
            overdue.sort(key=lambda x: -x["days_overdue"])
            # 결재·확인 필요 — 미해결 결정 이월 로그 + 최신 브리프 decision_summary 병합
            decisions, titles = [], set()
            if _load_open_decisions:
                try:
                    for d in _load_open_decisions(14):
                        t = (d.get("decision_needed") or "").strip()
                        if t and t not in titles:
                            titles.add(t)
                            decisions.append({"title": t, "who": d.get("source_employee") or "",
                                              "project": d.get("project") or "",
                                              "age": d.get("age_days", 0)})
                except Exception:
                    pass
            try:
                bfiles = sorted(f for f in BRIEF_DIR.glob("daily-brief-2*.json")
                                if re.fullmatch(r"daily-brief-\d{4}-\d{2}-\d{2}", f.stem))
                if bfiles:
                    bd = json.loads(bfiles[-1].read_text(encoding="utf-8"))
                    for it in bd.get("decision_summary") or []:
                        t = (it.get("title") or "").strip()
                        if t and t not in titles:
                            titles.add(t)
                            proj = (it.get("project") or "")
                            proj = "" if str(proj).lower() in ("none", "null") else proj
                            decisions.append({"title": t, "who": it.get("source_employee") or "",
                                              "project": proj, "age": 0})
            except Exception:
                pass
            return self._send(200, {"ok": True, "overdue": overdue, "decisions": decisions,
                                    "approvals": approvals, "unassigned": unassigned})
        if path == "/api/lead-home":
            # 리더 홈 — 팀 Todo(이번주 분장) + 진행중 프로젝트 + 팀원 오늘 보고 카드(HTML fragment)
            uid = (q.get("user") or [""])[0]
            if not load_users().get(uid):
                return self._send(401, {"ok": False, "error": "unknown user"})
            teams = lead_teams_of(uid)
            if not teams:
                return self._send(403, {"ok": False, "error": "리더 전용"})
            today = datetime.date.today()
            # 2주 윈도우(지난주 월~이번주 일) — 주 바뀐 직후 미완료 분장이 사라지는 문제 방지
            ws = today - datetime.timedelta(days=today.weekday() + 7)
            we = today - datetime.timedelta(days=today.weekday()) + datetime.timedelta(days=6)
            assigns = []
            approvals = []
            unassigned = []  # filament 반영 — 담당 미지정(팀 귀속 또는 팀 공란) 배정 큐
            for a in _assign_read():
                # 승인 대기(완료)는 날짜 무관 전체 — 리더 승인 큐
                if (a.get("status") or "") == "완료" and (a.get("team") in teams or emp_team(a.get("assignee")) in teams):
                    approvals.append(a)
                ds = (a.get("date") or "").strip()[:10]
                try:
                    d = datetime.date.fromisoformat(ds)
                except ValueError:
                    continue
                if not (ws <= d <= we):
                    continue
                if a.get("status") in _ASSIGN_HIDDEN_STATES:
                    continue
                if not (a.get("assignee") or "").strip() and a.get("status") in _ST.ASSIGN_OPEN_STATES \
                        and (a.get("team") in teams or not (a.get("team") or "").strip()):
                    unassigned.append(a)
                    continue
                if a.get("team") in teams or emp_team(a.get("assignee")) in teams:
                    assigns.append(a)
            assigns.sort(key=lambda a: (a["status"] == "완료", a["status"] != "진행중",
                                        a.get("deadline") or "9999", a.get("assignee") or ""))
            # 받은 업무 출처 라벨 (L열 등록자) — 과거 행(공란)은 종전 표기 유지 (2026-07-20)
            for a in assigns:
                by = (a.get("by") or "").strip()
                if not by or is_admin(by):
                    a["src"] = "대표 지시"
                elif by == uid:
                    a["src"] = "본인 등록"
                elif is_leader(by):
                    a["src"] = f"{by} 리더 이관"
                else:
                    a["src"] = f"{by} 등록"
            projs = []
            for p in load_projects():
                if not (project_teams(p) & set(teams)):
                    continue
                ru = _ST.task_rollup(p.get("tasks"))  # G3 — 태스크 파생 진행률 롤업
                end = (p.get("end") or p.get("dday") or "").strip()
                if end and end < today.isoformat():
                    continue  # 종료된 프로젝트 제외
                projs.append({"id": p.get("id"), "name": p.get("name"), "pm": p.get("pm"),
                              "dday": p.get("dday"), "task_done": ru["done"], "task_total": ru["total"],
                              "percent": ru["percent"]})
            projs.sort(key=lambda p: p.get("dday") or "9999")
            brief_html, brief_date = "", ""
            try:
                dba = _dba()
                for t in teams:
                    ds = _brief_biz_dates(t)
                    if ds and ds[-1] > brief_date:
                        brief_date = ds[-1]
                blocks = []
                for t in teams:
                    td = _brief_json(brief_date, t) if brief_date else None
                    if not td:
                        continue
                    hl, chips, cards = _team_block_html(dba, td)
                    if not cards:
                        cards = '<div class="muted">이 날짜 팀 보고 없음</div>'
                    blocks.append(f'<div class="tb-block"><div class="tb-head">{dba._esc(t)}'
                                  f'<span class="cnt">보고 {td.get("active_people", 0)}명</span></div>{hl}{chips}{cards}</div>')
                brief_html = "".join(blocks)
            except Exception as e:
                brief_html = f'<div class="muted">보고 로드 실패: {e}</div>'
            return self._send(200, {"ok": True, "teams": teams, "assignments": assigns,
                                    "projects": projs, "approvals": approvals,
                                    "unassigned": unassigned,
                                    "brief_html": brief_html, "brief_date": brief_date,
                                    "daily": _person_workbrief(uid, [a for a in assigns if a.get("assignee") == uid])})
        if path == "/api/projects":
            uid = (q.get("user") or [""])[0]
            if not load_users().get(uid): return self._send(401, {"ok": False, "error": "unknown user"})
            vis = []
            for p in load_projects():
                if can_view(uid, p):
                    q = dict(p); q["canManage"] = can_manage(uid, p); q["canEdit"] = can_edit(uid, p)
                    q["rollup"] = _ST.task_rollup(p.get("tasks"))  # G3 — 태스크 파생 진행률
                    vis.append(q)
            return self._send(200, {"ok": True, "projects": vis, "admin": is_admin(uid),
                                    "canCreate": is_admin(uid) or is_leader(uid)})
        if path == "/api/project":
            uid = (q.get("user") or [""])[0]; pid = (q.get("id") or [""])[0]
            p = get_project(pid)
            if not p: return self._send(404, {"ok": False, "error": "not found"})
            if not can_view(uid, p): return self._send(403, {"ok": False, "error": "forbidden"})
            if _sync_assign_status(p): save_project(p)  # 분장 시트(SSOT) 상태 lazy 반영
            return self._send(200, {"ok": True, "project": p, "canEdit": can_edit(uid, p), "canManage": can_manage(uid, p),
                                    "assignments": _project_assignments(p.get("name") or "", p.get("aliases"), p.get("id") or ""),
                                    "memory": _memory_hub(p)})  # B1 — ARISA 메모리 링크백
        if path == "/api/project/open-assigns":
            # 아카이브 모달 사전 경고용 — 열린 분장(미착수·진행중, 전 기간)
            uid = (q.get("user") or [""])[0]
            if not load_users().get(uid):
                return self._send(401, {"ok": False, "error": "unknown user"})
            p = get_project((q.get("id") or [""])[0])
            if not p:
                return self._send(404, {"ok": False, "error": "not found"})
            oa = _open_assigns(p)
            return self._send(200, {"ok": True, "count": len(oa), "list": oa[:5]})
        if path == "/api/assign-history":
            # 분장 상태 전이 이력 (G7) — 로그인 사용자 누구나 (팀 투명성)
            uid = (q.get("user") or [""])[0]
            if not load_users().get(uid):
                return self._send(401, {"ok": False, "error": "unknown user"})
            hist = _load_st_history(task=(q.get("task") or [""])[0],
                                    assignee=(q.get("assignee") or [""])[0],
                                    pid=(q.get("pid") or [""])[0], limit=30)
            return self._send(200, {"ok": True, "history": hist})
        if path == "/api/project/doc":
            # 업로드 자료(회의록) 원문 조회 — 프로젝트 열람 권한자
            uid = (q.get("user") or [""])[0]
            pid = (q.get("id") or [""])[0]
            ts = (q.get("ts") or [""])[0]
            p = get_project(pid)
            if not p: return self._send(404, {"ok": False, "error": "프로젝트 없음"})
            if not can_view(uid, p): return self._send(403, {"ok": False, "error": "열람 권한 없음"})
            if not re.fullmatch(r"\d{8}-\d{6}", ts):
                return self._send(400, {"ok": False, "error": "ts 형식"})
            f = DOC_DIR / _safe(pid) / f"{ts}.md"
            if not f.exists(): return self._send(404, {"ok": False, "error": "자료 파일 없음"})
            meta = next((d for d in (p.get("docs") or []) if d.get("ts") == ts), {})
            return self._send(200, {"ok": True, "title": meta.get("title") or ts,
                                    "by": meta.get("by") or "", "ts": ts,
                                    "text": f.read_text(encoding="utf-8")})
        if path == "/api/project/memory-doc":
            # ARISA 메모리 원문 열람(B1) — 프로젝트 열람 권한자, 새 탭 HTML
            uid = (q.get("user") or [""])[0]
            p = get_project((q.get("id") or [""])[0])
            if not p:
                return self._send(404, "<h1>프로젝트 없음</h1>".encode("utf-8"), "text/html; charset=utf-8")
            if not can_view(uid, p):
                return self._send(403, "<h1>열람 권한 없음</h1>".encode("utf-8"), "text/html; charset=utf-8")
            f = (q.get("f") or [""])[0]
            # 경로 이탈 방지 — 메모리 폴더 직속 .md 또는 01_meeting_logs/ 하위 .md만
            if not re.fullmatch(r"(?:01_meeting_logs/)?[\w가-힣.\- ()]+\.(?:md|html)", f) or ".." in f:
                return self._send(400, "<h1>파일명 형식</h1>".encode("utf-8"), "text/html; charset=utf-8")
            d = _memory_dir_for(p)
            fp = (d / f) if d else None
            if not (fp and fp.exists()):
                return self._send(404, "<h1>기록 파일 없음</h1>".encode("utf-8"), "text/html; charset=utf-8")
            import html as _h
            if f.endswith(".html"):
                return self._send(200, fp.read_text(encoding="utf-8").encode("utf-8"), "text/html; charset=utf-8")
            page = ("<!DOCTYPE html><html lang=\"ko\"><head><meta charset=\"utf-8\">"
                    "<meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">"
                    f"<title>{_h.escape(f)} · {_h.escape(p.get('name') or '')}</title>"
                    "<link rel=\"stylesheet\" href=\"https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable.min.css\">"
                    "<style>body{background:#1A1A1A;color:#F5F0EB;font-family:'Pretendard Variable',sans-serif;"
                    "max-width:860px;margin:0 auto;padding:40px 24px}pre{white-space:pre-wrap;line-height:1.7;font-size:13.5px}"
                    "h1{font-size:16px;color:#6C5CE7;font-weight:600}</style></head><body>"
                    f"<h1>🗂 {_h.escape(p.get('name') or '')} — {_h.escape(f)}</h1>"
                    f"<pre>{_h.escape(fp.read_text(encoding='utf-8'))}</pre></body></html>")
            return self._send(200, page.encode("utf-8"), "text/html; charset=utf-8")
        if path == "/api/brief-comments":
            uid = (q.get("user") or [""])[0]
            if not load_users().get(uid): return self._send(401, {"ok": False, "error": "unknown user"})
            ds = (q.get("date") or [""])[0]
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", ds): return self._send(400, {"ok": False, "error": "date 형식"})
            return self._send(200, {"ok": True, "comments": _load_comments(ds)})
        return self._send(404, {"ok": False, "error": "not found"})

    def do_POST(self):
        path = urlparse(self.path).path
        if path.startswith("/arisa2"):
            return self._proxy_arisa2("POST")
        b = self._body()
        if path == "/api/login":
            uid = b.get("id", ""); pin = b.get("pin", "")
            users = load_users()
            u = users.get(uid)
            if not u: return self._send(401, {"ok": False, "error": "등록되지 않은 이름입니다."})
            cur = str(u.get("pin") or "")
            pin_set = False
            if cur == "":
                # PIN 미설정 → 첫 로그인 시 본인이 설정
                if len(str(pin)) < 4:
                    return self._send(400, {"ok": False, "error": "첫 로그인입니다. PIN을 4자 이상으로 설정해주세요."})
                set_pin(uid, pin); pin_set = True
            elif cur != str(pin):
                return self._send(401, {"ok": False, "error": "ID 또는 PIN이 올바르지 않습니다."})
            return self._send(200, {"ok": True, "name": uid, "role": u.get("role", "직원"),
                                    "admin": is_admin(uid), "lead_teams": lead_teams_of(uid), "pin_set": pin_set})
        if path == "/api/set-pin":
            # PIN 변경(자가설정) — 현재 PIN 검증 후 새 PIN 저장
            uid = b.get("id", ""); cur = b.get("pin", ""); new = b.get("new_pin", "")
            if not auth(uid, cur): return self._send(401, {"ok": False, "error": "현재 PIN이 올바르지 않습니다."})
            if len(str(new)) < 4: return self._send(400, {"ok": False, "error": "새 PIN은 4자 이상이어야 합니다."})
            set_pin(uid, new)
            return self._send(200, {"ok": True})
        if path == "/api/simulator/review":
            sim_mode = b.get("mode", "daily")
            fields = b.get("fields") or {}
            if not any((fields.get(k) or "").strip() for k in fields):
                return self._send(400, {"ok": False, "error": "최소 1개 항목을 입력해주세요."})
            user_msg = _build_simulator_input(sim_mode, fields)
            if sim_mode == "brief":
                result = _call_claude(SIMULATOR_BRIEF_PROMPT, user_msg)
            else:
                # 채점은 봇과 동일: 공유 루브릭 + gpt-4o-mini temp 0.3 (2026-07-20 갭 해소)
                result = _call_llm_json(_sim_daily_prompt(), user_msg,
                                        openai_model="gpt-4o-mini", openai_only=True)
                if result and isinstance(result.get("scores"), dict):
                    result = _report_score.validate_scores(result)
                    t = result.get("total", 0)
                    result["grade"] = ("S" if t >= 90 else "A" if t >= 75 else
                                       "B" if t >= 60 else "C" if t >= 40 else "D")
            if not result:
                return self._send(500, {"ok": False, "error": "AI 리뷰에 실패했습니다. API 키를 확인하세요."})
            return self._send(200, {"ok": True, "result": result})
        if path == "/api/simulator/draft":
            sim_mode = b.get("mode", "daily")
            text = (b.get("text") or "").strip()
            if not text:
                return self._send(400, {"ok": False, "error": "텍스트를 입력해주세요."})
            user_msg = f"[MODE:{sim_mode}]\n{text}"
            result = _call_claude(SIMULATOR_DRAFT_PROMPT, user_msg)
            if not result:
                return self._send(500, {"ok": False, "error": "AI 드래프트 생성에 실패했습니다."})
            return self._send(200, {"ok": True, "fields": result})
        # 이하 쓰기: user+pin 검증
        uid = b.get("user", ""); pin = b.get("pin", "")
        if not auth(uid, pin): return self._send(401, {"ok": False, "error": "인증 실패"})
        if path == "/api/assign-parse":
            # 자유 텍스트 → AI to-do 항목화 (담당자 미지정 — 리더가 이후 지정)
            if not (is_admin(uid) or is_leader(uid)):
                return self._send(403, {"ok": False, "error": "분장 권한 없음"})
            items = _llm_todo(b.get("text") or "")
            return self._send(200, {"ok": True, "items": items})
        if path == "/api/assign-from-plan":
            # 주간업무계획.xlsx(base64) → 파싱 → AI 항목화 → 분장 검토 초안 (대표·리더 전용)
            if not (is_admin(uid) or is_leader(uid)):
                return self._send(403, {"ok": False, "error": "분장 권한 없음"})
            import base64, tempfile
            try:
                raw = base64.b64decode((b.get("b64") or "").split(",")[-1])
            except Exception:
                return self._send(400, {"ok": False, "error": "파일 디코드 실패"})
            if not raw or len(raw) > 10 * 1024 * 1024:
                return self._send(400, {"ok": False, "error": "빈 파일 또는 10MB 초과"})
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
                tf.write(raw)
                tmp = tf.name
            try:
                rows = _parse_weekly_plan(tmp)
            finally:
                try: os.unlink(tmp)
                except OSError: pass
            if not rows:
                return self._send(200, {"ok": False, "error": "표를 읽지 못했습니다 — '프로젝트명/금주/차주' 헤더가 있는 주간업무계획 형식인지 확인하세요"})
            lines = []
            for r in rows:
                seg = f"[{r['project']}]" if r.get("project") else "[프로젝트 미지정]"
                if r.get("status"): seg += f" (상황: {r['status']})"
                if r.get("this"): seg += f" 금주: {r['this']}"
                if r.get("next"): seg += f" / 차주: {r['next']}"
                if r.get("due"): seg += f" / 예상 런칭: {r['due']}"
                lines.append(seg)
            items = _llm_todo("\n".join(lines), max_items=30)
            return self._send(200, {"ok": True, "items": items, "rows": len(rows)})
        if path == "/api/brief-comment":
            # 브리프 항목 코멘트 — 대표·리더. 저장(시스템 피드백) + 보고자 텔레그램 회신
            if not (is_admin(uid) or is_leader(uid)):
                return self._send(403, {"ok": False, "error": "코멘트 권한 없음(대표·리더)"})
            ds = (b.get("date") or "").strip()
            item = (b.get("item") or "").strip()
            src = (b.get("src") or "").strip()
            text = (b.get("text") or "").strip()
            if not (re.fullmatch(r"\d{4}-\d{2}-\d{2}", ds) and item and text):
                return self._send(400, {"ok": False, "error": "date·item·text 필수"})
            if len(text) > 1000:
                return self._send(400, {"ok": False, "error": "코멘트는 1000자 이내"})
            tg = False
            if src and src != uid:
                cid = _tg_chat_id(src)
                if cid:
                    detail = _brief_item_source(ds, item, src)
                    quoted = f"▸ {item}" + (f"\n{detail}" if detail else "")
                    tg = _tg_send(cid, f"💬 {uid} 님의 코멘트 ({ds} 브리프)\n\n"
                                       f"📌 보고하신 내용\n{quoted}\n\n"
                                       f"💬 코멘트\n{text}\n\n"
                                       f"— 아리사 OS 브리프에서 확인할 수 있어요")
            cmt = {"date": ds, "item": item, "src": src, "author": uid, "text": text,
                   "ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), "tg_sent": tg}
            arr = _load_comments(ds)
            arr.append(cmt)
            _save_comments(ds, arr)
            return self._send(200, {"ok": True, "comment": cmt, "tg_sent": tg})
        if path == "/api/assign-project-check":
            # 분장 그룹 프로젝트명 → 기존 매칭/신규 판별 + 유사 후보 (대표·리더 전용)
            if not (is_admin(uid) or is_leader(uid)):
                return self._send(403, {"ok": False, "error": "분장 권한 없음"})
            projs = load_projects()
            res = []
            for name in (b.get("names") or []):
                name = (name or "").strip()
                if not name: continue
                hit = next((p for p in projs if _match_project_p(name, p)), None)
                if hit:
                    res.append({"name": name, "matched": {"id": hit.get("id"), "name": hit.get("name"),
                                                          "archived": bool(hit.get("archived"))}})
                else:
                    # P2 — 신규명은 네이밍 규칙 자동 정리 결과·오류를 함께 반환 (UI 안내용)
                    cleaned = _NM.clean_project_name(name)
                    ok_n, err_n = _NM.validate_project_name(cleaned)
                    res.append({"name": name, "isNew": True, "candidates": _similar_projects(name, projs),
                                "cleaned": cleaned if cleaned != name else "",
                                "nameError": "" if ok_n else err_n})
            return self._send(200, {"ok": True, "results": res,
                                    "all": [{"id": p.get("id"), "name": p.get("name") or p.get("id")} for p in projs]})
        if path == "/api/assign-commit":
            # 편집·담당자 지정 완료된 항목 일괄 등록
            # + 프로젝트 연동: projectActions(create=포트폴리오 카드 자동 생성 / merge=기존 정식 명칭으로 치환)
            #   등록 성공 항목은 대상 프로젝트 tasks에 영구 반영(akey 중복 방지)
            items = b.get("items") or []
            actions = {}
            for a in (b.get("projectActions") or []):
                n = (a.get("name") or "").strip()
                if n: actions[n] = a
            projs = load_projects()
            users_known = set(load_users().keys())
            created, merged, proj_errs = [], [], []
            name_map = {}   # 그룹명 → 정식 프로젝트명
            target = {}     # 정식 프로젝트명 → 프로젝트 dict
            for n, a in actions.items():
                if a.get("action") == "merge":
                    tp = get_project(a.get("mergeId") or "")
                    if not tp:
                        proj_errs.append(f"{n}: 합칠 프로젝트를 찾지 못함"); continue
                    name_map[n] = tp.get("name") or n
                    target[name_map[n]] = tp
                    merged.append(tp.get("name"))
                elif a.get("action") == "create":
                    if not (is_admin(uid) or is_leader(uid)):
                        proj_errs.append(f"{n}: 생성은 대표·팀 리더만"); continue
                    # P2 — 네이밍 규칙: 자동 정리 후 검증 (정리된 이름이 공식 명칭이 됨)
                    nc = _NM.clean_project_name(n)
                    ok_n, err_n = _NM.validate_project_name(nc)
                    if not ok_n:
                        proj_errs.append(f"{n}: 프로젝트명 규칙 위반 — {err_n}"); continue
                    pm = (a.get("pm") or uid).strip()
                    if pm not in users_known:
                        proj_errs.append(f"{n}: PM({pm}) 계정 없음"); continue
                    today = datetime.date.today()
                    grp = [it for it in items if (it.get("project") or "").strip() == n]
                    dls = sorted(d for d in ((it.get("deadline") or "").strip() for it in grp) if d)
                    start = (a.get("start") or "").strip() or today.isoformat()
                    end = (a.get("end") or "").strip() or (dls[-1] if dls else "")
                    mem = [pm] + [m for m in sorted({(it.get("assignee") or "").strip() for it in grp})
                                  if m in users_known and m != pm]
                    pid = _safe(nc) + "-" + today.strftime("%Y%m%d")
                    np = get_project(pid)  # 같은 날 재커밋 시 재사용
                    if not np:
                        np = {"id": pid, "name": nc, "desc": [], "pm": pm,
                              "start": start, "end": end, "dday": end, "members": mem,
                              "brief": {"name": nc, "pm": pm, "status": "In Progress",
                                        "start": start, "end": end, "dday": end,
                                        "summary": "업무분장 등록에서 자동 생성"},
                              "tasks": [], "issues": [], "origin": "assign",
                              "aliases": [n] if n != nc else []}
                        save_project(np)
                        created.append(nc)
                    name_map[n] = np.get("name") or n
                    target[name_map[n]] = np
            added, errs, ok_items = 0, list(proj_errs), []
            for it in items:
                assignee = (it.get("assignee") or "").strip()
                task = (it.get("task") or "").strip()
                if not (assignee and task):
                    errs.append(f"'{task[:16] or '항목'}': 담당자·업무 누락"); continue
                if is_admin(uid):
                    if assignee not in (set(load_emp().get("team_leads", {}).values()) | {uid}):
                        errs.append(f"{assignee}: 대표는 리더·본인에게만 배분"); continue
                elif not (emp_team(assignee) in set(lead_teams_of(uid))
                          or assignee in set(load_emp().get("team_leads", {}).values())):
                    # 자기 팀원 외에 타팀 리더 이관 허용 (2026-07-20)
                    errs.append(f"{assignee}: 자기 팀원 또는 타팀 리더(이관)에게만 분장 가능"); continue
                pn = (it.get("project") or "").strip()
                pn = name_map.get(pn, pn)
                ok, msg = _assign_append(assignee, task, (it.get("deadline") or "").strip(),
                                         (it.get("priority") or "일반").strip(), uid,
                                         project=pn)
                if ok:
                    added += 1
                    it2 = dict(it); it2["project"] = pn
                    ok_items.append(it2)
                else: errs.append(msg or "append 실패")
            # tasks 영구 반영 — 생성/합치기 대상 + 기존 매칭 프로젝트
            tasks_synced = 0
            by_proj = {}
            for it in ok_items:
                pn = (it.get("project") or "").strip()
                if pn: by_proj.setdefault(pn, []).append(it)
            for pn, its in by_proj.items():
                tp = target.get(pn) or next((p for p in projs if _match_project_p(pn, p)), None)
                if not tp: continue
                n_add = _append_assign_tasks(tp, its)
                if n_add:
                    save_project(tp); tasks_synced += n_add
            return self._send(200, {"ok": added > 0, "added": added, "errors": errs,
                                    "created": created, "merged": merged, "tasks_synced": tasks_synced})
        if path == "/api/assign":
            # (단건 — 하위호환) 업무분장 입력. 대표=리더·본인, 리더=자기 팀원+타팀 리더(이관)
            assignee = (b.get("assignee") or "").strip()
            task = (b.get("task") or "").strip()
            if not (assignee and task): return self._send(400, {"ok": False, "error": "담당자·업무 필수"})
            if is_admin(uid):
                if assignee not in (set(load_emp().get("team_leads", {}).values()) | {uid}):
                    return self._send(403, {"ok": False, "error": "대표는 리더·본인에게만 배분할 수 있습니다"})
            elif not (emp_team(assignee) in set(lead_teams_of(uid))
                      or assignee in set(load_emp().get("team_leads", {}).values())):
                return self._send(403, {"ok": False, "error": "자기 팀원 또는 타팀 리더(이관)에게만 분장할 수 있습니다"})
            ok, msg = _assign_append(assignee, task, (b.get("deadline") or "").strip(),
                                     (b.get("priority") or "일반").strip(), uid)
            return self._send(200 if ok else 500, {"ok": ok, "error": msg})
        if path == "/api/assign-self":
            # 본인 분장 등록 — 개인 브리프 '새로 해야 할 일' 제안 수락용 (assignee=본인 강제)
            task = (b.get("task") or "").strip()
            if not task:
                return self._send(400, {"ok": False, "error": "업무 내용 필수"})
            ok, msg = _assign_append(uid, task, (b.get("deadline") or "").strip(),
                                     (b.get("priority") or "일반").strip(), uid,
                                     project=(b.get("project") or "").strip())
            return self._send(200 if ok else 500, {"ok": ok, "error": msg})
        if path == "/api/assign-edit":
            # 분장 내용 수정(업무·프로젝트·마감) — 본인 + 대표·해당 팀 리더. 잘못 등록 정정용
            if not (_asgws and DAILY_SHEET):
                return self._send(500, {"ok": False, "error": "시트 미설정"})
            try:
                row = int(b.get("row") or 0)
            except (TypeError, ValueError):
                row = 0
            new_task = (b.get("new_task") or "").strip()
            new_project_raw = b.get("new_project")   # 키 생략 = 변경 안 함 (담당자만 변경·미지정 배정용)
            new_dl_raw = b.get("new_deadline")
            if row < 2 or not new_task:
                return self._send(400, {"ok": False, "error": "row·업무 내용 확인"})
            try:
                cur = _asgws.values_get(DAILY_SHEET, f"{ASSIGN_TAB}!A{row}:K{row}", retries=2, timeout=20)
            except Exception:
                cur = []
            r = (list(cur[0]) + [""] * 11)[:11] if cur else [""] * 11
            assignee = (r[3] or "").strip()
            old_task = (r[4] or "").strip()
            new_project = (r[1] or "").strip() if new_project_raw is None else str(new_project_raw).strip()
            new_dl = (r[5] or "").strip() if new_dl_raw is None else str(new_dl_raw).strip()
            if new_dl != (r[5] or "").strip() and new_dl and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", new_dl):
                return self._send(400, {"ok": False, "error": "마감일 형식(YYYY-MM-DD)"})
            if not old_task or old_task != (b.get("task") or "").strip() or assignee != (b.get("assignee") or "").strip():
                return self._send(409, {"ok": False, "error": "행 내용이 달라졌습니다 — 새로고침 후 다시 시도"})
            # 담당자 변경(filament 반영) — 미지정 큐 배정·리더 인라인 편집 공용
            new_assignee = (b.get("new_assignee") or "").strip()
            if not assignee:
                # 미지정 행 — 배정은 대표·리더만
                if not (is_admin(uid) or is_leader(uid)):
                    return self._send(403, {"ok": False, "error": "미지정 배정 권한 없음(대표·리더)"})
            elif not (assignee == uid or _can_approve(uid, assignee)):
                return self._send(403, {"ok": False, "error": "수정 권한 없음(본인·대표·해당 팀 리더)"})
            if new_assignee and new_assignee != assignee and not load_emp().get("by_name", {}).get(new_assignee):
                return self._send(400, {"ok": False, "error": "담당자가 명부에 없습니다"})
            final_assignee = new_assignee or assignee
            old_project = (r[1] or "").strip()
            try:
                for col, old_v, new_v in (("B", old_project, new_project), ("E", old_task, new_task),
                                          ("F", (r[5] or "").strip(), new_dl)):
                    if new_v != old_v:
                        if not _asgws.values_update(DAILY_SHEET, f"{ASSIGN_TAB}!{col}{row}", [[new_v]], timeout=20):
                            return self._send(500, {"ok": False, "error": f"{col}열 업데이트 실패"})
                if final_assignee != assignee:  # D=담당자, C=팀(명부 기준 재산정)
                    if not _asgws.values_update(DAILY_SHEET, f"{ASSIGN_TAB}!D{row}", [[final_assignee]], timeout=20):
                        return self._send(500, {"ok": False, "error": "D열(담당자) 업데이트 실패"})
                    _asgws.values_update(DAILY_SHEET, f"{ASSIGN_TAB}!C{row}", [[emp_team(final_assignee) or ""]], timeout=20)
                if new_project != old_project:  # G1 — 프로젝트 변경 시 ID Relation(K) 재확정
                    _asgws.values_update(DAILY_SHEET, f"{ASSIGN_TAB}!K{row}", [[_resolve_pid(new_project)]], timeout=20)
            except Exception as e:
                return self._send(500, {"ok": False, "error": str(e)[:80]})
            # 포트폴리오 akey 정합 — 기존 반영분 제거 후 새 값으로 재기록 (상태는 시트 lazy sync가 유지)
            removed = _remove_done_task({"date": (r[0] or "").strip(), "project": old_project,
                                         "assignee": assignee, "task": old_task,
                                         "pid": (r[10] or "").strip()})  # r은 편집 전 읽음 → 구 프로젝트 pid
            pn = new_project or old_project
            if removed and pn:
                tp = _find_project_for_assign({"project": pn, "pid": _resolve_pid(pn)})
                if tp:
                    st = (r[7] or "미착수").strip()
                    tp.setdefault("tasks", []).append({
                        "division": emp_team(final_assignee) or "", "task": new_task, "owner": final_assignee,
                        "start": (r[0] or "").strip()[:10], "end": new_dl,
                        "status": _ASSIGN_ST_MAP.get(st, "Not Started"),
                        "progress": _ASSIGN_PROG_MAP.get(st, 0),
                        "akey": _akey((r[0] or "").strip(), new_task, final_assignee)})
                    save_project(tp)
            return self._send(200, {"ok": True})
        if path == "/api/assign-status":
            # 분장 상태 변경 — 완료(본인 체크)·진행중(진행 표시/반려 복귀)·승인(리더·대표)
            if not (_asgws and DAILY_SHEET):
                return self._send(500, {"ok": False, "error": "시트 미설정"})
            try:
                row = int(b.get("row") or 0)
            except (TypeError, ValueError):
                row = 0
            new_st = (b.get("status") or "").strip()
            if row < 2 or new_st not in _ST.ASSIGN_STATES:
                return self._send(400, {"ok": False, "error": "row·status 확인"})
            # 행 재조회 — 수동 행 삭제 등으로 어긋났으면 거부
            try:
                cur = _asgws.values_get(DAILY_SHEET, f"{ASSIGN_TAB}!A{row}:K{row}", retries=2, timeout=20)
            except Exception:
                cur = []
            r = (list(cur[0]) + [""] * 11)[:11] if cur else [""] * 11
            assignee = (r[3] or "").strip()
            task = (r[4] or "").strip()
            if not task or task != (b.get("task") or "").strip() or assignee != (b.get("assignee") or "").strip():
                return self._send(409, {"ok": False, "error": "행 내용이 달라졌습니다 — 새로고침 후 다시 시도"})
            if new_st == "승인":
                if not _can_approve(uid, assignee):
                    return self._send(403, {"ok": False, "error": "승인 권한 없음(대표·해당 팀 리더)"})
            elif new_st == "삭제":
                # 대표·해당 팀 리더 + 본인(잘못 등록 정리) — 본인 삭제는 아래에서 리더·대표에 알림
                if not (_can_approve(uid, assignee) or assignee == uid):
                    return self._send(403, {"ok": False, "error": "삭제 권한 없음(대표·팀 리더·본인)"})
            elif not (assignee == uid or _can_approve(uid, assignee)):
                return self._send(403, {"ok": False, "error": "본인 분장만 변경할 수 있습니다"})
            try:
                ok = _asgws.values_update(DAILY_SHEET, f"{ASSIGN_TAB}!H{row}", [[new_st]], timeout=20)
            except Exception as e:
                return self._send(500, {"ok": False, "error": str(e)[:80]})
            if not ok:
                return self._send(500, {"ok": False, "error": "시트 업데이트 실패"})
            _log_st("dashboard", uid, new_st, from_status=_ST.norm_assign_status(r[7]), row=row,
                    date=(r[0] or "").strip(), project=(r[1] or "").strip(), pid=(r[10] or "").strip(),
                    task=task, assignee=assignee)  # G5 — 상태 전이 이력
            recorded = False
            notified = 0
            assign = {"date": (r[0] or "").strip(), "project": (r[1] or "").strip(),
                      "assignee": assignee, "task": task, "deadline": (r[5] or "").strip(),
                      "pid": (r[10] or "").strip()}  # G1 — ID Relation 우선 매칭용
            if new_st == "승인":
                recorded = _record_done_task(assign, uid)
            elif new_st == "삭제":
                _remove_done_task(assign)  # 포트폴리오에 반영돼 있었으면 함께 제거
                if not _can_approve(uid, assignee):
                    # 직원 본인 삭제 — 리더·대표에게 투명하게 공유 (시트에는 삭제 표시로 보존)
                    _notify_approvers(assignee, f"🗑 {uid} 본인 분장 삭제\n▸ [{assign['project'] or '기타'}] {task}\n\n"
                                                f"잘못 등록된 분장 정리 — 시트에 삭제 표시로 보존됩니다.")
            elif new_st == "완료" and b.get("notify"):
                notified = _notify_approvers(assignee, f"✅ {assignee} 완료 보고\n▸ [{assign['project'] or '기타'}] {task}\n\n"
                                                       f"아리사 OS 내 업무 '완료 승인 대기'에서 승인해주세요.")
            return self._send(200, {"ok": True, "status": new_st, "portfolio_recorded": recorded,
                                    "notified": notified})
        if path == "/api/project/doc-analyze":
            # 프로젝트 자료(회의록) 업로드 + AI 브리프 갱신 제안 (diff) — PM·대표
            pid = (b.get("id") or "").strip()
            p = get_project(pid)
            if not p:
                return self._send(404, {"ok": False, "error": "프로젝트 없음"})
            if not can_edit(uid, p):
                return self._send(403, {"ok": False, "error": "자료 업데이트 권한 없음(담당 PM·대표만)"})
            text = (b.get("text") or "").strip()
            if len(text) < 20:
                return self._send(400, {"ok": False, "error": "자료 내용이 너무 짧습니다(20자 이상)"})
            truncated = len(text) > 30000
            text = text[:30000]
            now = datetime.datetime.now()
            title = (b.get("title") or "").strip() or f"자료 {now.strftime('%Y-%m-%d')}"
            ts = now.strftime("%Y%m%d-%H%M%S")
            # ① 자료 원문 저장 (분석 실패해도 보존)
            ddir = DOC_DIR / _safe(pid)
            ddir.mkdir(parents=True, exist_ok=True)
            (ddir / f"{ts}.md").write_text(f"# {title}\n({uid} · {now.strftime('%Y-%m-%d %H:%M')})\n\n{text}",
                                           encoding="utf-8")
            p.setdefault("docs", []).append({"ts": ts, "title": title, "by": uid, "chars": len(text)})
            save_project(p)
            # ② LLM 갱신 제안
            brief = p.get("brief") or {}
            cur = {k: (brief.get(k) if k not in ("start", "end", "dday", "status")
                       else (brief.get(k) or p.get(k) or "")) for k in _BRIEF_AI_FIELDS}
            sys_p = ("당신은 프로젝트 브리프 관리자다. 새 자료(회의록·문서)에 명시적 근거가 있는 필드만 갱신을 제안한다. "
                     "자료에 근거가 없으면 절대 제안하지 않는다. 기존 값이 더 구체적이면 유지한다. 날짜는 YYYY-MM-DD. "
                     "status는 " + "/".join(_ST.BRIEF_STATES) + " 중 하나. "
                     "갱신 대상 필드와 의미: " + ", ".join(f"{k}({v})" for k, v in _BRIEF_AI_FIELDS.items()) + ". "
                     "텍스트 필드는 기존 내용에 새 정보를 통합한 완성된 최신 값을 after로 작성(단순 요약 금지, 한국어). "
                     '반드시 JSON만 반환: {"changes":[{"field":"...","after":"...","basis":"자료 속 근거 한 줄"}]}')
            user_p = ("[현재 브리프]\n" + json.dumps(cur, ensure_ascii=False) +
                      "\n\n[새 자료: " + title + "]\n" + text)
            res = _call_llm_json(sys_p, user_p, max_tokens=3000)
            if res is None:
                return self._send(200, {"ok": True, "doc": {"ts": ts, "title": title}, "changes": [],
                                        "error": "AI 분석 실패 — 자료는 저장되었습니다. 잠시 후 다시 시도하세요."})
            changes = []
            for c in (res.get("changes") or []):
                f = (c.get("field") or "").strip()
                if f not in _BRIEF_AI_FIELDS:
                    continue
                after = str(c.get("after") or "").strip()
                before = str(cur.get(f) or "").strip()
                if not after or after == before:
                    continue
                changes.append({"field": f, "label": _BRIEF_AI_FIELDS[f], "before": before,
                                "after": after, "basis": str(c.get("basis") or "")[:200]})
            return self._send(200, {"ok": True, "doc": {"ts": ts, "title": title},
                                    "changes": changes, "truncated": truncated})
        if path == "/api/project/doc-apply":
            # diff 미리보기에서 선택한 변경 적용 + 변경 로그 기록 — PM·대표
            pid = (b.get("id") or "").strip()
            p = get_project(pid)
            if not p:
                return self._send(404, {"ok": False, "error": "프로젝트 없음"})
            if not can_edit(uid, p):
                return self._send(403, {"ok": False, "error": "적용 권한 없음(담당 PM·대표만)"})
            brief = p.setdefault("brief", {})
            entry = {"ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), "by": uid,
                     "doc": (b.get("docTitle") or "").strip(), "changes": []}
            for c in (b.get("changes") or []):
                f = (c.get("field") or "").strip()
                if f not in _BRIEF_AI_FIELDS:
                    continue
                after = str(c.get("after") or "").strip()
                before = str(brief.get(f) or p.get(f) or "").strip()
                if not after or after == before:
                    continue
                entry["changes"].append({"field": f, "label": _BRIEF_AI_FIELDS[f],
                                         "before": before, "after": after,
                                         "basis": str(c.get("basis") or "")[:200]})
                brief[f] = after
                if f in ("start", "end", "dday", "status"):
                    p[f] = after   # 카드·간트가 쓰는 top-level 동기화
            if not entry["changes"]:
                return self._send(200, {"ok": True, "applied": 0, "project": p})
            p.setdefault("changelog", []).append(entry)
            p["changelog"] = p["changelog"][-100:]  # 최근 100건 유지
            save_project(p)
            return self._send(200, {"ok": True, "applied": len(entry["changes"]), "project": p})
        if path == "/api/assign-bulk-delete":
            # 체크리스트 일괄 삭제 — 스냅샷 1회 검증 후 행별 '삭제' 마킹 (+포트폴리오 제거)
            items = b.get("items") or []
            if not items or len(items) > 200:
                return self._send(400, {"ok": False, "error": "items 확인(1~200건)"})
            if not (_asgws and DAILY_SHEET):
                return self._send(500, {"ok": False, "error": "시트 미설정"})
            snap = {a["row"]: a for a in _assign_read()}
            deleted, errs = 0, []
            for it in items:
                try:
                    row = int(it.get("row") or 0)
                except (TypeError, ValueError):
                    row = 0
                a = snap.get(row)
                label = ((it.get("task") or "")[:16]) or f"행{row}"
                if not a or a.get("task") != (it.get("task") or "").strip() \
                        or a.get("assignee") != (it.get("assignee") or "").strip():
                    errs.append(f"{label}: 행 불일치 — 새로고침 필요"); continue
                if not (_can_approve(uid, a.get("assignee")) or a.get("assignee") == uid):
                    errs.append(f"{label}: 권한 없음"); continue
                try:
                    ok = _asgws.values_update(DAILY_SHEET, f"{ASSIGN_TAB}!H{row}", [["삭제"]], timeout=20)
                except Exception:
                    ok = False
                if not ok:
                    errs.append(f"{label}: 시트 업데이트 실패"); continue
                _log_st("dashboard-bulk-delete", uid, "삭제", from_status=a.get("status") or "", row=row,
                        date=a.get("date") or "", project=a.get("project") or "", pid=a.get("pid") or "",
                        task=a.get("task") or "", assignee=a.get("assignee") or "")  # G5
                _remove_done_task(a)
                deleted += 1
            return self._send(200, {"ok": deleted > 0 or not errs, "deleted": deleted, "errors": errs})
        if path == "/api/project/save":
            p = b.get("project") or {}
            if not p.get("id"): return self._send(400, {"ok": False, "error": "id 없음"})
            existing = get_project(p["id"])
            if existing and not can_edit(uid, existing):
                return self._send(403, {"ok": False, "error": "수정 권한 없음(담당 PM·대표만)"})
            if not existing and not (is_admin(uid) or is_leader(uid)):
                return self._send(403, {"ok": False, "error": "생성은 대표·팀 리더만"})
            if not existing:
                # P2 — 네이밍 규칙(Team Ops Guide 2부-①): 자동 정리 + 검증 + 중복(유사) 확인
                cleaned = _NM.clean_project_name(p.get("name"))
                ok_n, err_n = _NM.validate_project_name(cleaned)
                if not ok_n:
                    return self._send(400, {"ok": False, "error": f"프로젝트명 규칙 위반 — {err_n}"})
                p["name"] = cleaned
                if p.get("brief"):
                    p["brief"]["name"] = cleaned
                if not b.get("force"):
                    dup = _find_dup_project(cleaned, load_projects())
                    if dup:
                        return self._send(200, {"ok": False, "dup": {"id": dup.get("id"),
                                                "name": dup.get("name") or dup.get("id")},
                                                "cleaned": cleaned})
            save_project(p)
            return self._send(200, {"ok": True, "name": p.get("name")})
        if path == "/api/project/members":
            # 멤버(열람 직원) 배정 전용 — 내용은 안 건드림. 대표·PM·해당 팀 리더만.
            pid = b.get("id", ""); members = b.get("members")
            p = get_project(pid)
            if not p: return self._send(404, {"ok": False, "error": "not found"})
            if not can_manage(uid, p):
                return self._send(403, {"ok": False, "error": "멤버 배정 권한 없음(대표·PM·해당 팀 리더만)"})
            if not isinstance(members, list):
                return self._send(400, {"ok": False, "error": "members 배열이 필요합니다"})
            known = set(load_users().keys())
            clean = [m for m in members if m in known]
            if p.get("pm") and p["pm"] not in clean: clean.append(p["pm"])  # PM은 항상 유지
            p["members"] = clean
            save_project(p)
            return self._send(200, {"ok": True, "members": clean})
        if path == "/api/project/archive":
            # P1 — 프로젝트 아카이브 라이프사이클 (Team Ops Guide 2부-④).
            # 완료 조건 3종(납품·정산·회고 2줄) 충족 시에만 아카이브. 삭제가 아니라 이동:
            # 파일은 그대로 두고 archived 메타를 얹어 활성 목록·매칭에서 분리한다. 복원은 대표만.
            pid = b.get("id", "")
            p = get_project(pid)
            if not p:
                return self._send(404, {"ok": False, "error": "not found"})
            if not (is_admin(uid) or (p.get("pm") or "") == uid):
                return self._send(403, {"ok": False, "error": "아카이브는 대표·담당 PM만"})
            if b.get("restore"):
                if not is_admin(uid):
                    return self._send(403, {"ok": False, "error": "복원은 대표만"})
                p.pop("archived", None)
                save_project(p)
                _archive_log(pid, p.get("name"), uid, "restore")
                return self._send(200, {"ok": True, "restored": True})
            if p.get("archived"):
                return self._send(400, {"ok": False, "error": "이미 아카이브된 프로젝트"})
            # 열린 분장 관문 (KBO 사례 재발 방지) — 있으면 명시 확인(force_open) 요구
            if not b.get("force_open"):
                oa = _open_assigns(p)
                if oa:
                    return self._send(200, {"ok": False, "openAssigns": len(oa),
                                            "openList": oa[:5],
                                            "error": f"열린 분장 {len(oa)}건 — 완료·승인 처리 후 아카이브하거나, 그래도 진행하려면 확인이 필요합니다"})
            retro_good = (b.get("retro_good") or "").strip()
            retro_bad = (b.get("retro_bad") or "").strip()
            missing = []
            if not b.get("delivery"): missing.append("산출물 납품")
            if not b.get("settlement"): missing.append("정산")
            if not retro_good: missing.append("회고(잘된 것)")
            if not retro_bad: missing.append("회고(아쉬운 것)")
            if missing:
                return self._send(400, {"ok": False,
                                        "error": "완료 조건 미충족: " + ", ".join(missing)})
            p["archived"] = {"date": datetime.date.today().isoformat(), "by": uid,
                             "delivery": True, "settlement": True,
                             "retro": {"good": retro_good, "bad": retro_bad}}
            p.setdefault("brief", {})["status"] = "Done"
            save_project(p)
            _archive_log(pid, p.get("name"), uid, "archive",
                         retro={"good": retro_good, "bad": retro_bad})
            return self._send(200, {"ok": True, "archived": p["archived"]})
        if path == "/api/project/delete":
            # 프로젝트 통째 삭제 — 카드(JSON) + 이 프로젝트로 등록된 분장(미착수/진행중/완료)도 '삭제' 마킹.
            # 승인된 분장(완료 이력)은 보존.
            pid = b.get("id", "")
            p = get_project(pid)
            if not (is_admin(uid) or (p and (p.get("pm") or "") == uid)):
                return self._send(403, {"ok": False, "error": "삭제는 대표·담당 PM만"})
            assigns_deleted = 0
            if p and _asgws and DAILY_SHEET:
                for a in _assign_read():
                    if (a.get("status") or "") in ("삭제", "승인"):
                        continue
                    ap = (a.get("project") or "").strip()
                    if not ap or not _match_project_p(ap, p):
                        continue
                    try:
                        if _asgws.values_update(DAILY_SHEET, f"{ASSIGN_TAB}!H{a['row']}", [["삭제"]], timeout=20):
                            assigns_deleted += 1
                            _log_st("project-delete", uid, "삭제", from_status=a.get("status") or "",
                                    row=a.get("row"), date=a.get("date") or "", project=ap,
                                    pid=a.get("pid") or "", task=a.get("task") or "",
                                    assignee=a.get("assignee") or "")  # G5
                    except Exception:
                        pass
            f = PROJ_DIR / f"{_safe(pid)}.json"
            if f.exists(): f.unlink()
            return self._send(200, {"ok": True, "assigns_deleted": assigns_deleted})
        return self._send(404, {"ok": False, "error": "not found"})

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1: PORT = int(sys.argv[1])
    PROJ_DIR.mkdir(parents=True, exist_ok=True)
    srv = ThreadingHTTPServer((HOST, PORT), H)
    print(f"▶ 대시보드 서버 http://{HOST}:{PORT}  (데이터: {DATA})")
    print(f"  프로젝트 {len(load_projects())}개 · 사용자 {len(load_users())}명")
    try: srv.serve_forever()
    except KeyboardInterrupt: srv.shutdown()
