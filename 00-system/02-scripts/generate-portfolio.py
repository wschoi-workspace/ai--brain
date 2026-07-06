#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""포트폴리오 대시보드 생성기.
project-master-prototype.html 에 window.SEED_PROJECTS / window.USERS / window.EMPLOYEES 를 주입해
로그인 → 프로젝트 목록 → 상세(현 대시보드) 3-뷰로 동작하는 standalone HTML을 만든다.

시드 프로젝트:
  - KBO   : KBO_팝업_변환.xlsx 파싱
  - 봉은사 : loadSample() 리터럴(아래 BONGEUNSA) 사용
새 프로젝트는 브라우저에서 직접 추가(localStorage)하므로 여기엔 시드만 넣는다.
출력: 00-system/01-templates/포트폴리오_대시보드.html
"""
import json, datetime as dt, re, base64
import openpyxl
from pathlib import Path

# 이관 가능: 스크립트 위치 기준 상대경로 (하드코딩 금지)
HERE = Path(__file__).resolve().parent           # .../00-system/02-scripts
BASE = HERE.parent / "01-templates"
SCR  = HERE
PROTO = BASE/"project-master-prototype.html"
OUT   = BASE/"포트폴리오_대시보드.html"
KBO_XLSX = BASE/"KBO_팝업_변환.xlsx"
ARANGJE_XLSX = BASE/"아랑재_변환.xlsx"
USERS_JSON = SCR/"portfolio-users.json"
EMP_JSON = SCR/"arisa-employees.json"

def ds(v):
    if isinstance(v,(dt.datetime,dt.date)):
        d=v.date() if isinstance(v,dt.datetime) else v
        return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
    return str(v).strip() if v is not None else ""

LMAP={"프로젝트명":"name","프로젝트 코드":"code","클라이언트":"client","pm":"pm","상태":"status",
      "프로젝트 유형":"type","위치":"location","시작일":"start","종료일":"end","오픈일":"dday",
      "운영기간":"opPeriod","프로젝트 개요":"summary","프로젝트 목표":"goal",
      "critical check points":"critical","risk & notice":"risk"}

def parse_xlsx_project(path):
    wb=openpyxl.load_workbook(path,data_only=True)
    brief,tasks,issues={},[],[]
    def header_row(ws):
        for r in range(1,min(ws.max_row,8)+1):
            row=[str(ws.cell(r,c).value or "").lower() for c in range(1,ws.max_column+1)]
            if any("task" in x or "업무" in x for x in row): return r
        return 2
    def col(ws,hr,pat):
        for c in range(1,ws.max_column+1):
            if re.search(pat,str(ws.cell(hr,c).value or ""),re.I): return c
        return None
    for sn in wb.sheetnames:
        low=sn.lower()
        if "brief" in low:
            ws=wb[sn]
            for r in range(1,ws.max_row+1):
                k=ws.cell(r,1).value; v=ws.cell(r,2).value
                if not k or v in (None,""): continue
                key=LMAP.get(str(k).strip().lower()) or LMAP.get(str(k).strip())
                if key: brief[key]= ds(v) if key in("start","end","dday") else (ds(v) if isinstance(v,(dt.date,dt.datetime)) else v)
            continue
        if "config" in low or "cover" in low or "커버" in low or "overview" in low: continue
        if any(x in sn for x in ("리스크","확정","issue","이슈")):
            ws=wb[sn]; hr=1
            for r in range(1,5):
                row=[str(ws.cell(r,c).value or "") for c in range(1,ws.max_column+1)]
                if any("issue" in x.lower() or "항목" in x for x in row): hr=r; break
            ic=col(ws,hr,r"issue|항목|확정"); dc=col(ws,hr,r"division|구분|팀"); sc=col(ws,hr,r"status|상태")
            for r in range(hr+1,ws.max_row+1):
                it=ws.cell(r,ic).value if ic else None
                if not it: continue
                st=str(ws.cell(r,sc).value or "").strip() if sc else ""
                st="Resolved" if ("해결" in st or "확정" in st) else ("In Progress" if "진행" in st else "Open")
                issues.append({"issue":str(it).strip(),"division":str(ws.cell(r,dc).value or "").strip() if dc else "","status":st})
            continue
        ws=wb[sn]; hr=header_row(ws)
        div=re.sub(r"^\d+[_\s.\-]*","",sn).strip()
        cc={k:col(ws,hr,p) for k,p in {"cat":r"category|카테고리","task":r"task|업무","owner":r"owner|담당",
            "start":r"start|시작","end":r"end|종료","status":r"status|상태","prog":r"progress|진행|%","pri":r"priority|우선"}.items()}
        for r in range(hr+1,ws.max_row+1):
            tk=ws.cell(r,cc["task"]).value if cc["task"] else None
            if not tk or not str(tk).strip(): continue
            prog=ws.cell(r,cc["prog"]).value if cc["prog"] else None
            tasks.append({
                "division":div,
                "category":str(ws.cell(r,cc["cat"]).value or "").strip() if cc["cat"] else "",
                "task":str(tk).strip(),
                "owner":str(ws.cell(r,cc["owner"]).value or "").strip() if cc["owner"] else "",
                "start":ds(ws.cell(r,cc["start"]).value) if cc["start"] else "",
                "end":ds(ws.cell(r,cc["end"]).value) if cc["end"] else "",
                "status":(str(ws.cell(r,cc["status"]).value or "").strip() or "Not Started") if cc["status"] else "Not Started",
                "priority":str(ws.cell(r,cc["pri"]).value or "").strip() if cc["pri"] else "",
                "progress":(int(prog) if isinstance(prog,(int,float)) else "") if prog not in (None,"") else "",
            })
    return {"brief":brief,"tasks":tasks,"issues":issues}

# ── 봉은사 (loadSample 리터럴) ──
BONGEUNSA={
 "brief":{"name":"봉은사 구 웨딩홀 리뉴얼","code":"PR-2026-15","client":"봉은사","pm":"최원석","status":"In Progress",
   "start":"2026-04-01","end":"2026-12-20","dday":"2026-12-20","type":"공간 리뉴얼 / 직영 운영","location":"서울 강남구 봉은사로 531",
   "opPeriod":"2026.12 오픈 ~ 상시","budget":185000,"actual":64000,
   "summary":"구 웨딩홀(8,868㎡)을 직영 정신문화 거점으로 전환. 현대적 깊이를 담은 라이프스타일 공간.",
   "goal":"불교의 빈자리 = '현대적 깊이'를 점유하는 직영 라이프스타일 거점 구축",
   "req":"직영 운영 원칙 · 정신문화 콘텐츠 중심 · 종교시설 규제 준수",
   "kpi":"월 방문객 12,000 / 콘텐츠 프로그램 월 8회 / 멤버십 1,500",
   "critical":"인허가 일정 · 시공 우기 회피 · 콘텐츠-공간 동시 오픈","risk":"종교시설 용도 규제, 시공 인허가 지연 리스크"},
 "tasks":[
   {"division":"Planning","task":"브랜드 컨셉 디벨롭","owner":"최원석","start":"2026-04-01","end":"2026-05-10","status":"Done","progress":100},
   {"division":"Planning","task":"콘텐츠 기획(정신문화 프로그램)","owner":"기획팀","start":"2026-05-01","end":"2026-07-15","status":"In Progress","progress":60},
   {"division":"Planning","task":"고객 경험 시나리오 설계","owner":"기획팀","start":"2026-06-15","end":"2026-08-30","status":"Not Started","progress":0},
   {"division":"Operation","task":"직영 운영조직 설계","owner":"운영팀","start":"2026-05-20","end":"2026-07-30","status":"In Progress","progress":35},
   {"division":"Operation","task":"운영 매뉴얼/교육","owner":"운영팀","start":"2026-09-01","end":"2026-11-30","status":"Not Started","progress":0},
   {"division":"Space","task":"실측 · 기존도면 검토","owner":"공간팀","start":"2026-04-05","end":"2026-05-05","status":"Done","progress":100},
   {"division":"Space","task":"기본설계 · 인허가 도면","owner":"공간팀","start":"2026-05-10","end":"2026-06-15","status":"In Progress","progress":75},
   {"division":"Space","task":"실시설계 · 시공","owner":"시공사","start":"2026-07-01","end":"2026-11-20","status":"Not Started","progress":0},
   {"division":"Space","task":"사이니지 · 설치","owner":"공간팀","start":"2026-11-01","end":"2026-12-10","status":"Not Started","progress":0},
   {"division":"Business","task":"견적 · 예산 확정","owner":"최원석","start":"2026-04-10","end":"2026-05-30","status":"In Progress","progress":80},
   {"division":"Business","task":"시공/콘텐츠 계약","owner":"최원석","start":"2026-06-01","end":"2026-07-10","status":"In Progress","progress":40},
   {"division":"Business","task":"클라이언트(봉은사) 커뮤니케이션","owner":"최원석","start":"2026-04-01","end":"2026-12-20","status":"In Progress","progress":45}],
 "issues":[
   {"issue":"종교시설 용도 인허가 검토 회신 지연","division":"Space","status":"Open"},
   {"issue":"콘텐츠 프로그램 운영 인력 채용 미정","division":"Operation","status":"In Progress"}]
}

# ── 시드 프로젝트 정의 (메타 + 데이터) ──
SOURCES=[
 {"id":"kbo-2026-asg","pm":"최원석","name":"CJ온스타일 × KBO 올스타전",
  "desc":["잠실 올스타전 연계 브랜드 팝업 (행사 7/10~7/11).","굿즈 판매·SNS 확산 극대화 — 5개 팀 74개 업무."],
  "data":parse_xlsx_project(KBO_XLSX)},
 {"id":"bongeunsa-2026-15","pm":"최원석","name":"봉은사 구 웨딩홀 리뉴얼",
  "desc":["구 웨딩홀(8,868㎡) → 직영 정신문화 거점 전환.","현대적 깊이를 담은 라이프스타일 공간 구축."],
  "data":BONGEUNSA},
 {"id":"arangje-2026","pm":"윤혜정","name":"아랑재 × Basket Club",
  "desc":["고덕 아르테온 1F 아랑재 Residential Club Lounge.","스페셜티 커피·시즌 셀렉션 — 4파트 30개 업무, 8월 오픈."],
  "data":parse_xlsx_project(ARANGJE_XLSX)},
]

users_cfg=json.loads(USERS_JSON.read_text(encoding="utf-8"))
USERS=users_cfg.get("users",{})
PMEM=users_cfg.get("project_members",{})
EMP=json.loads(EMP_JSON.read_text(encoding="utf-8")).get("by_name",{})

SEED=[]
for s in SOURCES:
    d=s["data"]; b=d["brief"]
    if not b.get("pm"): b["pm"]=s["pm"]
    SEED.append({
        "id":s["id"], "name":s.get("name") or b.get("name",""), "desc":s["desc"], "pm":b.get("pm",s["pm"]),
        "start":b.get("start",""), "end":b.get("end",""), "dday":b.get("dday",""),
        "members":PMEM.get(s["id"],[]),
        "brief":b, "tasks":d["tasks"], "issues":d["issues"],
    })

TEMPLATE_XLSX = BASE/"project-brief-template.xlsx"
template_b64 = base64.b64encode(TEMPLATE_XLSX.read_bytes()).decode() if TEMPLATE_XLSX.exists() else ""

html=PROTO.read_text(encoding="utf-8")
inject=("<script>"
        "window.SEED_PROJECTS="+json.dumps(SEED,ensure_ascii=False)+";"
        "window.USERS="+json.dumps(USERS,ensure_ascii=False)+";"
        "window.EMPLOYEES="+json.dumps(EMP,ensure_ascii=False)+";"
        "window.TEMPLATE_B64="+json.dumps(template_b64)+";"
        "</script>\n</head>")
html=html.replace("</head>",inject,1)
html=html.replace("<title>Project Master — 검증용 프로토타입 v0.1</title>",
                  "<title>프로젝트 포트폴리오 — Project Rent</title>",1)
OUT.write_text(html,encoding="utf-8")

# ── 서버용 시드 데이터(_data) — 기존 프로젝트 파일은 덮어쓰지 않음(PM 수정 보존) ──
DATA=BASE/"_data"; (DATA/"projects").mkdir(parents=True,exist_ok=True)
# users.json은 최초 1회만 생성 — 운영 중 PIN 자가설정·계정 단일화(symlink→arisa2) 보존
_UP=DATA/"users.json"
if _UP.exists() or _UP.is_symlink():
    print("서버 users.json 존재 — 보존(덮어쓰지 않음)")
else:
    _UP.write_text(json.dumps({"users":USERS,"_employees":list(EMP.keys())},ensure_ascii=False,indent=2),encoding="utf-8")
seeded=0
for p in SEED:
    f=DATA/"projects"/(re.sub(r"[^A-Za-z0-9가-힣_\-]","_",p["id"])[:80]+".json")
    if not f.exists(): f.write_text(json.dumps(p,ensure_ascii=False,indent=2),encoding="utf-8"); seeded+=1
print("서버 시드: users.json +", seeded, "개 프로젝트 신규(_data/projects) · 기존은 보존")
print("저장:",OUT)
print("시드 프로젝트:",[(p["id"],len(p["tasks"]),len(p["issues"]),len(p["members"])) for p in SEED])
print("템플릿 내장:",len(template_b64),"bytes(b64)" if template_b64 else "없음")
print("사용자:",list(USERS.keys()))
print("직원(EMPLOYEES):",len(EMP),"명")
