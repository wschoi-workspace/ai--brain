#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Project Master — 기획 템플릿 생성기 (마스터 입력형)
산출물: 00-system/01-templates/project-brief-template.xlsx

구조
  00_Cover   : ★ 마스터 업무표 — PM이 모든 업무를 '파트' 컬럼과 함께 한 곳에 입력(기획).
               대시보드에 올리면 파트별 자동 분배 + 팀/담당자별 to-do 다운로드 가능.
  00_Brief   : 프로젝트 기본정보(라벨/값) — 대시보드 파서 호환.
  Config     : 드롭다운 값 정의.

흐름: PM이 00_Cover에 전체 업무 입력 → 대시보드 '엑셀 불러오기/프로젝트 추가' →
      '업무분장 분배'로 팀별·담당자별 to-do 엑셀 생성.
"""
import datetime as dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ACCENT="6C5CE7"; ACCENT_D="5A4BD1"; HEAD_FG="FFFFFF"; REQ_FILL="EDE9FE"; LINE="D9D5E8"; GREY="7A7570"
thin=Side(style="thin",color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def style_header(ws,row,ncols,fill=ACCENT):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name="Pretendard",bold=True,color=HEAD_FG,size=10)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border

STATUS=["Not Started","In Progress","Done","Hold"]
PROJ_STATUS=["Not Started","In Progress","On Track","At Risk","Done","Hold"]
PRIORITY=["High","Mid","Low"]
PARTS=["Planning","Operation","Space","Business"]

# 마스터 표 컬럼 (★=필수)  — 대시보드가 읽는 키와 호환(파트=division, 업무=task, 담당자=owner...)
MASTER_COLS=[("No.",6,False),("파트",14,True),("업무(Task) ★",34,True),("담당자 ★",12,True),
             ("시작 ★",13,True),("종료 ★",13,True),("상태 ★",13,True),("진행률(%)",11,False),
             ("우선",9,False),("카테고리",16,False),("비고",24,False)]

# 봉은사 샘플(파트 분산) — 작성법 예시
SAMPLE=[
 ("Planning","브랜드 컨셉 디벨롭","최원석",dt.date(2026,4,1),dt.date(2026,5,10),"Done",100,"High","Concept",""),
 ("Planning","콘텐츠 기획(정신문화 프로그램)","기획팀",dt.date(2026,5,1),dt.date(2026,7,15),"In Progress",60,"High","Content",""),
 ("Planning","고객 경험 시나리오 설계","기획팀",dt.date(2026,6,15),dt.date(2026,8,30),"Not Started",0,"Mid","CX",""),
 ("Operation","직영 운영조직 설계","운영팀",dt.date(2026,5,20),dt.date(2026,7,30),"In Progress",35,"High","Staff",""),
 ("Operation","운영 매뉴얼/교육","운영팀",dt.date(2026,9,1),dt.date(2026,11,30),"Not Started",0,"Mid","Manual",""),
 ("Space","실측 · 기존도면 검토","공간팀",dt.date(2026,4,5),dt.date(2026,5,5),"Done",100,"High","Drawing",""),
 ("Space","기본설계 · 인허가 도면","공간팀",dt.date(2026,5,10),dt.date(2026,6,15),"In Progress",75,"High","Design",""),
 ("Space","실시설계 · 시공","시공사",dt.date(2026,7,1),dt.date(2026,11,20),"Not Started",0,"High","Construction",""),
 ("Business","견적 · 예산 확정","최원석",dt.date(2026,4,10),dt.date(2026,5,30),"In Progress",80,"High","Estimate",""),
 ("Business","시공/콘텐츠 계약","최원석",dt.date(2026,6,1),dt.date(2026,7,10),"In Progress",40,"High","Contract",""),
]

wb=Workbook()

# ─────────────────────── 00_Cover (마스터 업무표) ───────────────────────
cv=wb.active; cv.title="00_Cover"; cv.sheet_properties.tabColor=ACCENT_D
for i,(h,w,req) in enumerate(MASTER_COLS,start=1):
    cv.column_dimensions[get_column_letter(i)].width=w
NC=len(MASTER_COLS)
cv.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC)
t=cv.cell(1,1,"📋 프로젝트 업무 마스터 — 여기에 모든 업무를 '파트'와 함께 입력하세요 (파트별·담당자별 자동 분배)")
t.font=Font(name="Pretendard",bold=True,color=HEAD_FG,size=12); t.fill=PatternFill("solid",fgColor=ACCENT_D)
t.alignment=Alignment(horizontal="left",vertical="center"); cv.row_dimensions[1].height=28
# 브리프 요약 스트립(00_Brief 링크)
cv.cell(2,1,"프로젝트").font=Font(bold=True,size=9,color=ACCENT)
cv.cell(2,2,"='00_Brief'!B3").font=Font(bold=True,size=10)
cv.cell(2,4,"PM").font=Font(bold=True,size=9,color=ACCENT); cv.cell(2,5,"='00_Brief'!B6")
cv.cell(2,6,"오픈일").font=Font(bold=True,size=9,color=ACCENT); cv.cell(2,7,"='00_Brief'!B13"); cv.cell(2,7).number_format="yyyy-mm-dd"
cv.cell(2,9,"← 기본정보는 00_Brief 시트에서").font=Font(italic=True,size=9,color=GREY)
# 헤더(3행)
for i,(h,w,req) in enumerate(MASTER_COLS,start=1):
    cv.cell(3,i,h)
style_header(cv,3,NC)
# 샘플 + 빈 행
rr=4
for idx,row in enumerate(SAMPLE,start=1):
    part,task,owner,sd,ed,st,pg,pr,cat,note=row
    vals=[idx,part,task,owner,sd,ed,st,pg,pr,cat,note]
    for ci,v in enumerate(vals,start=1):
        cell=cv.cell(rr,ci,v); cell.font=Font(name="Pretendard",size=10); cell.border=border
        if isinstance(v,dt.date): cell.number_format="yyyy-mm-dd"
        if MASTER_COLS[ci-1][2]: cell.fill=PatternFill("solid",fgColor=REQ_FILL)
    rr+=1
for blank in range(rr,rr+40):
    for ci,(h,w,req) in enumerate(MASTER_COLS,start=1):
        cell=cv.cell(blank,ci); cell.border=border
        if req: cell.fill=PatternFill("solid",fgColor=REQ_FILL)
last=rr+40
# 드롭다운: 파트(B)/상태(G)/우선(I), 진행률(H)
dv_part=DataValidation(type="list",formula1='"%s"'%",".join(PARTS),allow_blank=True,showErrorMessage=False)  # 커스텀 파트 허용
cv.add_data_validation(dv_part); dv_part.add(f"B4:B{last}")
dv_st=DataValidation(type="list",formula1='"%s"'%",".join(STATUS),allow_blank=True)
cv.add_data_validation(dv_st); dv_st.add(f"G4:G{last}")
dv_pr=DataValidation(type="list",formula1='"%s"'%",".join(PRIORITY),allow_blank=True)
cv.add_data_validation(dv_pr); dv_pr.add(f"I4:I{last}")
dv_pg=DataValidation(type="whole",operator="between",formula1=0,formula2=100,allow_blank=True)
cv.add_data_validation(dv_pg); dv_pg.add(f"H4:H{last}")
cv.freeze_panes="A4"

# ─────────────────────── 00_Brief (프로젝트 기본정보) ───────────────────────
bw=wb.create_sheet("00_Brief"); bw.sheet_properties.tabColor=ACCENT
bw.column_dimensions["A"].width=22; bw.column_dimensions["B"].width=52; bw.column_dimensions["C"].width=30
bw.merge_cells("A1:C1")
t=bw["A1"]; t.value="프로젝트 브리프 — PM 작성 (라벨은 수정 금지, B열 값만 입력)"
t.font=Font(name="Pretendard",bold=True,color=HEAD_FG,size=11); t.fill=PatternFill("solid",fgColor=ACCENT_D)
t.alignment=Alignment(horizontal="left",vertical="center"); bw.row_dimensions[1].height=24
BRIEF=[
 ("── 기본정보 ──","",""),
 ("프로젝트명","봉은사 구 웨딩홀 리뉴얼","필수"),
 ("프로젝트 코드","PR-2026-15","필수 · 예: PR-2026-15"),
 ("클라이언트","봉은사","필수"),
 ("PM","최원석","필수"),
 ("상태","In Progress","드롭다운"),
 ("프로젝트 유형","공간 리뉴얼 / 직영 운영",""),
 ("위치","서울 강남구 봉은사로 531",""),
 ("── 일정 ──","",""),
 ("시작일",dt.date(2026,4,1),"필수 · YYYY-MM-DD"),
 ("종료일",dt.date(2026,12,20),"필수 · YYYY-MM-DD"),
 ("오픈일",dt.date(2026,12,20),"필수(D-Day) · YYYY-MM-DD"),
 ("운영기간","2026.12 오픈 ~ 상시",""),
 ("── 예산(만원) ──","",""),
 ("총 예산",185000,"숫자만(만원)"),
 ("총 집행",64000,"숫자만(만원)"),
 ("── 내용 ──","",""),
 ("프로젝트 개요","구 웨딩홀(8,868㎡)을 직영 정신문화 거점으로 전환. 현대적 깊이를 담은 라이프스타일 공간.",""),
 ("프로젝트 목표","불교의 빈자리 = '현대적 깊이'를 점유하는 직영 라이프스타일 거점 구축",""),
 ("클라이언트 요구사항","직영 운영 원칙 · 정신문화 콘텐츠 중심 · 종교시설 규제 준수",""),
 ("Success KPI","월 방문객 12,000 / 콘텐츠 프로그램 월 8회 / 멤버십 1,500",""),
 ("Critical Check Points","인허가 일정 · 시공 우기 회피 · 콘텐츠-공간 동시 오픈",""),
 ("Risk & Notice","종교시설 용도 규제, 시공 인허가 지연 리스크",""),
]
r=2; sec_fill=PatternFill("solid",fgColor="EDE9FE")
for label,val,note in BRIEF:
    a=bw.cell(r,1,label); b=bw.cell(r,2,val); c=bw.cell(r,3,note)
    if label.startswith("──"):
        a.font=Font(name="Pretendard",bold=True,color=ACCENT,size=10); a.fill=sec_fill; b.fill=sec_fill; c.fill=sec_fill
    else:
        a.font=Font(name="Pretendard",bold=True,size=10); b.font=Font(name="Pretendard",size=10)
        c.font=Font(name="Pretendard",italic=True,color=GREY,size=9)
        if isinstance(val,dt.date): b.number_format="yyyy-mm-dd"
        a.border=border; b.border=border
    r+=1
dv_ps=DataValidation(type="list",formula1='"%s"'%",".join(PROJ_STATUS),allow_blank=True)
bw.add_data_validation(dv_ps); dv_ps.add("B7")
bw.freeze_panes="A2"

# ─────────────────────────── Config ───────────────────────────
cf=wb.create_sheet("Config"); cf.sheet_properties.tabColor=GREY
cf["A1"]="Config — 드롭다운 값 정의 (참조용)"; cf["A1"].font=Font(name="Pretendard",bold=True,size=10,color=ACCENT)
cf["A3"]="파트";   cf["B3"]=", ".join(PARTS)+"  (커스텀 입력 허용)"
cf["A4"]="Status"; cf["B4"]=", ".join(STATUS)
cf["A5"]="Project Status"; cf["B5"]=", ".join(PROJ_STATUS)
cf["A6"]="Priority"; cf["B6"]=", ".join(PRIORITY)
for r in range(3,7): cf.cell(r,1).font=Font(name="Pretendard",bold=True,size=10)
cf.column_dimensions["A"].width=16; cf.column_dimensions["B"].width=60

out=str(Path(__file__).resolve().parent.parent/"01-templates"/"project-brief-template.xlsx")  # 이관 가능: 상대경로
wb.save(out)
print("saved:",out)
print("sheets:",wb.sheetnames)
