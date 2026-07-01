#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""KBO_팝업_변환.xlsx → 데이터 내장 미리보기 HTML 생성.
project-master-prototype.html 에 window.EMBEDDED_PROJECT(JSON)를 주입해
열자마자 KBO 그룹 간트가 렌더되는 standalone 미리보기를 만든다.
출력: 00-system/01-templates/KBO_팝업_대시보드_미리보기.html
"""
import json, datetime as dt, re
import openpyxl
from pathlib import Path

BASE=Path.home()/"do-better-workspace"/"00-system"/"01-templates"
XLSX=BASE/"KBO_팝업_변환.xlsx"
PROTO=BASE/"project-master-prototype.html"
OUT=BASE/"KBO_팝업_대시보드_미리보기.html"

def ds(v):
    if isinstance(v,(dt.datetime,dt.date)):
        d=v.date() if isinstance(v,dt.datetime) else v
        return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
    return str(v).strip() if v is not None else ""

LMAP={"프로젝트명":"name","프로젝트 코드":"code","클라이언트":"client","pm":"pm","상태":"status",
      "프로젝트 유형":"type","위치":"location","시작일":"start","종료일":"end","오픈일":"dday",
      "운영기간":"opPeriod","프로젝트 개요":"summary","프로젝트 목표":"goal",
      "critical check points":"critical","risk & notice":"risk"}

wb=openpyxl.load_workbook(XLSX,data_only=True)
state={"brief":{},"tasks":[],"issues":[]}

# 00_Brief
bw=wb["00_Brief"]
for r in range(1,bw.max_row+1):
    k=bw.cell(r,1).value; v=bw.cell(r,2).value
    if not k or v in (None,""): continue
    key=LMAP.get(str(k).strip().lower()) or LMAP.get(str(k).strip())
    if key: state["brief"][key]=ds(v) if key in("start","end","dday") else (v if not isinstance(v,(dt.date,dt.datetime)) else ds(v))

def header_row(ws):
    for r in range(1,min(ws.max_row,8)+1):
        row=[str(ws.cell(r,c).value or "").lower() for c in range(1,ws.max_column+1)]
        if any("task" in x or "업무" in x for x in row): return r
    return 2

def col(ws,hr,pat):
    for c in range(1,ws.max_column+1):
        if re.search(pat,str(ws.cell(hr,c).value or ""),re.I): return c
    return None

for sn in wb.sheetnames:
    if sn=="00_Brief" or "config" in sn.lower(): continue
    if any(x in sn for x in ("리스크","확정","issue","이슈")):
        ws=wb[sn]; hr=1
        for r in range(1,5):
            row=[str(ws.cell(r,c).value or "") for c in range(1,ws.max_column+1)]
            if any("issue" in x.lower() or "항목" in x for x in row): hr=r; break
        ic=col(ws,hr,r"issue|항목|확정"); dc=col(ws,hr,r"division|구분|팀"); sc=col(ws,hr,r"status|상태")
        for r in range(hr+1,ws.max_row+1):
            it=ws.cell(r,ic).value if ic else None
            if not it: continue
            st=str(ws.cell(r,sc).value or "").strip() if sc else ""
            st="Resolved" if ("해결" in st or "확정" in st) else ("In Progress" if "진행" in st else "Open")
            state["issues"].append({"issue":str(it).strip(),"division":str(ws.cell(r,dc).value or "").strip() if dc else "","status":st})
        continue
    ws=wb[sn]; hr=header_row(ws)
    div=re.sub(r"^\d+[_\s.\-]*","",sn).strip()
    cc={k:col(ws,hr,p) for k,p in {"cat":r"category|카테고리","task":r"task|업무","owner":r"owner|담당",
        "start":r"start|시작","end":r"end|종료","status":r"status|상태","prog":r"progress|진행|%","pri":r"priority|우선"}.items()}
    for r in range(hr+1,ws.max_row+1):
        tk=ws.cell(r,cc["task"]).value if cc["task"] else None
        if not tk or not str(tk).strip(): continue
        prog=ws.cell(r,cc["prog"]).value if cc["prog"] else None
        state["tasks"].append({
            "division":div,
            "category":str(ws.cell(r,cc["cat"]).value or "").strip() if cc["cat"] else "",
            "task":str(tk).strip(),
            "owner":str(ws.cell(r,cc["owner"]).value or "").strip() if cc["owner"] else "",
            "start":ds(ws.cell(r,cc["start"]).value) if cc["start"] else "",
            "end":ds(ws.cell(r,cc["end"]).value) if cc["end"] else "",
            "status":(str(ws.cell(r,cc["status"]).value or "").strip() or "Not Started") if cc["status"] else "Not Started",
            "priority":str(ws.cell(r,cc["pri"]).value or "").strip() if cc["pri"] else "",
            "progress":(int(prog) if isinstance(prog,(int,float)) else "") if prog not in (None,"") else "",
        })

html=PROTO.read_text(encoding="utf-8")
inject=("<script>window.EMBEDDED_PROJECT="
        +json.dumps(state,ensure_ascii=False)
        +";</script>\n</head>")
html=html.replace("</head>",inject,1)
# 미리보기 표시: 타이틀/버전 라벨만 살짝 교체(선택)
html=html.replace("검증용 프로토타입 v0.1 · 단일 프로젝트","KBO 팝업 미리보기 · 데이터 내장",1)
OUT.write_text(html,encoding="utf-8")
print("저장:",OUT)
print("brief keys:",list(state["brief"].keys()))
print("tasks:",len(state["tasks"]),"issues:",len(state["issues"]))
print("divisions:",sorted(set(t["division"] for t in state["tasks"])))
