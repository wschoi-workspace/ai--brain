#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""KBO 팝업 간트(상세) 엑셀 → project-master 대시보드 표준 포맷으로 변환.
입력: ~/Downloads/KBO_팝업_간트차트_v2_상세.xlsx
출력: 00-system/01-templates/KBO_팝업_변환.xlsx  (00_Brief + 팀별시트 + 확정리스크)
상태/진행률은 원본에 없어 '시작·종료일 vs 오늘(2026-06-23)'로 자동 추정한다.
"""
import datetime as dt, re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC="/Users/choi_ai/Downloads/KBO_팝업_간트차트_v2_상세.xlsx"
OUT="/Users/choi_ai/do-better-workspace/00-system/01-templates/KBO_팝업_변환.xlsx"
TODAY=dt.date(2026,6,23)
YEAR=2026
ACCENT="6C5CE7"; ACCENT_D="5A4BD1"; REQ_FILL="EDE9FE"; LINE="D9D5E8"
thin=Side(style="thin",color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)
PRI={"상":"High","중":"Mid","하":"Low"}
TEAMS=["전팀","사업팀","공간팀","운영팀","기획팀"]

def parse_date(v):
    if v is None or v=="": return None
    if isinstance(v,(dt.datetime,dt.date)):
        d=v.date() if isinstance(v,dt.datetime) else v
        return d.replace(year=YEAR) if d.year<2000 else d
    s=str(v).strip()
    m=re.search(r"(\d{1,2})\s*[/.\-월]\s*(\d{1,2})",s)
    if m: return dt.date(YEAR,int(m.group(1)),int(m.group(2)))
    m=re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})",s)
    if m: return dt.date(int(m.group(1)),int(m.group(2)),int(m.group(3)))
    return None

def status_of(s,e):
    if e and e<TODAY: return "Done"
    if s and s<=TODAY and (not e or e>=TODAY): return "In Progress"
    return "Not Started"

# ── 원본 읽기 ──
wb=openpyxl.load_workbook(SRC,data_only=True)
g=wb["간트차트(상세)"]
# 헤더행 탐색 (업무/시작/종료 포함)
hrow=None
for r in range(1,12):
    row=[str(g.cell(r,c).value or "") for c in range(1,10)]
    if any("업무" in x for x in row) and any("시작" in x for x in row):
        hrow=r; break
hrow=hrow or 4
def colidx(kw):
    for c in range(1,12):
        if kw in str(g.cell(hrow,c).value or ""): return c
    return None
C={k:colidx(v) for k,v in {"team":"팀","task":"업무","owner":"담당","detail":"세부","cat":"구분","pri":"우선","start":"시작","end":"종료"}.items()}

tasks_by_team={t:[] for t in TEAMS}
all_dates=[]
for r in range(hrow+1, g.max_row+1):
    team=str(g.cell(r,C["team"]).value or "").strip()
    task=str(g.cell(r,C["task"]).value or "").strip()
    if team not in TEAMS or not task: continue          # 그룹헤더/빈행 skip
    task=re.sub(r"\s*[●○]\s*$","",task).strip()          # 확정표시 ● 제거
    s=parse_date(g.cell(r,C["start"]).value); e=parse_date(g.cell(r,C["end"]).value)
    if s: all_dates.append(s)
    if e: all_dates.append(e)
    tasks_by_team[team].append({
        "cat":str(g.cell(r,C["cat"]).value or "").strip(),
        "task":task,
        "owner":str(g.cell(r,C["owner"]).value or "").strip(),
        "start":s,"end":e,
        "status":status_of(s,e),
        "pri":PRI.get(str(g.cell(r,C["pri"]).value or "").strip(),""),
        "note":str(g.cell(r,C["detail"]).value or "").strip(),
    })

total=sum(len(v) for v in tasks_by_team.values())
dmin=min(all_dates) if all_dates else TODAY
dmax=max(all_dates) if all_dates else TODAY
title=str(g.cell(1,1).value or "CJ온스타일 × KBO 올스타전").strip()
subtitle=str(g.cell(2,1).value or "").strip()

# 리스크 시트
risks=[]
if "확정필요·리스크" in wb.sheetnames:
    rs=wb["확정필요·리스크"]
    rh=None
    for r in range(1,6):
        row=[str(rs.cell(r,c).value or "") for c in range(1,8)]
        if any("항목" in x for x in row) or any("확정" in x for x in row[1:]): rh=r; break
    rh=rh or 2
    def rcol(kw):
        for c in range(1,8):
            if kw in str(rs.cell(rh,c).value or ""): return c
        return None
    ri={"cat":rcol("구분"),"item":rcol("항목") or 2,"owner":rcol("담당"),"due":rcol("마감") or rcol("기준"),"st":rcol("상태")}
    for r in range(rh+1, rs.max_row+1):
        item=str(rs.cell(r,ri["item"]).value or "").strip()
        if not item: continue
        st=str(rs.cell(r,ri["st"]).value or "").strip() if ri["st"] else ""
        risks.append({
            "issue":item,
            "div":str(rs.cell(r,ri["cat"]).value or "").strip() if ri["cat"] else "",
            "owner":str(rs.cell(r,ri["owner"]).value or "").strip() if ri["owner"] else "",
            "due":str(rs.cell(r,ri["due"]).value or "").strip() if ri["due"] else "",
            "status":"Open" if ("미확정" in st or not st) else ("Resolved" if "확정" in st else "In Progress"),
        })

# ── 출력 워크북 ──
out=Workbook()
def hdr(ws,row,n,fill=ACCENT):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name="Pretendard",bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border

# 00_Brief
bw=out.active; bw.title="00_Brief"; bw.sheet_properties.tabColor=ACCENT
bw.column_dimensions["A"].width=22; bw.column_dimensions["B"].width=64
BRIEF=[
    ("프로젝트명",title),("프로젝트 코드","KBO-2026-ASG"),("클라이언트","CJ온스타일 / KBO"),
    ("PM",""),("상태","In Progress"),("프로젝트 유형","브랜드 팝업 / 행사 운영"),
    ("위치","잠실 제2주차장 (예정)"),
    ("시작일",dmin),("종료일",dmax),("오픈일",dt.date(2026,7,10)),
    ("운영기간",subtitle or "행사 7/10(금)~7/11(토)"),
    ("프로젝트 개요",f"{title} — {subtitle}. 잠실 올스타전 연계 브랜드 팝업 운영 ({dmin:%m/%d}~{dmax:%m/%d}, 총 {total}개 업무 / 5개 팀)."),
    ("프로젝트 목표","올스타전 현장 팝업으로 브랜드 경험·굿즈 판매·SNS 확산 극대화"),
    ("Critical Check Points","바코드·단가표 확보 / 부지·전기 인입 확정 / 반입·철거 동선 / SNS 폴리시"),
    ("Risk & Notice",f"확정필요·리스크 {len(risks)}건 (대부분 미확정) — 별도 시트 참조"),
]
bw.cell(1,1,"프로젝트 브리프 (자동 변환 — KBO 팝업 간트)").font=Font(bold=True,color="FFFFFF",size=11)
bw.cell(1,1).fill=PatternFill("solid",fgColor=ACCENT_D); bw.merge_cells("A1:B1")
for i,(k,v) in enumerate(BRIEF,start=2):
    a=bw.cell(i,1,k); b=bw.cell(i,2,v)
    a.font=Font(bold=True,size=10); b.font=Font(size=10); a.border=border; b.border=border
    if isinstance(v,dt.date): b.number_format="yyyy-mm-dd"
bw.freeze_panes="A2"

# 팀 시트
COLS=[("No.",6),("Category",16),("Task ★",30),("Owner ★",12),("Start ★",13),("End ★",13),
      ("Status ★",13),("Progress(%) ★",12),("Priority",10),("Note",40)]
for ti,team in enumerate([t for t in TEAMS if tasks_by_team[t]],start=1):
    ws=out.create_sheet(f"{ti:02d}_{team}"); ws.sheet_properties.tabColor=ACCENT
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(COLS))
    info=ws.cell(1,1,f"{team} — {len(tasks_by_team[team])}건 · ★필수(대시보드 계산) · 상태·진행률은 날짜기준 자동추정")
    info.font=Font(bold=True,color="FFFFFF",size=10); info.fill=PatternFill("solid",fgColor=ACCENT_D)
    for ci,(h,w) in enumerate(COLS,start=1):
        ws.cell(2,ci,h); ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width=w
    hdr(ws,2,len(COLS))
    rr=3
    for idx,t in enumerate(tasks_by_team[team],start=1):
        vals=[idx,t["cat"],t["task"],t["owner"],t["start"],t["end"],t["status"],"",t["pri"],t["note"]]
        for ci,v in enumerate(vals,start=1):
            cell=ws.cell(rr,ci,v); cell.font=Font(size=10); cell.border=border
            if isinstance(v,dt.date): cell.number_format="yyyy-mm-dd"
            if COLS[ci-1][0].endswith("★"): cell.fill=PatternFill("solid",fgColor=REQ_FILL)
        rr+=1
    ws.freeze_panes="A3"

# 확정·리스크 시트
if risks:
    iw=out.create_sheet("08_확정리스크"); iw.sheet_properties.tabColor="7A7570"
    ICOLS=[("Issue",46),("Division",14),("Owner",12),("Due",18),("Status",12)]
    for ci,(h,w) in enumerate(ICOLS,start=1):
        iw.cell(1,ci,h); iw.column_dimensions[openpyxl.utils.get_column_letter(ci)].width=w
    hdr(iw,1,len(ICOLS))
    for r,it in enumerate(risks,start=2):
        for ci,v in enumerate([it["issue"],it["div"],it["owner"],it["due"],it["status"]],start=1):
            c=iw.cell(r,ci,v); c.font=Font(size=10); c.border=border
    iw.freeze_panes="A2"

out.save(OUT)
print("저장:",OUT)
print("시트:",out.sheetnames)
print("팀별 업무:",{t:len(v) for t,v in tasks_by_team.items()},"총",total)
print("기간:",dmin,"~",dmax,"/ 리스크:",len(risks),"건")
