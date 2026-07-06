#!/usr/bin/env python3
from pathlib import Path
# -*- coding: utf-8 -*-
"""아랑재 × Basket 원본 타임라인(주차 채움색 막대) → 표준 마스터 포맷으로 변환.
입력: ~/Downloads/아랑재_바스켓_프로젝트타임라인.xlsx
출력: 00-system/01-templates/아랑재_변환.xlsx  (00_Brief + 파트별 시트)

원본은 시작/종료 컬럼이 없고, 주차열(W1~W8)의 셀 채움색(█)으로 기간을 표시한다.
→ 각 업무의 '채워진 첫 주 시작일 ~ 마지막 주 종료일'을 start/end로 산출.
상태/진행률은 원본에 없어 날짜기준(오늘 2026-06-23) 자동 추정.
"""
import datetime as dt, re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC=str(Path.home()/"Downloads")+"/아랑재_바스켓_프로젝트타임라인.xlsx"
OUT=str(Path.home()/"do-better-workspace")+"/00-system/01-templates/아랑재_변환.xlsx"
YEAR=2026; TODAY=dt.date(2026,6,23)
ACCENT="6C5CE7"; ACCENT_D="5A4BD1"; REQ_FILL="EDE9FE"; LINE="D9D5E8"
thin=Side(style="thin",color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def status_of(s,e):
    if e and e<TODAY: return "Done"
    if s and s<=TODAY and (not e or e>=TODAY): return "In Progress"
    return "Not Started"

def is_filled(cell):
    f=cell.fill
    if not f or f.patternType!="solid": return False
    rgb=getattr(f.fgColor,"rgb",None)
    if rgb in (None,"00000000","FFFFFFFF"): return False
    return True

wb=openpyxl.load_workbook(SRC)            # 채움색 필요 → data_only=False
ws=wb["파트별 타임라인"]
# 헤더행(파트/업무/담당 + 주차) 탐색
hrow=None
for r in range(1,8):
    row=[str(ws.cell(r,c).value or "") for c in range(1,5)]
    if any("업무" in x for x in row) and any("담당" in x for x in row): hrow=r; break
hrow=hrow or 3
# 주차열 → (시작일,종료일)
weeks={}  # col -> (date,date)
for c in range(4,ws.max_column+1):
    h=str(ws.cell(hrow,c).value or "")
    m=re.search(r"(\d{1,2})/(\d{1,2})\s*~\s*(\d{1,2})/(\d{1,2})",h)
    if m:
        sm,sd,em,ed=map(int,m.groups())
        weeks[c]=(dt.date(YEAR,sm,sd),dt.date(YEAR,em,ed))

# 업무 수집 (파트별)
from collections import OrderedDict
parts=OrderedDict()
for r in range(hrow+1,ws.max_row+1):
    part=ws.cell(r,1).value; task=ws.cell(r,2).value
    if not part or not task or "■" in str(part): continue
    part=str(part).strip(); task=str(task).strip()
    owner=str(ws.cell(r,3).value or "").strip()
    fw=[weeks[c] for c in weeks if is_filled(ws.cell(r,c))]
    if fw:
        s=min(w[0] for w in fw); e=max(w[1] for w in fw)
    else:
        s=e=None
    parts.setdefault(part,[]).append({"task":task,"owner":owner,"start":s,"end":e,"status":status_of(s,e)})

total=sum(len(v) for v in parts.values())

# 개요 → brief
ov=wb["프로젝트 개요"]
od={}
for r in range(1,ov.max_row+1):
    a=ov.cell(r,1).value; b=ov.cell(r,2).value
    if a and b: od[str(a).strip()]=str(b).strip()
title=str(ov.cell(1,1).value or "아랑재 × Basket Club").strip()

BRIEF=[
 ("프로젝트명","아랑재 × Basket Club"),
 ("프로젝트 코드","PR-2026-ARANGJE"),
 ("클라이언트","프로젝트렌트(주) · 자체운영"),
 ("PM","윤혜정"),
 ("상태","In Progress"),
 ("프로젝트 유형","리테일·F&B 공간 / 운영"),
 ("위치","고덕 아르테온 1F"),
 ("시작일",dt.date(2026,6,23)),
 ("종료일",dt.date(2026,8,17)),
 ("오픈일",dt.date(2026,8,17)),
 ("운영기간",od.get("운영 시간","평일 09:00~20:00 / 주말 09:00~18:00, 연중무휴")),
 ("프로젝트 개요",od.get("프로젝트","고덕아르테온 1F 아랑재 Residential Club Lounge")+" — "+od.get("컨셉","스페셜티 커피 + 시즌 셀렉션 + 커뮤니티 라운지")),
 ("프로젝트 목표","8월 3주차(8/11~8/17) 오픈 — 스페셜티 커피 라운지 구축"),
 ("클라이언트 요구사항",od.get("운영 인력","매니저 1인 + 바리스타 1~2인")+" / 초기투자 "+od.get("초기 투자","4,500~7,000만")),
 ("Critical Check Points","현장 실사 일정 확정 · 메뉴/레시피 정리 · 운영 매뉴얼 초안 (1순위 과제)"),
 ("Risk & Notice","세부 일정 현장 실사 후 확정 · 설비/하드웨어·인허가·채용 일정"),
]

out=Workbook()
def hdr(wss,row,n,fill=ACCENT):
    for c in range(1,n+1):
        cell=wss.cell(row=row,column=c)
        cell.font=Font(name="Pretendard",bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border

bw=out.active; bw.title="00_Brief"; bw.sheet_properties.tabColor=ACCENT
bw.column_dimensions["A"].width=22; bw.column_dimensions["B"].width=66
bw.cell(1,1,"프로젝트 브리프 (자동 변환 — 아랑재 타임라인)").font=Font(bold=True,color="FFFFFF",size=11)
bw.cell(1,1).fill=PatternFill("solid",fgColor=ACCENT_D); bw.merge_cells("A1:B1")
for i,(k,v) in enumerate(BRIEF,start=2):
    a=bw.cell(i,1,k); b=bw.cell(i,2,v)
    a.font=Font(bold=True,size=10); b.font=Font(size=10); a.border=border; b.border=border
    if isinstance(v,dt.date): b.number_format="yyyy-mm-dd"
bw.freeze_panes="A2"

COLS=[("No.",6),("Category",14),("Task ★",34),("Owner ★",12),("Start ★",13),("End ★",13),("Status ★",13),("Progress(%) ★",12),("Note",24)]
def safe_sheet(name): return re.sub(r'[\\/?*\[\]:]',"·",name)[:28]
for ti,(part,rows) in enumerate(parts.items(),start=1):
    wss=out.create_sheet(f"{ti:02d}_{safe_sheet(part)}"); wss.sheet_properties.tabColor=ACCENT
    wss.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(COLS))
    info=wss.cell(1,1,f"{part} — {len(rows)}건 · 상태·진행률은 날짜기준 자동추정 · 기간=주차 채움 범위")
    info.font=Font(bold=True,color="FFFFFF",size=10); info.fill=PatternFill("solid",fgColor=ACCENT_D)
    for ci,(h,w) in enumerate(COLS,start=1):
        wss.cell(2,ci,h); wss.column_dimensions[get_column_letter(ci)].width=w
    hdr(wss,2,len(COLS)); rr=3
    for idx,t in enumerate(rows,start=1):
        vals=[idx,"",t["task"],t["owner"],t["start"],t["end"],t["status"],"",""]
        for ci,v in enumerate(vals,start=1):
            cell=wss.cell(rr,ci,v); cell.font=Font(size=10); cell.border=border
            if isinstance(v,dt.date): cell.number_format="yyyy-mm-dd"
            if COLS[ci-1][0].endswith("★"): cell.fill=PatternFill("solid",fgColor=REQ_FILL)
        rr+=1
    wss.freeze_panes="A3"

out.save(OUT)
print("저장:",OUT)
print("파트별:",{k:len(v) for k,v in parts.items()},"총",total)
print("주차:",{c:(s.isoformat(),e.isoformat()) for c,(s,e) in weeks.items()})
