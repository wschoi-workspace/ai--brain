"""Onboard Flask Blueprint — /onboard/* endpoints.

페이지:
  GET  /onboard/admin                     admin 대시보드
  GET  /onboard/<token>                   본인 자가입력 폼 (토큰 검증)

API (admin):
  POST /onboard/api/admin/candidate       후보자 등록 + 토큰 URL 메일 발송
  GET  /onboard/api/list                  목록 조회
  GET  /onboard/api/get/<no>              단건 조회 (주민번호 마스킹)
  GET  /onboard/api/preview/<no>          계약서 PDF preview (Chrome headless)
  POST /onboard/api/admin/sign_request/<no>  모두싸인 다중 서명 요청
  POST /onboard/api/admin/setup/<no>      5연쇄 수동 fallback (자동 = webhook 트리거)
  POST /onboard/api/admin/notify_done/<no>  4대보험 회신 처리완료 체크
  POST /onboard/api/admin/cancel/<no>     취소·반려

API (본인):
  POST /onboard/<token>/submit            본인 자가입력 제출

API (공개):
  GET  /onboard/healthz                   진단
"""
from __future__ import annotations
import json
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from flask import (Blueprint, jsonify, render_template, request, send_file,
                    abort, redirect, url_for, session)

from . import __version__, store, token as token_mod, contract_builder
from . import candidate_mail, candidate_form


bp = Blueprint('onboard', __name__, template_folder='../templates')


# ===== 권한 헬퍼 =====

def _load_employees() -> dict:
    path = os.path.join(os.environ.get("DATA_DIR", "/data"), "employees.json")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _get_actor(actor_email: str) -> Optional[dict]:
    if not actor_email:
        return None
    return _load_employees().get(actor_email.lower().strip())


def _is_admin(actor_email: str) -> bool:
    actor = _get_actor(actor_email)
    return bool(actor) and actor.get("role") == "admin"


def _parse_int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _row_to_candidate(row: dict, secure: Optional[dict] = None) -> dict:
    """onboard row(+secure) → contract_builder candidate dict.

    연봉·고정OT시간·기간제여부·연장조건까지 전달해 계약서가 기간제·자동분해를
    그대로 반영하도록 한다.
    """
    secure = secure or {}
    # 생년월일이 row에 없으면 보안 주민번호 앞 6자리에서 유도
    # (구 수동 흐름 row는 일반 컬럼이 비어 있고 secure에만 데이터가 있다)
    birth = row.get("생년월일")
    if not birth:
        ssn = str(secure.get("주민번호") or "").replace("-", "")
        if len(ssn) >= 7 and ssn[:7].isdigit():
            century = "19" if ssn[6] in "1256" else "20"
            birth = f"{century}{ssn[:2]}-{ssn[2:4]}-{ssn[4:6]}"
    return {
        "이름": row.get("이름"), "회사": row.get("회사"),
        "팀": row.get("팀"), "직책": row.get("직책"),
        "이메일": row.get("이메일"), "전화번호": row.get("연락처"),
        "생년월일": birth,
        "주소": secure.get("주소") or row.get("주소"),
        "입사일": row.get("입사예정일"),
        "연봉": row.get("연봉"),
        "고정OT시간": row.get("고정OT시간"),
        "기본급": row.get("기본급"), "식대": row.get("식대"),
        "고정연장": row.get("고정연장수당"),
        "수습기간": row.get("수습기간") or "3개월",
        "기간제여부": row.get("기간제여부", "Y"),
        "연장조건": row.get("연장조건"),
        "근무시간": row.get("근무시간") or "09:00 ~ 18:00",
        "근무지": row.get("근무지"),
        "근무요일": row.get("근무요일"),
        "담당업무": row.get("담당업무"),
        "유급휴일": row.get("유급휴일"),
    }


# ===== 페이지 =====

@bp.route('/admin', methods=['GET'])
def admin_page():
    return render_template('onboard.html', mode='admin', version=__version__)


@bp.route('/admin/v2', methods=['GET'])
def admin_v2_page():
    """v2 통합 대시보드 — KPI + 액션 큐 + 통합 타임라인 + 상세 패널.

    포털 로그인 세션(auth gate가 admin 보장)의 email을 주입해 프론트가
    localStorage 미설정 상태에서도 즉시 데이터를 로드하게 한다."""
    return render_template('onboard_admin_v2.html', version=__version__,
                           session_email=(session.get("email") or ""))


@bp.route('/settings', methods=['GET'])
def settings_page():
    """설정 페이지 — admin 전용. 7 카테고리 (세무사·회사 R/F·대행기관·모두싸인·메일·규칙)."""
    return render_template('settings.html', version=__version__)


# ===== Settings API =====

@bp.route('/api/admin/master-sheet/diagnose', methods=['GET'])
def api_master_sheet_diagnose():
    """인사기록 마스터 시트 매핑 진단. 첫 입사 dogfood 전 호출 권장.

    응답:
      - sheet_total_headers: 시트가 가진 헤더 수
      - mapped_count: append 가능한 매핑 컬럼 수
      - mapped: 매핑된 라벨·컬럼·source 키 리스트
      - unmapped_in_sheet: 시트에는 있으나 매핑 안 되는 라벨 (참고용)
      - missing_in_sheet: 매핑 가능하지만 시트에 없는 라벨 (시트 추가 권고)
      - first_empty_row: 다음 신입사원이 들어갈 행 (5~15 안)
      - empty_row_error: 가득 찼을 때 사유
    """
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    try:
        import master_sync
    except ImportError as e:
        return jsonify({"ok": False, "error": f"master_sync import 실패: {e!r}"}), 500

    result = master_sync.diagnose_master_mapping()
    return jsonify(result)


@bp.route('/api/admin/master-sheet/append-test/<int:no>', methods=['POST'])
def api_master_sheet_append_test(no):
    """admin 수동 trigger — 특정 onboarding No로 마스터 시트 append 1회 강제 실행.

    body: {actor_email}
    dogfood·트러블슈팅 용도. 5연쇄 webhook이 실패했을 때 admin이 수동 재시도.
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "onboarding row not found"}), 404

    secure = store.get_secure_data(no, actor=actor_email) or {}

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    try:
        import master_sync
    except ImportError as e:
        return jsonify({"ok": False, "error": f"master_sync import 실패: {e!r}"}), 500

    result = master_sync.append_new_employee(
        row, secure=secure, actor=actor_email,
        update_if_exists=bool(body.get("update")))
    return jsonify(result)


@bp.route('/api/admin/secure/audit', methods=['GET'])
def api_secure_audit():
    """주민번호·계좌번호 평문 접근 감사 로그 조회. admin 전용.

    Query: ?limit=N (default 100), ?source=onboard|daily-worker, ?actor=email, ?no=int
    """
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    import secure_vault as sv

    try:
        limit = int(request.args.get("limit") or 100)
    except ValueError:
        limit = 100
    source = (request.args.get("source") or "").strip() or None
    actor = (request.args.get("actor") or "").strip() or None
    no_raw = (request.args.get("no") or "").strip()
    try:
        no_filter = int(no_raw) if no_raw else None
    except ValueError:
        no_filter = None

    logs = sv.read_audit_log(limit=limit, source=source, actor=actor, no=no_filter)
    # 본 호출 자체도 감사 기록 (관찰자 효과)
    sv.audit_log(actor=actor_email, action="audit_log_view",
                 source="admin_endpoint",
                 field=f"limit={limit}, source={source}, actor={actor}, no={no_filter}")
    return jsonify({
        "ok": True, "count": len(logs), "logs": logs,
        "secure_enabled": sv.is_enabled(),
    })


@bp.route('/api/admin/secure/migrate', methods=['POST'])
def api_secure_migrate():
    """기존 평문 보안 데이터를 AES 암호화로 일괄 마이그레이션 (1회 호출). admin 전용.

    body: {actor_email}
    onboarding-secure.xlsx 모든 행의 주민번호·계좌번호 → 암호화. 멱등.
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403
    result = store.migrate_secure_to_encrypted(actor=f"migration_by_{actor_email}")
    return jsonify(result)


@bp.route('/api/settings', methods=['GET'])
def api_settings_get():
    """현재 설정 조회 (시크릿은 마스킹). admin 전용."""
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    import settings as cfg

    return jsonify({
        "ok": True,
        "settings": cfg.export_for_admin(),
        "schema": cfg.schema(),
    })


@bp.route('/api/settings', methods=['POST'])
def api_settings_save():
    """설정 갱신. 시크릿 필드는 빈 문자열이면 기존 값 유지."""
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    updates = body.get("settings") or {}
    if not isinstance(updates, dict):
        return jsonify({"error": "settings 필드는 dict"}), 400

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    import settings as cfg
    result = cfg.save(updates, actor_email=actor_email)
    return jsonify(result)


# ===== 통합 진단 (settings 페이지의 '진단' 탭에서 사용) =====

@bp.route('/api/diagnostics', methods=['GET'])
def api_diagnostics():
    """3 모듈 healthz + settings 누락 + 다음 액션 종합 진단."""
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    import settings as cfg

    # ===== 모듈별 healthz 직접 호출 =====
    modules = {}

    # onboard
    try:
        deps = {}
        for mod_name in ("openpyxl", "jinja2", "jwt", "pypdf"):
            try:
                m = __import__(mod_name)
                deps[mod_name] = getattr(m, "__version__", "ok")
            except Exception as e:
                deps[mod_name] = f"ERR: {e!r}"
        templates = [p.name for p in contract_builder.list_templates()]
        modules["onboard"] = {"ok": True, "deps": deps, "templates": templates}
    except Exception as e:
        modules["onboard"] = {"ok": False, "error": repr(e)}

    # esign
    try:
        from esign import client as esign_client
        modules["esign"] = {
            "ok": True, "enabled": esign_client.is_enabled(),
            "mode": "production" if esign_client.is_enabled() else "mock",
        }
    except Exception as e:
        modules["esign"] = {"ok": False, "error": repr(e)}

    # offboard
    try:
        from offboard import __version__ as off_ver
        modules["offboard"] = {"ok": True, "version": off_ver}
    except Exception as e:
        modules["offboard"] = {"ok": False, "error": repr(e)}

    # ===== Settings 7 카테고리별 누락 필드 =====
    schema = cfg.schema()
    settings_status = {}
    critical_missing = []  # Ready to deploy 판정에 영향
    optional_missing = []

    # 카테고리별 critical 필드 정의 (운영 필수)
    CRITICAL = {
        "semusa": ["email"],
        "company_R": ["biznum", "4insure_id", "addr"],
        "company_F": [],  # F사 사용 안 하면 비워도 됨
        "modusign": [],   # ESIGN_ENABLED=false면 Mock으로 운영 가능
        "mail_notify": ["smtp_user", "smtp_pass"],
        "automation_rules": ["onboard_token_secret"],
        "agency": [],     # 양식 제출 시점에 받음 가능
    }

    for cat, info in schema.items():
        filled = 0
        missing = []
        for field in info["fields"]:
            key = f"{cat}.{field['key']}"
            value = cfg.get(key)
            is_secret = key in cfg.SECRET_FIELDS
            is_set = bool(value) and (not is_secret or value not in ("", None))
            if is_set:
                filled += 1
            else:
                missing.append(field["key"])
                full_field = {
                    "cat": cat, "cat_title": info["title"],
                    "key": field["key"], "label": field["label"],
                    "critical": field["key"] in CRITICAL.get(cat, []),
                }
                if field["key"] in CRITICAL.get(cat, []):
                    critical_missing.append(full_field)
                else:
                    optional_missing.append(full_field)
        settings_status[cat] = {
            "title": info["title"],
            "total": len(info["fields"]),
            "filled": filled,
            "missing": missing,
            "complete": len(missing) == 0,
        }

    # ===== employees.json admin 수 =====
    emps = _load_employees()
    admin_count = sum(1 for e in emps.values() if e.get("role") == "admin")

    # ===== Drive·Sheets 연동 진단 =====
    google_env = {
        "GOOGLE_SHEET_ID": bool(os.environ.get("GOOGLE_SHEET_ID")),
        "GOOGLE_CREDENTIALS_JSON": bool(os.environ.get("GOOGLE_CREDENTIALS_JSON")),
        "GOOGLE_DRIVE_ROOT_FOLDER_ID": bool(os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID")),
        "HR_MASTER_SHEET_ID": bool(os.environ.get("HR_MASTER_SHEET_ID")),
    }

    # ===== Ready to deploy 판정 =====
    ready = (
        len(critical_missing) == 0
        and admin_count > 0
        and all(google_env.values())
        and all(m.get("ok") for m in modules.values())
    )

    # ===== 다음 액션 권장 (우선순위 순) =====
    next_actions = []
    if not all(google_env.values()):
        miss = [k for k, v in google_env.items() if not v]
        next_actions.append({
            "priority": 1,
            "title": "Google 인프라 환경변수 설정",
            "detail": f"Fly secrets에 {', '.join(miss)} 등록 필요",
            "where": "fly secrets set ... --app rent-hr-portal",
        })
    if admin_count == 0:
        next_actions.append({
            "priority": 1,
            "title": "관리자 계정 등록",
            "detail": "employees.json에 role=admin 계정 1명 이상 필요",
            "where": "/onboard/settings → 관리자 권한",
        })
    if critical_missing:
        for m in critical_missing[:5]:
            next_actions.append({
                "priority": 2,
                "title": f"[{m['cat_title']}] {m['label']} 입력",
                "detail": f"운영 필수 필드. 미입력 시 해당 흐름 동작 불가",
                "where": f"/onboard/settings → {m['cat_title']} 탭",
            })
    if optional_missing:
        for m in optional_missing[:3]:
            next_actions.append({
                "priority": 3,
                "title": f"[{m['cat_title']}] {m['label']} (선택)",
                "detail": "운영 시 입력 권장",
                "where": f"/onboard/settings → {m['cat_title']} 탭",
            })
    if not ready:
        pass  # 위 액션으로 커버
    else:
        next_actions.append({
            "priority": 4,
            "title": "✓ 배포 준비 완료",
            "detail": "운영 필수 항목 모두 충족. fly deploy 가능",
            "where": "flyctl deploy --app rent-hr-portal",
        })

    return jsonify({
        "ok": True,
        "ready_to_deploy": ready,
        "summary": {
            "critical_missing": len(critical_missing),
            "optional_missing": len(optional_missing),
            "admin_count": admin_count,
            "google_env_ok": all(google_env.values()),
            "modules_ok": all(m.get("ok") for m in modules.values()),
        },
        "modules": modules,
        "settings_status": settings_status,
        "google_env": google_env,
        "next_actions": next_actions,
    })


# ===== 관리자 권한 (employees.json role 조작) =====

@bp.route('/api/settings/admins', methods=['GET'])
def api_settings_admins_list():
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    emps = _load_employees()
    admins = [{"email": e, "name": v.get("name", ""), "company": v.get("company", "")}
              for e, v in emps.items() if v.get("role") == "admin"]
    others = [{"email": e, "name": v.get("name", ""), "company": v.get("company", "")}
              for e, v in emps.items() if v.get("role") != "admin"
              and not v.get("blocked")]
    return jsonify({"ok": True, "admins": admins, "candidates": others})


@bp.route('/api/settings/admins/<string:target_email>', methods=['POST', 'DELETE'])
def api_settings_admins_change(target_email):
    """target_email의 role을 admin으로 부여(POST) 또는 staff로 강등(DELETE)."""
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") if request.method == "POST"
                   else (request.args.get("actor_email") or "")).lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    target = target_email.lower().strip()
    if target == actor_email and request.method == "DELETE":
        return jsonify({"error": "본인 권한은 강등 불가 (다른 admin이 처리)"}), 409

    path = os.path.join(os.environ.get("DATA_DIR", "/data"), "employees.json")
    if not os.path.exists(path):
        return jsonify({"error": "employees.json 없음"}), 404

    with open(path, "r", encoding="utf-8") as f:
        emps = json.load(f)
    if target not in emps:
        return jsonify({"error": f"{target} not in employees.json"}), 404

    new_role = "admin" if request.method == "POST" else "staff"
    emps[target]["role"] = new_role
    with open(path, "w", encoding="utf-8") as f:
        json.dump(emps, f, ensure_ascii=False, indent=2)

    print(f"[settings.audit] {actor_email} changed {target} role → {new_role}", flush=True)
    return jsonify({"ok": True, "target": target, "new_role": new_role})


# ===== Dashboard 통합 API =====

@bp.route('/api/dashboard', methods=['GET'])
def api_dashboard():
    """v2 대시보드용 통합 통계.

    Returns: {
      kpi: {today, urgent, in_progress, paused},
      action_queue: {urgent:[], in_progress:[], active_employees, completed_resignations},
      timeline: [cases with progress %],
      months_checked
    }
    """
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    from datetime import date as date_cls, timedelta
    today = date_cls.today()
    # 전체 월 스캔 — 이번달+전월만 보던 종전 방식은 과거 월 파일에 등록된
    # 재직자 소급 서명(예: 2026-07 윤혜정 케이스)의 진행을 아예 놓쳤다.
    # 월 파일 수는 운영 개월 수만큼만 늘어나므로 전량 읽어도 가볍다.
    months = store.list_onboard_months()
    if not months:
        months = [today.strftime("%Y-%m"),
                  (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")]

    # ===== onboard rows =====
    onboard_rows = []
    for m in months:
        onboard_rows.extend(store.read_onboard_rows(month=m))

    # ===== offboard rows =====
    offboard_rows = []
    try:
        portal_dir = os.path.dirname(os.path.dirname(__file__))
        if portal_dir not in sys.path:
            sys.path.insert(0, portal_dir)
        from offboard import store as off_store
        try:
            off_months = off_store.list_resignation_months()
        except Exception:
            off_months = months
        for m in off_months:
            offboard_rows.extend(off_store.read_resignation_rows(month=m))
    except Exception:
        pass

    # ===== KPI 계산 =====
    def _days_from_today(d_str):
        from datetime import datetime
        if not d_str:
            return None
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
            try:
                d = datetime.strptime(str(d_str).strip(), fmt).date()
                return (d - today).days
            except ValueError:
                continue
        return None

    onb_active = [r for r in onboard_rows
                  if r.get("상태") not in ("반려", "종결", "신고완료")]
    off_active = [r for r in offboard_rows
                  if r.get("상태") not in ("반려", "종결")]

    # 소급 서명 진행 — 재직자 사후 계약(force sign_request)은 라이프사이클
    # 상태(신고완료 등)를 보존하므로 위 활성 필터에 걸리지 않는다.
    # doc_status가 서명 진행 중이면 상태와 무관하게 추적 대상에 포함.
    _active_ids = {id(r) for r in onb_active}
    retro_signing = [r for r in onboard_rows
                     if r.get("doc_status") in ("sign_requested", "sign_partial")
                     and id(r) not in _active_ids]
    for r in retro_signing:
        r["_retro"] = True
    onb_active = onb_active + retro_signing

    # 오늘 액션 = 본인입력대기·계약서검토·서명대기 (입사) + 대기·처리실패 (퇴사)
    # + 소급 서명 진행 (재직자 사후 계약 — 회신 추적 필요)
    today_action_onb = [r for r in onb_active
                        if r.get("상태") in ("본인입력대기", "계약서검토",
                                              "서명대기")
                        or r.get("_retro")]
    today_action_off = [r for r in off_active
                        if r.get("상태") in ("대기", "처리실패", "신고대기")]

    # 긴급 = D-7 임박 또는 무회신 D+5
    urgent = []
    for r in onb_active:
        days = _days_from_today(r.get("입사예정일"))
        if days is not None and 0 <= days <= 7 and r.get("doc_status") == "candidate_registered":
            urgent.append({**r, "_kind": "onb", "_days": days, "_reason": "본인입력 무회신"})
    for r in off_active:
        days = _days_from_today(r.get("확정퇴사일") or r.get("퇴사희망일"))
        if days is not None and -7 <= days <= 0 and r.get("상태") == "대기":
            urgent.append({**r, "_kind": "off", "_days": days, "_reason": "승인 대기"})

    in_progress = onb_active + off_active
    paused = [r for r in onboard_rows + offboard_rows
              if r.get("상태") in ("보류",) or r.get("doc_status") == "paused"]

    # ===== 액션 큐 (긴급 + 진행중) =====
    def _to_card(r, kind):
        if kind == "onb":
            return {
                "id": f"ONB-{r.get('No', 0):04d}",
                "no": r.get("No"),
                "kind": "onb",
                "name": r.get("이름"),
                "company": r.get("회사약자", "R"),
                "date": str(r.get("입사예정일") or ""),
                "owner": r.get("팀") or "—",
                "status": r.get("상태"),
                "doc_status": r.get("doc_status"),
                "days": _days_from_today(r.get("입사예정일")),
                "retro": bool(r.get("_retro")),
            }
        else:
            return {
                "id": f"RES-{r.get('No', 0):04d}",
                "no": r.get("No"),
                "kind": "off",
                "name": r.get("이름"),
                "company": r.get("회사약자", "R"),
                "date": str(r.get("확정퇴사일") or r.get("퇴사희망일") or ""),
                "owner": "최원석",
                "status": r.get("상태"),
                "days": _days_from_today(r.get("확정퇴사일") or r.get("퇴사희망일")),
            }

    urgent_cards = ([_to_card(r, r.get("_kind", "onb")) for r in urgent[:5]])
    inprogress_cards = (
        [_to_card(r, "onb") for r in today_action_onb if r not in urgent[:5]][:5]
        + [_to_card(r, "off") for r in today_action_off if r not in urgent[:5]][:5]
    )

    # ===== 재직자 (employees.json 활동 직원) =====
    active_employees = []
    try:
        emps = _load_employees()
        for email, e in emps.items():
            if not e.get("blocked") and e.get("resignation_status") != "resigned" and not e.get("hidden"):
                active_employees.append({
                    "email": email,
                    "name": e.get("name", ""),
                    "company": e.get("company", ""),
                    "team": e.get("team", ""),
                    "position": e.get("position", ""),
                    "role": e.get("role", ""),
                    "join_date": e.get("join_date", ""),
                })
        # 이름순 정렬 (관리 편의)
        active_employees.sort(key=lambda x: (x.get("company", ""), x.get("name", "")))
    except Exception:
        pass

    # ===== 퇴사 이력 =====
    completed_resignations_rows = [r for r in offboard_rows
                                    if r.get("상태") == "종결"
                                    or r.get("상태") == "신고완료"]
    # 명단 (대시보드 카드용, 최근 퇴사일 순)
    completed_list = []
    for r in completed_resignations_rows:
        completed_list.append({
            "no": r.get("No"),
            "id": f"RES-{int(r.get('No', 0)):04d}",
            "name": r.get("이름", ""),
            "company": r.get("회사", ""),
            "team": r.get("팀", ""),
            "leave_date": str(r.get("확정퇴사일") or r.get("퇴사희망일") or ""),
            "status": r.get("상태", ""),
        })
    completed_list.sort(key=lambda x: x.get("leave_date", ""), reverse=True)

    # ===== 타임라인 (입사·퇴사 활성 케이스 6개) =====
    def _progress_pct(r, kind):
        if kind == "onb":
            stage_map = {
                "candidate_registered": 10,
                "candidate_input": 25,
                "contract_ready": 40,
                "sign_requested": 50,
                "sign_partial": 60,
                "sign_completed": 70,
                "setup_done": 85,
                "insurance_sent": 95,
                "insurance_done": 100,
            }
            return stage_map.get(r.get("doc_status"), 0)
        else:
            stage_map = {
                "대기": 10, "처리중": 30, "처리완료": 50,
                "회수완료": 65, "신고대기": 75,
                "신고완료": 85, "정산등록": 95, "종결": 100,
            }
            return stage_map.get(r.get("상태"), 0)

    timeline = []
    for r in (in_progress[:6]):
        kind = "off" if r in off_active else "onb"
        card = _to_card(r, kind)
        card["progress"] = _progress_pct(r, kind)
        timeline.append(card)

    return jsonify({
        "ok": True,
        "today": today.isoformat(),
        "months_checked": months,
        "kpi": {
            "today": len(today_action_onb) + len(today_action_off),
            "urgent": len(urgent),
            "in_progress": len(in_progress),
            "paused": len(paused),
            "active_employees": len(active_employees),
            "completed_resignations": len(completed_resignations_rows),
        },
        "action_queue": {
            "urgent": urgent_cards,
            "in_progress": inprogress_cards,
            "active_employees_summary": {
                "R": len([e for e in active_employees if "프로젝트렌트" in e.get("company", "")]),
                "F": len([e for e in active_employees if "필라멘트앤코" in e.get("company", "")]),
                "total": len(active_employees),
            },
            "active_employees": active_employees,  # 전체 명단 (대시보드 카드용)
            "completed_resignations_summary": {
                "total": len(completed_resignations_rows),
                "recent_30d": len([r for r in completed_resignations_rows
                                    if _days_from_today(r.get("확정퇴사일") or "")
                                    and -30 <= (_days_from_today(r.get("확정퇴사일") or "") or 0) <= 0]),
            },
            "completed_resignations": completed_list[:20],  # 최근 20건 (이름순)
        },
        "timeline": timeline,
    })


@bp.route('/<string:tok>', methods=['GET'])
def candidate_page(tok):
    """본인 자가입력 폼 — 토큰 검증 후 노출."""
    try:
        payload, row = token_mod.verify_and_lookup(tok)
    except token_mod.TokenError as e:
        # 가벼운 정적 안내 페이지 — onboard.html(React admin 앱)은 후보자 환경에서
        # 마운트되지 않아 빈 화면이 됨. 만료·재사용·교체 시 명확한 안내를 보장.
        return render_template('token_error.html',
                               error=str(e),
                               version=__version__), 400
    # 순수 HTML 폼 (React/Babel 의존 없음 — 모든 기기·브라우저에서 렌더)
    return render_template('self_input.html',
                           token=tok,
                           onboard_no=row.get("No"),
                           candidate={
                               "이름": row.get("이름"),
                               "이메일": row.get("이메일"),
                               "회사": row.get("회사"),
                               "팀": row.get("팀") or "",
                               "직책": row.get("직책") or "",
                               "입사예정일": str(row.get("입사예정일") or ""),
                               "연락처": row.get("연락처") or "",
                               "근무시간": row.get("근무시간") or "",
                               "근무지": row.get("근무지") or "",
                           },
                           version=__version__)


# ===== 급여 써머리 (등록 폼 [2] 단계) =====

@bp.route('/api/admin/salary_preview', methods=['POST'])
def api_salary_preview():
    """연봉 → 기본급/식대/고정연장 분해 + 4대보험·소득세·실수령액·사업주부담.

    body: actor_email, 연봉(원), [고정OT(기본20)], [부양가족수(기본1)]
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    annual = _parse_int(body.get("연봉"), 0)
    if annual <= 0:
        return jsonify({"error": "연봉(원)을 입력하세요"}), 400

    from . import salary_calc
    result = salary_calc.compute(
        annual,
        default_ot=_parse_int(body.get("고정OT"), 20),
        dependents=_parse_int(body.get("부양가족수"), 1),
    )
    return jsonify(result)


# ===== Admin: 후보자 등록 + 토큰 메일 =====

@bp.route('/api/admin/candidate', methods=['POST'])
def api_admin_candidate():
    """admin 후보자 등록 + 토큰 URL 메일 발송.

    Required body:
      actor_email, 이메일, 이름, 회사, 입사예정일
    Optional:
      회사약자, 팀, 직책, 연락처, 수습기간_종료일
      send_mail (bool, default True), dry_run (bool, default False)
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    required = ("이메일", "이름", "회사", "입사예정일")
    missing = [k for k in required if not body.get(k)]
    if missing:
        return jsonify({"error": f"필수 누락: {missing}"}), 400

    company_raw = body.get("회사", "")
    회사약자 = body.get("회사약자") or (
        "R" if ("프로젝트렌트" in company_raw or company_raw.upper() == "R") else "F"
    )

    # 급여 자동 계산 (연봉 입력 시 salary_calc로 분해)
    annual = _parse_int(body.get("연봉"), 0)
    wage_fields = {}
    if annual > 0:
        from . import salary_calc
        sal = salary_calc.compute(
            annual, default_ot=_parse_int(body.get("고정OT"), 20),
            dependents=_parse_int(body.get("부양가족수"), 1))
        wage_fields = {
            "연봉": annual, "고정OT시간": sal["고정OT시간"],
            "기본급": sal["기본급"], "식대": sal["식대"],
            "고정연장수당": sal["고정연장수당"],
        }

    # 계약 형태: 기간제 기본 ON (첫 계약)
    ft_raw = body.get("기간제여부", True)
    is_fixed = ft_raw in (True, "true", "True", 1, "1", "Y", "y")

    row = {
        "이메일": body["이메일"].lower().strip(),
        "이름": body["이름"].strip(),
        "회사": company_raw,
        "회사약자": 회사약자,
        "팀": body.get("팀", ""),
        "직책": body.get("직책", ""),
        "입사예정일": body["입사예정일"],
        "수습기간_종료일": body.get("수습기간_종료일", ""),
        "수습기간": body.get("수습기간") or "3개월",
        "기간제여부": "Y" if is_fixed else "N",
        "연장조건": body.get("연장조건", ""),
        "연락처": body.get("연락처", ""),
        # 근무조건 (매장직/본사직 차이 반영)
        "근무지": body.get("근무지", ""),
        "근무시간": body.get("근무시간", ""),
        "근무요일": body.get("근무요일", ""),
        "담당업무": body.get("담당업무", ""),
        "유급휴일": body.get("유급휴일", ""),
        "doc_status": "candidate_registered",
        "상태": "본인입력대기",
        **wage_fields,
    }
    no = store.append_onboard_row(row)

    # 토큰 발급 + xlsx에 저장
    tok, jti, exp = token_mod.issue_token(no, row["이메일"])
    store.update_onboard_row(no, {"토큰_jti": jti, "토큰_만료": str(exp)})

    token_url = token_mod.build_token_url(tok)

    # 토큰 메일 발송
    send_mail = body.get("send_mail", True)
    dry_run = body.get("dry_run", False)
    mail_result = None
    if send_mail:
        try:
            draft = candidate_mail.send_token_mail(
                to_email=row["이메일"],
                candidate_name=row["이름"],
                company=row["회사"],
                join_date=row["입사예정일"],
                token_url=token_url,
                dry_run=dry_run,
            )
            mail_result = {
                "to": draft.to, "subject": draft.subject,
                "sent": draft.sent, "dry_run": draft.dry_run,
                "error": draft.error,
            }
        except Exception as e:
            mail_result = {"error": repr(e)}

    return jsonify({
        "ok": True, "no": no, "id": f"ONB-{no:04d}",
        "token_url": token_url,
        "expires_in_days": int((exp - int(datetime.now().timestamp())) / 86400),
        "mail": mail_result,
    }), 201


# ===== 본인 자가입력 제출 =====

@bp.route('/<string:tok>/submit', methods=['POST'])
def api_candidate_submit(tok):
    """본인 자가입력 제출 — 토큰 검증 → 저장 → 토큰 무효화 → admin 알림.

    Required body:
      생년월일, 주민번호, 주소, 은행, 계좌번호, 국적, 직종
    Optional:
      부양가족수, 20세이하_부양가족수, 비상연락처, 비상관계
    """
    try:
        payload, row = token_mod.verify_and_lookup(tok)
    except token_mod.TokenError as e:
        return jsonify({"error": str(e)}), 401

    body = request.get_json(force=True, silent=True) or {}
    no = row["No"]

    # 통합 검증 — 주민번호 체크섬·계좌·전화·날짜·생년월일↔주민번호 cross check
    result = candidate_form.validate_self_input(body)
    if not result.ok:
        return jsonify({"error": "검증 실패", "errors": result.errors}), 400

    n = result.normalized
    updates = {
        "생년월일": n.get("생년월일", body.get("생년월일")),
        "주민번호": n.get("주민번호", body.get("주민번호")),  # store가 보안 xlsx로 분리
        "주소": n.get("주소", body.get("주소", "")),
        "개인이메일": n.get("개인이메일", body.get("개인이메일", "")),
        "은행": n.get("은행", body.get("은행", "")),
        "계좌번호": n.get("계좌번호", body.get("계좌번호")),
        "국적": n.get("국적", body.get("국적", "")),
        "직종": n.get("직종", body.get("직종", "")),
        "직급": n.get("직급", body.get("직급", "")),
        "부양가족수": n.get("부양가족수", 0),
        "20세이하_부양가족수": n.get("20세이하_부양가족수", 0),
        "근무시간": body.get("근무시간", "09:00 ~ 18:00"),
        "근무지": n.get("근무지", body.get("근무지", "")),
        "자가입력_완료일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "doc_status": "candidate_input",
        "상태": "계약서검토",
    }
    # [픽스 2026-07-14] 저장 성공을 확인한 뒤에만 토큰 무효화 — 보안시트 저장 실패 시
    #   주민번호 영구 소실 + 토큰 소진(재입력 불가)되던 버그 방지. 실패 시 토큰 유지·재제출 안내.
    saved = store.update_onboard_row(no, updates, month=row.get("_month"))
    if not saved:
        return jsonify({
            "error": "일시적 저장 오류로 제출이 완료되지 않았습니다. 잠시 후 다시 제출해주세요. "
                     "(관리자에게 자동 통보되었습니다)"
        }), 500

    # 토큰 무효화 (1회용)
    token_mod.invalidate_token(no, month=row.get("_month"))

    # 카카오워크 admin 알림 (실패 격리)
    try:
        portal_dir = os.path.dirname(os.path.dirname(__file__))
        if portal_dir not in sys.path:
            sys.path.insert(0, portal_dir)
        import kakaowork_api
        if hasattr(kakaowork_api, 'send_onboard_alert'):
            kakaowork_api.send_onboard_alert("candidate_input", {**row, "No": no})
    except Exception as e:
        print(f"[onboard] 카카오워크 알림 실패: {e}", flush=True)

    return jsonify({"ok": True, "no": no})


# ===== READ APIs =====

@bp.route('/api/list', methods=['GET'])
def api_list():
    """입사 row 목록. admin 전용. 주민번호는 항상 마스킹."""
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    month = request.args.get("month")
    status = request.args.get("status")
    doc_status = request.args.get("doc_status")

    rows = store.read_onboard_rows(month=month, status=status, doc_status=doc_status)
    return jsonify({"rows": rows, "count": len(rows)})


@bp.route('/api/get/<int:no>', methods=['GET'])
def api_get(no):
    """단건 조회. admin 전용. 주민번호는 마스킹.

    Query param include_secure=1 + admin 권한 → 평문 주민번호 동봉 (감사 로그 대상)."""
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404

    result = dict(row)
    if request.args.get("include_secure") == "1":
        secure = store.get_secure_data(no, actor=actor_email)
        if secure:
            result["_secure"] = {
                "주민번호": secure.get("주민번호"),
                "계좌번호": secure.get("계좌번호"),
                "은행": secure.get("은행"),
            }
            print(f"[onboard.audit] {actor_email} accessed secure data of No={no}",
                  flush=True)
    return jsonify(result)


# ===== 계약서 PDF preview =====

@bp.route('/api/preview/<int:no>', methods=['GET'])
def api_preview(no):
    """계약서 PDF preview — admin이 서명 요청 전 검토용.

    Query: doc_type=labor_contract (default) | nda | non_compete | privacy_consent
    """
    actor_email = (request.args.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404

    if row.get("doc_status") not in ("candidate_input", "contract_ready",
                                      "sign_requested", "sign_partial",
                                      "sign_completed"):
        return jsonify({"error": "본인 자가입력 미완료"}), 409

    doc_type = request.args.get("doc_type", "labor_contract")
    template_name = contract_builder.pick_template(doc_type)
    template_path = contract_builder.TEMPLATE_DIR / template_name
    if not template_path.exists():
        return jsonify({"error": f"템플릿 없음: {template_name}"}), 404

    # 보안 데이터 병합 (admin 전용)
    secure = store.get_secure_data(no, actor=actor_email) or {}
    candidate = _row_to_candidate(row, secure)

    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False,
                                          prefix=f"onboard_preview_{no}_") as f:
            pdf_path = Path(f.name)
        contract_builder.build_contract_pdf(template_name, candidate, pdf_path)
    except Exception as e:
        return jsonify({"error": f"PDF 생성 실패: {e!r}"}), 500

    fname = f"{doc_type}_{row.get('회사약자', 'R')}_{row.get('이름', '')}.pdf"
    return send_file(str(pdf_path), mimetype="application/pdf",
                     as_attachment=False, download_name=fname)


# ===== 모두싸인 서명 요청 =====

@bp.route('/api/admin/sign_request/<int:no>', methods=['POST'])
def api_admin_sign_request(no):
    """모두싸인 다중 서명자 (본인 1차 + 대표 2차) 요청.

    body:
      actor_email, doc_type ('labor_contract' default),
      ceo_email (옵션 — 미설정 시 env CEO_EMAIL)
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404
    # force=True: 소급 서명 — 수동(모두싸인) 시절 처리돼 doc_status가 안 넘어간
    # 재직자의 계약서를 사후 발송할 때 admin이 명시적으로 우회 (2026-07-19, 윤혜정 케이스)
    force = bool(body.get("force"))
    if row.get("doc_status") != "candidate_input" and not force:
        return jsonify({"error": f"본인 자가입력 미완료 (현재: {row.get('doc_status')})"}), 409

    # 기본 묶음 = 근로계약서 + 보안서약서(NDA) + 개인정보 동의서 (+ 매장직이면 매장 부속규정)
    is_store = "매장" in ((row.get("근무지") or "") + (row.get("담당업무") or ""))
    default_docs = ["labor_contract", "nda", "privacy_consent"]
    if is_store:
        default_docs.append("store_annex")
    # 팀장급 이상은 경업금지약정서 자동 포함 (2026-07-19 사용자 규칙)
    # 판정: 직책·직급에 리더 키워드 포함 (Lead·팀장·실장·본부장·이사·Head·Director)
    _pos = ((row.get("직책") or "") + " " + (row.get("직급") or "")).lower()
    if any(k in _pos for k in ("lead", "리드", "팀장", "실장", "본부장", "이사",
                               "head", "director")):
        default_docs.append("non_compete")
    doc_types = body.get("doc_types") or default_docs
    if isinstance(doc_types, str):
        doc_types = [doc_types]

    secure = store.get_secure_data(no, actor=actor_email) or {}
    if force and not str(secure.get("주민번호") or "").strip():
        return jsonify({"error": "소급 서명(force) 불가 — 보안 자가입력 데이터(주민번호) 없음"}), 409
    candidate = _row_to_candidate(row, secure)

    # 모두싸인 클라이언트
    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    from esign import client as esign_client, store as esign_store

    # DocuSeal provider면 계약서에 서명 텍스트 태그를 삽입해 PDF 생성
    active_provider = esign_client.get_active_provider()
    candidate["docuseal_tags"] = (active_provider == "docuseal")

    ceo_email = body.get("ceo_email") or os.environ.get(
        "CEO_EMAIL", "ws.choi@project-rent.com")
    ceo_name = os.environ.get(f"CEO_NAME_{row.get('회사약자', 'R')}", "최원석")
    signers = [
        {"name": row.get("이름"), "email": row.get("이메일"),
         "role": "signer", "order": 1, "auth": "email"},
        {"name": ceo_name, "email": ceo_email,
         "role": "signer", "order": 2, "auth": "email"},
    ]

    DOC_LABEL = {"labor_contract": "근로계약서", "nda": "비밀유지서약서(NDA)",
                 "non_compete": "경업금지약정서", "privacy_consent": "개인정보동의서",
                 "store_annex": "매장운영직 부속규정"}
    ID_KEY = {"labor_contract": "근로계약서", "nda": "NDA", "privacy_consent": "개인정보"}

    cli = esign_client.get_client()
    docs = []
    id_updates = {}
    for dt in doc_types:
        template_name = contract_builder.pick_template(dt)
        if not (contract_builder.TEMPLATE_DIR / template_name).exists():
            docs.append({"doc_type": dt, "error": f"템플릿 없음: {template_name}"})
            continue
        try:
            with tempfile.NamedTemporaryFile(
                    suffix=".pdf", delete=False,
                    prefix=f"onboard_sign_{no}_{dt}_") as f:
                pdf_path = Path(f.name)
            contract_builder.build_contract_pdf(template_name, candidate, pdf_path)
        except Exception as e:
            docs.append({"doc_type": dt, "error": f"PDF 생성 실패: {e!r}"})
            continue
        label = DOC_LABEL.get(dt, dt)
        title = f"{label} · {row.get('회사', '')} · {row.get('이름', '')}"
        try:
            resp = cli.create_document(
                file_path=str(pdf_path), signers=signers, title=title,
                metadata={"type": dt, "onboard_no": no, "doc_type": dt},
                expire_days=7,
            )
        except Exception as e:
            docs.append({"doc_type": dt, "error": f"서명 요청 실패: {e!r}"})
            continue
        esign_no = esign_store.append_esign_row({
            "doc_type": dt, "참조_No": no,
            "esign_id": resp.get("id"), "title": title,
            "서명자_1_이름": signers[0]["name"],
            "서명자_1_이메일": signers[0]["email"], "서명자_1_상태": "pending",
            "서명자_2_이름": signers[1]["name"],
            "서명자_2_이메일": signers[1]["email"], "서명자_2_상태": "pending",
            "전체_상태": resp.get("status", "pending"),
        })
        id_updates[f"esign_id_{ID_KEY.get(dt, dt)}"] = resp.get("id")
        docs.append({
            "doc_type": dt, "label": label, "esign_id": resp.get("id"),
            "esign_no": esign_no, "signing_url": resp.get("signing_url"),
        })

    sent = [d for d in docs if d.get("esign_id")]
    if not sent:
        return jsonify({"error": "발송된 문서 없음", "docs": docs}), 500

    # 소급 서명(이미 셋업·신고 완료된 재직자)이면 라이프사이클 상태를
    # 서명대기로 되돌리지 않는다 — doc_status·esign_id 추적만 갱신 (2026-07-19)
    row_updates = {"doc_status": "sign_requested", **id_updates}
    if not (row.get("세무사_발송일시")
            or str(row.get("상태") or "").strip() in ("셋업완료", "세무사발송", "신고완료", "종결")):
        row_updates["상태"] = "서명대기"
    store.update_onboard_row(no, row_updates)

    is_mock = not esign_client.is_enabled()
    out = {
        "ok": True, "no": no, "docs": docs, "sent_count": len(sent),
        "mode": "production" if not is_mock else "mock",
    }
    if is_mock:
        out["mock_hint"] = ("각 doc의 esign_id로 "
                            "/api/esign/mock_complete/<id> POST → webhook 시뮬")
    return jsonify(out)


# ===== 모두싸인 완료본 업로드 → Drive 저장 =====

@bp.route('/api/admin/upload_signed/<int:no>', methods=['POST'])
def api_admin_upload_signed(no):
    """모두싸인에서 받은 서명 완료본 PDF를 입사자 Drive 폴더에 저장 + 링크 기록.

    multipart/form-data:
      actor_email, doc_type (labor_contract | nda, default labor_contract), file
    """
    actor_email = (request.form.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404

    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "file 누락"}), 400
    doc_type = (request.form.get("doc_type") or "labor_contract").strip()

    from . import drive_deploy
    try:
        with tempfile.NamedTemporaryFile(
                suffix=".pdf", delete=False,
                prefix=f"signed_{no}_{doc_type}_") as tmp:
            pdf_path = tmp.name
        f.save(pdf_path)
    except Exception as e:
        return jsonify({"error": f"파일 저장 실패: {e!r}"}), 500

    if doc_type == "nda":
        link = drive_deploy.upload_signed_nda(no, pdf_path, row)
        updates = {"NDA_drive_link": link}
    else:
        link = drive_deploy.upload_signed_contract(no, pdf_path, row)
        updates = {"계약서_drive_link": link, "계약서_방식": "upload"}

    if not link:
        # silent 금지 — Drive 비활성/실패를 명확히 통보
        return jsonify({
            "ok": False, "drive_enabled": False,
            "error": "Drive 업로드 실패 — Drive 연동 비활성 또는 권한/폴더 문제. "
                     "Fly secret(GOOGLE_DRIVE_ROOT_FOLDER_ID·GOOGLE_CREDENTIALS_JSON) 확인 필요",
        }), 502

    store.update_onboard_row(no, updates)
    label = "NDA" if doc_type == "nda" else "근로계약서"
    return jsonify({"ok": True, "no": no, "doc_type": doc_type,
                    "label": label, "drive_link": link})


# ===== 5연쇄 수동 fallback =====

@bp.route('/api/admin/setup/<int:no>', methods=['POST'])
def api_admin_setup(no):
    """5연쇄 셋업 수동 trigger (webhook이 자동 호출하지 못한 경우 fallback).

    Day 12 setup_5chain.py 구현 전까지는 status만 업데이트.
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404

    try:
        from . import setup_5chain
        result = setup_5chain.run(no, row, actor_email=actor_email)
        return jsonify(result)
    except ImportError:
        # Day 12 미구현 — stub
        store.update_onboard_row(no, {
            "doc_status": "setup_done",
            "상태": "셋업완료",
            "처리메모": f"수동 셋업 by {actor_email} ({datetime.now().strftime('%H:%M')})",
        })
        return jsonify({"ok": True, "stub": True,
                        "note": "setup_5chain.py Day 12 구현 예정"})


# ===== 입사 확정 (서명 완료 → 인사카드 + 4대보험 연쇄) =====

@bp.route('/api/admin/confirm/<int:no>', methods=['POST'])
def api_admin_confirm(no):
    """입사 확정 — 모두싸인 수동 발송 케이스의 자동화 진입점.

    서명 완료본을 받아 on_contract_signed 연쇄(인사카드 마스터 append + employees.json
    + Drive + 4대보험 신고서 생성 + 카카오워크)를 admin이 수동 트리거한다.
    선택적으로 4대보험 세무사 메일을 즉시 발송(또는 dry_run 미리보기)한다.

    multipart/form-data (완료본 첨부 시) 또는 application/json:
      actor_email (required)
      send_insurance_now (bool, default False): 4대보험 세무사 메일 즉시 처리
      insurance_dry_run (bool, default True): 즉시 처리 시 미리보기만 (실발송하려면 False)
      세무사 (optional): 세무사 이메일 직접 지정 (settings.semusa.email 미등록 시)
      file (optional): 모두싸인 완료본 PDF (있으면 Drive 저장 + 세무사 메일 첨부)
    """
    is_multipart = bool(request.content_type and 'multipart' in request.content_type)
    if is_multipart:
        actor_email = (request.form.get("actor_email") or "").lower().strip()
        send_now = (request.form.get("send_insurance_now") or "").lower() in ("true", "1", "on", "yes")
        dry_run = (request.form.get("insurance_dry_run") or "true").lower() in ("true", "1", "on", "yes")
        세무사 = (request.form.get("세무사") or "").strip() or None
        f = request.files.get("file")
    else:
        body = request.get_json(force=True, silent=True) or {}
        actor_email = (body.get("actor_email") or "").lower().strip()
        send_now = bool(body.get("send_insurance_now"))
        dry_run = body.get("insurance_dry_run", True) not in (False, "false", "False", 0, "0")
        세무사 = (body.get("세무사") or "").strip() or None
        f = None

    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404

    # 1) 완료본 PDF 임시 저장 (있으면 on_contract_signed가 Drive 업로드 + 세무사 메일 첨부)
    signed_pdf_path = ""
    if f and f.filename:
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False,
                                              prefix=f"confirm_{no}_") as tmp:
                signed_pdf_path = tmp.name
            f.save(signed_pdf_path)
        except Exception as e:
            return jsonify({"error": f"완료본 저장 실패: {e!r}"}), 500

    # 2) on_contract_signed 연쇄: Drive + setup_5chain(인사카드 마스터 등) + 4대보험 신고서
    from . import tasks
    chain = tasks.on_contract_signed(no, signed_pdf_path)

    # 3) 입사 확정 표시
    store.update_onboard_row(no, {
        "입사확정일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "입사확정_by": actor_email,
    })

    # 4) 4대보험 세무사 메일 (옵션: 즉시 발송 또는 dry_run 미리보기)
    insurance_mail = None
    if send_now:
        row2 = store.get_onboard_row(no) or row
        extra = [signed_pdf_path] if signed_pdf_path else []
        insurance_mail = tasks._send_acquisition_mail(
            no, row2, dry_run=dry_run, 세무사=세무사, extra_attachments=extra)
        if insurance_mail.get("sent") and not dry_run:
            store.update_onboard_row(no, {
                "세무사_발송일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "doc_status": "insurance_sent",
                "상태": "신고대기",
            })

    print(f"[onboard.confirm] {actor_email} confirmed No={no} "
          f"send_now={send_now} dry_run={dry_run}", flush=True)
    return jsonify({
        "ok": bool(chain.get("ok")),
        "no": no,
        "chain": chain,
        "insurance_mail": insurance_mail,
    })


# ===== [복구 전용] 깨진 onboard xlsx 정리 =====

@bp.route('/api/admin/_repair_onboard/<month>', methods=['POST'])
def api_admin_repair_onboard(month):
    """헤더 불일치로 데이터 정렬이 어긋난 월별 onboard xlsx를 진단/정리.

    body: actor_email (required), confirm_clear (bool) — true면 헤더(ONBOARD_HEADERS)만
    남기고 데이터 전체 제거. 미지정 시 진단만 반환 (안전장치).
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    import os
    from openpyxl import load_workbook, Workbook
    path = store.get_onboard_xlsx_path(month)
    if not os.path.exists(path):
        return jsonify({"error": f"{month} 파일 없음", "path": path}), 404

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    data_rows = []
    for r in range(2, ws.max_row + 1):
        data_rows.append({h: ws.cell(row=r, column=i + 1).value
                          for i, h in enumerate(headers) if h is not None})
    diag = {
        "month": month,
        "header_count": len(headers),
        "expected_header_count": len(store.ONBOARD_HEADERS),
        "header_match": headers == store.ONBOARD_HEADERS,
        "data_rows": len(data_rows),
        "rows_preview": [{"No": d.get("No"), "이름": d.get("이름"),
                          "doc_status": d.get("doc_status")} for d in data_rows],
    }
    if not body.get("confirm_clear"):
        return jsonify({"diagnose_only": True,
                        "hint": "데이터 제거하려면 confirm_clear=true",
                        **diag})

    # 헤더만 ONBOARD_HEADERS로 새 파일 작성 (깨진 데이터 전체 제거)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "입사"
    ws2.append(store.ONBOARD_HEADERS)
    wb2.save(path)
    print(f"[onboard.repair] {actor_email} cleared {month}: "
          f"{len(data_rows)} rows removed, header normalized", flush=True)
    return jsonify({"ok": True, "cleared_rows": len(data_rows),
                    "before": diag, "new_header_count": len(store.ONBOARD_HEADERS)})


@bp.route('/api/admin/_migrate_headers/<month>', methods=['POST'])
def api_admin_migrate_headers(month):
    """월별 onboard xlsx 헤더를 현행 ONBOARD_HEADERS로 정규화 (데이터 보존).

    각 행을 '컬럼명 → 값' dict로 읽어 ONBOARD_HEADERS 순서로 재작성하므로
    단순 헤더 덮어쓰기와 달리 데이터 정렬이 깨지지 않는다.
    body: actor_email (필수), confirm (bool) — 미지정이면 dry_run(진단만).
    confirm=true면 .bak 백업 후 실제 재작성.
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    import shutil
    from openpyxl import load_workbook, Workbook
    path = store.get_onboard_xlsx_path(month)
    if not os.path.exists(path):
        return jsonify({"error": f"{month} 파일 없음", "path": path}), 404

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    actual_headers = [c.value for c in ws[1]]

    if actual_headers == store.ONBOARD_HEADERS:
        return jsonify({"ok": True, "already_aligned": True,
                        "header_count": len(actual_headers)})

    # 데이터 행을 컬럼명 기준 dict로 (빈 행 제외)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        non_empty = False
        for i, h in enumerate(actual_headers):
            if not h:
                continue
            v = ws.cell(row=r, column=i + 1).value
            row[h] = v
            if v not in (None, ""):
                non_empty = True
        if non_empty:
            rows.append(row)

    old_set = set(h for h in actual_headers if h)
    new_set = set(store.ONBOARD_HEADERS)
    added = sorted(new_set - old_set)
    removed = sorted(old_set - new_set)  # 비어있어야 데이터 손실 0

    if not body.get("confirm"):
        return jsonify({
            "ok": True, "dry_run": True,
            "current_header_count": len(old_set),
            "target_header_count": len(store.ONBOARD_HEADERS),
            "data_rows": len(rows),
            "added_columns": added,
            "removed_columns": removed,
            "preview_names": [r.get("이름") for r in rows],
            "hint": "confirm=true 로 실제 마이그레이션 (.bak 백업 생성)",
        })

    backup = f"{path}.bak"
    shutil.copy2(path, backup)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "입사"
    ws2.append(store.ONBOARD_HEADERS)
    for row in rows:
        ws2.append([row.get(h, "") for h in store.ONBOARD_HEADERS])
    wb2.save(path)

    print(f"[onboard.migrate] {actor_email} migrated {month}: "
          f"{len(rows)} rows, {len(old_set)}->{len(store.ONBOARD_HEADERS)} cols, "
          f"removed={removed}", flush=True)
    return jsonify({
        "ok": True, "migrated": True,
        "backup": os.path.basename(backup),
        "rows_preserved": len(rows),
        "added_columns": added,
        "removed_columns": removed,
        "new_header_count": len(store.ONBOARD_HEADERS),
    })


@bp.route('/api/admin/_recover_submit/<int:no>', methods=['POST'])
def api_admin_recover_submit(no):
    """제출됐으나 헤더 버그로 본 onboard 시트에 미반영된 케이스 복구.

    보안저장소(get_secure_data)의 주민·주소·은행·계좌를 본 시트에 복원하고
    생년월일은 주민번호로 도출, doc_status=candidate_input 전환.
    국적·직종은 본 시트 전용(보안저장소에 없음)이라 body 또는 기본값(한국/사무직).
    body: actor_email(필수), [국적], [직종], [직급]
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404

    secure = store.get_secure_data(no, actor=actor_email)
    if not secure:
        return jsonify({"error": "보안저장소에 데이터 없음 — 복구 불가"}), 404

    jumin = (secure.get("주민번호") or "").strip()
    addr = (secure.get("주소") or "").strip()
    bank = (secure.get("은행") or "").strip()
    acct = (secure.get("계좌번호") or "").strip()

    # 주민번호로 생년월일 도출 (성별코드 1·2=1900s, 3·4=2000s, 5·6=1900s 외국인)
    birth = ""
    digits = jumin.replace("-", "").replace(" ", "")
    if len(digits) >= 7:
        yy, mm, dd, g = digits[0:2], digits[2:4], digits[4:6], digits[6]
        century = "19" if g in ("1", "2", "5", "6") else "20"
        birth = f"{century}{yy}-{mm}-{dd}"

    updates = {
        "생년월일": birth,
        "주민번호_마스킹": store.mask_juminno(jumin),  # 평문은 이미 보안저장소에 있음
        "주소": addr,
        "은행": bank,
        "계좌번호": acct,
        "국적": (body.get("국적") or "").strip() or "한국",
        "직종": (body.get("직종") or "").strip() or "사무직",
        "직급": (body.get("직급") or "").strip() or row.get("직급") or "",
        "자가입력_완료일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " (복구)",
        "doc_status": "candidate_input",
        "상태": "계약서검토",
    }
    ok = store.update_onboard_row(no, updates, month=row.get("_month"))
    if not ok:
        return jsonify({"error": "복원 실패 (update_onboard_row False)"}), 500

    print(f"[onboard.recover] {actor_email} recovered No={no} from secure store", flush=True)
    updated = store.get_onboard_row(no) or {}
    return jsonify({
        "ok": True, "no": no, "recovered": True,
        "doc_status": updated.get("doc_status"),
        "fields": {k: updated.get(k) for k in
                   ("생년월일", "주민번호_마스킹", "주소", "은행", "계좌번호", "국적", "직종", "직급")},
    })


@bp.route('/api/admin/_move_resigned', methods=['POST'])
def api_admin_move_resigned():
    """인사기록 마스터의 직원을 '퇴사자' 탭으로 이동 (인사기록카드에서 제거).

    body: actor_email(필수), names(이름 리스트), confirm(bool).
    confirm 미지정이면 dry_run(대상만 조회). confirm=true면 실제 이동.
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403
    names = body.get("names") or []
    if isinstance(names, str):
        names = [names]
    if not names:
        return jsonify({"error": "names 필요 (이동할 이름 리스트)"}), 400

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    try:
        import master_sync
    except ImportError as e:
        return jsonify({"ok": False, "error": f"master_sync import 실패: {e!r}"}), 500

    result = master_sync.move_to_resigned(
        names, actor=actor_email, confirm=bool(body.get("confirm")))
    return jsonify(result)


@bp.route('/api/admin/_move_resigned_from/<worksheet>', methods=['POST'])
def api_admin_move_resigned_from(worksheet):
    """임의 워크시트(예: 인사평가 이력)의 직원 행을 '퇴사자' 탭으로 이관.

    body: actor_email(필수), names(이름 리스트), confirm(bool), name_col(int, 기본 1).
    confirm 미지정이면 dry_run(워크시트 목록·첫 5행·매칭 행). confirm=true면 실제 이관.
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403
    names = body.get("names") or []
    if isinstance(names, str):
        names = [names]
    if not names:
        return jsonify({"error": "names 필요"}), 400
    try:
        name_col = int(body.get("name_col") or 1)
    except (TypeError, ValueError):
        name_col = 1

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    try:
        import master_sync
    except ImportError as e:
        return jsonify({"ok": False, "error": f"master_sync import 실패: {e!r}"}), 500

    result = master_sync.move_rows_to_resigned(
        worksheet, names, actor=actor_email,
        confirm=bool(body.get("confirm")), name_col=name_col)
    return jsonify(result)


# ===== 4대보험 회신 처리완료 =====

@bp.route('/api/admin/notify_done/<int:no>', methods=['POST'])
def api_admin_notify_done(no):
    """세무사 4대보험 취득 회신 완료 처리 (admin 수동 체크)."""
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    store.update_onboard_row(no, {
        "세무사_회신_체크일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "doc_status": "insurance_done",
        "상태": "신고완료",
    })
    return jsonify({"ok": True})


@bp.route('/api/admin/cancel/<int:no>', methods=['POST'])
def api_admin_cancel(no):
    """후보자 취소·반려."""
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    memo = body.get("memo") or "관리자 취소"
    store.update_onboard_row(no, {
        "doc_status": "reject",
        "상태": "반려",
        "처리메모": memo,
        "토큰_jti": "",  # 토큰 즉시 무효화
    })
    return jsonify({"ok": True})


# ===== Admin 토큰 재발급 =====

@bp.route('/api/admin/reissue_token/<int:no>', methods=['POST'])
def api_admin_reissue(no):
    """토큰 만료·분실 시 admin이 재발급 + 메일 재발송."""
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404
    if row.get("doc_status") not in ("candidate_registered",):
        return jsonify({"error": f"본인 자가입력 이미 완료/진행됨 (현재: {row.get('doc_status')})"}), 409

    tok, jti, exp = token_mod.issue_token(no, row.get("이메일"))
    store.update_onboard_row(no, {"토큰_jti": jti, "토큰_만료": str(exp)})
    token_url = token_mod.build_token_url(tok)

    dry_run = bool(body.get("dry_run", False))
    draft = candidate_mail.send_token_mail(
        to_email=row.get("이메일"), candidate_name=row.get("이름"),
        company=row.get("회사"), join_date=str(row.get("입사예정일") or ""),
        token_url=token_url, dry_run=dry_run,
    )
    return jsonify({
        "ok": True, "no": no, "token_url": token_url,
        "mail": {"sent": draft.sent, "dry_run": draft.dry_run, "error": draft.error},
    })


# ===== Admin 계약 필드 직접 수정 =====

@bp.route('/api/admin/edit/<int:no>', methods=['POST'])
def api_admin_edit(no):
    """admin이 입사 row의 계약 관련 필드를 직접 수정.

    화이트리스트 필드만 허용 — 주민번호·계좌 등 보안 필드는 제외(별도 경로).
    body = {actor_email, updates: {필드: 값, ...}}
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403

    row = store.get_onboard_row(no)
    if not row:
        return jsonify({"error": "not found"}), 404

    ALLOWED = {
        "팀", "직책", "직급", "직종", "회사",
        "근무시간", "근무지", "근무요일", "담당업무", "유급휴일",
        "수습기간", "기간제여부", "연장조건",
        "연락처",  # 개인 연락처(휴대폰) — admin 보정 허용
    }
    raw = body.get("updates") or {}
    if not isinstance(raw, dict):
        return jsonify({"error": "updates 필드는 dict"}), 400
    updates = {k: v for k, v in raw.items() if k in ALLOWED}
    rejected = [k for k in raw if k not in ALLOWED]
    if not updates:
        return jsonify({"error": "허용된 수정 필드 없음",
                        "allowed": sorted(ALLOWED), "rejected": rejected}), 400

    ok = store.update_onboard_row(no, updates, month=row.get("_month"))
    if not ok:
        return jsonify({"error": "업데이트 실패 (행 없음)"}), 500

    print(f"[onboard.audit] {actor_email} edited No={no} "
          f"fields={list(updates)}", flush=True)
    updated = store.get_onboard_row(no)
    return jsonify({
        "ok": True, "no": no,
        "updated_fields": list(updates),
        "rejected_fields": rejected,
        "row": updated,
    })


# ===== 단계계약 (2026-07-19) =====

RECONTRACT_KINDS = {
    "renewal_fixed": "근로계약서(기간제 갱신)",
    "convert_permanent": "근로계약서(무기계약 전환)",
    "salary_change": "임금 변경 합의서",
}

EXTERNAL_KINDS = {
    "outsourcing": "업무위탁계약서(개인외주)",
    "daily_labor": "일용근로계약서",
}


def _load_employees_json() -> dict:
    import json as _json
    path = os.path.join(os.environ.get("DATA_DIR", "/data"), "employees.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}


def _esign_signers_pair(name: str, email: str, company_abbr: str = "R"):
    ceo_email = os.environ.get("CEO_EMAIL", "ws.choi@project-rent.com")
    ceo_name = os.environ.get(f"CEO_NAME_{company_abbr}", "최원석")
    return [
        {"name": name, "email": email, "role": "signer", "order": 1, "auth": "email"},
        {"name": ceo_name, "email": ceo_email, "role": "signer", "order": 2, "auth": "email"},
    ]


@bp.route('/api/admin/recontract', methods=['POST'])
def api_admin_recontract():
    """재직자 재계약 발송 — 기간제 갱신 / 무기 전환 / 임금 변경 합의.

    body: {actor_email, employee_email, kind: renewal_fixed|convert_permanent|salary_change,
           연봉(만원, 필수), 계약시작일(YYYY-MM-DD, 기본 오늘), 고정OT시간(기본 20),
           변경전_기본급/변경전_식대/변경전_고정연장(salary_change 비교표용, 옵션)}
    서명: 직원 1차 → 대표 2차. 완료 webhook은 Drive 저장·알림만 (5연쇄 없음).
    """
    from datetime import datetime as _dt
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403
    kind = (body.get("kind") or "").strip()
    if kind not in RECONTRACT_KINDS:
        return jsonify({"error": f"kind는 {list(RECONTRACT_KINDS)} 중 하나"}), 400
    emp_email = (body.get("employee_email") or "").lower().strip()
    emp = _load_employees_json().get(emp_email)
    if not emp:
        return jsonify({"error": f"재직 직원 아님: {emp_email}"}), 404
    annual_man = int(body.get("연봉") or 0)
    if annual_man <= 0:
        return jsonify({"error": "연봉(만원 단위) 필수"}), 400
    start = (body.get("계약시작일") or _dt.now().strftime("%Y-%m-%d")).strip()

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    from esign import client as esign_client, store as esign_store
    is_docuseal = esign_client.get_active_provider() == "docuseal"

    company = emp.get("company") or "프로젝트렌트"
    abbr = "F" if "필라멘트" in company else "R"
    candidate = {
        "이름": emp.get("name", ""), "회사": company,
        "팀": emp.get("team", ""), "직책": emp.get("position", ""),
        "이메일": emp_email, "생년월일": emp.get("birth_date", ""),
        "주소": emp.get("address", ""), "전화번호": emp.get("phone", ""),
        "입사일": start,  # 계약서상 계약 시작일 = 갱신/시행일
        "연봉": annual_man * 10000,
        "고정OT시간": int(body.get("고정OT시간") or 20),
        "수습기간": "해당 없음",
        "기간제여부": "Y" if kind == "renewal_fixed" else "N",
        "변경전_기본급": body.get("변경전_기본급") or 0,
        "변경전_식대": body.get("변경전_식대") or 0,
        "변경전_고정연장": body.get("변경전_고정연장") or 0,
        "docuseal_tags": is_docuseal,
    }
    doc_type = "salary_change" if kind == "salary_change" else "labor_contract"
    label = RECONTRACT_KINDS[kind]
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False,
                                         prefix=f"recontract_{kind}_") as f:
            pdf_path = Path(f.name)
        contract_builder.build_contract_pdf(
            contract_builder.pick_template(doc_type), candidate, pdf_path)
    except Exception as e:
        return jsonify({"error": f"PDF 생성 실패: {e!r}"}), 500

    signers = _esign_signers_pair(candidate["이름"], emp_email, abbr)
    title = f"{label} · {company} · {candidate['이름']}"
    try:
        resp = esign_client.get_client().create_document(
            file_path=str(pdf_path), signers=signers, title=title,
            metadata={"type": "recontract", "doc_type": doc_type,
                      "kind": kind, "employee_email": emp_email},
            expire_days=7,
        )
    except Exception as e:
        return jsonify({"error": f"서명 요청 실패: {e!r}"}), 500

    esign_no = esign_store.append_esign_row({
        "doc_type": f"recontract_{kind}", "참조_No": 0,
        "esign_id": resp.get("id"), "title": title,
        "서명자_1_이름": signers[0]["name"], "서명자_1_이메일": signers[0]["email"],
        "서명자_1_상태": "pending",
        "서명자_2_이름": signers[1]["name"], "서명자_2_이메일": signers[1]["email"],
        "서명자_2_상태": "pending",
        "전체_상태": resp.get("status", "pending"),
    })
    return jsonify({"ok": True, "kind": kind, "label": label,
                    "esign_id": resp.get("id"), "esign_no": esign_no,
                    "signing_url": resp.get("signing_url"),
                    "mode": "production" if esign_client.is_enabled() else "mock"})


@bp.route('/api/admin/contract_send', methods=['POST'])
def api_admin_contract_send():
    """외부인 단기계약 발송 — 개인외주 업무위탁 / 일용근로.

    body: {actor_email, kind: outsourcing|daily_labor, name, email, phone?,
           회사(기본 프로젝트렌트), work_desc, period_start, period_end,
           fee_total(외주 총액 원) | wage_desc(일용 임금 문구, 예 '일당 120,000원'),
           pay_terms?, work_place?, daily_hours?}
    서명: 상대방 1차 → 대표 2차. 완료 webhook은 Drive 저장·알림만.
    """
    body = request.get_json(force=True, silent=True) or {}
    actor_email = (body.get("actor_email") or "").lower().strip()
    if not _is_admin(actor_email):
        return jsonify({"error": "admin 권한 필요"}), 403
    kind = (body.get("kind") or "").strip()
    if kind not in EXTERNAL_KINDS:
        return jsonify({"error": f"kind는 {list(EXTERNAL_KINDS)} 중 하나"}), 400
    name = (body.get("name") or "").strip()
    email = (body.get("email") or "").lower().strip()
    work_desc = (body.get("work_desc") or "").strip()
    period_start = (body.get("period_start") or "").strip()
    period_end = (body.get("period_end") or "").strip()
    missing = [k for k, v in [("name", name), ("email", email),
                              ("work_desc", work_desc),
                              ("period_start", period_start),
                              ("period_end", period_end)] if not v]
    if missing:
        return jsonify({"error": f"필수 누락: {missing}"}), 400
    if kind == "outsourcing" and not int(body.get("fee_total") or 0):
        return jsonify({"error": "fee_total(위탁 대금, 원) 필수"}), 400
    if kind == "daily_labor" and not (body.get("wage_desc") or "").strip():
        return jsonify({"error": "wage_desc(임금 문구) 필수 — 예: '일당 120,000원'"}), 400

    portal_dir = os.path.dirname(os.path.dirname(__file__))
    if portal_dir not in sys.path:
        sys.path.insert(0, portal_dir)
    from esign import client as esign_client, store as esign_store
    is_docuseal = esign_client.get_active_provider() == "docuseal"

    company = body.get("회사") or "프로젝트렌트"
    abbr = "F" if "필라멘트" in company else "R"
    ctx = contract_builder.build_context(
        {"이름": name, "회사": company, "docuseal_tags": is_docuseal})
    ctx.update({
        "contractor_name": name,
        "contractor_contact": body.get("phone") or email,
        "work_desc": work_desc,
        "period_start": period_start, "period_end": period_end,
        "fee_total": int(body.get("fee_total") or 0),
        "wage_desc": body.get("wage_desc") or "",
        "pay_terms": body.get("pay_terms") or "",
        "work_place_short": body.get("work_place") or "",
        "daily_hours": body.get("daily_hours") or "",
        "break_desc": body.get("break_desc") or "",
    })
    label = EXTERNAL_KINDS[kind]
    try:
        html = contract_builder.get_env().get_template(
            contract_builder.pick_template(kind)).render(**ctx)
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False,
                                         prefix=f"external_{kind}_") as f:
            pdf_path = Path(f.name)
        contract_builder.pdf_renderer.html_string_to_pdf(html, pdf_path)
    except Exception as e:
        return jsonify({"error": f"PDF 생성 실패: {e!r}"}), 500

    signers = _esign_signers_pair(name, email, abbr)
    title = f"{label} · {company} · {name}"
    try:
        resp = esign_client.get_client().create_document(
            file_path=str(pdf_path), signers=signers, title=title,
            metadata={"type": kind, "doc_type": kind, "contractor_name": name},
            expire_days=7,
        )
    except Exception as e:
        return jsonify({"error": f"서명 요청 실패: {e!r}"}), 500

    esign_no = esign_store.append_esign_row({
        "doc_type": kind, "참조_No": 0,
        "esign_id": resp.get("id"), "title": title,
        "서명자_1_이름": name, "서명자_1_이메일": email, "서명자_1_상태": "pending",
        "서명자_2_이름": signers[1]["name"], "서명자_2_이메일": signers[1]["email"],
        "서명자_2_상태": "pending",
        "전체_상태": resp.get("status", "pending"),
    })
    return jsonify({"ok": True, "kind": kind, "label": label,
                    "esign_id": resp.get("id"), "esign_no": esign_no,
                    "signing_url": resp.get("signing_url"),
                    "mode": "production" if esign_client.is_enabled() else "mock"})


# ===== 진단 =====

@bp.route('/healthz', methods=['GET'])
def healthz():
    deps = {}
    for mod_name in ("openpyxl", "jinja2", "jwt", "pypdf"):
        try:
            m = __import__(mod_name)
            deps[mod_name] = getattr(m, "__version__", "ok")
        except Exception as e:
            deps[mod_name] = f"ERR: {e!r}"

    env_required = ("ONBOARD_TOKEN_SECRET", "BASE_URL",
                    "SMTP_USER", "SMTP_PASS", "GOOGLE_SHEET_ID")
    env_status = {k: bool(os.environ.get(k)) for k in env_required}

    templates = []
    try:
        templates = [p.name for p in contract_builder.list_templates()]
    except Exception:
        pass

    return jsonify({
        "ok": True, "module": "onboard", "version": __version__,
        "env": env_status,
        "env_missing": [k for k, v in env_status.items() if not v],
        "deps": deps,
        "templates": templates,
        "endpoints": [
            "GET  /onboard/admin",
            "GET  /onboard/<token>",
            "POST /onboard/<token>/submit",
            "POST /onboard/api/admin/candidate",
            "GET  /onboard/api/list",
            "GET  /onboard/api/get/<no>",
            "GET  /onboard/api/preview/<no>",
            "POST /onboard/api/admin/sign_request/<no>",
            "POST /onboard/api/admin/setup/<no>",
            "POST /onboard/api/admin/notify_done/<no>",
            "POST /onboard/api/admin/cancel/<no>",
            "POST /onboard/api/admin/reissue_token/<no>",
            "GET  /onboard/healthz",
        ],
    })
