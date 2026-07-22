#!/usr/bin/env python3
"""엑셀 템플릿 생성 (정기업무 + POS)"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl 필요. pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
os.makedirs(TEMPLATE_DIR, exist_ok=True)

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def create_regular_tasks_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "정기업무"

    headers = ["업무명", "주기", "날짜(일)", "금액(원)", "비고"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    # 샘플 데이터
    samples = [
        ["임대료 송금 (성수점)", "매월", 25, 3500000, "우리은행 자동이체 확인"],
        ["임대료 송금 (북촌점)", "매월", 25, 4200000, "신한은행 수동 송금"],
        ["부가세 예정신고", "매분기", 25, None, "정세무사 사무실 연락"],
        ["직원 급여 이체", "매월", 10, None, ""],
        ["4대보험료 납부", "매월", 10, None, "자동이체"],
        ["매장 소방점검", "매분기", 15, 150000, "성수점+북촌점 동시"],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # 열 너비
    widths = [28, 10, 10, 15, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    path = os.path.join(TEMPLATE_DIR, "regular-tasks-template.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


def create_pos_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "매장매출"

    headers = ["날짜", "매장명", "매출액", "객수", "객단가", "VOC건수",
               "재고부족품목", "이슈사항", "업무요청내용", "요청긴급도(상/중/하)", "처리상태(대기/진행/완료)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    # 샘플 데이터
    samples = [
        ["2026-04-11", "성수점", 2450000, 82, 29878, 2, "오리지널 커피포 10T", "해외결제 POS 오류 1건", "에어컨 필터 교체 요청", "중", "대기"],
        ["2026-04-11", "북촌점", 1750000, 58, 30172, 0, "", "", "화장실 수전 교체", "하", "완료"],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    widths = [12, 10, 12, 8, 10, 8, 20, 25, 25, 18, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    path = os.path.join(TEMPLATE_DIR, "pos-daily-template.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")


if __name__ == "__main__":
    print("엑셀 템플릿 생성:")
    create_regular_tasks_template()
    create_pos_template()
    print("완료!")
