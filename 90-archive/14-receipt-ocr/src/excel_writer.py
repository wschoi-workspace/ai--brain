#!/usr/bin/env python3
"""
Excel Writer - 영수증 OCR 결과를 .xlsx로 저장
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("date", "날짜", 12),
    ("merchant", "상호", 25),
    ("business_number", "사업자번호", 15),
    ("total_amount", "총금액", 12),
    ("payment_method", "결제수단", 10),
    ("address", "주소", 35),
    ("filename", "원본파일명", 25),
    ("error", "오류", 20),
]


def write_excel(results: list[dict], output_path: Path) -> Path:
    """
    OCR 결과 리스트를 Excel 파일로 저장.

    Args:
        results: process_folder()가 반환한 dict 리스트
        output_path: 저장 경로 (.xlsx)

    Returns:
        저장된 파일 경로
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "영수증_정리"

    # 헤더 스타일
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="E7E6E6")
    header_align = Alignment(horizontal="center", vertical="center")

    # 헤더 행
    for col_idx, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터 행
    for row_idx, result in enumerate(results, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            value = result.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # 금액 컬럼: 천단위 구분자
            if key == "total_amount" and isinstance(value, (int, float)):
                cell.number_format = "#,##0"

            # 오류 행은 붉은색 표시
            if result.get("error"):
                cell.font = Font(color="C00000")

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 저장
    output_path = Path(output_path)
    wb.save(output_path)
    return output_path


def generate_output_filename(input_folder: Path) -> Path:
    """입력 폴더 기준 타임스탬프 파일명 생성."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return input_folder / f"receipts_{timestamp}.xlsx"


# CLI 테스트용
if __name__ == "__main__":
    sample = [
        {
            "date": "2026-04-05", "merchant": "스타벅스 강남점",
            "business_number": "123-45-67890", "total_amount": 15000,
            "payment_method": "카드", "address": "서울 강남구 테헤란로 1",
            "filename": "test.jpg", "error": None,
        },
        {
            "date": None, "merchant": None, "business_number": None,
            "total_amount": None, "payment_method": None, "address": None,
            "filename": "fail.jpg", "error": "API 타임아웃",
        },
    ]
    out = write_excel(sample, Path("./test_output.xlsx"))
    print(f"저장 완료: {out}")
